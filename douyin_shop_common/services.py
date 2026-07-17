"""抖音店铺公共服务层——读取 CSV/xlsx，写入 SQLite，导出 Excel。

所有函数接受 ShopConfig 对象，内含表名前缀，
供香娜露儿和幕莲蔓两个店铺复用相同逻辑。
"""

from __future__ import annotations

import re
import sqlite3
import unicodedata
import zipfile
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from flask import current_app
from werkzeug.datastructures import FileStorage

from .table_schemas import (
    ORDERS_COLUMN_MAPPING,
    ORDERS_COLUMN_TYPES,
    ORDERS_REQUIRED_COLUMNS,
    ORDERS_DEDUP_KEY_COLUMNS,
    ORDERS_TEXT_SOURCE_COLUMNS,
    OVERSEAS_ORDERS_COLUMN_MAPPING,
    OVERSEAS_ORDERS_COLUMN_TYPES,
    OVERSEAS_ORDERS_REQUIRED_COLUMNS,
    OVERSEAS_ORDERS_DEDUP_KEY_COLUMNS,
    OVERSEAS_ORDERS_TEXT_SOURCE_COLUMNS,
    FUND_FLOW_COLUMN_MAPPING,
    FUND_FLOW_COLUMN_TYPES,
    FUND_FLOW_REQUIRED_COLUMNS,
    FUND_FLOW_DEDUP_KEY_COLUMNS,
    FUND_FLOW_TEXT_SOURCE_COLUMNS,
    OVERSEAS_FUND_FLOW_COLUMN_MAPPING,
    OVERSEAS_FUND_FLOW_COLUMN_TYPES,
    OVERSEAS_FUND_FLOW_REQUIRED_COLUMNS,
    OVERSEAS_FUND_FLOW_DEDUP_KEY_COLUMNS,
    OVERSEAS_FUND_FLOW_TEXT_SOURCE_COLUMNS,
    COMMISSION_COLUMN_MAPPING,
    COMMISSION_COLUMN_TYPES,
    COMMISSION_REQUIRED_COLUMNS,
    COMMISSION_DEDUP_KEY_COLUMNS,
    COMMISSION_TEXT_SOURCE_COLUMNS,
    MERCHANT_COLUMN_MAPPING,
    MERCHANT_COLUMN_TYPES,
    MERCHANT_REQUIRED_COLUMNS,
    MERCHANT_DEDUP_KEY_COLUMNS,
    MERCHANT_TEXT_SOURCE_COLUMNS,
    DATA_STATUS_CONFIG,
    EXPORT_TABLE_CONFIG,
)


@dataclass(frozen=True)
class ShopConfig:
    shop_name: str       # 蓝图名，如 'douyin_chantelle'
    display_name: str    # 页面显示名，如 '香娜露儿（抖音）'
    module_key: str      # 权限 key，如 'douyin_shop_chantelle'
    table_prefix: str    # 表名前缀，如 'dy_chantelle'
    fund_flow_format: str = 'standard'
    order_format: str = 'standard'

    @property
    def orders_table(self) -> str:
        return f'{self.table_prefix}_orders'

    @property
    def fund_flow_table(self) -> str:
        return f'{self.table_prefix}_fund_flow'

    @property
    def commission_table(self) -> str:
        return f'{self.table_prefix}_commission'

    @property
    def merchant_table(self) -> str:
        return f'{self.table_prefix}_merchant'

    @property
    def data_status_table(self) -> str:
        return f'{self.table_prefix}_data_status'


# ===================== 工具函数 =====================

def _is_overseas_fund_flow(config: ShopConfig) -> bool:
    return config.fund_flow_format == 'overseas'


def _is_overseas_orders(config: ShopConfig) -> bool:
    return config.order_format == 'overseas'


def _orders_column_mapping(config: ShopConfig) -> dict[str, str]:
    if _is_overseas_orders(config):
        return OVERSEAS_ORDERS_COLUMN_MAPPING
    return ORDERS_COLUMN_MAPPING


def _orders_column_types(config: ShopConfig) -> dict[str, str]:
    if _is_overseas_orders(config):
        return OVERSEAS_ORDERS_COLUMN_TYPES
    return ORDERS_COLUMN_TYPES


def _orders_required_columns(config: ShopConfig) -> list[str]:
    if _is_overseas_orders(config):
        return OVERSEAS_ORDERS_REQUIRED_COLUMNS
    return ORDERS_REQUIRED_COLUMNS


def _orders_dedup_key_columns(config: ShopConfig) -> list[str]:
    if _is_overseas_orders(config):
        return OVERSEAS_ORDERS_DEDUP_KEY_COLUMNS
    return ORDERS_DEDUP_KEY_COLUMNS


def _orders_text_source_columns(config: ShopConfig) -> set[str]:
    if _is_overseas_orders(config):
        return OVERSEAS_ORDERS_TEXT_SOURCE_COLUMNS
    return ORDERS_TEXT_SOURCE_COLUMNS


def _fund_flow_column_mapping(config: ShopConfig) -> dict[str, str]:
    if _is_overseas_fund_flow(config):
        return OVERSEAS_FUND_FLOW_COLUMN_MAPPING
    return FUND_FLOW_COLUMN_MAPPING


def _fund_flow_column_types(config: ShopConfig) -> dict[str, str]:
    if _is_overseas_fund_flow(config):
        return OVERSEAS_FUND_FLOW_COLUMN_TYPES
    return FUND_FLOW_COLUMN_TYPES


def _fund_flow_required_columns(config: ShopConfig) -> list[str]:
    if _is_overseas_fund_flow(config):
        return OVERSEAS_FUND_FLOW_REQUIRED_COLUMNS
    return FUND_FLOW_REQUIRED_COLUMNS


def _fund_flow_dedup_key_columns(config: ShopConfig) -> list[str]:
    if _is_overseas_fund_flow(config):
        return OVERSEAS_FUND_FLOW_DEDUP_KEY_COLUMNS
    return FUND_FLOW_DEDUP_KEY_COLUMNS


def _fund_flow_text_source_columns(config: ShopConfig) -> set[str]:
    if _is_overseas_fund_flow(config):
        return OVERSEAS_FUND_FLOW_TEXT_SOURCE_COLUMNS
    return FUND_FLOW_TEXT_SOURCE_COLUMNS


def get_export_table_config(config: ShopConfig) -> dict[str, dict[str, Any]]:
    export_config = {key: dict(value) for key, value in EXPORT_TABLE_CONFIG.items()}
    export_config['orders'] = dict(export_config['orders'])
    export_config['orders']['column_mapping'] = _orders_column_mapping(config)
    export_config['orders']['column_types'] = _orders_column_types(config)
    export_config['fund_flow'] = dict(export_config['fund_flow'])
    export_config['fund_flow']['column_mapping'] = _fund_flow_column_mapping(config)
    export_config['fund_flow']['column_types'] = _fund_flow_column_types(config)
    return export_config

def _normalize_header(value: Any) -> str:
    return unicodedata.normalize('NFKC', str(value)).strip()


def _clean_text_value(value: Any) -> Any:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return format(value, '.0f')
        return format(value, 'f').rstrip('0').rstrip('.')
    text = str(value).strip()
    if text.startswith("'"):
        text = text[1:].strip()
    return text


def _clean_numeric_value(value: Any) -> Any:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (int, float)):
        return value
    text = re.sub(r'[,¥￥%\s]', '', str(value).strip())
    try:
        return float(text)
    except ValueError:
        return None


def _get_database_path() -> Path:
    db_path = current_app.config.get('DATABASE_PATH')
    if db_path:
        return Path(db_path)
    return Path(current_app.root_path) / 'data' / 'main.db'


def _get_upload_source_filename(file_obj: Any) -> str:
    return str(getattr(file_obj, 'filename', '') or '').strip()


def _read_upload_source_bytes(file_obj: Any) -> bytes:
    path = getattr(file_obj, 'path', None)
    if path:
        return Path(path).read_bytes()
    return file_obj.read()


def _reset_upload_source(file_obj: Any) -> None:
    stream = getattr(file_obj, 'stream', None)
    if stream is not None:
        stream.seek(0)


def _get_excel_display_width(value: Any) -> int:
    if value is None:
        return 0
    text = str(value)
    return sum(2 if ord(ch) > 127 else 1 for ch in text)


def _auto_adjust_excel_columns(worksheet) -> None:
    min_width, max_width = 10, 40
    for column_index, column_cells in enumerate(worksheet.iter_cols(), start=1):
        max_w = max((_get_excel_display_width(c.value) for c in column_cells), default=0)
        adjusted = min(max(max_w + 2, min_width), max_width)
        worksheet.column_dimensions[get_column_letter(column_index)].width = adjusted


# ===================== 建表 & 状态表 =====================

def _build_create_table_sql(table_name: str, column_types: dict[str, str]) -> str:
    col_defs = [f'{col} {typ}' for col, typ in column_types.items()]
    return (
        f'CREATE TABLE IF NOT EXISTS {table_name} (\n'
        f'    id INTEGER PRIMARY KEY AUTOINCREMENT,\n'
        f'    ' + ',\n    '.join(col_defs) + '\n);'
    )


def _get_existing_columns(conn: sqlite3.Connection, table_name: str) -> list[str]:
    rows = conn.execute(f'PRAGMA table_info({table_name})').fetchall()
    return [row[1] for row in rows]


def _sync_table_columns(conn: sqlite3.Connection, table_name: str, column_types: dict[str, str]) -> list[str]:
    existing = set(_get_existing_columns(conn, table_name))
    added: list[str] = []
    for col, typ in column_types.items():
        if col not in existing:
            conn.execute(f'ALTER TABLE {table_name} ADD COLUMN {col} {typ}')
            added.append(col)
    if added:
        conn.commit()
    return added


def _ensure_table(conn: sqlite3.Connection, table_name: str, column_types: dict[str, str]) -> tuple[bool, str]:
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,)
    ).fetchone()
    if row is None:
        conn.execute(_build_create_table_sql(table_name, column_types))
        conn.commit()
        return True, f'已自动创建表：{table_name}'
    added = _sync_table_columns(conn, table_name, column_types)
    if added:
        return False, f'表 {table_name} 已存在，补充字段：{", ".join(added)}'
    return False, f'表 {table_name} 已存在'


def _ensure_data_status_table(conn: sqlite3.Connection, status_table: str) -> None:
    conn.execute(f'''
        CREATE TABLE IF NOT EXISTS {status_table} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            table_key TEXT NOT NULL UNIQUE,
            table_name TEXT NOT NULL,
            record_count INTEGER DEFAULT 0,
            min_date TEXT,
            max_date TEXT,
            last_import_time TEXT
        )
    ''')
    for key, cfg in DATA_STATUS_CONFIG.items():
        conn.execute(
            f'INSERT OR IGNORE INTO {status_table} (table_key, table_name, record_count) VALUES (?,?,0)',
            (key, cfg['table_name']),
        )
    conn.commit()


def ensure_all_tables(config: ShopConfig) -> None:
    db_path = _get_database_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db_path) as conn:
        _ensure_table(conn, config.orders_table, _orders_column_types(config))
        _ensure_table(conn, config.fund_flow_table, _fund_flow_column_types(config))
        _ensure_table(conn, config.commission_table, COMMISSION_COLUMN_TYPES)
        _ensure_table(conn, config.merchant_table, MERCHANT_COLUMN_TYPES)
        _ensure_data_status_table(conn, config.data_status_table)


def _update_data_status(table_key: str, config: ShopConfig) -> str:
    cfg = DATA_STATUS_CONFIG.get(table_key)
    if not cfg:
        return ''
    table_map = {
        'orders': config.orders_table,
        'fund_flow': config.fund_flow_table,
        'commission': config.commission_table,
        'merchant': config.merchant_table,
    }
    source_table = table_map.get(table_key, '')
    if not source_table:
        return ''
    try:
        db_path = _get_database_path()
        with sqlite3.connect(db_path) as conn:
            _ensure_data_status_table(conn, config.data_status_table)
            conn.execute(f'''
                UPDATE {config.data_status_table}
                SET
                    table_name = ?,
                    record_count = (SELECT COUNT(*) FROM {source_table}),
                    min_date = (SELECT MIN({cfg["date_field"]}) FROM {source_table}),
                    max_date = (SELECT MAX({cfg["date_field"]}) FROM {source_table}),
                    last_import_time = datetime('now', 'localtime')
                WHERE table_key = ?
            ''', (cfg['table_name'], table_key))
            conn.commit()
        return ''
    except Exception as exc:
        return f'提示：数据已写入，但刷新数据状态失败，请刷新页面确认。错误：{exc}'


def get_data_status_rows(config: ShopConfig) -> list[dict]:
    """读取数据状态行；首次访问时自动建表并植入 4 条空行（与 wechat_shop 行为一致）。"""
    db_path = _get_database_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        with sqlite3.connect(db_path) as conn:
            conn.row_factory = sqlite3.Row
            _ensure_table(conn, config.orders_table, _orders_column_types(config))
            _ensure_table(conn, config.fund_flow_table, _fund_flow_column_types(config))
            _ensure_table(conn, config.commission_table, COMMISSION_COLUMN_TYPES)
            _ensure_table(conn, config.merchant_table, MERCHANT_COLUMN_TYPES)
            _ensure_data_status_table(conn, config.data_status_table)
            rows = conn.execute(f'''
                SELECT table_key, table_name, record_count, min_date, max_date, last_import_time
                FROM {config.data_status_table}
                ORDER BY CASE table_key
                    WHEN 'orders' THEN 1 WHEN 'fund_flow' THEN 2
                    WHEN 'commission' THEN 3 WHEN 'merchant' THEN 4 ELSE 9 END
            ''').fetchall()
            return [dict(r) for r in rows]
    except Exception:
        return []


# ===================== 防重键辅助 =====================

def _build_dedup_key(row: pd.Series, key_columns: list[str]) -> str | None:
    parts: list[str] = []
    for col in key_columns:
        val = _clean_text_value(row.get(col))
        if val is None or str(val).strip() == '':
            return None
        parts.append(str(val).strip())
    return '||'.join(parts)


def _get_existing_keys(conn: sqlite3.Connection, table_name: str, key_columns: list[str]) -> set[str]:
    cols_sql = ', '.join(key_columns)
    rows = conn.execute(f'SELECT {cols_sql} FROM {table_name}').fetchall()
    keys: set[str] = set()
    for row in rows:
        row_dict = dict(zip(key_columns, row))
        key = _build_dedup_key(pd.Series(row_dict), key_columns)
        if key:
            keys.add(key)
    return keys


def _deduplicate_df(
    df: pd.DataFrame,
    conn: sqlite3.Connection,
    table_name: str,
    key_columns: list[str],
) -> tuple[pd.DataFrame, int, int]:
    if not all(col in df.columns for col in key_columns):
        return df, 0, 0

    work = df.copy()
    work['_dedup_key'] = work.apply(lambda r: _build_dedup_key(r, key_columns), axis=1)
    valid_mask = work['_dedup_key'].notna() & (work['_dedup_key'].astype(str).str.strip() != '')

    batch_valid = work[valid_mask]
    batch_unique = batch_valid.drop_duplicates(subset=['_dedup_key'], keep='first')
    batch_dups = int(len(batch_valid) - len(batch_unique))

    empty_key = work[~valid_mask]
    combined = pd.concat([batch_unique, empty_key], ignore_index=True)

    existing_keys = _get_existing_keys(conn, table_name, key_columns)
    if not existing_keys:
        return combined.drop(columns=['_dedup_key'], errors='ignore'), batch_dups, 0

    valid_after = combined['_dedup_key'].notna() & (combined['_dedup_key'].astype(str).str.strip() != '')
    valid_df = combined[valid_after]
    no_key_df = combined[~valid_after]
    filtered = valid_df[~valid_df['_dedup_key'].isin(existing_keys)]
    db_dups = int(len(valid_df) - len(filtered))
    final = pd.concat([filtered, no_key_df], ignore_index=True)
    return final.drop(columns=['_dedup_key'], errors='ignore'), batch_dups, db_dups


# ===================== DataFrame 准备 =====================

def _prepare_df_for_db(df: pd.DataFrame, column_types: dict[str, str], text_fields: set[str]) -> pd.DataFrame:
    db_cols = list(column_types.keys())
    out = df.copy()
    for col in db_cols:
        if col not in out.columns:
            out[col] = None
    out = out[db_cols]
    for col, typ in column_types.items():
        if col in text_fields:
            out[col] = out[col].apply(_clean_text_value)
        elif typ == 'REAL':
            out[col] = out[col].apply(_clean_numeric_value)
        elif typ == 'INTEGER':
            out[col] = (
                out[col].apply(_clean_numeric_value)
                .apply(lambda x: int(x) if x is not None else None)
            )
        else:
            out[col] = out[col].apply(_clean_text_value)
    return out


# ===================== CSV 读取辅助 =====================

def _build_csv_dtype_mapping(header_row: list[str], text_source_cols: set[str]) -> dict[str, str]:
    normalized_text = {_normalize_header(c) for c in text_source_cols}
    return {
        col: 'string'
        for col in header_row
        if _normalize_header(col) in normalized_text
    }


def _read_orders_csv(file_bytes: bytes) -> pd.DataFrame:
    """读取订单 CSV，第2行起是数据，需要 strip 所有值。"""
    buf = BytesIO(file_bytes)
    # 先读取表头以构建 dtype
    header_df = pd.read_csv(buf, nrows=0, encoding='utf-8-sig')
    buf.seek(0)
    dtype_map = _build_csv_dtype_mapping(header_df.columns.tolist(), ORDERS_TEXT_SOURCE_COLUMNS)
    df = pd.read_csv(buf, dtype=dtype_map if dtype_map else None,
                     encoding='utf-8-sig', low_memory=False)
    df.columns = [_normalize_header(c) for c in df.columns]
    return df


def _read_fund_flow_csv(
    file_bytes: bytes,
    text_source_cols: set[str],
    skip_description_row: bool = True,
) -> pd.DataFrame:
    """读取资金结算 CSV；标准抖音文件跳过第2行字段说明，海外文件不跳过。"""
    buf = BytesIO(file_bytes)
    header_df = pd.read_csv(buf, nrows=0, encoding='utf-8-sig')
    buf.seek(0)
    dtype_map = _build_csv_dtype_mapping(header_df.columns.tolist(), text_source_cols)
    skiprows = [1] if skip_description_row else None
    df = pd.read_csv(buf, skiprows=skiprows, dtype=dtype_map if dtype_map else None,
                     encoding='utf-8-sig', low_memory=False)
    df.columns = [_normalize_header(c) for c in df.columns]
    return df


def _read_commission_xlsx(file_bytes: bytes) -> pd.DataFrame:
    buf = BytesIO(file_bytes)
    header_df = pd.read_excel(buf, nrows=0)
    buf.seek(0)
    dtype_map = _build_csv_dtype_mapping(header_df.columns.tolist(), COMMISSION_TEXT_SOURCE_COLUMNS)
    df = pd.read_excel(buf, dtype=dtype_map if dtype_map else None)
    df.columns = [_normalize_header(c) for c in df.columns]
    return df


def _read_merchant_xlsx(file_bytes: bytes) -> pd.DataFrame:
    buf = BytesIO(file_bytes)
    header_df = pd.read_excel(buf, nrows=0)
    buf.seek(0)
    dtype_map = _build_csv_dtype_mapping(header_df.columns.tolist(), MERCHANT_TEXT_SOURCE_COLUMNS)
    df = pd.read_excel(buf, dtype=dtype_map if dtype_map else None)
    df.columns = [_normalize_header(c) for c in df.columns]
    return df


# ===================== 写库通用函数 =====================

def _write_df_to_table(
    dfs: list[pd.DataFrame],
    config: ShopConfig,
    table_name: str,
    column_types: dict[str, str],
    key_columns: list[str],
    text_fields: set[str],
    table_key: str,
    label: str,
) -> tuple[int, str]:
    if not dfs:
        return 0, f'没有可写入的{label}数据'
    db_path = _get_database_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)

    prepared = [_prepare_df_for_db(df, column_types, text_fields) for df in dfs]
    merged = pd.concat(prepared, ignore_index=True)

    with sqlite3.connect(db_path) as conn:
        _ensure_table(conn, table_name, column_types)
        deduped, batch_dups, db_dups = _deduplicate_df(merged, conn, table_name, key_columns)
        if deduped.empty:
            parts = [f'没有可写入的新{label}数据']
            if batch_dups:
                parts.append(f'本次文件内重复已跳过：{batch_dups} 条')
            if db_dups:
                parts.append(f'数据库已存在已跳过：{db_dups} 条')
            return 0, '；'.join(parts)
        deduped.to_sql(table_name, conn, if_exists='append', index=False)

    status_warn = _update_data_status(table_key, config)
    parts = [f'成功写入 {len(deduped)} 行{label}数据']
    if batch_dups:
        parts.append(f'本次文件内重复已跳过：{batch_dups} 条')
    if db_dups:
        parts.append(f'数据库已存在已跳过：{db_dups} 条')
    if status_warn:
        parts.append(status_warn)
    return int(len(deduped)), '；'.join(parts)


# ===================== 导入入口函数 =====================

def _file_is_csv(filename: str) -> bool:
    return filename.lower().endswith('.csv')


def _file_is_excel(filename: str) -> bool:
    return filename.lower().endswith(('.xlsx', '.xls'))


def _file_is_zip(filename: str) -> bool:
    return filename.lower().endswith('.zip')


def _safe_password_from_filename(filename: str) -> str:
    stem = Path(filename).stem
    if len(stem) < 6:
        raise ValueError(f'订单表文件名不足 6 位，无法提取打开密码：{filename}')
    return stem[-6:]


def _excel_buffer_from_upload(file_bytes: bytes, filename: str) -> BytesIO:
    """读取 Excel；如文件加密，则使用文件名最后 6 位作为密码解密。"""
    raw_buffer = BytesIO(file_bytes)
    try:
        import msoffcrypto
        from msoffcrypto.exceptions import DecryptionError
    except ImportError:
        raw_buffer.seek(0)
        return raw_buffer

    try:
        office_file = msoffcrypto.OfficeFile(raw_buffer)
        if not office_file.is_encrypted():
            raw_buffer.seek(0)
            return raw_buffer

        decrypted = BytesIO()
        office_file.load_key(password=_safe_password_from_filename(filename))
        office_file.decrypt(decrypted)
        decrypted.seek(0)
        return decrypted
    except DecryptionError as exc:
        raise ValueError(f'文件已加密但密码校验失败，请确认文件名最后 6 位是否为打开密码：{filename}') from exc
    except Exception:
        raw_buffer.seek(0)
        return raw_buffer


def _iter_order_upload_sources(file_obj: Any, config: ShopConfig) -> list[tuple[str, bytes]]:
    filename = _get_upload_source_filename(file_obj)
    file_bytes = _read_upload_source_bytes(file_obj)
    if not _file_is_zip(filename):
        return [(filename, file_bytes)]
    if not _is_overseas_orders(config):
        raise ValueError('只有海外旗舰订单表支持上传 ZIP 压缩包')

    sources: list[tuple[str, bytes]] = []
    try:
        with zipfile.ZipFile(BytesIO(file_bytes)) as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                inner_name = info.filename.replace('\\', '/').split('/')[-1]
                if not inner_name or inner_name.startswith('.') or inner_name.startswith('~$'):
                    continue
                if not (_file_is_csv(inner_name) or _file_is_excel(inner_name)):
                    continue
                sources.append((inner_name, zf.read(info)))
    except zipfile.BadZipFile as exc:
        raise ValueError(f'ZIP 文件无法读取：{filename}') from exc

    if not sources:
        raise ValueError(f'ZIP 文件中未找到可导入的订单文件（支持 .csv/.xlsx/.xls）：{filename}')
    return sources


def _build_precheck_failed(title: str, summaries: list, invalid: list, failed: list) -> dict:
    parts = [f'{title}预检未通过，未写入数据库']
    for s in summaries:
        parts += ['', f"文件：{s['filename']}", f"行数：{s['row_count']}",
                  f"列数：{s['column_count']}", f"列名：{'，'.join(s['columns'])}"]
    if failed:
        parts += ['', '预检失败：'] + [f"- {f['filename']}（{f['error']}）" for f in failed]
    if invalid:
        parts += ['', f"无效文件：{'，'.join(invalid)}"]
    return {
        'success': False, 'message': '\n'.join(parts),
        'file_count': len(summaries), 'files': summaries,
        'invalid_files': invalid, 'failed_files': failed,
        'precheck_failed': True, 'written_rows': 0,
    }


def import_orders_files(files: list, config: ShopConfig) -> dict[str, Any]:
    source_mapping = _orders_column_mapping(config)
    column_types = _orders_column_types(config)
    text_source_cols = _orders_text_source_columns(config)
    col_mapping = {_normalize_header(k): v for k, v in source_mapping.items()}
    required = [_normalize_header(c) for c in _orders_required_columns(config)]
    text_fields = {v for k, v in source_mapping.items()
                   if _normalize_header(k) in {_normalize_header(c) for c in text_source_cols}}

    valid, invalid, failed, summaries, prepared = [], [], [], [], []
    base_cols: list[str] | None = None

    for file_obj in files:
        fn = _get_upload_source_filename(file_obj)
        if not fn:
            invalid.append('未命名文件')
            continue
        if not (_file_is_csv(fn) or _file_is_excel(fn) or (_is_overseas_orders(config) and _file_is_zip(fn))):
            invalid.append(fn)
            continue
        valid.append(file_obj)

    if not valid:
        return {'success': False, 'message': '没有有效的文件（支持 .csv / .xlsx）',
                'file_count': 0, 'files': [], 'invalid_files': invalid, 'failed_files': failed}

    for file_obj in valid:
        try:
            for fn, file_bytes in _iter_order_upload_sources(file_obj, config):
                if _file_is_csv(fn):
                    df = _read_orders_csv(file_bytes)
                else:
                    buf = _excel_buffer_from_upload(file_bytes, fn)
                    df = pd.read_excel(buf)
                    df.columns = [_normalize_header(c) for c in df.columns]

                cur_cols = df.columns.tolist()
                missing_req = [c for c in required if c not in cur_cols]
                if missing_req:
                    failed.append({'filename': fn, 'error': f'缺少必需字段：{", ".join(missing_req)}'})
                    continue

                df = df.rename(columns=col_mapping)
                summary = {'filename': fn, 'row_count': len(df), 'column_count': len(df.columns),
                           'columns': df.columns.tolist()}

                if base_cols is None:
                    base_cols = cur_cols
                    summaries.append(summary)
                    prepared.append(df)
                else:
                    if set(cur_cols) != set(base_cols):
                        failed.append({'filename': fn, 'error': '列结构与首个文件不一致'})
                        summaries.append(summary)
                        continue
                    summaries.append(summary)
                    prepared.append(df)
        except Exception as exc:
            failed.append({'filename': _get_upload_source_filename(file_obj), 'error': str(exc)})
        finally:
            _reset_upload_source(file_obj)

    if failed:
        return _build_precheck_failed('订单导入', summaries, invalid, failed)

    if not prepared:
        return {'success': False, 'message': '读取失败，没有可导入的数据',
                'file_count': 0, 'files': summaries, 'invalid_files': invalid, 'failed_files': failed}

    try:
        written, msg = _write_df_to_table(
            prepared, config, config.orders_table, column_types,
            _orders_dedup_key_columns(config), text_fields, 'orders', '订单',
        )
    except Exception as exc:
        return {'success': False, 'message': f'写入订单数据库失败：{exc}',
                'file_count': len(summaries), 'files': summaries,
                'invalid_files': invalid, 'failed_files': failed}

    message = f'成功读取 {len(summaries)} 个订单文件\n{msg}'
    return {'success': True, 'message': message, 'file_count': len(summaries),
            'files': summaries, 'invalid_files': invalid, 'failed_files': failed,
            'written_rows': written}


def import_fund_flow_files(files: list, config: ShopConfig) -> dict[str, Any]:
    source_mapping = _fund_flow_column_mapping(config)
    column_types = _fund_flow_column_types(config)
    text_source_cols = _fund_flow_text_source_columns(config)
    col_mapping = {_normalize_header(k): v for k, v in source_mapping.items()}
    required = [_normalize_header(c) for c in _fund_flow_required_columns(config)]
    text_fields = {v for k, v in source_mapping.items()
                   if _normalize_header(k) in {_normalize_header(c) for c in text_source_cols}}

    valid, invalid, failed, summaries, prepared = [], [], [], [], []

    for file_obj in files:
        fn = _get_upload_source_filename(file_obj)
        if not fn:
            invalid.append('未命名文件')
            continue
        if not (_file_is_csv(fn) or _file_is_excel(fn)):
            invalid.append(fn)
            continue
        valid.append(file_obj)

    if not valid:
        return {'success': False, 'message': '没有有效的文件（支持 .csv）',
                'file_count': 0, 'files': [], 'invalid_files': invalid, 'failed_files': failed}

    for file_obj in valid:
        fn = _get_upload_source_filename(file_obj)
        try:
            file_bytes = _read_upload_source_bytes(file_obj)
            if _file_is_csv(fn):
                df = _read_fund_flow_csv(
                    file_bytes,
                    text_source_cols,
                    skip_description_row=not _is_overseas_fund_flow(config),
                )
            else:
                buf = BytesIO(file_bytes)
                skiprows = [1] if not _is_overseas_fund_flow(config) else None
                df = pd.read_excel(buf, skiprows=skiprows)
                df.columns = [_normalize_header(c) for c in df.columns]

            cur_cols = df.columns.tolist()
            missing_req = [c for c in required if c not in cur_cols]
            if missing_req:
                failed.append({'filename': fn, 'error': f'缺少必需字段：{", ".join(missing_req)}'})
                continue

            df = df.rename(columns=col_mapping)
            summaries.append({'filename': fn, 'row_count': len(df), 'column_count': len(df.columns),
                               'columns': df.columns.tolist()})
            prepared.append(df)
        except Exception as exc:
            failed.append({'filename': fn, 'error': str(exc)})
        finally:
            _reset_upload_source(file_obj)

    if failed:
        return _build_precheck_failed('资金结算导入', summaries, invalid, failed)

    if not prepared:
        return {'success': False, 'message': '读取失败，没有可导入的数据',
                'file_count': 0, 'files': summaries, 'invalid_files': invalid, 'failed_files': failed}

    try:
        written, msg = _write_df_to_table(
            prepared, config, config.fund_flow_table, column_types,
            _fund_flow_dedup_key_columns(config), text_fields, 'fund_flow', '资金结算',
        )
    except Exception as exc:
        return {'success': False, 'message': f'写入资金结算数据库失败：{exc}',
                'file_count': len(summaries), 'files': summaries,
                'invalid_files': invalid, 'failed_files': failed}

    message = f'成功读取 {len(summaries)} 个资金结算文件\n{msg}'
    return {'success': True, 'message': message, 'file_count': len(summaries),
            'files': summaries, 'invalid_files': invalid, 'failed_files': failed,
            'written_rows': written}


def import_commission_files(files: list, config: ShopConfig) -> dict[str, Any]:
    col_mapping = {_normalize_header(k): v for k, v in COMMISSION_COLUMN_MAPPING.items()}
    required = [_normalize_header(c) for c in COMMISSION_REQUIRED_COLUMNS]
    text_fields = {v for k, v in COMMISSION_COLUMN_MAPPING.items()
                   if _normalize_header(k) in {_normalize_header(c) for c in COMMISSION_TEXT_SOURCE_COLUMNS}}

    valid, invalid, failed, summaries, prepared = [], [], [], [], []

    for file_obj in files:
        fn = _get_upload_source_filename(file_obj)
        if not fn:
            invalid.append('未命名文件')
            continue
        if not _file_is_excel(fn):
            invalid.append(fn)
            continue
        valid.append(file_obj)

    if not valid:
        return {'success': False, 'message': '没有有效的 Excel 文件（.xlsx/.xls）',
                'file_count': 0, 'files': [], 'invalid_files': invalid, 'failed_files': failed}

    for file_obj in valid:
        fn = _get_upload_source_filename(file_obj)
        try:
            file_bytes = _read_upload_source_bytes(file_obj)
            df = _read_commission_xlsx(file_bytes)
            cur_cols = df.columns.tolist()
            missing_req = [c for c in required if c not in cur_cols]
            if missing_req:
                failed.append({'filename': fn, 'error': f'缺少必需字段：{", ".join(missing_req)}'})
                continue
            df = df.rename(columns=col_mapping)
            summaries.append({'filename': fn, 'row_count': len(df), 'column_count': len(df.columns),
                               'columns': df.columns.tolist()})
            prepared.append(df)
        except Exception as exc:
            failed.append({'filename': fn, 'error': str(exc)})
        finally:
            _reset_upload_source(file_obj)

    if failed:
        return _build_precheck_failed('佣金明细导入', summaries, invalid, failed)

    if not prepared:
        return {'success': False, 'message': '读取失败，没有可导入的数据',
                'file_count': 0, 'files': summaries, 'invalid_files': invalid, 'failed_files': failed}

    try:
        written, msg = _write_df_to_table(
            prepared, config, config.commission_table, COMMISSION_COLUMN_TYPES,
            COMMISSION_DEDUP_KEY_COLUMNS, text_fields, 'commission', '佣金明细',
        )
    except Exception as exc:
        return {'success': False, 'message': f'写入佣金明细数据库失败：{exc}',
                'file_count': len(summaries), 'files': summaries,
                'invalid_files': invalid, 'failed_files': failed}

    message = f'成功读取 {len(summaries)} 个佣金明细文件\n{msg}'
    return {'success': True, 'message': message, 'file_count': len(summaries),
            'files': summaries, 'invalid_files': invalid, 'failed_files': failed,
            'written_rows': written}


def import_merchant_files(files: list, config: ShopConfig) -> dict[str, Any]:
    col_mapping = {_normalize_header(k): v for k, v in MERCHANT_COLUMN_MAPPING.items()}
    required = [_normalize_header(c) for c in MERCHANT_REQUIRED_COLUMNS]
    text_fields = {v for k, v in MERCHANT_COLUMN_MAPPING.items()
                   if _normalize_header(k) in {_normalize_header(c) for c in MERCHANT_TEXT_SOURCE_COLUMNS}}

    valid, invalid, failed, summaries, prepared = [], [], [], [], []

    for file_obj in files:
        fn = _get_upload_source_filename(file_obj)
        if not fn:
            invalid.append('未命名文件')
            continue
        if not _file_is_excel(fn):
            invalid.append(fn)
            continue
        valid.append(file_obj)

    if not valid:
        return {'success': False, 'message': '没有有效的 Excel 文件（.xlsx/.xls）',
                'file_count': 0, 'files': [], 'invalid_files': invalid, 'failed_files': failed}

    for file_obj in valid:
        fn = _get_upload_source_filename(file_obj)
        try:
            file_bytes = _read_upload_source_bytes(file_obj)
            df = _read_merchant_xlsx(file_bytes)
            if df.empty:
                summaries.append({'filename': fn, 'row_count': 0, 'column_count': len(df.columns),
                                   'columns': df.columns.tolist()})
                continue
            cur_cols = df.columns.tolist()
            missing_req = [c for c in required if c not in cur_cols]
            if missing_req:
                failed.append({'filename': fn, 'error': f'缺少必需字段：{", ".join(missing_req)}'})
                continue
            df = df.rename(columns=col_mapping)
            summaries.append({'filename': fn, 'row_count': len(df), 'column_count': len(df.columns),
                               'columns': df.columns.tolist()})
            prepared.append(df)
        except Exception as exc:
            failed.append({'filename': fn, 'error': str(exc)})
        finally:
            _reset_upload_source(file_obj)

    if failed:
        return _build_precheck_failed('招商明细导入', summaries, invalid, failed)

    if not prepared:
        msg = f'成功读取 {len(summaries)} 个文件，本月无招商数据，不写入数据库' if summaries else '读取失败，没有可导入的数据'
        return {'success': True if summaries else False, 'message': msg,
                'file_count': len(summaries), 'files': summaries,
                'invalid_files': invalid, 'failed_files': failed, 'written_rows': 0}

    try:
        written, msg = _write_df_to_table(
            prepared, config, config.merchant_table, MERCHANT_COLUMN_TYPES,
            MERCHANT_DEDUP_KEY_COLUMNS, text_fields, 'merchant', '招商明细',
        )
    except Exception as exc:
        return {'success': False, 'message': f'写入招商明细数据库失败：{exc}',
                'file_count': len(summaries), 'files': summaries,
                'invalid_files': invalid, 'failed_files': failed}

    message = f'成功读取 {len(summaries)} 个招商明细文件\n{msg}'
    return {'success': True, 'message': message, 'file_count': len(summaries),
            'files': summaries, 'invalid_files': invalid, 'failed_files': failed,
            'written_rows': written}


# ===================== 导出 =====================

def _normalize_export_dt(value: str | None) -> str | None:
    if not value:
        return None
    text = str(value).strip().replace('T', ' ').replace('/', '-')
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
        try:
            from datetime import datetime
            dt = datetime.strptime(text, fmt)
            if fmt == '%Y-%m-%d':
                return dt.strftime('%Y-%m-%d 00:00:00')
            if fmt == '%Y-%m-%d %H:%M':
                return dt.strftime('%Y-%m-%d %H:%M:00')
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except ValueError:
            continue
    return text


def _is_datetime_column(col: str) -> bool:
    c = col.lower()
    return c.endswith('_at') or c.endswith('_time') or c in ('settlement_time',)


def _is_numeric_column(col: str, column_types: dict[str, str]) -> bool:
    return str(column_types.get(col, '')).upper() in ('REAL', 'INTEGER', 'NUMERIC', 'FLOAT')


def _build_filter_sql(
    filter_conditions: list[dict],
    allowed_fields: set[str],
    column_types: dict[str, str],
) -> tuple[list[str], list[Any]]:
    from datetime import datetime
    parts: list[str] = []
    params: list[Any] = []

    for cond in filter_conditions:
        if not isinstance(cond, dict):
            continue
        field = str(cond.get('field') or '').strip()
        op = str(cond.get('operator') or '').strip().lower()
        logic = 'OR' if str(cond.get('logic') or '').strip().lower() == 'or' else 'AND'
        raw_val = cond.get('value')
        val_text = '' if raw_val is None else str(raw_val).strip()

        if not field or field not in allowed_fields:
            continue
        if op not in {'eq', 'ne', 'contains', 'not_contains', 'gt', 'gte', 'lt', 'lte',
                      'is_empty', 'is_not_empty'}:
            continue

        is_dt = _is_datetime_column(field)
        is_num = _is_numeric_column(field, column_types)
        dt_expr = f"REPLACE(CAST({field} AS TEXT), '/', '-')"
        clause = ''
        p: list[Any] = []

        def _dt(v: str, boundary: str) -> str:
            v = v.replace('/', '-')
            for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
                try:
                    dt2 = datetime.strptime(v, fmt)
                    if fmt == '%Y-%m-%d':
                        return dt2.strftime('%Y-%m-%d 23:59:59' if boundary == 'end' else '%Y-%m-%d 00:00:00')
                    return dt2.strftime('%Y-%m-%d %H:%M:00' if fmt == '%Y-%m-%d %H:%M' else '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    pass
            return v

        if op == 'eq' and val_text:
            if is_dt:
                clause = f'({dt_expr} >= ? AND {dt_expr} <= ?)'; p = [_dt(val_text, 'start'), _dt(val_text, 'end')]
            elif is_num:
                nv = _clean_numeric_value(val_text)
                clause = f'CAST({field} AS REAL) = ?'; p = [nv if nv is not None else val_text]
            else:
                clause = f'CAST({field} AS TEXT) = ?'; p = [val_text]
        elif op == 'ne' and val_text:
            if is_dt:
                clause = f'({dt_expr} < ? OR {dt_expr} > ?)'; p = [_dt(val_text, 'start'), _dt(val_text, 'end')]
            else:
                clause = f'CAST({field} AS TEXT) <> ?'; p = [val_text]
        elif op == 'contains' and val_text:
            clause = f'CAST({field} AS TEXT) LIKE ?'; p = [f'%{val_text}%']
        elif op == 'not_contains' and val_text:
            clause = f'CAST({field} AS TEXT) NOT LIKE ?'; p = [f'%{val_text}%']
        elif op == 'gt' and val_text:
            if is_dt:
                clause = f'{dt_expr} > ?'; p = [_dt(val_text, 'end')]
            elif is_num:
                nv = _clean_numeric_value(val_text)
                clause = f'CAST({field} AS REAL) > ?'; p = [nv if nv is not None else val_text]
            else:
                clause = f'CAST({field} AS TEXT) > ?'; p = [val_text]
        elif op == 'gte' and val_text:
            if is_dt:
                clause = f'{dt_expr} >= ?'; p = [_dt(val_text, 'start')]
            elif is_num:
                nv = _clean_numeric_value(val_text)
                clause = f'CAST({field} AS REAL) >= ?'; p = [nv if nv is not None else val_text]
            else:
                clause = f'CAST({field} AS TEXT) >= ?'; p = [val_text]
        elif op == 'lt' and val_text:
            if is_dt:
                clause = f'{dt_expr} < ?'; p = [_dt(val_text, 'start')]
            elif is_num:
                nv = _clean_numeric_value(val_text)
                clause = f'CAST({field} AS REAL) < ?'; p = [nv if nv is not None else val_text]
            else:
                clause = f'CAST({field} AS TEXT) < ?'; p = [val_text]
        elif op == 'lte' and val_text:
            if is_dt:
                clause = f'{dt_expr} <= ?'; p = [_dt(val_text, 'end')]
            elif is_num:
                nv = _clean_numeric_value(val_text)
                clause = f'CAST({field} AS REAL) <= ?'; p = [nv if nv is not None else val_text]
            else:
                clause = f'CAST({field} AS TEXT) <= ?'; p = [val_text]
        elif op == 'is_empty':
            clause = f"({field} IS NULL OR TRIM(CAST({field} AS TEXT)) = '')"
        elif op == 'is_not_empty':
            clause = f"({field} IS NOT NULL AND TRIM(CAST({field} AS TEXT)) <> '')"

        if not clause:
            continue
        parts.append(clause if not parts else f'{logic} {clause}')
        params.extend(p)

    return parts, params


def export_data_to_excel(
    table_key: str,
    start_time: str | None,
    end_time: str | None,
    selected_fields: list[str],
    filter_conditions: list[dict] | None,
    config: ShopConfig,
) -> tuple[BytesIO, str]:
    cfg = get_export_table_config(config).get(table_key)
    if not cfg:
        raise ValueError('不支持的导出表类型')
    if not selected_fields:
        raise ValueError('请至少选择一个导出字段')

    allowed = set(cfg['column_types'].keys())
    bad = [f for f in selected_fields if f not in allowed]
    if bad:
        raise ValueError(f"无效导出字段：{', '.join(bad)}")

    start_text = _normalize_export_dt(start_time)
    end_text = _normalize_export_dt(end_time)
    if start_text and end_text and start_text > end_text:
        raise ValueError('开始时间不能大于结束时间')

    table_name = {
        'orders': config.orders_table,
        'fund_flow': config.fund_flow_table,
        'commission': config.commission_table,
        'merchant': config.merchant_table,
    }[table_key]

    db_path = _get_database_path()
    fields_sql = ', '.join(selected_fields)
    dt_expr = f"REPLACE(CAST({cfg['date_field']} AS TEXT), '/', '-')"
    where: list[str] = []
    params: list[Any] = []
    if start_text:
        where.append(f'{dt_expr} >= ?'); params.append(start_text)
    if end_text:
        where.append(f'{dt_expr} <= ?'); params.append(end_text)

    if filter_conditions:
        fparts, fparams = _build_filter_sql(filter_conditions, allowed, cfg['column_types'])
        if fparts:
            where.append('(' + ' '.join(fparts) + ')')
            params.extend(fparams)

    sql = f"SELECT {fields_sql} FROM {table_name}"
    if where:
        sql += ' WHERE ' + ' AND '.join(where)
    sql += f" ORDER BY {cfg['date_field']} ASC, id ASC"

    with sqlite3.connect(db_path) as conn:
        exists = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,)
        ).fetchone()
        if not exists:
            raise ValueError(f'数据表不存在：{table_name}')
        df = pd.read_sql_query(sql, conn, params=params)

    # 中文列名映射
    rev_map = {v: k for k, v in cfg['column_mapping'].items()}
    df.columns = [rev_map.get(c, c) for c in df.columns]

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='导出结果')
        _auto_adjust_excel_columns(writer.sheets['导出结果'])
    output.seek(0)

    safe = re.sub(r'[\\/:*?"<>|\s]+', '_', f'{cfg["label"]}')
    s = re.sub(r'[\\/:*?"<>|\s]+', '_', start_text or '全部时间')
    e = re.sub(r'[\\/:*?"<>|\s]+', '_', end_text or '全部时间')
    return output, f'{safe}_{s}_到_{e}.xlsx'

# ===================== 佣金导出 =====================

DOUYIN_COMMISSION_AMOUNT_COLUMNS = {
    '佣金金额', '应开金额', '达人佣金', '招商服务费', '佣金(元)', '招商服务费(元)',
    '实际佣金支出', '实际服务费收入',
}

DETAIL_EXPORT_MODE_LABELS = {
    'all': '全部',
    'exempt': '豁免',
    'non_exempt': '不豁免',
}


def _table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,),
    ).fetchone()
    return row is not None


def _get_douyin_alias_nicknames(conn: sqlite3.Connection, keyword: str | None) -> list[str]:
    text = str(keyword or '').strip()
    if not text:
        return []
    if not _table_exists(conn, 'customer') or not _table_exists(conn, 'customer_alias'):
        return []

    rows = conn.execute(
        """
        SELECT DISTINCT TRIM(c.short_name) AS nickname
        FROM customer_alias ca
        JOIN customer c ON c.id = ca.customer_id
        WHERE TRIM(ca.alias) = ?
          AND TRIM(COALESCE(c.short_name, '')) <> ''
        ORDER BY nickname
        """,
        (text,),
    ).fetchall()
    return [str(row['nickname']).strip() for row in rows if str(row['nickname']).strip()]


def _normalize_order_no(val: Any) -> str:
    return str(val or '').strip().lstrip("'")


def _normalize_summary_id(value: Any) -> str:
    return str(value or '').strip().lstrip("'")


def _join_summary_ids(values: Any) -> str:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        text = _normalize_summary_id(value)
        if text and text not in seen:
            seen.add(text)
            result.append(text)
    return ';'.join(sorted(result))


def _normalize_commission_date(value: str | None, boundary: str) -> str:
    text = str(value or '').strip().replace('/', '-')
    if not text:
        raise ValueError('请选择佣金导出的开始日期和结束日期')
    try:
        dt = datetime.strptime(text[:10], '%Y-%m-%d')
    except ValueError as exc:
        raise ValueError('佣金导出日期格式不正确，请使用 YYYY-MM-DD') from exc
    return dt.strftime('%Y-%m-%d')


def _normalize_commission_date_range(start_date: str | None, end_date: str | None) -> tuple[str, str]:
    start_text = _normalize_commission_date(start_date, 'start')
    end_text = _normalize_commission_date(end_date, 'end')
    if start_text > end_text:
        raise ValueError('佣金导出开始日期不能晚于结束日期')
    return start_text, end_text


def _month_values_between(start_text: str, end_text: str) -> list[tuple[int, int]]:
    start_dt = datetime.strptime(start_text[:10], '%Y-%m-%d')
    end_dt = datetime.strptime(end_text[:10], '%Y-%m-%d')
    values: list[tuple[int, int]] = []
    year, month = start_dt.year, start_dt.month
    while (year, month) <= (end_dt.year, end_dt.month):
        values.append((year, month))
        if month == 12:
            year += 1
            month = 1
        else:
            month += 1
    return values


def _build_commission_month_text(start_text: str, end_text: str, short_year: bool = False) -> str:
    months = _month_values_between(start_text, end_text)
    if not months:
        return ''
    has_multiple_years = len({year for year, _month in months}) > 1
    previous_year: int | None = None
    parts: list[str] = []
    for year, month in months:
        year_text = str(year)[-2:] if short_year else str(year)
        if has_multiple_years or previous_year != year:
            parts.append(f'{year_text}年{month}月')
        else:
            parts.append(f'{month}月')
        previous_year = year
    return ' '.join(parts)


def _safe_download_part(value: Any, fallback: str = '未命名') -> str:
    text = str(value or '').strip()
    text = re.sub(r'[\\/:*?"<>|\r\n\t]+', '_', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text or fallback


def _build_commission_zip_name(prefix: str, start_text: str, end_text: str) -> str:
    month_text = _safe_download_part(_build_commission_month_text(start_text, end_text))
    return f'{prefix}_{month_text}.zip'


def _write_dataframe_excel(
    sheets: list[tuple[str, pd.DataFrame]],
    amount_columns: set[str] | None = None,
    text_columns: set[str] | None = None,
) -> BytesIO:
    amount_columns = amount_columns or set()
    text_columns = text_columns or set()
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in sheets:
            safe_sheet = re.sub(r'[\[\]:*?/\\]', '_', str(sheet_name))[:31] or 'Sheet1'
            df.to_excel(writer, index=False, sheet_name=safe_sheet)
            worksheet = writer.sheets[safe_sheet]
            for column_index, column_name in enumerate(df.columns, start=1):
                column_letter = get_column_letter(column_index)
                if column_name in amount_columns:
                    for cell in worksheet[column_letter][1:]:
                        cell.number_format = '0.00'
                if column_name in text_columns:
                    for cell in worksheet[column_letter]:
                        cell.number_format = '@'
            _auto_adjust_excel_columns(worksheet)
    output.seek(0)
    return output


def _commission_date_expr(alias: str = '') -> str:
    prefix = f'{alias}.' if alias else ''
    return f"SUBSTR(REPLACE(CAST({prefix}settlement_time AS TEXT), '/', '-'), 1, 10)"


def _build_name_filter_sql(
    field_sql: str,
    keyword: str,
    alias_nicknames: list[str],
    params: list[Any],
) -> str:
    filters = [f'{field_sql} LIKE ?']
    params.append(f'%{keyword}%')
    if alias_nicknames:
        placeholders = ', '.join('?' for _ in alias_nicknames)
        filters.append(f'{field_sql} IN ({placeholders})')
        params.extend(alias_nicknames)
    return '(' + ' OR '.join(filters) + ')'


def _ensure_douyin_commission_tables(conn: sqlite3.Connection, config: ShopConfig) -> None:
    if not _table_exists(conn, config.fund_flow_table):
        raise ValueError(f'数据表不存在：{config.fund_flow_table}，请先导入资金结算表')


def _overseas_creator_lookup_from_sql(config: ShopConfig) -> str:
    return f"""
        {config.fund_flow_table} f
        LEFT JOIN (
            SELECT
                sub_order_no,
                MAX(NULLIF(TRIM(influencer_nickname), '')) AS order_influencer_name,
                MAX(NULLIF(TRIM(influencer_id), '')) AS order_influencer_id
            FROM {config.orders_table}
            WHERE COALESCE(sub_order_no, '') <> ''
            GROUP BY sub_order_no
        ) os ON f.sub_order_no = os.sub_order_no
        LEFT JOIN (
            SELECT
                main_order_no,
                MAX(NULLIF(TRIM(influencer_nickname), '')) AS order_influencer_name,
                MAX(NULLIF(TRIM(influencer_id), '')) AS order_influencer_id
            FROM {config.orders_table}
            WHERE COALESCE(main_order_no, '') <> ''
            GROUP BY main_order_no
        ) om ON f.order_no = om.main_order_no
    """


def _query_creator_summary(
    conn: sqlite3.Connection,
    start_text: str,
    end_text: str,
    keyword: str,
    alias_nicknames: list[str],
    config: ShopConfig,
) -> pd.DataFrame:
    params: list[Any] = [start_text, end_text]
    table_sql = config.fund_flow_table
    name_expr = 'influencer_name'
    id_expr = 'influencer_id'
    date_expr = _commission_date_expr()
    commission_expr = 'influencer_commission'
    if _is_overseas_fund_flow(config) and _table_exists(conn, config.orders_table):
        table_sql = _overseas_creator_lookup_from_sql(config)
        name_expr = "COALESCE(NULLIF(TRIM(f.influencer_name), ''), os.order_influencer_name, om.order_influencer_name)"
        id_expr = "COALESCE(NULLIF(TRIM(f.influencer_id), ''), os.order_influencer_id, om.order_influencer_id)"
        date_expr = _commission_date_expr('f')
        commission_expr = 'f.influencer_commission'
    where = [
        f'{date_expr} >= ?',
        f'{date_expr} <= ?',
        f'ABS(CAST({commission_expr} AS REAL)) > 0.0001',
        f"{name_expr} IS NOT NULL",
        f"TRIM({name_expr}) <> ''",
        f"TRIM({name_expr}) <> '-'",
    ]
    if keyword:
        where.append(_build_name_filter_sql(name_expr, keyword, alias_nicknames, params))
    sql = f"""
        SELECT
            TRIM({name_expr}) AS name,
            {id_expr} AS influencer_id,
            SUM(CAST({commission_expr} AS REAL)) AS net_amount
        FROM {table_sql}
        WHERE {' AND '.join(where)}
        GROUP BY TRIM({name_expr}), {id_expr}
        ORDER BY SUM(CAST({commission_expr} AS REAL)) ASC
    """
    df = pd.read_sql_query(sql, conn, params=params)
    if df.empty:
        return pd.DataFrame(columns=['达人名称', '达人ID', '佣金金额', 'net_amount'])
    grouped = (
        df.groupby('name', dropna=False)
        .agg(
            达人ID=('influencer_id', _join_summary_ids),
            net_amount=('net_amount', 'sum'),
        )
        .reset_index()
    )
    grouped['net_amount'] = pd.to_numeric(grouped['net_amount'], errors='coerce').fillna(0)
    # 取反：资金流水里支出为负，取反后正常支出显示为正数；
    # 若某达人当期退款超过费用（净值为正），取反后显示为负数——表示净收入/无需开票。
    # 这保证：sum(佣金金额) = abs(达人佣金列总和)，满足核对等式。
    grouped['佣金金额'] = -grouped['net_amount']
    grouped = grouped.sort_values('net_amount', ascending=True)
    return grouped.rename(columns={'name': '达人名称'})[['达人名称', '达人ID', '佣金金额', 'net_amount']]


def _load_merchant_order_map(conn: sqlite3.Connection, config: ShopConfig) -> dict[str, dict[str, str]]:
    if not _table_exists(conn, config.merchant_table):
        return {}
    rows = conn.execute(
        f"""
        SELECT order_id, issuing_institution, group_campaign_id
        FROM {config.merchant_table}
        WHERE order_id IS NOT NULL
          AND TRIM(CAST(order_id AS TEXT)) <> ''
        """
    ).fetchall()
    result: dict[str, dict[str, str]] = {}
    for row in rows:
        order_id = _normalize_order_no(row['order_id'])
        leader_name = str(row['issuing_institution'] or '').strip()
        leader_id = _normalize_summary_id(row['group_campaign_id'])
        if order_id and leader_name and leader_name != '-' and order_id not in result:
            result[order_id] = {'name': leader_name, 'id': leader_id}
    return result


def _name_matches_keyword(name: str, keyword: str, alias_nicknames: list[str]) -> bool:
    if not keyword:
        return True
    return keyword in name or name in alias_nicknames


def _query_leader_fee_rows(
    conn: sqlite3.Connection,
    start_text: str,
    end_text: str,
    config: ShopConfig,
    selected_columns: list[str] | None = None,
) -> pd.DataFrame:
    cols = selected_columns or ['settlement_time', 'order_no', 'merchant_recruitment_fee']
    columns_sql = ', '.join(cols)
    sql = f"""
        SELECT {columns_sql}
        FROM {config.fund_flow_table}
        WHERE {_commission_date_expr()} >= ?
          AND {_commission_date_expr()} <= ?
          AND ABS(CAST(merchant_recruitment_fee AS REAL)) > 0.0001
        ORDER BY settlement_time ASC, order_no ASC
    """
    return pd.read_sql_query(sql, conn, params=[start_text, end_text])


def _query_leader_summary(
    conn: sqlite3.Connection,
    start_text: str,
    end_text: str,
    keyword: str,
    alias_nicknames: list[str],
    config: ShopConfig,
) -> pd.DataFrame:
    merchant_map = _load_merchant_order_map(conn, config)
    if not merchant_map:
        return pd.DataFrame(columns=['团长名称', '团长ID', '佣金金额', 'net_amount', 'matched_rows'])

    rows_df = _query_leader_fee_rows(conn, start_text, end_text, config)
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows_df.itertuples(index=False):
        leader_info = merchant_map.get(_normalize_order_no(row.order_no)) or {}
        leader_name = leader_info.get('name', '')
        if not leader_name or not _name_matches_keyword(leader_name, keyword, alias_nicknames):
            continue
        bucket = grouped.setdefault(
            leader_name,
            {'团长名称': leader_name, '_ids': set(), 'net_amount': 0.0, 'matched_rows': 0},
        )
        leader_id = leader_info.get('id', '')
        if leader_id:
            bucket['_ids'].add(leader_id)
        bucket['net_amount'] += float(row.merchant_recruitment_fee or 0)
        bucket['matched_rows'] += 1

    if not grouped:
        return pd.DataFrame(columns=['团长名称', '团长ID', '佣金金额', 'net_amount', 'matched_rows'])
    for bucket in grouped.values():
        bucket['团长ID'] = ';'.join(sorted(bucket.pop('_ids')))
    df = pd.DataFrame(grouped.values())
    # 取反：同达人逻辑。sum(佣金金额) = abs(招商服务费列已匹配部分总和)，差额 = 未匹配行的绝对值。
    df['佣金金额'] = -pd.to_numeric(df['net_amount'], errors='coerce').fillna(0)
    return df.sort_values('佣金金额', ascending=False)[['团长名称', '团长ID', '佣金金额', 'net_amount', 'matched_rows']]


def _query_unmatched_leader_rows(
    conn: sqlite3.Connection,
    start_text: str,
    end_text: str,
    config: ShopConfig,
) -> pd.DataFrame:
    rows_df = _query_leader_fee_rows(conn, start_text, end_text, config)
    merchant_map = _load_merchant_order_map(conn, config)
    if not merchant_map:
        unmatched_df = rows_df.copy()
    else:
        mask = ~rows_df['order_no'].map(lambda value: _normalize_order_no(value) in merchant_map)
        unmatched_df = rows_df[mask].copy()
    if unmatched_df.empty:
        return pd.DataFrame(columns=['结算时间', '订单号', '招商服务费'])
    unmatched_df = unmatched_df.rename(
        columns={
            'settlement_time': '结算时间',
            'order_no': '订单号',
            'merchant_recruitment_fee': '招商服务费',
        }
    )
    return unmatched_df[['结算时间', '订单号', '招商服务费']]


def _build_invoice_import_df(creator_df: pd.DataFrame, leader_df: pd.DataFrame) -> pd.DataFrame:
    # 佣金金额 = -net_amount，全部放入发票导入，包括当期净值为正（退款超费用）者。
    # 是否开票由业务人员在发票模块里判断，代码不自行过滤。
    rows: list[dict[str, Any]] = []
    for _idx, row in creator_df.iterrows():
        rows.append({'达人/客户': row['达人名称'], '应开金额': float(row['佣金金额'] or 0)})
    for _idx, row in leader_df.iterrows():
        rows.append({'达人/客户': row['团长名称'], '应开金额': float(row['佣金金额'] or 0)})
    df = pd.DataFrame(rows, columns=['达人/客户', '应开金额'])
    if not df.empty:
        df = df.sort_values('应开金额', ascending=False)
    return df


def _fund_flow_chinese_columns(df: pd.DataFrame, config: ShopConfig) -> pd.DataFrame:
    reverse_map = {v: k for k, v in _fund_flow_column_mapping(config).items()}
    return df.rename(columns={col: reverse_map.get(col, col) for col in df.columns})


def _commission_chinese_columns(df: pd.DataFrame) -> pd.DataFrame:
    reverse_map = {v: k for k, v in COMMISSION_COLUMN_MAPPING.items()}
    return df.rename(columns={col: reverse_map.get(col, col) for col in df.columns})


def _merchant_chinese_columns(df: pd.DataFrame) -> pd.DataFrame:
    reverse_map = {v: k for k, v in MERCHANT_COLUMN_MAPPING.items()}
    return df.rename(columns={col: reverse_map.get(col, col) for col in df.columns})


def _normalize_detail_export_mode(value: str | None) -> str:
    mode = str(value or 'all').strip() or 'all'
    if mode not in DETAIL_EXPORT_MODE_LABELS:
        raise ValueError('明细表导出口径不正确，请选择全部、豁免或不豁免')
    return mode


def _detail_export_mode_label(mode: str) -> str:
    return DETAIL_EXPORT_MODE_LABELS.get(mode, DETAIL_EXPORT_MODE_LABELS['all'])


def _detail_export_brand_key(config: ShopConfig) -> str | None:
    if config.table_prefix == 'dy_chantelle':
        return 'chantelle'
    if config.table_prefix == 'dy_mulianman':
        return 'mulianman'
    return None


def _ensure_douyin_detail_commission_tables(conn: sqlite3.Connection, config: ShopConfig) -> None:
    if not _table_exists(conn, config.commission_table):
        raise ValueError(f'数据表不存在：{config.commission_table}，请先导入佣金订单明细')


def _load_exemption_ranges(conn: sqlite3.Connection, config: ShopConfig) -> dict[str, list[tuple[str, str]]]:
    brand = _detail_export_brand_key(config)
    if not brand or not _table_exists(conn, 'creator_exemptions'):
        return {}
    rows = conn.execute(
        """
        SELECT creator_uid, start_date, end_date
        FROM creator_exemptions
        WHERE brand = ?
          AND TRIM(COALESCE(creator_uid, '')) <> ''
          AND TRIM(COALESCE(start_date, '')) <> ''
          AND TRIM(COALESCE(end_date, '')) <> ''
        """,
        (brand,),
    ).fetchall()
    result: dict[str, list[tuple[str, str]]] = {}
    for row in rows:
        uid = _normalize_summary_id(row['creator_uid'])
        start_date = str(row['start_date'] or '').strip()[:10].replace('/', '-')
        end_date = str(row['end_date'] or '').strip()[:10].replace('/', '-')
        if uid and start_date and end_date:
            result.setdefault(uid, []).append((start_date, end_date))
    return result


def _is_exempted_by_ranges(identity: Any, date_value: Any, ranges: dict[str, list[tuple[str, str]]]) -> bool:
    identity_text = _normalize_summary_id(identity)
    if not identity_text:
        return False
    date_text = str(date_value or '').strip()[:10].replace('/', '-')
    if not date_text:
        return False
    return any(start_date <= date_text <= end_date for start_date, end_date in ranges.get(identity_text, []))


def _apply_detail_exemption_filter(
    df: pd.DataFrame,
    identity_col: str,
    date_col: str,
    ranges: dict[str, list[tuple[str, str]]],
    mode: str,
) -> pd.DataFrame:
    if df.empty:
        out = df.copy()
        out['exemption_status'] = []
        return out
    out = df.copy()
    exempt_mask = out.apply(
        lambda row: _is_exempted_by_ranges(row.get(identity_col), row.get(date_col), ranges),
        axis=1,
    )
    out['exemption_status'] = exempt_mask.map(lambda value: '豁免' if value else '不豁免')
    if mode == 'exempt':
        out = out[exempt_mask].copy()
    elif mode == 'non_exempt':
        out = out[~exempt_mask].copy()
    return out


def _apply_detail_keyword_filter(
    df: pd.DataFrame,
    name_col: str,
    identity_col: str,
    keyword: str,
    alias_nicknames: list[str],
) -> pd.DataFrame:
    if df.empty or not keyword:
        return df
    names = df[name_col].fillna('').astype(str)
    identities = df[identity_col].fillna('').astype(str)
    alias_set = set(alias_nicknames)
    mask = (
        names.str.contains(keyword, regex=False)
        | identities.str.contains(keyword, regex=False)
        | names.isin(alias_set)
    )
    return df[mask].copy()


def _order_creator_lookup_join_sql(config: ShopConfig) -> str:
    return f"""
        LEFT JOIN (
            SELECT
                sub_order_no,
                product_id,
                MAX(NULLIF(TRIM(influencer_nickname), '')) AS order_influencer_name,
                MAX(NULLIF(TRIM(influencer_id), '')) AS order_influencer_id
            FROM {config.orders_table}
            WHERE COALESCE(sub_order_no, '') <> ''
            GROUP BY sub_order_no, product_id
        ) os ON c.order_id = os.sub_order_no AND c.product_id = os.product_id
        LEFT JOIN (
            SELECT
                main_order_no,
                product_id,
                MAX(NULLIF(TRIM(influencer_nickname), '')) AS order_influencer_name,
                MAX(NULLIF(TRIM(influencer_id), '')) AS order_influencer_id
            FROM {config.orders_table}
            WHERE COALESCE(main_order_no, '') <> ''
            GROUP BY main_order_no, product_id
        ) om ON c.order_id = om.main_order_no AND c.product_id = om.product_id
    """


def _query_detail_source_creator_rows(
    conn: sqlite3.Connection,
    start_text: str,
    end_text: str,
    keyword: str,
    alias_nicknames: list[str],
    config: ShopConfig,
) -> pd.DataFrame:
    columns_sql = ', '.join(f'c.{col}' for col in COMMISSION_COLUMN_TYPES.keys())
    if _table_exists(conn, config.orders_table):
        join_sql = _order_creator_lookup_join_sql(config)
        name_expr = "COALESCE(os.order_influencer_name, om.order_influencer_name, NULLIF(TRIM(c.author_account), ''))"
        id_expr = "COALESCE(os.order_influencer_id, om.order_influencer_id)"
    else:
        join_sql = ''
        name_expr = "NULLIF(TRIM(c.author_account), '')"
        id_expr = "NULL"
    sql = f"""
        SELECT
            {columns_sql},
            {name_expr} AS detail_creator_name,
            {id_expr} AS detail_creator_id
        FROM {config.commission_table} c
        {join_sql}
        WHERE {_commission_date_expr('c')} >= ?
          AND {_commission_date_expr('c')} <= ?
          AND ABS(CAST(c.actual_commission AS REAL)) > 0.0001
        ORDER BY detail_creator_name ASC, c.settlement_time ASC, c.order_id ASC
    """
    df = pd.read_sql_query(sql, conn, params=[start_text, end_text])
    if df.empty:
        return df
    df['detail_creator_name'] = df['detail_creator_name'].fillna('').astype(str).str.strip()
    df['detail_creator_id'] = df['detail_creator_id'].fillna('').astype(str).str.strip()
    return _apply_detail_keyword_filter(df, 'detail_creator_name', 'detail_creator_id', keyword, alias_nicknames)


def _query_detail_source_leader_rows(
    conn: sqlite3.Connection,
    start_text: str,
    end_text: str,
    keyword: str,
    alias_nicknames: list[str],
    config: ShopConfig,
) -> pd.DataFrame:
    if not _table_exists(conn, config.merchant_table):
        return pd.DataFrame(columns=list(MERCHANT_COLUMN_TYPES.keys()) + ['detail_leader_name', 'detail_leader_id'])
    columns_sql = ', '.join(f'm.{col}' for col in MERCHANT_COLUMN_TYPES.keys())
    sql = f"""
        SELECT
            {columns_sql},
            NULLIF(TRIM(m.issuing_institution), '') AS detail_leader_name,
            NULLIF(TRIM(m.group_campaign_id), '') AS detail_leader_id
        FROM {config.merchant_table} m
        WHERE {_commission_date_expr('m')} >= ?
          AND {_commission_date_expr('m')} <= ?
          AND ABS(CAST(m.actual_service_income AS REAL)) > 0.0001
        ORDER BY detail_leader_name ASC, m.settlement_time ASC, m.order_id ASC
    """
    df = pd.read_sql_query(sql, conn, params=[start_text, end_text])
    if df.empty:
        return df
    df['detail_leader_name'] = df['detail_leader_name'].fillna('').astype(str).str.strip()
    df['detail_leader_id'] = df['detail_leader_id'].fillna('').astype(str).str.strip()
    return _apply_detail_keyword_filter(df, 'detail_leader_name', 'detail_leader_id', keyword, alias_nicknames)


def _build_detail_creator_summary(rows_df: pd.DataFrame) -> pd.DataFrame:
    if rows_df.empty:
        return pd.DataFrame(columns=['达人名称', '达人ID', '佣金金额'])
    work = rows_df[rows_df['detail_creator_name'].fillna('').astype(str).str.strip() != ''].copy()
    if work.empty:
        return pd.DataFrame(columns=['达人名称', '达人ID', '佣金金额'])
    work['actual_commission'] = pd.to_numeric(work['actual_commission'], errors='coerce').fillna(0)
    grouped = (
        work.groupby('detail_creator_name', dropna=False)
        .agg(
            达人ID=('detail_creator_id', _join_summary_ids),
            佣金金额=('actual_commission', 'sum'),
        )
        .reset_index()
        .rename(columns={'detail_creator_name': '达人名称'})
    )
    return grouped.sort_values('佣金金额', ascending=False)[['达人名称', '达人ID', '佣金金额']]


def _build_detail_leader_summary(rows_df: pd.DataFrame) -> pd.DataFrame:
    if rows_df.empty:
        return pd.DataFrame(columns=['团长名称', '团长ID', '佣金金额'])
    work = rows_df[rows_df['detail_leader_name'].fillna('').astype(str).str.strip() != ''].copy()
    if work.empty:
        return pd.DataFrame(columns=['团长名称', '团长ID', '佣金金额'])
    work['actual_service_income'] = pd.to_numeric(work['actual_service_income'], errors='coerce').fillna(0)
    grouped = (
        work.groupby('detail_leader_name', dropna=False)
        .agg(
            团长ID=('detail_leader_id', _join_summary_ids),
            佣金金额=('actual_service_income', 'sum'),
        )
        .reset_index()
        .rename(columns={'detail_leader_name': '团长名称'})
    )
    return grouped.sort_values('佣金金额', ascending=False)[['团长名称', '团长ID', '佣金金额']]


def _build_detail_invoice_import_df(creator_df: pd.DataFrame, leader_df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for _idx, row in creator_df.iterrows():
        rows.append({'达人/客户': row['达人名称'], '应开金额': float(row['佣金金额'] or 0)})
    for _idx, row in leader_df.iterrows():
        rows.append({'达人/客户': row['团长名称'], '应开金额': float(row['佣金金额'] or 0)})
    df = pd.DataFrame(rows, columns=['达人/客户', '应开金额'])
    if not df.empty:
        df = df.sort_values('应开金额', ascending=False)
    return df


def _build_unmatched_detail_creator_rows(rows_df: pd.DataFrame) -> pd.DataFrame:
    if rows_df.empty:
        return pd.DataFrame(columns=['结算时间', '订单id', '商品id', '达人名称', '实际佣金支出', '原因'])
    mask = (
        rows_df['detail_creator_id'].fillna('').astype(str).str.strip() == ''
    ) | (
        rows_df['detail_creator_name'].fillna('').astype(str).str.strip() == ''
    )
    work = rows_df[mask].copy()
    if work.empty:
        return pd.DataFrame(columns=['结算时间', '订单id', '商品id', '达人名称', '实际佣金支出', '原因'])
    work['原因'] = work.apply(
        lambda row: '缺少达人ID' if str(row.get('detail_creator_id') or '').strip() == '' else '缺少达人名称',
        axis=1,
    )
    work = work.rename(
        columns={
            'settlement_time': '结算时间',
            'order_id': '订单id',
            'product_id': '商品id',
            'detail_creator_name': '达人名称',
            'actual_commission': '实际佣金支出',
        }
    )
    return work[['结算时间', '订单id', '商品id', '达人名称', '实际佣金支出', '原因']]


def _build_unmatched_detail_leader_rows(rows_df: pd.DataFrame) -> pd.DataFrame:
    if rows_df.empty:
        return pd.DataFrame(columns=['结算时间', '订单id', '商品id', '团长名称', '团长ID', '实际服务费收入', '原因'])
    mask = (
        rows_df['detail_leader_id'].fillna('').astype(str).str.strip() == ''
    ) | (
        rows_df['detail_leader_name'].fillna('').astype(str).str.strip() == ''
    )
    work = rows_df[mask].copy()
    if work.empty:
        return pd.DataFrame(columns=['结算时间', '订单id', '商品id', '团长名称', '团长ID', '实际服务费收入', '原因'])
    work['原因'] = work.apply(
        lambda row: '缺少团长ID' if str(row.get('detail_leader_id') or '').strip() == '' else '缺少团长名称',
        axis=1,
    )
    work = work.rename(
        columns={
            'settlement_time': '结算时间',
            'order_id': '订单id',
            'product_id': '商品id',
            'detail_leader_name': '团长名称',
            'detail_leader_id': '团长ID',
            'actual_service_income': '实际服务费收入',
        }
    )
    return work[['结算时间', '订单id', '商品id', '团长名称', '团长ID', '实际服务费收入', '原因']]


def _detail_creator_export_df(group_df: pd.DataFrame) -> pd.DataFrame:
    detail_df = group_df.copy().rename(
        columns={
            'detail_creator_name': '达人名称',
            'detail_creator_id': '达人ID',
            'exemption_status': '豁免状态',
        }
    )
    detail_df = _commission_chinese_columns(detail_df)
    front_cols = ['达人名称', '达人ID', '豁免状态']
    remaining = [col for col in detail_df.columns if col not in front_cols]
    return detail_df[front_cols + remaining]


def _detail_leader_export_df(group_df: pd.DataFrame) -> pd.DataFrame:
    detail_df = group_df.copy().rename(
        columns={
            'detail_leader_name': '团长名称',
            'detail_leader_id': '团长ID',
            'exemption_status': '豁免状态',
        }
    )
    detail_df = _merchant_chinese_columns(detail_df)
    front_cols = ['团长名称', '团长ID', '豁免状态']
    remaining = [col for col in detail_df.columns if col not in front_cols]
    return detail_df[front_cols + remaining]


def _load_detail_source_export_data(
    start_text: str,
    end_text: str,
    nickname_query: str | None,
    exemption_mode: str,
    config: ShopConfig,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    keyword = str(nickname_query or '').strip()
    db_path = _get_database_path()
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_douyin_detail_commission_tables(conn, config)
        alias_nicknames = _get_douyin_alias_nicknames(conn, keyword)
        exemption_ranges = _load_exemption_ranges(conn, config)
        creator_rows = _query_detail_source_creator_rows(conn, start_text, end_text, keyword, alias_nicknames, config)
        leader_rows = _query_detail_source_leader_rows(conn, start_text, end_text, keyword, alias_nicknames, config)

    creator_rows = _apply_detail_exemption_filter(
        creator_rows, 'detail_creator_id', 'settlement_time', exemption_ranges, exemption_mode
    )
    leader_rows = _apply_detail_exemption_filter(
        leader_rows, 'detail_leader_id', 'settlement_time', exemption_ranges, exemption_mode
    )
    creator_df = _build_detail_creator_summary(creator_rows)
    leader_df = _build_detail_leader_summary(leader_rows)
    unmatched_creator_df = _build_unmatched_detail_creator_rows(creator_rows)
    unmatched_leader_df = _build_unmatched_detail_leader_rows(leader_rows)
    return creator_rows, leader_rows, creator_df, leader_df, unmatched_creator_df, unmatched_leader_df


def _query_creator_detail_rows(
    conn: sqlite3.Connection,
    start_text: str,
    end_text: str,
    keyword: str,
    alias_nicknames: list[str],
    config: ShopConfig,
) -> pd.DataFrame:
    column_types = _fund_flow_column_types(config)
    columns_sql = ', '.join(column_types.keys())
    params: list[Any] = [start_text, end_text]
    table_sql = config.fund_flow_table
    name_expr = 'influencer_name'
    id_expr = 'influencer_id'
    date_expr = _commission_date_expr()
    commission_expr = 'influencer_commission'
    if _is_overseas_fund_flow(config) and _table_exists(conn, config.orders_table):
        table_sql = _overseas_creator_lookup_from_sql(config)
        name_expr = "COALESCE(NULLIF(TRIM(f.influencer_name), ''), os.order_influencer_name, om.order_influencer_name)"
        id_expr = "COALESCE(NULLIF(TRIM(f.influencer_id), ''), os.order_influencer_id, om.order_influencer_id)"
        date_expr = _commission_date_expr('f')
        commission_expr = 'f.influencer_commission'
        select_parts: list[str] = []
        for col in column_types.keys():
            if col == 'influencer_name':
                select_parts.append(f'{name_expr} AS influencer_name')
            elif col == 'influencer_id':
                select_parts.append(f'{id_expr} AS influencer_id')
            else:
                select_parts.append(f'f.{col}')
        columns_sql = ', '.join(select_parts)
    where = [
        f'{date_expr} >= ?',
        f'{date_expr} <= ?',
        f'ABS(CAST({commission_expr} AS REAL)) > 0.0001',
        f"{name_expr} IS NOT NULL",
        f"TRIM({name_expr}) <> ''",
        f"TRIM({name_expr}) <> '-'",
    ]
    if keyword:
        where.append(_build_name_filter_sql(name_expr, keyword, alias_nicknames, params))
    sql = f"""
        SELECT {columns_sql}
        FROM {table_sql}
        WHERE {' AND '.join(where)}
        ORDER BY influencer_name ASC, settlement_time ASC, order_no ASC
    """
    return pd.read_sql_query(sql, conn, params=params)


def _query_leader_detail_rows(
    conn: sqlite3.Connection,
    start_text: str,
    end_text: str,
    keyword: str,
    alias_nicknames: list[str],
    config: ShopConfig,
) -> pd.DataFrame:
    merchant_map = _load_merchant_order_map(conn, config)
    if not merchant_map:
        return pd.DataFrame()

    rows_df = _query_leader_fee_rows(
        conn,
        start_text,
        end_text,
        config,
        selected_columns=list(_fund_flow_column_types(config).keys()),
    )
    if rows_df.empty:
        return rows_df

    rows_df['leader_name'] = rows_df['order_no'].map(
        lambda value: (merchant_map.get(_normalize_order_no(value)) or {}).get('name', '')
    )
    rows_df = rows_df[
        rows_df['leader_name'].map(lambda value: bool(value) and _name_matches_keyword(value, keyword, alias_nicknames))
    ].copy()
    if rows_df.empty:
        return rows_df
    return rows_df.sort_values(['leader_name', 'settlement_time', 'order_no'])


def _unique_zip_name(used_names: dict[str, int], filename: str) -> str:
    used_count = used_names.get(filename, 0)
    used_names[filename] = used_count + 1
    if not used_count:
        return filename
    stem = filename[:-5] if filename.lower().endswith('.xlsx') else filename
    return f'{stem}_{used_count + 1}.xlsx'


def export_commission_summary_zip(
    start_date: str | None,
    end_date: str | None,
    nickname_query: str | None,
    config: ShopConfig,
) -> tuple[BytesIO, str]:
    start_text, end_text = _normalize_commission_date_range(start_date, end_date)
    keyword = str(nickname_query or '').strip()
    db_path = _get_database_path()
    month_text = _build_commission_month_text(start_text, end_text)
    safe_month = _safe_download_part(month_text)

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_douyin_commission_tables(conn, config)
        alias_nicknames = _get_douyin_alias_nicknames(conn, keyword)
        creator_df = _query_creator_summary(conn, start_text, end_text, keyword, alias_nicknames, config)
        leader_df = _query_leader_summary(conn, start_text, end_text, keyword, alias_nicknames, config)
        unmatched_df = _query_unmatched_leader_rows(conn, start_text, end_text, config)

    summary_excel = _write_dataframe_excel(
        [
            ('达人汇总', creator_df[['达人名称', '达人ID', '佣金金额']]),
            ('团长汇总', leader_df[['团长名称', '团长ID', '佣金金额']]),
        ],
        amount_columns={'佣金金额'},
        text_columns={'达人名称', '达人ID', '团长名称', '团长ID'},
    )
    invoice_df = _build_invoice_import_df(creator_df, leader_df)
    invoice_excel = _write_dataframe_excel(
        [('应开金额导入', invoice_df)],
        amount_columns={'应开金额'},
        text_columns={'达人/客户'},
    )

    archive = BytesIO()
    with zipfile.ZipFile(archive, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f'{safe_month}佣金汇总.xlsx', summary_excel.getvalue())
        zf.writestr(f'{safe_month}应开金额导入.xlsx', invoice_excel.getvalue())
        if not unmatched_df.empty:
            unmatched_excel = _write_dataframe_excel(
                [('未匹配招商订单', unmatched_df)],
                amount_columns={'招商服务费'},
                text_columns={'订单号'},
            )
            zf.writestr(f'{safe_month}未匹配招商订单.xlsx', unmatched_excel.getvalue())
    archive.seek(0)
    return archive, _build_commission_zip_name(f'{config.display_name}佣金汇总', start_text, end_text)


def _write_streaming_excel(df: pd.DataFrame, sheet_name: str) -> BytesIO:
    output = BytesIO()
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title=sheet_name)
    worksheet.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        worksheet.append([None if pd.isna(value) else value for value in row])
    workbook.save(output)
    output.seek(0)
    return output


def export_store_self_sale_zip_file(
    output_path: str | Path,
    start_date: str | None,
    end_date: str | None,
    config: ShopConfig,
    chunk_size: int = 50000,
) -> str:
    start_text, end_text = _normalize_commission_date_range(start_date, end_date)
    db_path = _get_database_path()
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    fund_columns = [
        'id', 'settlement_time', 'order_no', 'sub_order_no', 'settlement_type',
        'settlement_amount', 'total_income', 'influencer_commission',
        'merchant_recruitment_fee',
    ]
    order_columns = [
        'id', 'main_order_no', 'sub_order_no', 'product_id', 'product_name',
        'product_quantity', 'order_payable_amount', 'traffic_source', 'order_status',
        'influencer_id', 'influencer_nickname',
    ]

    def normalized_key(series: pd.Series) -> pd.Series:
        return series.fillna('').astype(str).str.strip().str.lstrip("'")

    with sqlite3.connect(db_path) as conn:
        selected_df = pd.read_sql_query(
            f"""SELECT {', '.join(fund_columns)} FROM {config.fund_flow_table}
                WHERE TRIM(COALESCE(settlement_type, '')) = '已结算'
                  AND REPLACE(CAST(settlement_time AS TEXT), '/', '-') >= ?
                  AND REPLACE(CAST(settlement_time AS TEXT), '/', '-') <= ?""",
            conn,
            params=[start_text, end_text],
        )
        all_fees_df = pd.read_sql_query(
            f"""SELECT order_no, sub_order_no, influencer_commission,
                       merchant_recruitment_fee
                FROM {config.fund_flow_table}""",
            conn,
        )
        orders_df = pd.read_sql_query(
            f"SELECT {', '.join(order_columns)} FROM {config.orders_table}",
            conn,
        )

    selected_df['order_key'] = normalized_key(selected_df['order_no'])
    selected_df['sub_key'] = normalized_key(selected_df['sub_order_no'])
    selected_df['_settlement_sort'] = selected_df['settlement_time'].fillna('').astype(str).str.replace('/', '-', regex=False)
    selected_df = selected_df.sort_values(['order_key', 'sub_key', '_settlement_sort', 'id'])
    selected_df['settled_rank'] = selected_df.groupby(['order_key', 'sub_key'], dropna=False).cumcount() + 1

    selected_keys = selected_df[['order_key', 'sub_key']].drop_duplicates()
    all_fees_df['order_key'] = normalized_key(all_fees_df['order_no'])
    all_fees_df['sub_key'] = normalized_key(all_fees_df['sub_order_no'])
    all_fees_df = all_fees_df.merge(selected_keys, on=['order_key', 'sub_key'], how='inner')
    for column in ('influencer_commission', 'merchant_recruitment_fee'):
        all_fees_df[column] = pd.to_numeric(all_fees_df[column], errors='coerce').fillna(0)
    fee_totals = all_fees_df.groupby(['order_key', 'sub_key'], as_index=False).agg(
        creator_fee_total=('influencer_commission', 'sum'),
        leader_fee_total=('merchant_recruitment_fee', 'sum'),
    )

    orders_df['order_key'] = normalized_key(orders_df['main_order_no'])
    orders_df['sub_key'] = normalized_key(orders_df['sub_order_no'])
    primary = selected_df[selected_df['settled_rank'] == 1].merge(
        orders_df, on=['order_key', 'sub_key'], how='left', suffixes=('_fund', '_order'), indicator=True,
    ).merge(fee_totals, on=['order_key', 'sub_key'], how='left')
    primary['exclusion_reason'] = ''
    primary.loc[primary['_merge'] == 'left_only', 'exclusion_reason'] = '未匹配订单'
    matched = primary['exclusion_reason'] == ''
    primary.loc[matched & ~primary['traffic_source'].fillna('').astype(str).str.strip().isin(['小店自卖', '-']), 'exclusion_reason'] = '非小店自卖流量来源'
    matched = primary['exclusion_reason'] == ''
    primary.loc[matched & ~primary['order_status'].fillna('').astype(str).str.strip().isin(['已完成', '已发货']), 'exclusion_reason'] = '订单状态不符合'
    matched = primary['exclusion_reason'] == ''
    has_fee = primary['creator_fee_total'].fillna(0).abs().gt(0.0001) | primary['leader_fee_total'].fillna(0).abs().gt(0.0001)
    primary.loc[matched & has_fee, 'exclusion_reason'] = '存在达人/团长费用'

    output_columns = {
        'order_payable_amount': '统计金额', 'settlement_time': '结算时间',
        'order_no': '资金结算_订单号', 'sub_order_no_fund': '资金结算_子订单号',
        'settlement_type': '资金结算_结算单类型', 'settlement_amount': '资金结算_结算金额',
        'total_income': '资金结算_收入合计', 'influencer_commission': '资金结算_达人佣金',
        'merchant_recruitment_fee': '资金结算_招商服务费',
        'creator_fee_total': '订单全部流水_达人佣金合计', 'leader_fee_total': '订单全部流水_招商服务费合计',
        'main_order_no': '订单_主订单编号', 'sub_order_no_order': '订单_子订单编号',
        'product_id': '订单_商品ID', 'product_name': '订单_商品名称',
        'product_quantity': '订单_商品数量', 'order_payable_amount_source': '订单_订单应付金额',
        'traffic_source': '订单_流量来源', 'order_status': '订单_订单状态',
        'influencer_id': '订单_达人ID', 'influencer_nickname': '订单_达人昵称',
    }
    primary['统计金额'] = pd.to_numeric(primary['order_payable_amount'], errors='coerce').fillna(0)
    primary['订单_订单应付金额'] = primary['order_payable_amount']
    rows_df = primary.rename(columns={k: v for k, v in output_columns.items() if k not in {'order_payable_amount', 'order_payable_amount_source'}})
    report_columns = ['exclusion_reason', '统计金额'] + list(output_columns.values())[1:]
    rows_df = rows_df.reindex(columns=report_columns)

    duplicates = selected_df[selected_df['settled_rank'] > 1].copy()
    if not duplicates.empty:
        duplicate_rows = pd.DataFrame({column: None for column in report_columns}, index=duplicates.index)
        duplicate_rows['exclusion_reason'] = '重复/补充已结算，不重复统计'
        duplicate_rows['统计金额'] = 0
        for source, target in {
            'settlement_time': '结算时间', 'order_no': '资金结算_订单号',
            'sub_order_no': '资金结算_子订单号', 'settlement_type': '资金结算_结算单类型',
            'settlement_amount': '资金结算_结算金额', 'total_income': '资金结算_收入合计',
            'influencer_commission': '资金结算_达人佣金',
            'merchant_recruitment_fee': '资金结算_招商服务费',
        }.items():
            duplicate_rows[target] = duplicates[source].values
        rows_df = pd.concat([rows_df, duplicate_rows], ignore_index=True)
    rows_df = rows_df.sort_values(['结算时间', '资金结算_订单号', '资金结算_子订单号'], na_position='last')

    detail_df = rows_df[rows_df['exclusion_reason'] == ''].copy()
    unmatched_df = rows_df[rows_df['exclusion_reason'] != ''].copy()
    detail_df = detail_df.drop(columns=['exclusion_reason']).rename(columns={'统计金额': '订单应付金额'})
    unmatched_df = unmatched_df.rename(columns={'exclusion_reason': '不计入原因'})
    detail_total = round(float(pd.to_numeric(detail_df['订单应付金额'], errors='coerce').fillna(0).sum()), 2)
    summary_df = pd.DataFrame([{
        '项目': '店铺自卖',
        '期间': _build_commission_month_text(start_text, end_text),
        '账期起点': start_text[:10],
        '账期终点': end_text[:10],
        '汇总金额': detail_total,
        '计入明细行数': len(detail_df),
        '未匹配/不计入行数': len(unmatched_df),
    }])
    if round(float(detail_df['订单应付金额'].sum()), 2) != detail_total:
        raise ValueError('店铺自卖汇总金额与来源明细订单应付金额合计不一致')

    with zipfile.ZipFile(output_file, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('店铺自卖_汇总.xlsx', _write_streaming_excel(summary_df, '店铺自卖').getvalue())
        for index in range(0, len(detail_df), chunk_size):
            part = index // chunk_size + 1
            zf.writestr(f'店铺自卖_来源明细_第{part:03d}部分.xlsx', _write_streaming_excel(detail_df.iloc[index:index + chunk_size], '来源明细').getvalue())
        for index in range(0, len(unmatched_df), chunk_size):
            part = index // chunk_size + 1
            zf.writestr(f'店铺自卖_未匹配_第{part:03d}部分.xlsx', _write_streaming_excel(unmatched_df.iloc[index:index + chunk_size], '未匹配').getvalue())
        if detail_df.empty:
            zf.writestr('店铺自卖_来源明细_无数据.xlsx', _write_streaming_excel(detail_df, '来源明细').getvalue())
        if unmatched_df.empty:
            zf.writestr('店铺自卖_未匹配_无数据.xlsx', _write_streaming_excel(unmatched_df, '未匹配').getvalue())

    safe_month = _safe_download_part(_build_commission_month_text(start_text, end_text))
    return f'{config.display_name}_店铺自卖_{safe_month}_{detail_total:.2f}.zip'


def export_commission_detail_zip(
    start_date: str | None,
    end_date: str | None,
    nickname_query: str | None,
    config: ShopConfig,
) -> tuple[BytesIO, str]:
    start_text, end_text = _normalize_commission_date_range(start_date, end_date)
    keyword = str(nickname_query or '').strip()
    db_path = _get_database_path()
    month_text = _build_commission_month_text(start_text, end_text, short_year=True)

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_douyin_commission_tables(conn, config)
        alias_nicknames = _get_douyin_alias_nicknames(conn, keyword)
        creator_rows = _query_creator_detail_rows(conn, start_text, end_text, keyword, alias_nicknames, config)
        leader_rows = _query_leader_detail_rows(conn, start_text, end_text, keyword, alias_nicknames, config)

    archive = BytesIO()
    used_names: dict[str, int] = {}
    with zipfile.ZipFile(archive, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        if not creator_rows.empty:
            for name, group_df in creator_rows.groupby('influencer_name', sort=True):
                amount_sum = float(pd.to_numeric(group_df['influencer_commission'], errors='coerce').fillna(0).sum())
                detail_df = _fund_flow_chinese_columns(group_df.copy(), config)
                detail_excel = _write_dataframe_excel(
                    [('明细', detail_df)],
                    amount_columns=DOUYIN_COMMISSION_AMOUNT_COLUMNS,
                    text_columns={'订单号', '子订单号', '商品ID', '达人ID'},
                )
                filename = _safe_download_part(f'{name} {-amount_sum:.2f} {month_text}') + '.xlsx'
                zf.writestr(_unique_zip_name(used_names, filename), detail_excel.getvalue())

        if not leader_rows.empty:
            for name, group_df in leader_rows.groupby('leader_name', sort=True):
                amount_sum = float(pd.to_numeric(group_df['merchant_recruitment_fee'], errors='coerce').fillna(0).sum())
                detail_df = group_df.copy().rename(columns={'leader_name': '团长名称'})
                detail_df = _fund_flow_chinese_columns(detail_df, config)
                detail_excel = _write_dataframe_excel(
                    [('明细', detail_df)],
                    amount_columns=DOUYIN_COMMISSION_AMOUNT_COLUMNS,
                    text_columns={'订单号', '子订单号', '商品ID', '达人ID'},
                )
                filename = _safe_download_part(f'团长{name} {-amount_sum:.2f} 招商佣金 {month_text}') + '.xlsx'
                zf.writestr(_unique_zip_name(used_names, filename), detail_excel.getvalue())

        if not used_names:
            empty_df = pd.DataFrame(columns=['提示'])
            empty_excel = _write_dataframe_excel([('明细', empty_df)])
            zf.writestr('无佣金明细.xlsx', empty_excel.getvalue())

    archive.seek(0)
    return archive, _build_commission_zip_name(f'{config.display_name}佣金明细', start_text, end_text)


def export_detail_source_commission_summary_zip(
    start_date: str | None,
    end_date: str | None,
    nickname_query: str | None,
    exemption_mode: str | None,
    config: ShopConfig,
) -> tuple[BytesIO, str]:
    start_text, end_text = _normalize_commission_date_range(start_date, end_date)
    mode = _normalize_detail_export_mode(exemption_mode)
    mode_label = _detail_export_mode_label(mode)
    month_text = _build_commission_month_text(start_text, end_text)
    safe_month = _safe_download_part(month_text)

    (
        _creator_rows,
        _leader_rows,
        creator_df,
        leader_df,
        unmatched_creator_df,
        unmatched_leader_df,
    ) = _load_detail_source_export_data(start_text, end_text, nickname_query, mode, config)

    summary_excel = _write_dataframe_excel(
        [
            ('达人汇总', creator_df),
            ('团长汇总', leader_df),
        ],
        amount_columns={'佣金金额'},
        text_columns={'达人名称', '达人ID', '团长名称', '团长ID'},
    )
    invoice_df = _build_detail_invoice_import_df(creator_df, leader_df)
    invoice_excel = _write_dataframe_excel(
        [('应开金额导入', invoice_df)],
        amount_columns={'应开金额'},
        text_columns={'达人/客户'},
    )

    archive = BytesIO()
    with zipfile.ZipFile(archive, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f'{safe_month}按明细表{mode_label}佣金汇总.xlsx', summary_excel.getvalue())
        zf.writestr(f'{safe_month}按明细表{mode_label}应开金额导入.xlsx', invoice_excel.getvalue())
        if not unmatched_creator_df.empty:
            unmatched_creator_excel = _write_dataframe_excel(
                [('未匹配达人ID', unmatched_creator_df)],
                amount_columns={'实际佣金支出'},
                text_columns={'订单id', '商品id'},
            )
            zf.writestr(f'{safe_month}按明细表{mode_label}未匹配达人ID.xlsx', unmatched_creator_excel.getvalue())
        if not unmatched_leader_df.empty:
            unmatched_leader_excel = _write_dataframe_excel(
                [('未匹配团长ID', unmatched_leader_df)],
                amount_columns={'实际服务费收入'},
                text_columns={'订单id', '商品id', '团长ID'},
            )
            zf.writestr(f'{safe_month}按明细表{mode_label}未匹配团长ID.xlsx', unmatched_leader_excel.getvalue())
    archive.seek(0)
    return archive, _build_commission_zip_name(f'{config.display_name}佣金汇总_按明细表_{mode_label}', start_text, end_text)


def export_detail_source_commission_detail_zip(
    start_date: str | None,
    end_date: str | None,
    nickname_query: str | None,
    exemption_mode: str | None,
    config: ShopConfig,
) -> tuple[BytesIO, str]:
    start_text, end_text = _normalize_commission_date_range(start_date, end_date)
    mode = _normalize_detail_export_mode(exemption_mode)
    mode_label = _detail_export_mode_label(mode)
    month_text = _build_commission_month_text(start_text, end_text, short_year=True)

    (
        creator_rows,
        leader_rows,
        _creator_df,
        _leader_df,
        unmatched_creator_df,
        unmatched_leader_df,
    ) = _load_detail_source_export_data(start_text, end_text, nickname_query, mode, config)

    archive = BytesIO()
    used_names: dict[str, int] = {}
    with zipfile.ZipFile(archive, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        if not creator_rows.empty:
            creator_rows = creator_rows[creator_rows['detail_creator_name'].fillna('').astype(str).str.strip() != '']
            for name, group_df in creator_rows.groupby('detail_creator_name', sort=True):
                amount_sum = float(pd.to_numeric(group_df['actual_commission'], errors='coerce').fillna(0).sum())
                detail_df = _detail_creator_export_df(group_df)
                detail_excel = _write_dataframe_excel(
                    [('明细', detail_df)],
                    amount_columns=DOUYIN_COMMISSION_AMOUNT_COLUMNS,
                    text_columns={'订单id', '商品id', '店铺id', '营销活动id', '阶梯计划ID', '达人ID'},
                )
                filename = _safe_download_part(f'{name} {amount_sum:.2f} {mode_label} 按明细表{month_text}') + '.xlsx'
                zf.writestr(_unique_zip_name(used_names, filename), detail_excel.getvalue())

        if not leader_rows.empty:
            leader_rows = leader_rows[leader_rows['detail_leader_name'].fillna('').astype(str).str.strip() != '']
            for name, group_df in leader_rows.groupby('detail_leader_name', sort=True):
                amount_sum = float(pd.to_numeric(group_df['actual_service_income'], errors='coerce').fillna(0).sum())
                detail_df = _detail_leader_export_df(group_df)
                detail_excel = _write_dataframe_excel(
                    [('明细', detail_df)],
                    amount_columns=DOUYIN_COMMISSION_AMOUNT_COLUMNS,
                    text_columns={'订单id', '商品id', '店铺id', '团长活动id', '团长ID'},
                )
                filename = _safe_download_part(f'团长{name} {amount_sum:.2f} 招商佣金 {mode_label} 按明细表{month_text}') + '.xlsx'
                zf.writestr(_unique_zip_name(used_names, filename), detail_excel.getvalue())

        if not unmatched_creator_df.empty:
            unmatched_creator_excel = _write_dataframe_excel(
                [('未匹配达人ID', unmatched_creator_df)],
                amount_columns={'实际佣金支出'},
                text_columns={'订单id', '商品id'},
            )
            zf.writestr('未匹配达人ID.xlsx', unmatched_creator_excel.getvalue())
        if not unmatched_leader_df.empty:
            unmatched_leader_excel = _write_dataframe_excel(
                [('未匹配团长ID', unmatched_leader_df)],
                amount_columns={'实际服务费收入'},
                text_columns={'订单id', '商品id', '团长ID'},
            )
            zf.writestr('未匹配团长ID.xlsx', unmatched_leader_excel.getvalue())

        if not used_names and unmatched_creator_df.empty and unmatched_leader_df.empty:
            empty_df = pd.DataFrame(columns=['提示'])
            empty_excel = _write_dataframe_excel([('明细', empty_df)])
            zf.writestr('无佣金明细.xlsx', empty_excel.getvalue())

    archive.seek(0)
    return archive, _build_commission_zip_name(f'{config.display_name}佣金明细_按明细表_{mode_label}', start_text, end_text)
