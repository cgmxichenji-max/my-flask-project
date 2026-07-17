from io import BytesIO
from pathlib import Path
from typing import Any
from datetime import datetime, timedelta
import sqlite3
import zipfile

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from flask import current_app
import re
from werkzeug.datastructures import FileStorage
from common.excel_utils import (
    is_excel_filename,
    normalize_header_text,
    normalize_columns,
    check_columns_match,
)

from wechat_shop.table_schemas import (
    ORDER_COLUMN_MAPPING,
    ORDER_REQUIRED_COLUMNS,
    ORDER_COLUMN_TYPES,
    WECHAT_ORDER_TABLE_NAME,
    FUND_FLOW_COLUMN_MAPPING,
    FUND_FLOW_REQUIRED_COLUMNS,
    FUND_FLOW_COLUMN_TYPES,
    WECHAT_FUND_FLOW_TABLE_NAME,
    AFTER_SALES_COLUMN_MAPPING,
    AFTER_SALES_REQUIRED_COLUMNS,
    AFTER_SALES_COLUMN_TYPES,
    WECHAT_AFTER_SALES_TABLE_NAME,
)



TEXT_SOURCE_COLUMNS = {
    '订单号',
    '交易单号',
    '快递单号',
    '收件人手机',
    '商品编码（平台）',
    '商品编码(平台)',
    '商品编码(自定义)',
    'SKU编码(自定义)',
    'sku编码(自定义)',
    '礼物单号',
    'gift_order_no',
    'custom_product_code',
    'custom_sku_code',
    '流水单号',
    '关联订单号',
    '关联售后单号',
    '关联提现单号',
    '关联保单号',
    '关联礼物单号',
    '售后单号',
    '订单编号',
    '发货物流单号',
    '退换货物流单号',
    '商家退换货联系人电话',
    '商品编码（平台）',
    '商品编码(自定义)',
}


ORDER_DEDUP_KEY_COLUMNS = [
    'order_no',
    'platform_product_code',
    'product_attributes',
]

FUND_FLOW_DEDUP_KEY_COLUMNS = [
    'flow_no',
    'booking_time',
    'transaction_type',
    'related_order_no',
]
FUND_FLOW_EMPTY_DEDUP_VALUE = '<EMPTY>'

AFTER_SALES_DEDUP_KEY_COLUMNS = [
    'after_sales_no',
    'after_sales_apply_time',
]


# ===================== 数据状态表定义与操作 =====================
DATA_STATUS_TABLE_NAME = 'wechat_shop_data_status'

DATA_STATUS_CONFIG = {
    'orders': {
        'table_name': '订单表',
        'source_table': WECHAT_ORDER_TABLE_NAME,
        'date_field': 'order_created_at',
    },
    'fund_flows': {
        'table_name': '资金流水表',
        'source_table': WECHAT_FUND_FLOW_TABLE_NAME,
        'date_field': 'booking_time',
    },
    'aftersales': {
        'table_name': '售后表',
        'source_table': WECHAT_AFTER_SALES_TABLE_NAME,
        'date_field': 'after_sales_apply_time',
    },
}

# === EXPORT_TABLE_CONFIG block inserted here ===

EXPORT_TABLE_CONFIG = {
    'orders': {
        'table_name': '订单表',
        'source_table': WECHAT_ORDER_TABLE_NAME,
        'date_field': 'order_created_at',
        'allowed_fields': list(ORDER_COLUMN_TYPES.keys()),
        'column_types': ORDER_COLUMN_TYPES,
    },
    'fund_flows': {
        'table_name': '资金流水表',
        'source_table': WECHAT_FUND_FLOW_TABLE_NAME,
        'date_field': 'booking_time',
        'allowed_fields': list(FUND_FLOW_COLUMN_TYPES.keys()),
        'column_types': FUND_FLOW_COLUMN_TYPES,
    },
    'aftersales': {
        'table_name': '售后表',
        'source_table': WECHAT_AFTER_SALES_TABLE_NAME,
        'date_field': 'after_sales_apply_time',
        'allowed_fields': list(AFTER_SALES_COLUMN_TYPES.keys()),
        'column_types': AFTER_SALES_COLUMN_TYPES,
    },
}

# === EXPORT_HEADER_MAPPING block inserted here ===
EXPORT_HEADER_MAPPING = {
    'orders': {english_name: chinese_name for chinese_name, english_name in ORDER_COLUMN_MAPPING.items()},
    'fund_flows': {english_name: chinese_name for chinese_name, english_name in FUND_FLOW_COLUMN_MAPPING.items()},
    'aftersales': {english_name: chinese_name for chinese_name, english_name in AFTER_SALES_COLUMN_MAPPING.items()},
}


def _ensure_data_status_seeded(conn: sqlite3.Connection) -> None:
    """确保数据状态表存在 3 条基础记录。"""
    cursor = conn.cursor()
    cursor.execute(
        f'''
        CREATE TABLE IF NOT EXISTS {DATA_STATUS_TABLE_NAME} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            table_key TEXT NOT NULL UNIQUE,
            table_name TEXT NOT NULL,
            record_count INTEGER DEFAULT 0,
            min_date TEXT,
            max_date TEXT,
            last_import_time TEXT
        );
        '''
    )

    for table_key, config in DATA_STATUS_CONFIG.items():
        cursor.execute(
            f'''
            INSERT OR IGNORE INTO {DATA_STATUS_TABLE_NAME}
            (table_key, table_name, record_count)
            VALUES (?, ?, 0)
            ''',
            (table_key, config['table_name']),
        )

    conn.commit()



def _update_data_status(table_key: str) -> None:
    """按表标识刷新当前数据状态。"""
    config = DATA_STATUS_CONFIG.get(table_key)
    if not config:
        return

    db_path = _get_database_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)

    with sqlite3.connect(db_path) as conn:
        _ensure_data_status_seeded(conn)
        cursor = conn.cursor()
        cursor.execute(
            f'''
            UPDATE {DATA_STATUS_TABLE_NAME}
            SET
                table_name = ?,
                record_count = (SELECT COUNT(*) FROM {config['source_table']}),
                min_date = (SELECT MIN({config['date_field']}) FROM {config['source_table']}),
                max_date = (SELECT MAX({config['date_field']}) FROM {config['source_table']}),
                last_import_time = datetime('now', 'localtime')
            WHERE table_key = ?
            ''',
            (config['table_name'], table_key),
        )
        conn.commit()


def _try_update_data_status(table_key: str) -> str:
    try:
        _update_data_status(table_key)
        return ''
    except Exception as exc:
        return f'提示：数据已写入，但刷新数据状态失败，请刷新页面确认。错误：{exc}'


def _build_import_precheck_failed_response(
    title: str,
    file_summaries: list[dict[str, Any]],
    invalid_files: list[str],
    failed_files: list[dict[str, str]],
) -> dict[str, Any]:
    message_parts: list[str] = [f'{title}预检未通过，未写入数据库']

    for summary in file_summaries:
        message_parts.append('')
        message_parts.append(_build_file_summary_text(summary))

    if failed_files:
        message_parts.append('')
        message_parts.append('预检失败：')
        for failed in failed_files:
            message_parts.append(f"- {failed['filename']}（{failed['error']}）")

    if invalid_files:
        message_parts.append('')
        message_parts.append(f"无效文件：{'，'.join(invalid_files)}")

    return {
        'success': False,
        'message': '\n'.join(message_parts),
        'file_count': len(file_summaries),
        'files': file_summaries,
        'invalid_files': invalid_files,
        'failed_files': failed_files,
        'precheck_failed': True,
        'written_rows': 0,
    }


def _build_file_summary_text(file_info: dict[str, Any]) -> str:
    """把单个文件摘要拼成前端当前文本框可直接显示的文字。"""
    lines = [
        f"文件：{file_info['filename']}",
        f"行数：{file_info['row_count']}",
        f"列数：{file_info['column_count']}",
        f"列名：{'，'.join(file_info['columns'])}",
    ]
    return '\n'.join(lines)




def _get_database_path() -> Path:
    """获取 SQLite 数据库路径，优先读取 Flask 配置，未配置时回退到 data/wechat_shop.db。"""
    db_path = current_app.config.get('DATABASE_PATH')
    if db_path:
        return Path(db_path)

    return Path(current_app.root_path) / 'data' / 'wechat_shop.db'


def _get_upload_source_filename(file_obj: Any) -> str:
    """兼容浏览器上传流和服务器暂存文件对象。"""
    return str(getattr(file_obj, 'filename', '') or '').strip()


def _read_upload_source_bytes(file_obj: Any) -> bytes:
    """读取 Excel 内容；优先读取暂存文件，避免导入阶段依赖请求流。"""
    path = getattr(file_obj, 'path', None)
    if path:
        return Path(path).read_bytes()
    return file_obj.read()


def _reset_upload_source(file_obj: Any) -> None:
    """旧上传流需要复位；暂存文件无需处理。"""
    stream = getattr(file_obj, 'stream', None)
    if stream is not None:
        stream.seek(0)

# === Begin export to excel helpers ===
EXPORT_ZIP_CHUNK_SIZE = 50000


def _normalize_export_datetime_text(value: str | None) -> str | None:
    """把前端传入的日期/时间文本统一转换为可用于 SQLite 比较的字符串。"""
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    normalized = text.replace('T', ' ').replace('/', '-').strip()
    formats = [
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M',
        '%Y-%m-%d',
    ]

    for fmt in formats:
        try:
            dt = datetime.strptime(normalized, fmt)
            if fmt == '%Y-%m-%d':
                return dt.strftime('%Y-%m-%d 00:00:00')
            if fmt == '%Y-%m-%d %H:%M':
                return dt.strftime('%Y-%m-%d %H:%M:00')
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except ValueError:
            continue

    return normalized




def _build_export_download_base_name(table_key: str, start_time: str | None, end_time: str | None) -> str:
    """生成导出文件基础名，不包含扩展名。"""
    table_name_map = {
        'orders': '订单表',
        'fund_flows': '资金流水表',
        'aftersales': '售后表',
    }
    table_name = table_name_map.get(table_key, '导出数据')

    def _safe_part(value: str | None) -> str:
        if not value:
            return '全部时间'
        return re.sub(r'[\\/:*?"<>|\s]+', '_', value)

    return f"{table_name}_{_safe_part(start_time)}_到_{_safe_part(end_time)}"


def _build_export_download_name(table_key: str, start_time: str | None, end_time: str | None) -> str:
    """生成导出 Excel 文件名。"""
    return f"{_build_export_download_base_name(table_key, start_time, end_time)}.xlsx"


def _build_export_zip_download_name(table_key: str, start_time: str | None, end_time: str | None) -> str:
    """生成导出 ZIP 文件名。"""
    return f"{_build_export_download_base_name(table_key, start_time, end_time)}.zip"


# ===== Excel导出列宽自适应辅助函数 =====
def _get_excel_display_width(value: Any) -> int:
    """按中英文混合文本估算 Excel 显示宽度。"""
    if value is None:
        return 0

    text = str(value)
    width = 0
    for ch in text:
        width += 2 if ord(ch) > 127 else 1
    return width



def _auto_adjust_excel_columns(worksheet) -> None:
    """按表头和单元格内容自动调整列宽，并限制最大宽度避免过宽。"""
    min_width = 10
    max_width = 40

    for column_index, column_cells in enumerate(worksheet.iter_cols(), start=1):
        max_display_width = 0

        for cell in column_cells:
            cell_width = _get_excel_display_width(cell.value)
            if cell_width > max_display_width:
                max_display_width = cell_width

        adjusted_width = min(max(max_display_width + 2, min_width), max_width)
        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def _build_export_headers(table_key: str, selected_fields: list[str]) -> list[str]:
    header_mapping = EXPORT_HEADER_MAPPING.get(table_key, {})
    headers: list[str] = []
    used_headers: dict[str, int] = {}

    for column_name in selected_fields:
        base_header = header_mapping.get(column_name, column_name)
        if base_header not in used_headers:
            used_headers[base_header] = 1
            headers.append(base_header)
        else:
            used_headers[base_header] += 1
            headers.append(f"{base_header}_{used_headers[base_header]}")

    return headers


def _write_rows_to_excel_bytes(headers: list[str], rows: list[tuple[Any, ...]]) -> BytesIO:
    output = BytesIO()
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title='导出结果')
    worksheet.append(headers)

    max_widths = [_get_excel_display_width(header) for header in headers]
    for row in rows:
        values = list(row)
        worksheet.append(values)
        for index, value in enumerate(values):
            if index >= len(max_widths):
                continue
            value_width = _get_excel_display_width(value)
            if value_width > max_widths[index]:
                max_widths[index] = value_width

    for column_index, width in enumerate(max_widths, start=1):
        column_letter = get_column_letter(column_index)
        worksheet.column_dimensions[column_letter].width = min(max(width + 2, 10), 40)

    workbook.save(output)
    output.seek(0)
    return output


def _build_export_query_parts(
    table_key: str,
    start_time: str | None,
    end_time: str | None,
    selected_fields: list[str],
    filter_conditions: list[dict[str, Any]] | None = None,
) -> tuple[dict[str, Any], str | None, str | None, str, list[Any]]:
    config = EXPORT_TABLE_CONFIG.get(table_key)
    if not config:
        raise ValueError('不支持的导出表类型')

    if not selected_fields:
        raise ValueError('请至少选择一个导出字段')

    allowed_fields = set(config['allowed_fields'])
    column_types = config.get('column_types', {})
    invalid_fields = [field for field in selected_fields if field not in allowed_fields]
    if invalid_fields:
        raise ValueError(f"存在无效导出字段：{', '.join(invalid_fields)}")

    start_text = _normalize_export_datetime_text(start_time)
    end_text = _normalize_export_datetime_text(end_time)

    if start_text and end_text and start_text > end_text:
        raise ValueError('开始时间不能大于结束时间')
    if filter_conditions is None:
        filter_conditions = []
    if not isinstance(filter_conditions, list):
        raise ValueError('筛选条件格式不正确')

    fields_sql = ', '.join(selected_fields)
    sql = f"SELECT {fields_sql} FROM {config['source_table']}"
    where_parts: list[str] = []
    params: list[Any] = []

    date_compare_expr = _build_datetime_compare_expr(config['date_field'])

    if start_text:
        where_parts.append(f"{date_compare_expr} >= ?")
        params.append(start_text)

    if end_text:
        where_parts.append(f"{date_compare_expr} <= ?")
        params.append(end_text)

    filter_sql_parts, filter_sql_params = _build_filter_sql_parts(
        filter_conditions=filter_conditions,
        allowed_fields=allowed_fields,
        column_types=column_types,
    )
    if filter_sql_parts:
        where_parts.append('(' + ' '.join(filter_sql_parts) + ')')
        params.extend(filter_sql_params)

    if where_parts:
        sql += ' WHERE ' + ' AND '.join(where_parts)

    return config, start_text, end_text, sql, params


def _normalize_filter_logic(value: Any) -> str:
    """把前端传来的逻辑连接符标准化为 SQL 可用值。"""
    text = str(value or '').strip().lower()
    return 'OR' if text == 'or' else 'AND'



def _normalize_filter_operator(value: Any) -> str:
    """把前端传来的运算符标准化。"""
    return str(value or '').strip().lower()


def _try_parse_numeric_value(value: Any) -> float | None:
    """尝试把筛选值解析为数字；失败则返回 None。"""
    text = str(value or '').strip()
    if text == '':
        return None

    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _is_numeric_column(column_name: str, column_types: dict[str, str]) -> bool:
    """判断字段是否为数值列。"""
    column_type = str(column_types.get(column_name, '')).upper()
    return column_type in {'REAL', 'INTEGER', 'NUMERIC', 'FLOAT', 'DECIMAL'}


def _is_datetime_column(column_name: str) -> bool:
    """按字段名判断是否为日期时间列。"""
    text = str(column_name or '').strip().lower()
    return text.endswith('_at') or text.endswith('_time') or text == 'booking_time'


def _normalize_filter_datetime_text(value: Any, boundary: str = 'start') -> str | None:
    """把筛选条件中的日期/时间文本标准化为可比较的字符串。"""
    text = str(value or '').strip()
    if text == '':
        return None

    normalized = text.replace('T', ' ').replace('/', '-').strip()
    formats = [
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d %H:%M',
        '%Y-%m-%d',
    ]

    for fmt in formats:
        try:
            dt = datetime.strptime(normalized, fmt)
            if fmt == '%Y-%m-%d':
                if boundary == 'end':
                    return dt.strftime('%Y-%m-%d 23:59:59')
                return dt.strftime('%Y-%m-%d 00:00:00')
            if fmt == '%Y-%m-%d %H:%M':
                return dt.strftime('%Y-%m-%d %H:%M:00')
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except ValueError:
            continue

    return normalized


def _build_datetime_compare_expr(column_name: str) -> str:
    """构造日期字段比较表达式，统一把库内的 / 替换为 -。"""
    return f"REPLACE(CAST({column_name} AS TEXT), '/', '-')"



def _build_filter_sql_parts(
    filter_conditions: list[dict[str, Any]],
    allowed_fields: set[str],
    column_types: dict[str, str],
) -> tuple[list[str], list[Any]]:
    """把筛选条件转换为 SQL 片段与参数。"""
    sql_parts: list[str] = []
    sql_params: list[Any] = []

    for raw_condition in filter_conditions:
        if not isinstance(raw_condition, dict):
            continue

        field_name = str(raw_condition.get('field') or '').strip()
        operator = _normalize_filter_operator(raw_condition.get('operator'))
        logic = _normalize_filter_logic(raw_condition.get('logic'))
        raw_value = raw_condition.get('value')
        value_text = '' if raw_value is None else str(raw_value).strip()

        if not field_name:
            continue
        if field_name not in allowed_fields:
            raise ValueError(f'存在非法筛选字段：{field_name}')

        if operator not in {
            'eq', 'ne', 'contains', 'not_contains',
            'gt', 'gte', 'lt', 'lte',
            'is_empty', 'is_not_empty',
        }:
            raise ValueError(f'存在非法筛选条件：{operator}')

        is_numeric_field = _is_numeric_column(field_name, column_types)
        is_datetime_field = _is_datetime_column(field_name)

        clause = ''
        params: list[Any] = []

        if operator == 'eq':
            if value_text == '':
                continue
            if is_datetime_field:
                start_value = _normalize_filter_datetime_text(value_text, 'start')
                end_value = _normalize_filter_datetime_text(value_text, 'end')
                compare_expr = _build_datetime_compare_expr(field_name)
                clause = f"({compare_expr} >= ? AND {compare_expr} <= ?)"
                params = [start_value, end_value]
            else:
                numeric_value = _try_parse_numeric_value(value_text) if is_numeric_field else None
                if is_numeric_field and numeric_value is not None:
                    clause = f"CAST({field_name} AS REAL) = ?"
                    params = [numeric_value]
                else:
                    clause = f"CAST({field_name} AS TEXT) = ?"
                    params = [value_text]
        elif operator == 'ne':
            if value_text == '':
                continue
            if is_datetime_field:
                start_value = _normalize_filter_datetime_text(value_text, 'start')
                end_value = _normalize_filter_datetime_text(value_text, 'end')
                compare_expr = _build_datetime_compare_expr(field_name)
                clause = f"({compare_expr} < ? OR {compare_expr} > ?)"
                params = [start_value, end_value]
            else:
                numeric_value = _try_parse_numeric_value(value_text) if is_numeric_field else None
                if is_numeric_field and numeric_value is not None:
                    clause = f"CAST({field_name} AS REAL) <> ?"
                    params = [numeric_value]
                else:
                    clause = f"CAST({field_name} AS TEXT) <> ?"
                    params = [value_text]
        elif operator == 'contains':
            if value_text == '':
                continue
            clause = f"CAST({field_name} AS TEXT) LIKE ?"
            params = [f"%{value_text}%"]
        elif operator == 'not_contains':
            if value_text == '':
                continue
            clause = f"CAST({field_name} AS TEXT) NOT LIKE ?"
            params = [f"%{value_text}%"]
        elif operator == 'gt':
            if value_text == '':
                continue
            if is_datetime_field:
                compare_value = _normalize_filter_datetime_text(value_text, 'end')
                compare_expr = _build_datetime_compare_expr(field_name)
                clause = f"{compare_expr} > ?"
                params = [compare_value]
            else:
                numeric_value = _try_parse_numeric_value(value_text) if is_numeric_field else None
                if is_numeric_field and numeric_value is not None:
                    clause = f"CAST({field_name} AS REAL) > ?"
                    params = [numeric_value]
                else:
                    clause = f"CAST({field_name} AS TEXT) > ?"
                    params = [value_text]
        elif operator == 'gte':
            if value_text == '':
                continue
            if is_datetime_field:
                compare_value = _normalize_filter_datetime_text(value_text, 'start')
                compare_expr = _build_datetime_compare_expr(field_name)
                clause = f"{compare_expr} >= ?"
                params = [compare_value]
            else:
                numeric_value = _try_parse_numeric_value(value_text) if is_numeric_field else None
                if is_numeric_field and numeric_value is not None:
                    clause = f"CAST({field_name} AS REAL) >= ?"
                    params = [numeric_value]
                else:
                    clause = f"CAST({field_name} AS TEXT) >= ?"
                    params = [value_text]
        elif operator == 'lt':
            if value_text == '':
                continue
            if is_datetime_field:
                compare_value = _normalize_filter_datetime_text(value_text, 'start')
                compare_expr = _build_datetime_compare_expr(field_name)
                clause = f"{compare_expr} < ?"
                params = [compare_value]
            else:
                numeric_value = _try_parse_numeric_value(value_text) if is_numeric_field else None
                if is_numeric_field and numeric_value is not None:
                    clause = f"CAST({field_name} AS REAL) < ?"
                    params = [numeric_value]
                else:
                    clause = f"CAST({field_name} AS TEXT) < ?"
                    params = [value_text]
        elif operator == 'lte':
            if value_text == '':
                continue
            if is_datetime_field:
                compare_value = _normalize_filter_datetime_text(value_text, 'end')
                compare_expr = _build_datetime_compare_expr(field_name)
                clause = f"{compare_expr} <= ?"
                params = [compare_value]
            else:
                numeric_value = _try_parse_numeric_value(value_text) if is_numeric_field else None
                if is_numeric_field and numeric_value is not None:
                    clause = f"CAST({field_name} AS REAL) <= ?"
                    params = [numeric_value]
                else:
                    clause = f"CAST({field_name} AS TEXT) <= ?"
                    params = [value_text]
        elif operator == 'is_empty':
            clause = f"({field_name} IS NULL OR TRIM(CAST({field_name} AS TEXT)) = '')"
        elif operator == 'is_not_empty':
            clause = f"({field_name} IS NOT NULL AND TRIM(CAST({field_name} AS TEXT)) <> '')"

        if not clause:
            continue

        if not sql_parts:
            sql_parts.append(clause)
        else:
            sql_parts.append(f"{logic} {clause}")
        sql_params.extend(params)

    return sql_parts, sql_params


def export_data_to_excel(
    table_key: str,
    start_time: str | None,
    end_time: str | None,
    selected_fields: list[str],
    filter_conditions: list[dict[str, Any]] | None = None,
) -> tuple[BytesIO, str]:
    """按表、时间范围、字段选择从数据库导出 Excel。"""
    db_path = _get_database_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)

    config, start_text, end_text, sql, params = _build_export_query_parts(
        table_key=table_key,
        start_time=start_time,
        end_time=end_time,
        selected_fields=selected_fields,
        filter_conditions=filter_conditions,
    )

    sql += f" ORDER BY {config['date_field']} ASC, id ASC"

    with sqlite3.connect(db_path) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
            (config['source_table'],),
        )
        table_exists = cursor.fetchone() is not None
        if not table_exists:
            raise ValueError(f"数据表不存在：{config['source_table']}")

        df = pd.read_sql_query(sql, conn, params=params)

    header_mapping = EXPORT_HEADER_MAPPING.get(table_key, {})
    renamed_columns: list[str] = []
    used_headers: dict[str, int] = {}

    for column_name in df.columns.tolist():
        base_header = header_mapping.get(column_name, column_name)
        if base_header not in used_headers:
            used_headers[base_header] = 1
            renamed_columns.append(base_header)
        else:
            used_headers[base_header] += 1
            renamed_columns.append(f"{base_header}_{used_headers[base_header]}")

    df.columns = renamed_columns

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='导出结果')
        worksheet = writer.sheets['导出结果']
        _auto_adjust_excel_columns(worksheet)
    output.seek(0)

    download_name = _build_export_download_name(table_key, start_text, end_text)
    return output, download_name


def export_data_to_zip(
    table_key: str,
    start_time: str | None,
    end_time: str | None,
    selected_fields: list[str],
    filter_conditions: list[dict[str, Any]] | None = None,
    chunk_size: int = EXPORT_ZIP_CHUNK_SIZE,
) -> tuple[BytesIO, str]:
    """按表、时间范围、字段选择导出 ZIP，内部 Excel 按固定行数切割。"""
    if chunk_size <= 0:
        raise ValueError('导出切割行数配置不正确')

    db_path = _get_database_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)

    config, start_text, end_text, sql, params = _build_export_query_parts(
        table_key=table_key,
        start_time=start_time,
        end_time=end_time,
        selected_fields=selected_fields,
        filter_conditions=filter_conditions,
    )
    ordered_sql = f"{sql} ORDER BY {config['date_field']} ASC, id ASC"
    headers = _build_export_headers(table_key, selected_fields)
    base_name = _build_export_download_base_name(table_key, start_text, end_text)

    archive = BytesIO()
    with sqlite3.connect(db_path) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
            (config['source_table'],),
        )
        table_exists = cursor.fetchone() is not None
        if not table_exists:
            raise ValueError(f"数据表不存在：{config['source_table']}")

        with zipfile.ZipFile(archive, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
            cursor.execute(ordered_sql, params)
            part_index = 1
            while True:
                rows = cursor.fetchmany(chunk_size)
                if not rows:
                    break

                excel_output = _write_rows_to_excel_bytes(headers, rows)
                part_name = f"{base_name}_第{part_index:03d}部分.xlsx"
                zf.writestr(part_name, excel_output.getvalue())
                part_index += 1

            if part_index == 1:
                empty_excel = _write_rows_to_excel_bytes(headers, [])
                zf.writestr(f"{base_name}_无数据.xlsx", empty_excel.getvalue())

    archive.seek(0)
    return archive, _build_export_zip_download_name(table_key, start_text, end_text)


COMMISSION_TRANSACTION_TYPES = ('达人佣金', '带货机构服务费')
STORE_SELF_SALE_TRANSACTION_TYPE = '技术服务费'
COMMISSION_UNMATCHED_NICKNAME = '未匹配带货账号昵称'
COMMISSION_DETAIL_COLUMNS = [
    '流水单号',
    '记账时间',
    '动账类型',
    '收支类型',
    '收支金额',
    '关联订单号',
    '带货账号昵称',
]


def _normalize_commission_date(value: str | None, boundary: str) -> str:
    text = str(value or '').strip()
    if not text:
        raise ValueError('请选择佣金导出的开始日期和结束日期')

    normalized = text.replace('/', '-')
    try:
        dt = datetime.strptime(normalized, '%Y-%m-%d')
    except ValueError as exc:
        raise ValueError('佣金导出日期格式不正确，请使用 YYYY-MM-DD') from exc

    if boundary == 'end':
        return dt.strftime('%Y-%m-%d 23:59:59')
    return dt.strftime('%Y-%m-%d 00:00:00')


def _normalize_commission_date_range(start_date: str | None, end_date: str | None) -> tuple[str, str]:
    start_text = _normalize_commission_date(start_date, 'start')
    end_text = _normalize_commission_date(end_date, 'end')
    if start_text > end_text:
        raise ValueError('佣金导出开始日期不能晚于结束日期')
    return start_text, end_text


def _ensure_commission_export_indexes(conn: sqlite3.Connection) -> None:
    cursor = conn.cursor()
    cursor.execute(
        f"""
        CREATE INDEX IF NOT EXISTS idx_wechat_fund_flow_booking_type
        ON {WECHAT_FUND_FLOW_TABLE_NAME}(booking_time, transaction_type)
        """
    )
    cursor.execute(
        f"""
        CREATE INDEX IF NOT EXISTS idx_wechat_fund_flow_related_order_no
        ON {WECHAT_FUND_FLOW_TABLE_NAME}(related_order_no)
        """
    )
    cursor.execute(
        f"""
        CREATE INDEX IF NOT EXISTS idx_wechat_orders_order_no
        ON {WECHAT_ORDER_TABLE_NAME}(order_no)
        """
    )
    cursor.execute(
        f"""
        CREATE INDEX IF NOT EXISTS idx_wechat_orders_promotion_account_nickname
        ON {WECHAT_ORDER_TABLE_NAME}(promotion_account_nickname)
        """
    )
    conn.commit()


def _table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,),
    ).fetchone()
    return row is not None


def _get_alias_nicknames(conn: sqlite3.Connection, nickname_query: str) -> list[str]:
    keyword = (nickname_query or '').strip()
    if not keyword:
        return []
    if not _table_exists(conn, 'customer') or not _table_exists(conn, 'customer_alias'):
        return []

    rows = conn.execute(
        """
        SELECT DISTINCT TRIM(c.short_name) AS nickname
        FROM customer_alias ca
        JOIN customer c ON c.id = ca.customer_id
        WHERE TRIM(ca.alias) = ?
          AND TRIM(COALESCE(c.short_name, '')) <> ''
        ORDER BY nickname
        """,
        (keyword,),
    ).fetchall()
    return [row['nickname'] for row in rows]


def _month_values_between(start_text: str, end_text: str) -> list[tuple[int, int]]:
    start_dt = datetime.strptime(start_text[:10], '%Y-%m-%d')
    end_dt = datetime.strptime(end_text[:10], '%Y-%m-%d')
    values: list[tuple[int, int]] = []
    year = start_dt.year
    month = start_dt.month
    while (year, month) <= (end_dt.year, end_dt.month):
        values.append((year, month))
        if month == 12:
            year += 1
            month = 1
        else:
            month += 1
    return values


def _build_commission_month_text(start_text: str, end_text: str, short_year: bool = False) -> str:
    months = _month_values_between(start_text, end_text)
    if not months:
        return ''

    parts: list[str] = []
    previous_year: int | None = None
    has_multiple_years = len({year for year, _month in months}) > 1

    for year, month in months:
        if short_year:
            if has_multiple_years or previous_year != year:
                parts.append(f"{str(year)[-2:]}年{month}月")
            else:
                parts.append(f"{month}月")
        else:
            if has_multiple_years or previous_year != year:
                parts.append(f"{year}年{month}月")
            else:
                parts.append(f"{month}月")
        previous_year = year

    return ' '.join(parts)


def _safe_download_part(value: Any) -> str:
    text = str(value or '').strip()
    text = re.sub(r'[\\/:*?"<>|\r\n\t]+', '_', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text or '未命名'


def _build_commission_zip_name(prefix: str, start_text: str, end_text: str) -> str:
    month_text = _build_commission_month_text(start_text, end_text)
    safe_month = _safe_download_part(month_text)
    return f"{prefix}_{safe_month}.zip"


def _build_store_self_sale_zip_name(month_text: str, total_amount: float) -> str:
    safe_month = _safe_download_part(month_text)
    safe_amount = _safe_download_part(f"{total_amount:.2f}")
    return f"店铺自卖_{safe_month}_{safe_amount}.zip"


def _build_commission_detail_filename(nickname: str, amount_sum: float, start_text: str, end_text: str) -> str:
    month_text = _build_commission_month_text(start_text, end_text, short_year=True)
    return _safe_download_part(f"{nickname} {amount_sum:.2f} 澳柯{month_text}份.xlsx")


def _write_dataframe_excel(
    sheets: list[tuple[str, pd.DataFrame]],
    amount_columns: set[str] | None = None,
    text_columns: set[str] | None = None,
) -> BytesIO:
    amount_columns = amount_columns or set()
    text_columns = text_columns or set()

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in sheets:
            safe_sheet_name = re.sub(r'[\[\]:*?/\\]', '_', sheet_name)[:31] or 'Sheet1'
            df.to_excel(writer, index=False, sheet_name=safe_sheet_name)
            worksheet = writer.sheets[safe_sheet_name]
            for column_index, column_name in enumerate(df.columns, start=1):
                column_letter = get_column_letter(column_index)
                if column_name in amount_columns:
                    for cell in worksheet[column_letter][1:]:
                        cell.number_format = '0.00'
                if column_name in text_columns:
                    for cell in worksheet[column_letter]:
                        cell.number_format = '@'
            _auto_adjust_excel_columns(worksheet)
    output.seek(0)
    return output


def _query_commission_rows(
    conn: sqlite3.Connection,
    start_text: str,
    end_text: str,
    nickname_query: str | None,
) -> pd.DataFrame:
    _ensure_commission_export_indexes(conn)

    if not _table_exists(conn, WECHAT_FUND_FLOW_TABLE_NAME):
        raise ValueError(f'数据表不存在：{WECHAT_FUND_FLOW_TABLE_NAME}')
    if not _table_exists(conn, WECHAT_ORDER_TABLE_NAME):
        raise ValueError(f'数据表不存在：{WECHAT_ORDER_TABLE_NAME}')

    keyword = (nickname_query or '').strip()
    alias_nicknames = _get_alias_nicknames(conn, keyword)

    where_parts = [
        "REPLACE(CAST(f.booking_time AS TEXT), '/', '-') >= ?",
        "REPLACE(CAST(f.booking_time AS TEXT), '/', '-') <= ?",
        "f.transaction_type IN (?, ?)",
    ]
    params: list[Any] = [start_text, end_text, *COMMISSION_TRANSACTION_TYPES]

    if keyword:
        nickname_filters = ["带货账号昵称 LIKE ?"]
        params.append(f"%{keyword}%")
        if alias_nicknames:
            placeholders = ', '.join('?' for _ in alias_nicknames)
            nickname_filters.append(f"带货账号昵称 IN ({placeholders})")
            params.extend(alias_nicknames)
        where_parts.append('(' + ' OR '.join(nickname_filters) + ')')

    sql = f"""
        WITH commission_rows AS (
            SELECT
                CAST(f.flow_no AS TEXT) AS 流水单号,
                f.booking_time AS 记账时间,
                f.transaction_type AS 动账类型,
                f.income_expense_type AS 收支类型,
                COALESCE(CAST(f.amount AS REAL), 0) AS 收支金额,
                CAST(f.related_order_no AS TEXT) AS 关联订单号,
                COALESCE(
                    (
                        SELECT
                            CASE
                                WHEN COUNT(DISTINCT TRIM(o_agency.promotion_account_nickname)) = 1
                                THEN MAX(TRIM(o_agency.promotion_account_nickname))
                                ELSE NULL
                            END
                        FROM {WECHAT_ORDER_TABLE_NAME} o_agency
                        WHERE f.transaction_type = '带货机构服务费'
                          AND o_agency.order_no = f.related_order_no
                          AND o_agency.promotion_fee_channel = '机构服务费'
                          AND COALESCE(o_agency.promotion_fee_amount, 0) <> 0
                          AND TRIM(o_agency.promotion_account_nickname) NOT IN ('', '-')
                    ),
                    (
                        SELECT TRIM(o_creator.promotion_account_nickname)
                        FROM {WECHAT_ORDER_TABLE_NAME} o_creator
                        WHERE f.transaction_type <> '带货机构服务费'
                          AND o_creator.order_no = f.related_order_no
                          AND TRIM(o_creator.promotion_account_nickname) NOT IN ('', '-')
                          AND ABS(COALESCE(o_creator.promotion_fee_amount, 0) - COALESCE(f.amount, 0)) < 0.01
                        ORDER BY o_creator.id ASC
                        LIMIT 1
                    ),
                    (
                        SELECT TRIM(o_fallback.promotion_account_nickname)
                        FROM {WECHAT_ORDER_TABLE_NAME} o_fallback
                        WHERE f.transaction_type <> '带货机构服务费'
                          AND o_fallback.order_no = f.related_order_no
                          AND TRIM(o_fallback.promotion_account_nickname) NOT IN ('', '-')
                        ORDER BY o_fallback.id ASC
                        LIMIT 1
                    ),
                    ?
                ) AS 带货账号昵称
            FROM {WECHAT_FUND_FLOW_TABLE_NAME} f
            WHERE REPLACE(CAST(f.booking_time AS TEXT), '/', '-') >= ?
              AND REPLACE(CAST(f.booking_time AS TEXT), '/', '-') <= ?
              AND f.transaction_type IN (?, ?)
        )
        SELECT *
        FROM commission_rows
        {"WHERE " + " AND ".join(where_parts[3:]) if len(where_parts) > 3 else ""}
        ORDER BY 记账时间 ASC, 流水单号 ASC, 关联订单号 ASC
    """
    query_params = [COMMISSION_UNMATCHED_NICKNAME, *params]
    df = pd.read_sql_query(sql, conn, params=query_params)
    if not df.empty:
        df['收支金额'] = pd.to_numeric(df['收支金额'], errors='coerce').fillna(0)
    return df


def _build_commission_summary_frames(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if df.empty:
        creator_df = pd.DataFrame(columns=['达人名称', '佣金之和'])
        agency_df = pd.DataFrame(columns=['带货机构名称', '佣金之和'])
        invoice_df = pd.DataFrame(columns=['达人/客户', '应开金额'])
        return creator_df, agency_df, invoice_df

    creator_df = (
        df[df['动账类型'] == '达人佣金']
        .groupby('带货账号昵称', as_index=False)['收支金额']
        .sum()
        .rename(columns={'带货账号昵称': '达人名称', '收支金额': '佣金之和'})
        .sort_values('佣金之和', ascending=False)
    )
    agency_df = (
        df[df['动账类型'] == '带货机构服务费']
        .groupby('带货账号昵称', as_index=False)['收支金额']
        .sum()
        .rename(columns={'带货账号昵称': '带货机构名称', '收支金额': '佣金之和'})
        .sort_values('佣金之和', ascending=False)
    )
    invoice_df = (
        df.groupby('带货账号昵称', as_index=False)['收支金额']
        .sum()
        .rename(columns={'带货账号昵称': '达人/客户', '收支金额': '应开金额'})
        .sort_values('应开金额', ascending=False)
    )
    return creator_df, agency_df, invoice_df


def _query_store_self_sale_joined_rows(
    conn: sqlite3.Connection,
    start_text: str,
    end_text: str,
) -> pd.DataFrame:
    _ensure_commission_export_indexes(conn)

    if not _table_exists(conn, WECHAT_FUND_FLOW_TABLE_NAME):
        raise ValueError(f'数据表不存在：{WECHAT_FUND_FLOW_TABLE_NAME}')
    if not _table_exists(conn, WECHAT_ORDER_TABLE_NAME):
        raise ValueError(f'数据表不存在：{WECHAT_ORDER_TABLE_NAME}')

    sql = f"""
        SELECT
            f.id AS fund_id,
            CAST(f.flow_no AS TEXT) AS 资金流水_流水单号,
            f.booking_time AS 资金流水_记账时间,
            f.transaction_type AS 资金流水_动账类型,
            f.income_expense_type AS 资金流水_收支类型,
            COALESCE(CAST(f.amount AS REAL), 0) AS 资金流水_收支金额,
            f.account_balance AS 资金流水_账户余额,
            CAST(f.related_order_no AS TEXT) AS 资金流水_关联订单号,
            CAST(f.related_after_sales_no AS TEXT) AS 资金流水_关联售后单号,
            CAST(f.related_withdrawal_no AS TEXT) AS 资金流水_关联提现单号,
            CAST(f.related_policy_no AS TEXT) AS 资金流水_关联保单号,
            CAST(f.related_gift_no AS TEXT) AS 资金流水_关联礼物单号,
            f.detail AS 资金流水_详情,
            o.id AS order_row_id,
            CAST(o.order_no AS TEXT) AS 订单_订单号,
            o.order_created_at AS 订单_订单下单时间,
            o.order_shipped_at AS 订单_订单发货时间,
            o.order_received_at AS 订单_订单确认收货时间,
            o.order_settled_at AS 订单_订单完成结算时间,
            o.order_status AS 订单_订单状态,
            o.delivery_method AS 订单_发货方式,
            o.recipient_name AS 订单_收件人姓名,
            o.recipient_address AS 订单_收件人地址,
            o.province AS 订单_省,
            o.city AS 订单_市,
            o.district AS 订单_区,
            CAST(o.recipient_phone AS TEXT) AS 订单_收件人手机,
            o.buyer_note AS 订单_买家备注,
            o.seller_note AS 订单_商家备注,
            o.label_color AS 订单_打标颜色,
            o.product_total_amount AS 订单_商品总价,
            o.order_paid_amount AS 订单_订单实际支付金额,
            o.order_received_amount AS 订单_订单实际收款金额,
            o.order_shipping_fee AS 订单_订单运费,
            o.product_discount_amount AS 订单_商品优惠,
            o.cross_store_discount_amount AS 订单_跨店优惠,
            o.product_price_adjustment AS 订单_商品改价,
            o.points_deduction_amount AS 订单_积分抵扣,
            o.payment_method AS 订单_支付方式,
            o.payment_at AS 订单_支付时间,
            CAST(o.transaction_no AS TEXT) AS 订单_交易单号,
            o.logistics_company AS 订单_物流公司,
            CAST(o.tracking_no AS TEXT) AS 订单_快递单号,
            o.technical_service_fee AS 订单_技术服务费,
            o.technical_service_fee_refund_popularity_card AS 订单_技术服务费_人气卡返还,
            o.shipping_insurance_estimated_fee AS 订单_运费险预计投保费用,
            o.promotion_method AS 订单_带货方式,
            o.promotion_account_type AS 订单_带货账号类型,
            o.promotion_account_nickname AS 订单_带货账号昵称,
            o.promotion_fee_type AS 订单_带货费用类型,
            o.promotion_fee_amount AS 订单_带货费用,
            o.promotion_fee_channel AS 订单_带货费用渠道,
            o.promotion_commission_rate AS 订单_带货佣金率,
            CAST(o.gift_order_no AS TEXT) AS 订单_礼物单号,
            o.product_name AS 订单_商品名称,
            CAST(o.platform_product_code AS TEXT) AS 订单_商品编码_平台,
            CAST(o.custom_product_code AS TEXT) AS 订单_商品编码_自定义,
            CAST(o.custom_sku_code AS TEXT) AS 订单_SKU编码_自定义,
            o.product_attributes AS 订单_商品属性,
            o.product_unit_price AS 订单_商品价格_单件,
            o.product_actual_unit_price AS 订单_商品实际价格_单件,
            o.product_actual_total_price AS 订单_商品实际价格_总共,
            o.is_presale AS 订单_是否预售,
            o.product_quantity AS 订单_商品数量,
            o.platform_coupon_discount_amount AS 订单_商品平台券优惠,
            o.average_shipping_fee_per_item AS 订单_商品平均运费,
            o.customization_info AS 订单_定制信息,
            o.customization_preview_image AS 订单_定制预览图,
            o.product_delivery_status AS 订单_商品发货,
            o.product_after_sales_status AS 订单_商品售后,
            o.product_refunded_amount AS 订单_商品已退款金额
        FROM {WECHAT_FUND_FLOW_TABLE_NAME} f
        LEFT JOIN {WECHAT_ORDER_TABLE_NAME} o ON o.order_no = f.related_order_no
        WHERE REPLACE(CAST(f.booking_time AS TEXT), '/', '-') >= ?
          AND REPLACE(CAST(f.booking_time AS TEXT), '/', '-') <= ?
          AND f.transaction_type = ?
        ORDER BY f.booking_time ASC, f.flow_no ASC, f.related_order_no ASC, o.id ASC
    """
    df = pd.read_sql_query(sql, conn, params=[start_text, end_text, STORE_SELF_SALE_TRANSACTION_TYPE])
    if not df.empty:
        for column_name in {
            '资金流水_收支金额',
            '订单_商品实际价格_总共',
            '订单_订单实际收款金额',
            '订单_订单实际支付金额',
            '订单_商品总价',
        }:
            if column_name in df.columns:
                df[column_name] = pd.to_numeric(df[column_name], errors='coerce').fillna(0)
    return df


def _build_store_self_sale_frames(
    joined_df: pd.DataFrame,
    start_text: str,
    end_text: str,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, float]:
    month_text = _build_commission_month_text(start_text, end_text)
    detail_rows: list[dict[str, Any]] = []
    unmatched_rows: list[dict[str, Any]] = []

    if joined_df.empty:
        summary_df = pd.DataFrame([{
            '项目': '店铺自卖',
            '期间': month_text,
            '账期起点': start_text[:10],
            '账期终点': end_text[:10],
            '汇总金额': 0.0,
            '计入明细行数': 0,
            '未匹配/不计入资金流水条数': 0,
        }])
        return summary_df, pd.DataFrame(), pd.DataFrame(), 0.0

    for _fund_id, group in joined_df.groupby('fund_id', sort=False):
        first_row = group.iloc[0]
        order_mask = group['order_row_id'].notna()
        order_count = int(order_mask.sum())
        self_mask = order_mask & (group['订单_带货方式'].fillna('').astype(str).str.strip() == '-')
        self_count = int(self_mask.sum())

        base_info = {
            '匹配订单行数': order_count,
            '自卖订单行数': self_count,
        }

        if order_count == 0:
            row = first_row.to_dict()
            row.update(base_info)
            row['不计入原因'] = '未匹配订单'
            unmatched_rows.append(row)
            continue

        if self_count == 0:
            row = first_row.to_dict()
            row.update(base_info)
            row['不计入原因'] = '无带货方式=-的订单行'
            unmatched_rows.append(row)
            continue

        self_rows = group[self_mask]
        for _index, self_row in self_rows.iterrows():
            row = self_row.to_dict()
            row.update(base_info)
            row['计入口径'] = '计入-自卖商品实际价格'
            row['统计金额'] = float(row.get('订单_商品实际价格_总共') or 0)
            detail_rows.append(row)

    detail_df = pd.DataFrame(detail_rows)
    unmatched_df = pd.DataFrame(unmatched_rows)

    total_amount = 0.0 if detail_df.empty else float(pd.to_numeric(detail_df['统计金额'], errors='coerce').fillna(0).sum())
    summary_df = pd.DataFrame([{
        '项目': '店铺自卖',
        '期间': month_text,
        '账期起点': start_text[:10],
        '账期终点': end_text[:10],
        '汇总金额': total_amount,
        '计入明细行数': int(len(detail_df)),
        '未匹配/不计入资金流水条数': int(len(unmatched_df)),
    }])

    preferred_detail_columns = [
        '计入口径', '统计金额', '匹配订单行数', '自卖订单行数',
        'fund_id', '资金流水_流水单号', '资金流水_记账时间', '资金流水_动账类型',
        '资金流水_收支类型', '资金流水_收支金额', '资金流水_账户余额',
        '资金流水_关联订单号', '资金流水_关联售后单号', '资金流水_关联提现单号',
        '资金流水_关联保单号', '资金流水_关联礼物单号', '资金流水_详情',
        'order_row_id', '订单_订单号', '订单_订单下单时间', '订单_订单发货时间',
        '订单_订单确认收货时间', '订单_订单完成结算时间', '订单_订单状态',
        '订单_发货方式', '订单_收件人姓名', '订单_收件人地址', '订单_省',
        '订单_市', '订单_区', '订单_收件人手机', '订单_买家备注', '订单_商家备注',
        '订单_打标颜色', '订单_商品总价', '订单_订单实际支付金额',
        '订单_订单实际收款金额', '订单_订单运费', '订单_商品优惠',
        '订单_跨店优惠', '订单_商品改价', '订单_积分抵扣', '订单_支付方式',
        '订单_支付时间', '订单_交易单号', '订单_物流公司', '订单_快递单号',
        '订单_技术服务费', '订单_技术服务费_人气卡返还', '订单_运费险预计投保费用',
        '订单_带货方式', '订单_带货账号类型', '订单_带货账号昵称',
        '订单_带货费用类型', '订单_带货费用', '订单_带货费用渠道',
        '订单_带货佣金率', '订单_礼物单号', '订单_商品名称', '订单_商品编码_平台',
        '订单_商品编码_自定义', '订单_SKU编码_自定义', '订单_商品属性',
        '订单_商品价格_单件', '订单_商品实际价格_单件', '订单_商品实际价格_总共',
        '订单_是否预售', '订单_商品数量', '订单_商品平台券优惠',
        '订单_商品平均运费', '订单_定制信息', '订单_定制预览图',
        '订单_商品发货', '订单_商品售后', '订单_商品已退款金额',
    ]
    preferred_unmatched_columns = [
        '不计入原因', '匹配订单行数', '自卖订单行数',
        *[column for column in preferred_detail_columns if column not in {'计入口径', '统计金额', '匹配订单行数', '自卖订单行数'}],
    ]

    if detail_df.empty:
        detail_df = pd.DataFrame(columns=preferred_detail_columns)
    else:
        detail_df = detail_df[[column for column in preferred_detail_columns if column in detail_df.columns]]

    if unmatched_df.empty:
        unmatched_df = pd.DataFrame(columns=preferred_unmatched_columns)
    else:
        unmatched_df = unmatched_df[[column for column in preferred_unmatched_columns if column in unmatched_df.columns]]

    return summary_df, detail_df, unmatched_df, total_amount


def _store_self_sale_amount_columns() -> set[str]:
    return {
        '汇总金额', '统计金额', '资金流水_收支金额', '资金流水_账户余额',
        '订单_商品总价', '订单_订单实际支付金额', '订单_订单实际收款金额',
        '订单_订单运费', '订单_商品优惠', '订单_跨店优惠', '订单_商品改价',
        '订单_积分抵扣', '订单_技术服务费', '订单_技术服务费_人气卡返还',
        '订单_运费险预计投保费用', '订单_带货费用', '订单_商品价格_单件',
        '订单_商品实际价格_单件', '订单_商品实际价格_总共',
        '订单_商品平台券优惠', '订单_商品平均运费', '订单_商品已退款金额',
    }


def _store_self_sale_text_columns() -> set[str]:
    return {
        '资金流水_流水单号', '资金流水_关联订单号', '资金流水_关联售后单号',
        '资金流水_关联提现单号', '资金流水_关联保单号', '资金流水_关联礼物单号',
        '订单_订单号', '订单_收件人手机', '订单_交易单号', '订单_快递单号',
        '订单_礼物单号', '订单_商品编码_平台', '订单_商品编码_自定义',
        '订单_SKU编码_自定义',
    }


def _iter_dataframe_chunks(df: pd.DataFrame, chunk_size: int):
    if chunk_size <= 0:
        raise ValueError('导出切割行数配置不正确')
    for start_index in range(0, len(df), chunk_size):
        yield df.iloc[start_index:start_index + chunk_size].copy()


def _write_dataframe_excel_streaming(df: pd.DataFrame, sheet_name: str) -> BytesIO:
    output = BytesIO()
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title=sheet_name)
    worksheet.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        worksheet.append([None if pd.isna(value) else value for value in row])
    workbook.save(output)
    output.seek(0)
    return output


def _write_chunked_dataframe_excels_to_zip(
    zf: zipfile.ZipFile,
    df: pd.DataFrame,
    sheet_name: str,
    filename_prefix: str,
    amount_columns: set[str],
    text_columns: set[str],
    chunk_size: int,
) -> None:
    if df.empty:
        excel = _write_dataframe_excel_streaming(df, sheet_name)
        zf.writestr(f"{filename_prefix}_无数据.xlsx", excel.getvalue())
        return

    for part_index, chunk_df in enumerate(_iter_dataframe_chunks(df, chunk_size), start=1):
        excel = _write_dataframe_excel_streaming(chunk_df, sheet_name)
        zf.writestr(f"{filename_prefix}_第{part_index:03d}部分.xlsx", excel.getvalue())


def _write_store_self_sale_zip(
    archive_target: BytesIO | str | Path,
    start_date: str | None,
    end_date: str | None,
    chunk_size: int = EXPORT_ZIP_CHUNK_SIZE,
) -> str:
    start_text, end_text = _normalize_commission_date_range(start_date, end_date)
    db_path = _get_database_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        joined_df = _query_store_self_sale_joined_rows(conn, start_text, end_text)

    summary_df, detail_df, unmatched_df, total_amount = _build_store_self_sale_frames(
        joined_df,
        start_text,
        end_text,
    )
    month_text = _build_commission_month_text(start_text, end_text)
    amount_columns = _store_self_sale_amount_columns()
    text_columns = _store_self_sale_text_columns()

    with zipfile.ZipFile(archive_target, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        summary_excel = _write_dataframe_excel(
            [('店铺自卖', summary_df)],
            amount_columns=amount_columns,
            text_columns=text_columns,
        )
        zf.writestr('店铺自卖_汇总.xlsx', summary_excel.getvalue())
        _write_chunked_dataframe_excels_to_zip(
            zf,
            detail_df,
            '来源明细',
            '店铺自卖_来源明细',
            amount_columns,
            text_columns,
            chunk_size,
        )
        _write_chunked_dataframe_excels_to_zip(
            zf,
            unmatched_df,
            '未匹配',
            '店铺自卖_未匹配',
            amount_columns,
            text_columns,
            chunk_size,
        )

    return _build_store_self_sale_zip_name(month_text, total_amount)


def export_store_self_sale_zip(
    start_date: str | None,
    end_date: str | None,
    chunk_size: int = EXPORT_ZIP_CHUNK_SIZE,
) -> tuple[BytesIO, str]:
    archive = BytesIO()
    download_name = _write_store_self_sale_zip(archive, start_date, end_date, chunk_size)
    archive.seek(0)
    return archive, download_name


def export_store_self_sale_zip_file(
    output_path: str | Path,
    start_date: str | None,
    end_date: str | None,
    chunk_size: int = EXPORT_ZIP_CHUNK_SIZE,
) -> str:
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    start_text, end_text = _normalize_commission_date_range(start_date, end_date)
    db_path = _get_database_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)

    start_datetime = datetime.strptime(start_text, '%Y-%m-%d %H:%M:%S')
    end_datetime = datetime.strptime(end_text, '%Y-%m-%d %H:%M:%S')
    total_amount = 0.0
    detail_row_count = 0
    unmatched_row_count = 0
    detail_part_index = 1
    unmatched_part_index = 1

    with sqlite3.connect(db_path) as conn, zipfile.ZipFile(
        output_file,
        'w',
        compression=zipfile.ZIP_DEFLATED,
    ) as zf:
        conn.row_factory = sqlite3.Row
        current_start = start_datetime
        while current_start <= end_datetime:
            if current_start.month == 12:
                next_month = datetime(current_start.year + 1, 1, 1)
            else:
                next_month = datetime(current_start.year, current_start.month + 1, 1)
            current_end = min(end_datetime, next_month - timedelta(seconds=1))

            joined_df = _query_store_self_sale_joined_rows(
                conn,
                current_start.strftime('%Y-%m-%d %H:%M:%S'),
                current_end.strftime('%Y-%m-%d %H:%M:%S'),
            )
            _summary_df, detail_df, unmatched_df, month_amount = _build_store_self_sale_frames(
                joined_df,
                current_start.strftime('%Y-%m-%d %H:%M:%S'),
                current_end.strftime('%Y-%m-%d %H:%M:%S'),
            )
            total_amount += month_amount
            detail_row_count += len(detail_df)
            unmatched_row_count += len(unmatched_df)

            for chunk_df in _iter_dataframe_chunks(detail_df, chunk_size):
                excel = _write_dataframe_excel_streaming(chunk_df, '来源明细')
                zf.writestr(
                    f'店铺自卖_来源明细_第{detail_part_index:03d}部分.xlsx',
                    excel.getvalue(),
                )
                detail_part_index += 1

            for chunk_df in _iter_dataframe_chunks(unmatched_df, chunk_size):
                excel = _write_dataframe_excel_streaming(chunk_df, '未匹配')
                zf.writestr(
                    f'店铺自卖_未匹配_第{unmatched_part_index:03d}部分.xlsx',
                    excel.getvalue(),
                )
                unmatched_part_index += 1

            del joined_df, detail_df, unmatched_df
            current_start = next_month

        if detail_part_index == 1:
            empty_excel = _write_dataframe_excel_streaming(pd.DataFrame(), '来源明细')
            zf.writestr('店铺自卖_来源明细_无数据.xlsx', empty_excel.getvalue())
        if unmatched_part_index == 1:
            empty_excel = _write_dataframe_excel_streaming(pd.DataFrame(), '未匹配')
            zf.writestr('店铺自卖_未匹配_无数据.xlsx', empty_excel.getvalue())

        month_text = _build_commission_month_text(start_text, end_text)
        summary_df = pd.DataFrame([{
            '项目': '店铺自卖',
            '期间': month_text,
            '账期起点': start_text[:10],
            '账期终点': end_text[:10],
            '汇总金额': total_amount,
            '计入明细行数': detail_row_count,
            '未匹配/不计入资金流水条数': unmatched_row_count,
        }])
        summary_excel = _write_dataframe_excel(
            [('店铺自卖', summary_df)],
            amount_columns=_store_self_sale_amount_columns(),
            text_columns=_store_self_sale_text_columns(),
        )
        zf.writestr('店铺自卖_汇总.xlsx', summary_excel.getvalue())

    return _build_store_self_sale_zip_name(month_text, total_amount)


def export_commission_summary_zip(
    start_date: str | None,
    end_date: str | None,
    nickname_query: str | None = None,
) -> tuple[BytesIO, str]:
    start_text, end_text = _normalize_commission_date_range(start_date, end_date)
    db_path = _get_database_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        df = _query_commission_rows(conn, start_text, end_text, nickname_query)

    creator_df, agency_df, invoice_df = _build_commission_summary_frames(df)
    month_text = _build_commission_month_text(start_text, end_text)
    safe_month_text = _safe_download_part(month_text)

    vba_excel = _write_dataframe_excel(
        [
            ('主播佣金', creator_df),
            ('团长佣金', agency_df),
        ],
        amount_columns={'佣金之和'},
    )

    invoice_export_df = invoice_df.copy()
    invoice_export_df['店铺/平台'] = '澳柯'
    invoice_export_df['期间'] = month_text
    invoice_export_df['归属'] = '澳柯'
    invoice_export_df['账期起点'] = start_text[:10]
    invoice_export_df['账期终点'] = end_text[:10]
    invoice_export_df = invoice_export_df[
        ['达人/客户', '应开金额', '店铺/平台', '期间', '归属', '账期起点', '账期终点']
    ]
    invoice_excel = _write_dataframe_excel(
        [('应开金额导入', invoice_export_df[['达人/客户', '应开金额']])],
        amount_columns={'应开金额'},
        text_columns={'达人/客户'},
    )

    archive = BytesIO()
    with zipfile.ZipFile(archive, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"澳柯视频号佣金汇总_老版_{safe_month_text}.xlsx", vba_excel.getvalue())
        zf.writestr(f"澳柯视频号应开金额导入_{safe_month_text}.xlsx", invoice_excel.getvalue())
    archive.seek(0)
    return archive, _build_commission_zip_name('澳柯视频号佣金汇总', start_text, end_text)


def export_commission_detail_zip(
    start_date: str | None,
    end_date: str | None,
    nickname_query: str | None = None,
) -> tuple[BytesIO, str]:
    start_text, end_text = _normalize_commission_date_range(start_date, end_date)
    db_path = _get_database_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        df = _query_commission_rows(conn, start_text, end_text, nickname_query)

    archive = BytesIO()
    used_names: dict[str, int] = {}
    with zipfile.ZipFile(archive, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        if df.empty:
            empty_excel = _write_dataframe_excel(
                [('明细', pd.DataFrame(columns=COMMISSION_DETAIL_COLUMNS))],
                amount_columns={'收支金额'},
                text_columns={'流水单号', '关联订单号', '带货账号昵称'},
            )
            zf.writestr('无匹配佣金明细.xlsx', empty_excel.getvalue())
        else:
            combined_detail_df = df[COMMISSION_DETAIL_COLUMNS].copy()
            combined_detail_excel = _write_dataframe_excel(
                [('明细', combined_detail_df)],
                amount_columns={'收支金额'},
                text_columns={'流水单号', '关联订单号', '带货账号昵称'},
            )
            month_text = _safe_download_part(
                _build_commission_month_text(start_text, end_text)
            )
            zf.writestr(
                f'澳柯视频号佣金合并明细总表_{month_text}.xlsx',
                combined_detail_excel.getvalue(),
            )

            grouped = df.groupby('带货账号昵称', sort=True)
            for nickname, group_df in grouped:
                amount_sum = float(group_df['收支金额'].sum())
                detail_df = group_df[COMMISSION_DETAIL_COLUMNS].copy()
                detail_excel = _write_dataframe_excel(
                    [('明细', detail_df)],
                    amount_columns={'收支金额'},
                    text_columns={'流水单号', '关联订单号', '带货账号昵称'},
                )
                filename = _build_commission_detail_filename(str(nickname), amount_sum, start_text, end_text)
                base_name = filename[:-5] if filename.lower().endswith('.xlsx') else filename
                candidate = filename
                used_count = used_names.get(candidate, 0)
                if used_count:
                    candidate = f"{base_name}_{used_count + 1}.xlsx"
                used_names[filename] = used_count + 1
                zf.writestr(candidate, detail_excel.getvalue())
    archive.seek(0)
    return archive, _build_commission_zip_name('澳柯视频号佣金明细', start_text, end_text)



def _build_create_table_sql(table_name: str, column_types: dict[str, str]) -> str:
    """根据字段类型定义生成 CREATE TABLE SQL。"""
    column_defs: list[str] = []

    for column_name, column_type in column_types.items():
        if column_name == 'order_no':
            column_defs.append(f'{column_name} {column_type} NOT NULL')
        else:
            column_defs.append(f'{column_name} {column_type}')

    columns_sql = ',\n    '.join(column_defs)
    return f'''CREATE TABLE IF NOT EXISTS {table_name} (\n    id INTEGER PRIMARY KEY AUTOINCREMENT,\n    {columns_sql}\n);'''


def _get_existing_table_columns(conn: sqlite3.Connection, table_name: str) -> list[str]:
    """读取 SQLite 现有表字段。"""
    cursor = conn.cursor()
    cursor.execute(f"PRAGMA table_info({table_name})")
    rows = cursor.fetchall()
    return [row[1] for row in rows]



def _sync_order_table_columns(conn: sqlite3.Connection) -> list[str]:
    """若订单表缺少新字段，则自动补列。"""
    existing_columns = _get_existing_table_columns(conn, WECHAT_ORDER_TABLE_NAME)
    added_columns: list[str] = []
    cursor = conn.cursor()

    for column_name, column_type in ORDER_COLUMN_TYPES.items():
        if column_name not in existing_columns:
            cursor.execute(
                f"ALTER TABLE {WECHAT_ORDER_TABLE_NAME} ADD COLUMN {column_name} {column_type}"
            )
            added_columns.append(column_name)

    if added_columns:
        conn.commit()

    return added_columns




def _clean_numeric_value(value: Any) -> Any:
    """清洗金额/比例类字段，尽量转成可写入 SQLite REAL 的值。"""
    if value is None or pd.isna(value):
        return None

    if isinstance(value, (int, float)):
        return value

    text = str(value).strip()
    if not text:
        return None

    text = text.replace(',', '')
    text = text.replace('¥', '')
    text = text.replace('￥', '')
    text = text.replace('%', '')
    text = re.sub(r'\s+', '', text)

    try:
        return float(text)
    except ValueError:
        return None


# ---------------- 新增文本字段清洗和 dtype mapping 辅助函数 ----------------

def _clean_text_value(value: Any) -> Any:
    """清洗文本类字段，尽量避免长数字被写成科学计数法。"""
    if value is None or pd.isna(value):
        return None

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if value.is_integer():
            return format(value, '.0f')
        return format(value, 'f').rstrip('0').rstrip('.')

    return str(value).strip()


def _build_text_dtype_mapping(excel_buffer: BytesIO) -> dict[str, str]:
    """先读取表头，为关键文本字段构造 dtype 映射，避免 Excel 长数字被自动转成浮点数。"""
    header_df = pd.read_excel(excel_buffer, nrows=0)
    excel_buffer.seek(0)

    normalized_text_columns = {normalize_header_text(col) for col in TEXT_SOURCE_COLUMNS}
    dtype_mapping: dict[str, str] = {}

    for raw_column_name in header_df.columns.tolist():
        if normalize_header_text(raw_column_name) in normalized_text_columns:
            dtype_mapping[raw_column_name] = 'string'

    return dtype_mapping


def _ensure_order_table_exists() -> tuple[bool, str]:
    """确保微信订单表存在；不存在则自动创建，已存在则自动补齐缺失字段。"""
    db_path = _get_database_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)

    table_exists = False
    with sqlite3.connect(db_path) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
            (WECHAT_ORDER_TABLE_NAME,),
        )
        table_exists = cursor.fetchone() is not None

        if not table_exists:
            create_sql = _build_create_table_sql(WECHAT_ORDER_TABLE_NAME, ORDER_COLUMN_TYPES)
            cursor.execute(create_sql)
            conn.commit()
            print(f'[wechat_shop] 数据库路径：{db_path}')
            return True, f'订单表不存在，已自动创建：{WECHAT_ORDER_TABLE_NAME}（数据库：{db_path.name}）'

        added_columns = _sync_order_table_columns(conn)
        print(f'[wechat_shop] 数据库路径：{db_path}')
        if added_columns:
            return False, f'订单表已存在：{WECHAT_ORDER_TABLE_NAME}，并已补齐字段：{", ".join(added_columns)}（数据库：{db_path.name}）'

    return False, f'订单表已存在：{WECHAT_ORDER_TABLE_NAME}（数据库：{db_path.name}）'


def _prepare_orders_dataframe_for_db(df: pd.DataFrame) -> pd.DataFrame:
    """按数据库表结构整理订单 DataFrame，确保字段齐全、顺序一致、类型尽量可落库。"""
    db_columns = list(ORDER_COLUMN_TYPES.keys())
    prepared_df = df.copy()

    for column_name in db_columns:
        if column_name not in prepared_df.columns:
            prepared_df[column_name] = None

    prepared_df = prepared_df[db_columns]

    for column_name, column_type in ORDER_COLUMN_TYPES.items():
        # 强制指定部分字段为文本（即使 schema 写成 REAL 也纠正）
        if column_name in {
            'transaction_no',
            'tracking_no',
            'gift_order_no',
            'custom_product_code',
            'custom_sku_code',
        }:
            prepared_df[column_name] = prepared_df[column_name].apply(_clean_text_value)
            continue

        # 修正费率字段：应为文本，不是数值
        if column_name == 'promoter_commission_rate':
            prepared_df[column_name] = prepared_df[column_name].apply(_clean_text_value)
            continue

        if column_type == 'REAL':
            prepared_df[column_name] = prepared_df[column_name].apply(_clean_numeric_value)
        elif column_type == 'INTEGER':
            prepared_df[column_name] = (
                prepared_df[column_name]
                .apply(_clean_numeric_value)
                .apply(lambda x: int(x) if x is not None else None)
            )
        else:
            prepared_df[column_name] = prepared_df[column_name].apply(_clean_text_value)

    return prepared_df




def _build_order_dedup_key(row: pd.Series) -> str | None:
    """构造商品维度防重键：订单号 + 平台商品编码 + 商品属性。

    其中：
    - 订单号、平台商品编码必须存在
    - 商品属性允许为空；为空时也视为一种合法商品形态
    """
    required_columns = ['order_no', 'platform_product_code']
    required_parts: list[str] = []

    for column_name in required_columns:
        value = row.get(column_name)
        cleaned_value = _clean_text_value(value)
        if cleaned_value is None or str(cleaned_value).strip() == '':
            return None
        required_parts.append(str(cleaned_value).strip())

    product_attributes = _clean_text_value(row.get('product_attributes'))
    product_attributes_part = '' if product_attributes is None else str(product_attributes).strip()

    return '||'.join([required_parts[0], required_parts[1], product_attributes_part])


def _get_existing_order_keys(conn: sqlite3.Connection) -> set[str]:
    """读取数据库中已存在的商品维度防重键，用于防重导入。"""
    cursor = conn.cursor()
    cursor.execute(
        f"SELECT order_no, platform_product_code, product_attributes FROM {WECHAT_ORDER_TABLE_NAME}"
    )
    rows = cursor.fetchall()

    existing_keys: set[str] = set()
    for row in rows:
        row_dict = {
            'order_no': row[0],
            'platform_product_code': row[1],
            'product_attributes': row[2],
        }
        dedup_key = _build_order_dedup_key(pd.Series(row_dict))
        if dedup_key:
            existing_keys.add(dedup_key)

    return existing_keys




def _deduplicate_orders_df(merged_df: pd.DataFrame, conn: sqlite3.Connection) -> tuple[pd.DataFrame, int, int]:
    """按商品维度防重：订单号 + 平台商品编码 + 商品属性。"""
    missing_key_columns = [col for col in ORDER_DEDUP_KEY_COLUMNS if col not in merged_df.columns]
    if missing_key_columns:
        return merged_df, 0, 0

    working_df = merged_df.copy()
    working_df['_dedup_key'] = working_df.apply(_build_order_dedup_key, axis=1)

    valid_key_mask = working_df['_dedup_key'].notna() & (working_df['_dedup_key'].astype(str).str.strip() != '')
    batch_valid_df = working_df[valid_key_mask].copy()
    batch_unique_df = batch_valid_df.drop_duplicates(subset=['_dedup_key'], keep='first')
    batch_duplicate_count = int(len(batch_valid_df) - len(batch_unique_df))

    empty_key_df = working_df[~valid_key_mask].copy()
    combined_df = pd.concat([batch_unique_df, empty_key_df], ignore_index=True)

    existing_keys = _get_existing_order_keys(conn)
    if not existing_keys:
        final_df = combined_df.drop(columns=['_dedup_key'], errors='ignore')
        return final_df, batch_duplicate_count, 0

    valid_after_batch_mask = combined_df['_dedup_key'].notna() & (combined_df['_dedup_key'].astype(str).str.strip() != '')
    valid_after_batch_df = combined_df[valid_after_batch_mask].copy()
    non_key_df = combined_df[~valid_after_batch_mask].copy()

    db_filtered_df = valid_after_batch_df[~valid_after_batch_df['_dedup_key'].isin(existing_keys)].copy()
    db_duplicate_count = int(len(valid_after_batch_df) - len(db_filtered_df))

    final_df = pd.concat([db_filtered_df, non_key_df], ignore_index=True)
    final_df = final_df.drop(columns=['_dedup_key'], errors='ignore')
    return final_df, batch_duplicate_count, db_duplicate_count



def _write_orders_to_db(dataframes: list[pd.DataFrame]) -> tuple[int, str]:
    """将订单数据写入 SQLite，返回写入行数和结果消息。"""
    if not dataframes:
        return 0, '没有可写入的数据'

    db_path = _get_database_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)

    prepared_frames = [_prepare_orders_dataframe_for_db(df) for df in dataframes]
    merged_df = pd.concat(prepared_frames, ignore_index=True)

    with sqlite3.connect(db_path) as conn:
        _sync_order_table_columns(conn)
        deduped_df, batch_duplicate_count, db_duplicate_count = _deduplicate_orders_df(merged_df, conn)

        if deduped_df.empty:
            message_parts = ['没有可写入的新数据']
            if batch_duplicate_count > 0:
                message_parts.append(f'本次文件内重复商品行已跳过：{batch_duplicate_count} 条')
            if db_duplicate_count > 0:
                message_parts.append(f'数据库中已存在商品行已跳过：{db_duplicate_count} 条')
            return 0, '；'.join(message_parts) + f'（数据库：{db_path.name}）'

        deduped_df.to_sql(WECHAT_ORDER_TABLE_NAME, conn, if_exists='append', index=False)

    message_parts = [f'成功写入 {len(deduped_df)} 行数据']
    if batch_duplicate_count > 0:
        message_parts.append(f'本次文件内重复商品行已跳过：{batch_duplicate_count} 条')
    if db_duplicate_count > 0:
        message_parts.append(f'数据库中已存在商品行已跳过：{db_duplicate_count} 条')

    return int(len(deduped_df)), '；'.join(message_parts) + f'（数据库：{db_path.name}）'


# ===================== 资金流水相关辅助函数 =====================

def _build_fund_flow_create_table_sql(table_name: str, column_types: dict[str, str]) -> str:
    """根据资金流水字段类型定义生成 CREATE TABLE SQL。"""
    column_defs: list[str] = []

    for column_name, column_type in column_types.items():
        column_defs.append(f'{column_name} {column_type}')

    columns_sql = ',\n    '.join(column_defs)
    return f'''CREATE TABLE IF NOT EXISTS {table_name} (\n    id INTEGER PRIMARY KEY AUTOINCREMENT,\n    {columns_sql}\n);'''



def _ensure_fund_flow_table_exists() -> tuple[bool, str]:
    """确保微信资金流水表存在；不存在则自动创建，已存在则自动补齐缺失字段。"""
    db_path = _get_database_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)

    table_exists = False
    with sqlite3.connect(db_path) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
            (WECHAT_FUND_FLOW_TABLE_NAME,),
        )
        table_exists = cursor.fetchone() is not None

        if not table_exists:
            create_sql = _build_fund_flow_create_table_sql(WECHAT_FUND_FLOW_TABLE_NAME, FUND_FLOW_COLUMN_TYPES)
            cursor.execute(create_sql)
            conn.commit()
            return True, f'资金流水表不存在，已自动创建：{WECHAT_FUND_FLOW_TABLE_NAME}（数据库：{db_path.name}）'

        existing_columns = _get_existing_table_columns(conn, WECHAT_FUND_FLOW_TABLE_NAME)
        added_columns: list[str] = []
        for column_name, column_type in FUND_FLOW_COLUMN_TYPES.items():
            if column_name not in existing_columns:
                cursor.execute(
                    f"ALTER TABLE {WECHAT_FUND_FLOW_TABLE_NAME} ADD COLUMN {column_name} {column_type}"
                )
                added_columns.append(column_name)

        if added_columns:
            conn.commit()
            return False, f'资金流水表已存在：{WECHAT_FUND_FLOW_TABLE_NAME}，并已补齐字段：{", ".join(added_columns)}（数据库：{db_path.name}）'

    return False, f'资金流水表已存在：{WECHAT_FUND_FLOW_TABLE_NAME}（数据库：{db_path.name}）'



def _prepare_fund_flow_dataframe_for_db(df: pd.DataFrame) -> pd.DataFrame:
    """按数据库表结构整理资金流水 DataFrame。"""
    db_columns = list(FUND_FLOW_COLUMN_TYPES.keys())
    prepared_df = df.copy()

    for column_name in db_columns:
        if column_name not in prepared_df.columns:
            prepared_df[column_name] = None

    prepared_df = prepared_df[db_columns]

    for column_name, column_type in FUND_FLOW_COLUMN_TYPES.items():
        if column_name in {
            'flow_no',
            'related_order_no',
            'related_after_sales_no',
            'related_withdrawal_no',
            'related_policy_no',
            'related_gift_no',
        }:
            prepared_df[column_name] = prepared_df[column_name].apply(_clean_text_value)
            continue

        if column_type == 'REAL':
            prepared_df[column_name] = prepared_df[column_name].apply(_clean_numeric_value)
        else:
            prepared_df[column_name] = prepared_df[column_name].apply(_clean_text_value)

    return prepared_df



def _build_fund_flow_dedup_key(row: pd.Series) -> str | None:
    """构造资金流水防重键：流水单号 + 记账时间 + 动帐类型 + 关联订单号。"""
    parts: list[str] = []

    for column_name in FUND_FLOW_DEDUP_KEY_COLUMNS:
        value = row.get(column_name)
        cleaned_value = _clean_text_value(value)
        if cleaned_value is None or str(cleaned_value).strip() == '':
            if column_name == 'related_order_no':
                cleaned_value = FUND_FLOW_EMPTY_DEDUP_VALUE
            else:
                return None
        parts.append(str(cleaned_value).strip())

    return '||'.join(parts)


def _ensure_fund_flow_dedup_index(conn: sqlite3.Connection) -> None:
    """Use SQLite as the final duplicate guard, including blank order numbers."""
    conn.execute(
        f"""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_wechat_fund_flow_dedup
        ON {WECHAT_FUND_FLOW_TABLE_NAME}(
            flow_no,
            booking_time,
            transaction_type,
            COALESCE(related_order_no, '')
        )
        """
    )
    conn.commit()



def _get_existing_fund_flow_keys(conn: sqlite3.Connection) -> set[str]:
    """读取数据库中已存在的资金流水防重键。"""
    cursor = conn.cursor()
    cursor.execute(
        f"SELECT flow_no, booking_time, transaction_type, related_order_no FROM {WECHAT_FUND_FLOW_TABLE_NAME}"
    )
    rows = cursor.fetchall()

    existing_keys: set[str] = set()
    for row in rows:
        row_dict = {
            'flow_no': row[0],
            'booking_time': row[1],
            'transaction_type': row[2],
            'related_order_no': row[3],
        }
        dedup_key = _build_fund_flow_dedup_key(pd.Series(row_dict))
        if dedup_key:
            existing_keys.add(dedup_key)

    return existing_keys



def _deduplicate_fund_flow_df(merged_df: pd.DataFrame, conn: sqlite3.Connection) -> tuple[pd.DataFrame, int, int]:
    """按资金流水维度防重：流水单号 + 记账时间 + 动帐类型 + 关联订单号。"""
    missing_key_columns = [col for col in FUND_FLOW_DEDUP_KEY_COLUMNS if col not in merged_df.columns]
    if missing_key_columns:
        return merged_df, 0, 0

    working_df = merged_df.copy()
    working_df['_dedup_key'] = working_df.apply(_build_fund_flow_dedup_key, axis=1)

    valid_key_mask = working_df['_dedup_key'].notna() & (working_df['_dedup_key'].astype(str).str.strip() != '')
    batch_valid_df = working_df[valid_key_mask].copy()
    batch_unique_df = batch_valid_df.drop_duplicates(subset=['_dedup_key'], keep='first')
    batch_duplicate_count = int(len(batch_valid_df) - len(batch_unique_df))

    empty_key_df = working_df[~valid_key_mask].copy()
    combined_df = pd.concat([batch_unique_df, empty_key_df], ignore_index=True)
    final_df = combined_df.drop(columns=['_dedup_key'], errors='ignore')
    return final_df, batch_duplicate_count, 0



def _write_fund_flow_to_db(dataframes: list[pd.DataFrame]) -> tuple[int, str]:
    """将资金流水数据写入 SQLite，返回写入行数和结果消息。"""
    if not dataframes:
        return 0, '没有可写入的资金流水数据'

    db_path = _get_database_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)

    prepared_frames = [_prepare_fund_flow_dataframe_for_db(df) for df in dataframes]
    merged_df = pd.concat(prepared_frames, ignore_index=True)

    with sqlite3.connect(db_path) as conn:
        existing_columns = _get_existing_table_columns(conn, WECHAT_FUND_FLOW_TABLE_NAME)
        for column_name, column_type in FUND_FLOW_COLUMN_TYPES.items():
            if column_name not in existing_columns:
                conn.execute(
                    f"ALTER TABLE {WECHAT_FUND_FLOW_TABLE_NAME} ADD COLUMN {column_name} {column_type}"
                )
        conn.commit()

        _ensure_fund_flow_dedup_index(conn)
        deduped_df, batch_duplicate_count, _ = _deduplicate_fund_flow_df(merged_df, conn)
        db_duplicate_count = 0

        if deduped_df.empty:
            message_parts = ['没有可写入的新资金流水数据']
            if batch_duplicate_count > 0:
                message_parts.append(f'本次文件内重复流水已跳过：{batch_duplicate_count} 条')
            if db_duplicate_count > 0:
                message_parts.append(f'数据库中已存在流水已跳过：{db_duplicate_count} 条')
            return 0, '；'.join(message_parts) + f'（数据库：{db_path.name}）'

        db_columns = list(FUND_FLOW_COLUMN_TYPES.keys())
        placeholders = ', '.join(['?'] * len(db_columns))
        column_sql = ', '.join(db_columns)
        insert_sql = (
            f"INSERT OR IGNORE INTO {WECHAT_FUND_FLOW_TABLE_NAME} "
            f"({column_sql}) VALUES ({placeholders})"
        )
        rows = [
            tuple(None if pd.isna(value) else value for value in row)
            for row in deduped_df[db_columns].itertuples(index=False, name=None)
        ]
        before_changes = conn.total_changes
        conn.executemany(insert_sql, rows)
        conn.commit()
        written_rows = int(conn.total_changes - before_changes)
        db_duplicate_count = int(len(rows) - written_rows)
        deduped_df = deduped_df.head(written_rows).copy()

    message_parts = [f'成功写入 {len(deduped_df)} 行资金流水数据']
    if batch_duplicate_count > 0:
        message_parts.append(f'本次文件内重复流水已跳过：{batch_duplicate_count} 条')
    if db_duplicate_count > 0:
        message_parts.append(f'数据库中已存在流水已跳过：{db_duplicate_count} 条')

    return int(len(deduped_df)), '；'.join(message_parts) + f'（数据库：{db_path.name}）'



def read_fund_flow_excel_files(files: list[FileStorage]) -> dict[str, Any]:
    """读取微信资金流水 Excel 文件，校验、映射、写库。"""
    valid_files: list[FileStorage] = []
    invalid_files: list[str] = []
    failed_files: list[dict[str, str]] = []
    file_summaries: list[dict[str, Any]] = []
    prepared_dataframes: list[pd.DataFrame] = []
    base_columns: list[str] | None = None
    base_filename: str | None = None
    has_structure_mismatch = False

    for file_obj in files:
        filename = _get_upload_source_filename(file_obj)
        if not filename:
            invalid_files.append('未命名文件')
            continue

        if not is_excel_filename(filename):
            invalid_files.append(filename)
            continue

        valid_files.append(file_obj)

    if not valid_files:
        return {
            'success': False,
            'message': '没有有效的Excel文件（.xlsx/.xls）',
            'file_count': 0,
            'files': [],
            'invalid_files': invalid_files,
            'failed_files': failed_files,
        }

    for file_obj in valid_files:
        filename = _get_upload_source_filename(file_obj)

        try:
            file_bytes = _read_upload_source_bytes(file_obj)
            excel_buffer = BytesIO(file_bytes)
            dtype_mapping = _build_text_dtype_mapping(excel_buffer)
            df = pd.read_excel(excel_buffer, dtype=dtype_mapping if dtype_mapping else None)
            df.columns = [normalize_header_text(col) for col in df.columns.tolist()]
            current_columns = normalize_columns(df.columns.tolist())

            normalized_required_columns = [normalize_header_text(col) for col in FUND_FLOW_REQUIRED_COLUMNS]
            normalized_column_mapping = {
                normalize_header_text(chinese_name): english_name
                for chinese_name, english_name in FUND_FLOW_COLUMN_MAPPING.items()
            }

            missing_required = [col for col in normalized_required_columns if col not in current_columns]
            if missing_required:
                failed_files.append({
                    'filename': filename,
                    'error': f"缺少必需字段：{', '.join(missing_required)}",
                })
                continue

            df = df.rename(columns=normalized_column_mapping)
            mapped_columns = normalize_columns(df.columns.tolist())

            current_summary = {
                'filename': filename,
                'row_count': int(len(df)),
                'column_count': int(len(df.columns)),
                'columns': mapped_columns,
            }

            if base_columns is None:
                base_columns = current_columns
                base_filename = filename
                file_summaries.append(current_summary)
                prepared_dataframes.append(df)
            else:
                columns_match, missing_columns, extra_columns = check_columns_match(
                    base_columns,
                    current_columns,
                )

                if not columns_match:
                    error_parts: list[str] = []
                    if missing_columns:
                        error_parts.append(f"缺少列：{', '.join(missing_columns)}")
                    if extra_columns:
                        error_parts.append(f"多出列：{', '.join(extra_columns)}")
                    if not missing_columns and not extra_columns:
                        error_parts.append('列名顺序不一致')

                    failed_files.append({
                        'filename': filename,
                        'error': f"列结构不一致（基准文件：{base_filename}；{'；'.join(error_parts)}）",
                    })
                    has_structure_mismatch = True
                    file_summaries.append(current_summary)
                    continue

                file_summaries.append(current_summary)
                prepared_dataframes.append(df)
        except Exception as exc:
            failed_files.append({
                'filename': filename,
                'error': str(exc),
            })
        finally:
            _reset_upload_source(file_obj)

    success_count = len(file_summaries)

    if has_structure_mismatch or failed_files:
        return _build_import_precheck_failed_response(
            '本次资金流水导入',
            file_summaries,
            invalid_files,
            failed_files,
        )

    if success_count == 0:
        message_parts: list[str] = ['资金流水文件已接收，但读取失败']

        if failed_files:
            message_parts.append('')
            message_parts.append('读取失败：')
            for failed in failed_files:
                message_parts.append(f"- {failed['filename']}（{failed['error']}）")

        if invalid_files:
            message_parts.append('')
            message_parts.append(f"无效文件：{'，'.join(invalid_files)}")

        return {
            'success': False,
            'message': '\n'.join(message_parts),
            'file_count': 0,
            'files': [],
            'invalid_files': invalid_files,
            'failed_files': failed_files,
        }

    try:
        table_created, table_message = _ensure_fund_flow_table_exists()
        written_rows, write_message = _write_fund_flow_to_db(prepared_dataframes)
    except Exception as exc:
        db_path = _get_database_path()
        return {
            'success': False,
            'message': f'写入资金流水数据库失败：{str(exc)}\n数据库路径：{db_path}',
            'file_count': success_count,
            'files': file_summaries,
            'invalid_files': invalid_files,
            'failed_files': failed_files,
        }

    status_warning = _try_update_data_status('fund_flows') if written_rows > 0 else ''
    message = f'成功读取 {success_count} 个资金流水文件'
    if table_message:
        message += f'\n{table_message}'
    if write_message:
        message += f'\n{write_message}'
    if status_warning:
        message += f'\n{status_warning}'

    return {
        'success': True,
        'message': message,
        'file_count': success_count,
        'files': file_summaries,
        'invalid_files': invalid_files,
        'failed_files': failed_files,
        'table_created': table_created,
        'table_message': table_message,
        'written_rows': written_rows,
        'write_message': write_message,
    }


def read_order_excel_files(files: list[FileStorage]) -> dict[str, Any]:
    """
    第一阶段：
    接收前端上传的多个订单 Excel 文件，
    先读取每个文件的基础信息，不做数据库写入。

    当前返回：
    - success
    - message
    - file_count
    - files
    - invalid_files
    - failed_files
    """
    valid_files: list[FileStorage] = []
    invalid_files: list[str] = []
    failed_files: list[dict[str, str]] = []
    file_summaries: list[dict[str, Any]] = []
    prepared_dataframes: list[pd.DataFrame] = []
    base_columns: list[str] | None = None
    base_filename: str | None = None
    has_structure_mismatch = False
    table_created = False
    table_message = ''

    for file_obj in files:
        filename = _get_upload_source_filename(file_obj)
        if not filename:
            invalid_files.append('未命名文件')
            continue

        if not is_excel_filename(filename):
            invalid_files.append(filename)
            continue

        valid_files.append(file_obj)

    if not valid_files:
        return {
            'success': False,
            'message': '没有有效的Excel文件（.xlsx/.xls）',
            'file_count': 0,
            'files': [],
            'invalid_files': invalid_files,
            'failed_files': failed_files,
        }

    for file_obj in valid_files:
        filename = _get_upload_source_filename(file_obj)

        try:
            file_bytes = _read_upload_source_bytes(file_obj)
            excel_buffer = BytesIO(file_bytes)
            dtype_mapping = _build_text_dtype_mapping(excel_buffer)
            df = pd.read_excel(excel_buffer, dtype=dtype_mapping if dtype_mapping else None)
            df.columns = [normalize_header_text(col) for col in df.columns.tolist()]
            current_columns = normalize_columns(df.columns.tolist())

            normalized_required_columns = [normalize_header_text(col) for col in ORDER_REQUIRED_COLUMNS]
            normalized_column_mapping = {
                normalize_header_text(chinese_name): english_name
                for chinese_name, english_name in ORDER_COLUMN_MAPPING.items()
            }

            # ===== 新增：必需字段校验 =====
            missing_required = [col for col in normalized_required_columns if col not in current_columns]
            if missing_required:
                failed_files.append({
                    'filename': filename,
                    'error': f"缺少必需字段：{', '.join(missing_required)}",
                })
                continue

            # ===== 新增：列名映射（中文 -> 英文） =====
            df = df.rename(columns=normalized_column_mapping)
            mapped_columns = normalize_columns(df.columns.tolist())

            current_summary = {
                'filename': filename,
                'row_count': int(len(df)),
                'column_count': int(len(df.columns)),
                'columns': mapped_columns,
            }

            if base_columns is None:
                base_columns = current_columns
                base_filename = filename
                file_summaries.append(current_summary)
                prepared_dataframes.append(df)
            else:
                columns_match, missing_columns, extra_columns = check_columns_match(
                    base_columns,
                    current_columns,
                )

                if not columns_match:
                    error_parts: list[str] = []
                    if missing_columns:
                        error_parts.append(f"缺少列：{', '.join(missing_columns)}")
                    if extra_columns:
                        error_parts.append(f"多出列：{', '.join(extra_columns)}")
                    if not missing_columns and not extra_columns:
                        error_parts.append('列名顺序不一致')

                    failed_files.append({
                        'filename': filename,
                        'error': f"列结构不一致（基准文件：{base_filename}；{'；'.join(error_parts)}）",
                    })
                    has_structure_mismatch = True
                    file_summaries.append(current_summary)
                    continue

                file_summaries.append(current_summary)
                prepared_dataframes.append(df)
        except Exception as exc:
            failed_files.append({
                'filename': filename,
                'error': str(exc),
            })
        finally:
            _reset_upload_source(file_obj)

    success_count = len(file_summaries)

    if has_structure_mismatch or failed_files:
        return _build_import_precheck_failed_response(
            '本次订单导入',
            file_summaries,
            invalid_files,
            failed_files,
        )

    if success_count == 0:
        message_parts: list[str] = ['文件已接收，但读取失败']

        if failed_files:
            message_parts.append('')
            message_parts.append('读取失败：')
            for failed in failed_files:
                message_parts.append(f"- {failed['filename']}（{failed['error']}）")

        if invalid_files:
            message_parts.append('')
            message_parts.append(f"无效文件：{'，'.join(invalid_files)}")

        return {
            'success': False,
            'message': '\n'.join(message_parts),
            'file_count': 0,
            'files': [],
            'invalid_files': invalid_files,
            'failed_files': failed_files,
        }

    try:
        table_created, table_message = _ensure_order_table_exists()
        written_rows, write_message = _write_orders_to_db(prepared_dataframes)
    except Exception as exc:
        db_path = _get_database_path()
        return {
            'success': False,
            'message': f'写入数据库失败：{str(exc)}\n数据库路径：{db_path}',
            'file_count': success_count,
            'files': file_summaries,
            'invalid_files': invalid_files,
            'failed_files': failed_files,
        }

    status_warning = _try_update_data_status('orders') if written_rows > 0 else ''
    message = f'成功读取 {success_count} 个订单文件'
    if table_message:
        message += f'\n{table_message}'
    if write_message:
        message += f'\n{write_message}'
    if status_warning:
        message += f'\n{status_warning}'

    return {
        'success': True,
        'message': message,
        'file_count': success_count,
        'files': file_summaries,
        'invalid_files': invalid_files,
        'failed_files': failed_files,
        'table_created': table_created,
        'table_message': table_message,
        'written_rows': written_rows,
        'write_message': write_message,
    }
#########################################################
# 售后导入相关辅助函数和主入口
#########################################################


def _build_after_sales_create_table_sql(table_name: str, column_types: dict[str, str]) -> str:
    """根据售后字段类型定义生成 CREATE TABLE SQL。"""
    column_defs: list[str] = []

    for column_name, column_type in column_types.items():
        column_defs.append(f'{column_name} {column_type}')

    columns_sql = ',\n    '.join(column_defs)
    return f'''CREATE TABLE IF NOT EXISTS {table_name} (\n    id INTEGER PRIMARY KEY AUTOINCREMENT,\n    {columns_sql}\n);'''



def _ensure_after_sales_table_exists() -> tuple[bool, str]:
    """确保微信售后表存在；不存在则自动创建，已存在则自动补齐缺失字段。"""
    db_path = _get_database_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)

    table_exists = False
    with sqlite3.connect(db_path) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
            (WECHAT_AFTER_SALES_TABLE_NAME,),
        )
        table_exists = cursor.fetchone() is not None

        if not table_exists:
            create_sql = _build_after_sales_create_table_sql(WECHAT_AFTER_SALES_TABLE_NAME, AFTER_SALES_COLUMN_TYPES)
            cursor.execute(create_sql)
            conn.commit()
            return True, f'售后表不存在，已自动创建：{WECHAT_AFTER_SALES_TABLE_NAME}（数据库：{db_path.name}）'

        existing_columns = _get_existing_table_columns(conn, WECHAT_AFTER_SALES_TABLE_NAME)
        added_columns: list[str] = []
        for column_name, column_type in AFTER_SALES_COLUMN_TYPES.items():
            if column_name not in existing_columns:
                cursor.execute(
                    f"ALTER TABLE {WECHAT_AFTER_SALES_TABLE_NAME} ADD COLUMN {column_name} {column_type}"
                )
                added_columns.append(column_name)

        if added_columns:
            conn.commit()
            return False, f'售后表已存在：{WECHAT_AFTER_SALES_TABLE_NAME}，并已补齐字段：{", ".join(added_columns)}（数据库：{db_path.name}）'

    return False, f'售后表已存在：{WECHAT_AFTER_SALES_TABLE_NAME}（数据库：{db_path.name}）'



def _prepare_after_sales_dataframe_for_db(df: pd.DataFrame) -> pd.DataFrame:
    """按数据库表结构整理售后 DataFrame。"""
    db_columns = list(AFTER_SALES_COLUMN_TYPES.keys())
    prepared_df = df.copy()

    for column_name in db_columns:
        if column_name not in prepared_df.columns:
            prepared_df[column_name] = None

    prepared_df = prepared_df[db_columns]

    for column_name, column_type in AFTER_SALES_COLUMN_TYPES.items():
        if column_name in {
            'after_sales_no',
            'platform_product_code',
            'custom_product_code',
            'custom_sku_code',
            'order_no',
            'delivery_tracking_no',
            'return_tracking_no',
            'merchant_contact_phone',
        }:
            prepared_df[column_name] = prepared_df[column_name].apply(_clean_text_value)
            continue

        if column_type == 'REAL':
            prepared_df[column_name] = prepared_df[column_name].apply(_clean_numeric_value)
        elif column_type == 'INTEGER':
            prepared_df[column_name] = (
                prepared_df[column_name]
                .apply(_clean_numeric_value)
                .apply(lambda x: int(x) if x is not None else None)
            )
        else:
            prepared_df[column_name] = prepared_df[column_name].apply(_clean_text_value)

    return prepared_df



def _build_after_sales_dedup_key(row: pd.Series) -> str | None:
    """构造售后防重键：售后单号 + 售后申请时间。"""
    parts: list[str] = []

    for column_name in AFTER_SALES_DEDUP_KEY_COLUMNS:
        value = row.get(column_name)
        cleaned_value = _clean_text_value(value)
        if cleaned_value is None or str(cleaned_value).strip() == '':
            return None
        parts.append(str(cleaned_value).strip())

    return '||'.join(parts)



def _get_existing_after_sales_keys(conn: sqlite3.Connection) -> set[str]:
    """读取数据库中已存在的售后防重键。"""
    cursor = conn.cursor()
    cursor.execute(
        f"SELECT after_sales_no, after_sales_apply_time FROM {WECHAT_AFTER_SALES_TABLE_NAME}"
    )
    rows = cursor.fetchall()

    existing_keys: set[str] = set()
    for row in rows:
        row_dict = {
            'after_sales_no': row[0],
            'after_sales_apply_time': row[1],
        }
        dedup_key = _build_after_sales_dedup_key(pd.Series(row_dict))
        if dedup_key:
            existing_keys.add(dedup_key)

    return existing_keys



def _deduplicate_after_sales_df(merged_df: pd.DataFrame, conn: sqlite3.Connection) -> tuple[pd.DataFrame, int, int]:
    """按售后维度防重：售后单号 + 售后申请时间。"""
    missing_key_columns = [col for col in AFTER_SALES_DEDUP_KEY_COLUMNS if col not in merged_df.columns]
    if missing_key_columns:
        return merged_df, 0, 0

    working_df = merged_df.copy()
    working_df['_dedup_key'] = working_df.apply(_build_after_sales_dedup_key, axis=1)

    valid_key_mask = working_df['_dedup_key'].notna() & (working_df['_dedup_key'].astype(str).str.strip() != '')
    batch_valid_df = working_df[valid_key_mask].copy()
    batch_unique_df = batch_valid_df.drop_duplicates(subset=['_dedup_key'], keep='first')
    batch_duplicate_count = int(len(batch_valid_df) - len(batch_unique_df))

    empty_key_df = working_df[~valid_key_mask].copy()
    combined_df = pd.concat([batch_unique_df, empty_key_df], ignore_index=True)

    existing_keys = _get_existing_after_sales_keys(conn)
    if not existing_keys:
        final_df = combined_df.drop(columns=['_dedup_key'], errors='ignore')
        return final_df, batch_duplicate_count, 0

    valid_after_batch_mask = combined_df['_dedup_key'].notna() & (combined_df['_dedup_key'].astype(str).str.strip() != '')
    valid_after_batch_df = combined_df[valid_after_batch_mask].copy()
    non_key_df = combined_df[~valid_after_batch_mask].copy()

    db_filtered_df = valid_after_batch_df[~valid_after_batch_df['_dedup_key'].isin(existing_keys)].copy()
    db_duplicate_count = int(len(valid_after_batch_df) - len(db_filtered_df))

    final_df = pd.concat([db_filtered_df, non_key_df], ignore_index=True)
    final_df = final_df.drop(columns=['_dedup_key'], errors='ignore')
    return final_df, batch_duplicate_count, db_duplicate_count



def _write_after_sales_to_db(dataframes: list[pd.DataFrame]) -> tuple[int, str]:
    """将售后数据写入 SQLite，返回写入行数和结果消息。"""
    if not dataframes:
        return 0, '没有可写入的售后数据'

    db_path = _get_database_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)

    prepared_frames = [_prepare_after_sales_dataframe_for_db(df) for df in dataframes]
    merged_df = pd.concat(prepared_frames, ignore_index=True)

    with sqlite3.connect(db_path) as conn:
        existing_columns = _get_existing_table_columns(conn, WECHAT_AFTER_SALES_TABLE_NAME)
        for column_name, column_type in AFTER_SALES_COLUMN_TYPES.items():
            if column_name not in existing_columns:
                conn.execute(
                    f"ALTER TABLE {WECHAT_AFTER_SALES_TABLE_NAME} ADD COLUMN {column_name} {column_type}"
                )
        conn.commit()

        deduped_df, batch_duplicate_count, db_duplicate_count = _deduplicate_after_sales_df(merged_df, conn)

        if deduped_df.empty:
            message_parts = ['没有可写入的新售后数据']
            if batch_duplicate_count > 0:
                message_parts.append(f'本次文件内重复售后已跳过：{batch_duplicate_count} 条')
            if db_duplicate_count > 0:
                message_parts.append(f'数据库中已存在售后已跳过：{db_duplicate_count} 条')
            return 0, '；'.join(message_parts) + f'（数据库：{db_path.name}）'

        deduped_df.to_sql(WECHAT_AFTER_SALES_TABLE_NAME, conn, if_exists='append', index=False)

    message_parts = [f'成功写入 {len(deduped_df)} 行售后数据']
    if batch_duplicate_count > 0:
        message_parts.append(f'本次文件内重复售后已跳过：{batch_duplicate_count} 条')
    if db_duplicate_count > 0:
        message_parts.append(f'数据库中已存在售后已跳过：{db_duplicate_count} 条')

    return int(len(deduped_df)), '；'.join(message_parts) + f'（数据库：{db_path.name}）'



def read_after_sales_excel_files(files: list[FileStorage]) -> dict[str, Any]:
    """读取微信售后 Excel 文件，校验、映射、写库。"""
    valid_files: list[FileStorage] = []
    invalid_files: list[str] = []
    failed_files: list[dict[str, str]] = []
    file_summaries: list[dict[str, Any]] = []
    prepared_dataframes: list[pd.DataFrame] = []
    base_columns: list[str] | None = None
    base_filename: str | None = None
    has_structure_mismatch = False

    for file_obj in files:
        filename = _get_upload_source_filename(file_obj)
        if not filename:
            invalid_files.append('未命名文件')
            continue

        if not is_excel_filename(filename):
            invalid_files.append(filename)
            continue

        valid_files.append(file_obj)

    if not valid_files:
        return {
            'success': False,
            'message': '没有有效的Excel文件（.xlsx/.xls）',
            'file_count': 0,
            'files': [],
            'invalid_files': invalid_files,
            'failed_files': failed_files,
        }

    for file_obj in valid_files:
        filename = _get_upload_source_filename(file_obj)

        try:
            file_bytes = _read_upload_source_bytes(file_obj)
            excel_buffer = BytesIO(file_bytes)
            dtype_mapping = _build_text_dtype_mapping(excel_buffer)
            df = pd.read_excel(excel_buffer, dtype=dtype_mapping if dtype_mapping else None)
            df.columns = [normalize_header_text(col) for col in df.columns.tolist()]
            current_columns = normalize_columns(df.columns.tolist())

            normalized_required_columns = [normalize_header_text(col) for col in AFTER_SALES_REQUIRED_COLUMNS]
            normalized_column_mapping = {
                normalize_header_text(chinese_name): english_name
                for chinese_name, english_name in AFTER_SALES_COLUMN_MAPPING.items()
            }

            missing_required = [col for col in normalized_required_columns if col not in current_columns]
            if missing_required:
                failed_files.append({
                    'filename': filename,
                    'error': f"缺少必需字段：{', '.join(missing_required)}",
                })
                continue

            df = df.rename(columns=normalized_column_mapping)
            mapped_columns = normalize_columns(df.columns.tolist())

            current_summary = {
                'filename': filename,
                'row_count': int(len(df)),
                'column_count': int(len(df.columns)),
                'columns': mapped_columns,
            }

            if base_columns is None:
                base_columns = current_columns
                base_filename = filename
                file_summaries.append(current_summary)
                prepared_dataframes.append(df)
            else:
                columns_match, missing_columns, extra_columns = check_columns_match(
                    base_columns,
                    current_columns,
                )

                if not columns_match:
                    error_parts: list[str] = []
                    if missing_columns:
                        error_parts.append(f"缺少列：{', '.join(missing_columns)}")
                    if extra_columns:
                        error_parts.append(f"多出列：{', '.join(extra_columns)}")
                    if not missing_columns and not extra_columns:
                        error_parts.append('列名顺序不一致')

                    failed_files.append({
                        'filename': filename,
                        'error': f"列结构不一致（基准文件：{base_filename}；{'；'.join(error_parts)}）",
                    })
                    has_structure_mismatch = True
                    file_summaries.append(current_summary)
                    continue

                file_summaries.append(current_summary)
                prepared_dataframes.append(df)
        except Exception as exc:
            failed_files.append({
                'filename': filename,
                'error': str(exc),
            })
        finally:
            _reset_upload_source(file_obj)

    success_count = len(file_summaries)

    if has_structure_mismatch or failed_files:
        return _build_import_precheck_failed_response(
            '本次售后导入',
            file_summaries,
            invalid_files,
            failed_files,
        )

    if success_count == 0:
        message_parts: list[str] = ['售后文件已接收，但读取失败']

        if failed_files:
            message_parts.append('')
            message_parts.append('读取失败：')
            for failed in failed_files:
                message_parts.append(f"- {failed['filename']}（{failed['error']}）")

        if invalid_files:
            message_parts.append('')
            message_parts.append(f"无效文件：{'，'.join(invalid_files)}")

        return {
            'success': False,
            'message': '\n'.join(message_parts),
            'file_count': 0,
            'files': [],
            'invalid_files': invalid_files,
            'failed_files': failed_files,
        }

    try:
        table_created, table_message = _ensure_after_sales_table_exists()
        written_rows, write_message = _write_after_sales_to_db(prepared_dataframes)
    except Exception as exc:
        db_path = _get_database_path()
        return {
            'success': False,
            'message': f'写入售后数据库失败：{str(exc)}\n数据库路径：{db_path}',
            'file_count': success_count,
            'files': file_summaries,
            'invalid_files': invalid_files,
            'failed_files': failed_files,
        }

    status_warning = _try_update_data_status('aftersales') if written_rows > 0 else ''
    message = f'成功读取 {success_count} 个售后文件'
    if table_message:
        message += f'\n{table_message}'
    if write_message:
        message += f'\n{write_message}'
    if status_warning:
        message += f'\n{status_warning}'

    return {
        'success': True,
        'message': message,
        'file_count': success_count,
        'files': file_summaries,
        'invalid_files': invalid_files,
        'failed_files': failed_files,
        'table_created': table_created,
        'table_message': table_message,
        'written_rows': written_rows,
        'write_message': write_message,
    }
