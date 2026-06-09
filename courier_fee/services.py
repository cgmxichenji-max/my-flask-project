"""快递费计算模块 — 底单记录服务层。"""

from __future__ import annotations

import sqlite3
import unicodedata
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from flask import current_app
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .table_schemas import (
    SHIPMENT_COLUMN_MAPPING,
    SHIPMENT_COLUMN_TYPES,
    SHIPMENT_REQUIRED_COLUMNS,
    SHIPMENT_TABLE_NAME,
    SHIPMENT_DISPLAY_COLUMNS,
    SHIPMENT_EXPORT_COLUMNS,
    PRICING_RULES_TABLE_NAME,
    PRICING_RULES_DEFAULTS,
    BASE_WEIGHT_OP_OPTIONS,
)


# ──────────────────────────────────────────────
# 工具
# ──────────────────────────────────────────────

def _db_path() -> Path:
    db = current_app.config.get('DATABASE_PATH')
    if db:
        return Path(db)
    return Path(current_app.root_path) / 'data' / 'main.db'


def _normalize(value: Any) -> str:
    """NFKC 归一化 + 去首尾空格，统一全半角/空格差异。"""
    return unicodedata.normalize('NFKC', str(value)).strip()


def _safe_text(v: Any) -> str | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip()
    return s if s else None


def _safe_num(v: Any, cast=float) -> Any:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        return cast(v)
    except (TypeError, ValueError):
        return None


def _parse_datetime_text(v: Any) -> str | None:
    """把 Excel 读出的日期值转成 'YYYY-MM-DD HH:MM:SS' 字符串。"""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d %H:%M:%S')
    text = str(v).strip()
    if not text:
        return None
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y/%m/%d %H:%M:%S',
                '%Y/%m/%d %H:%M', '%Y-%m-%d', '%Y/%m/%d'):
        try:
            return datetime.strptime(text, fmt).strftime('%Y-%m-%d %H:%M:%S')
        except ValueError:
            pass
    return None  # 无法解析为日期时返回 None（非日期行跳过）


def _format_dt_display(v: str | None) -> str:
    if not v:
        return ''
    return v


# ──────────────────────────────────────────────
# 建表 / 数据状态
# ──────────────────────────────────────────────

def _ensure_shipment_table(conn: sqlite3.Connection) -> None:
    col_defs = ',\n    '.join(
        f'{col} {typ}' for col, typ in SHIPMENT_COLUMN_TYPES.items()
    )
    conn.execute(f'''
        CREATE TABLE IF NOT EXISTS {SHIPMENT_TABLE_NAME} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            {col_defs}
        )
    ''')
    # 确保索引存在（tracking_no 查询高频）
    conn.execute(f'''
        CREATE INDEX IF NOT EXISTS idx_{SHIPMENT_TABLE_NAME}_tracking_no
        ON {SHIPMENT_TABLE_NAME}(tracking_no)
    ''')
    conn.execute(f'''
        CREATE INDEX IF NOT EXISTS idx_{SHIPMENT_TABLE_NAME}_ship_time
        ON {SHIPMENT_TABLE_NAME}(ship_time)
    ''')
    conn.execute(f'''
        CREATE INDEX IF NOT EXISTS idx_{SHIPMENT_TABLE_NAME}_order_no
        ON {SHIPMENT_TABLE_NAME}(order_no)
    ''')
    conn.commit()


def get_data_status() -> dict:
    """返回底单记录的数据状态：总记录数、时间范围、最近导入时间。"""
    with sqlite3.connect(_db_path()) as conn:
        _ensure_shipment_table(conn)
        # 只统计 ship_time 格式为标准日期的行（排除非日期的异常值）
        row = conn.execute(f'''
            SELECT
                COUNT(*)        AS cnt,
                MIN(ship_time)  AS min_dt,
                MAX(ship_time)  AS max_dt
            FROM {SHIPMENT_TABLE_NAME}
            WHERE ship_time LIKE '20__-__-__%'
        ''').fetchone()
        status_row = conn.execute('''
            SELECT last_import_time FROM courier_fee_import_status
            ORDER BY id DESC LIMIT 1
        ''').fetchone() if _import_status_table_exists(conn) else None

    return {
        'record_count':      row[0] or 0,
        'min_date':          _format_dt_display(row[1]),
        'max_date':          _format_dt_display(row[2]),
        'last_import_time':  status_row[0] if status_row else '',
    }


def _import_status_table_exists(conn: sqlite3.Connection) -> bool:
    r = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name='courier_fee_import_status'"
    ).fetchone()
    return bool(r)


def _ensure_import_status_table(conn: sqlite3.Connection) -> None:
    conn.execute('''
        CREATE TABLE IF NOT EXISTS courier_fee_import_status (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            last_import_time TEXT NOT NULL,
            files_count      INTEGER DEFAULT 0,
            rows_written     INTEGER DEFAULT 0,
            rows_skipped     INTEGER DEFAULT 0
        )
    ''')
    conn.commit()


def _record_import_status(conn: sqlite3.Connection,
                           files_count: int, rows_written: int, rows_skipped: int) -> None:
    _ensure_import_status_table(conn)
    conn.execute('''
        INSERT INTO courier_fee_import_status
            (last_import_time, files_count, rows_written, rows_skipped)
        VALUES (datetime('now','localtime'), ?, ?, ?)
    ''', (files_count, rows_written, rows_skipped))
    conn.commit()


# ──────────────────────────────────────────────
# 导入
# ──────────────────────────────────────────────

def _parse_one_file(file_obj: Any) -> tuple[pd.DataFrame, list[str]]:
    """
    读取单个 Excel 文件，做列名归一化和字段映射，
    返回 (mapped_df, warning_list)。
    """
    filename = getattr(file_obj, 'filename', '') or getattr(file_obj, 'name', '') or 'unknown'
    warnings: list[str] = []

    raw_bytes = file_obj.read() if hasattr(file_obj, 'read') else Path(file_obj.path).read_bytes()
    df = pd.read_excel(BytesIO(raw_bytes), dtype=str, header=0)

    # 归一化列名
    df.columns = [_normalize(c) for c in df.columns]

    # 检查必填列
    missing = [c for c in SHIPMENT_REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f'{filename}：缺少必填列 {missing}')

    # 保留已知列，映射到英文字段名
    mapped_rows: list[dict] = []
    skipped_junk = 0
    nondate_kept = 0

    for _, row in df.iterrows():
        ship_time_raw = row.get('发货时间', None)
        ship_time = _parse_datetime_text(ship_time_raw)

        tracking_no = _safe_text(row.get('快递单号', None))
        order_no    = _safe_text(row.get('订单编号', None))

        # 发货时间无法解析为日期时，不再整行丢弃：
        #   只要有快递单号或订单编号（如「已占用单号但未发货」「已回收面单」等退货/占号场景），
        #   仍保留该行（发货时间留空），以便账单按快递单号回填发货内容。
        if not ship_time:
            if not tracking_no and not order_no:
                skipped_junk += 1
                continue
            nondate_kept += 1

        r: dict[str, Any] = {}
        for cn, en in SHIPMENT_COLUMN_MAPPING.items():
            v = row.get(cn, None)
            col_type = SHIPMENT_COLUMN_TYPES.get(en, 'TEXT')
            if en == 'ship_time':
                r[en] = ship_time or ''
            elif en == 'apply_tracking_time':
                r[en] = _parse_datetime_text(v)
            elif col_type == 'INTEGER':
                r[en] = _safe_num(v, int)
            elif col_type == 'REAL':
                r[en] = _safe_num(v, float)
            else:
                r[en] = _safe_text(v)
        mapped_rows.append(r)

    if skipped_junk:
        warnings.append(f'{filename}：{skipped_junk} 行无快递单号且无订单编号，已跳过')
    if nondate_kept:
        warnings.append(f'{filename}：{nondate_kept} 行发货时间非日期（退货/占号等），发货时间留空已保留')

    return pd.DataFrame(mapped_rows), warnings


def read_shipment_excel_files(staged_files: list) -> dict[str, Any]:
    """
    解析并导入多个底单记录 Excel 文件。
    去重逻辑：tracking_no 已存在则跳过（INSERT OR IGNORE 基于去重键）。
    """
    all_warnings: list[str] = []
    all_dfs: list[pd.DataFrame] = []
    parse_errors: list[str] = []

    for sf in staged_files:
        try:
            df, warns = _parse_one_file(sf)
            all_dfs.append(df)
            all_warnings.extend(warns)
        except Exception as exc:
            fname = getattr(sf, 'filename', str(sf))
            parse_errors.append(f'{fname}：{exc}')

    if parse_errors:
        return {
            'success': False,
            'message': '部分文件解析失败：\n' + '\n'.join(parse_errors),
        }

    if not all_dfs:
        return {'success': False, 'message': '所有文件均无有效数据行'}

    combined = pd.concat(all_dfs, ignore_index=True)
    total_rows = len(combined)

    # ── 内存内预去重：相同 tracking_no 只保留第一行 ──────
    # tracking_no 非空的行按 tracking_no 去重
    has_track = combined['tracking_no'].notna() & (combined['tracking_no'] != '')
    dedup_track = combined[has_track].drop_duplicates(subset=['tracking_no'], keep='first')
    no_track = combined[~has_track].drop_duplicates(
        subset=['order_no', 'ship_time'], keep='first'
    )
    combined = pd.concat([dedup_track, no_track], ignore_index=True)
    in_mem_dedup = total_rows - len(combined)
    if in_mem_dedup:
        all_warnings.append(f'批次内重复 {in_mem_dedup} 行已去重')

    # ── 写入数据库 ──────────────────────────────────────
    fields = list(SHIPMENT_COLUMN_TYPES.keys())
    placeholders = ', '.join(['?'] * len(fields))
    col_names = ', '.join(fields)

    rows_written = 0
    rows_skipped = 0

    with sqlite3.connect(_db_path()) as conn:
        _ensure_shipment_table(conn)

        # 取已有 tracking_no 集合做快速去重
        existing_tracks: set[str] = set()
        for (t,) in conn.execute(
            f'SELECT tracking_no FROM {SHIPMENT_TABLE_NAME} WHERE tracking_no IS NOT NULL'
        ):
            existing_tracks.add(t)

        # 取已有 (order_no, ship_time) 集合（仅对 tracking_no 为空的行）
        existing_pairs: set[tuple] = set()
        for (o, s) in conn.execute(
            f'SELECT order_no, ship_time FROM {SHIPMENT_TABLE_NAME} '
            f'WHERE tracking_no IS NULL OR tracking_no = ""'
        ):
            existing_pairs.add((o, s))

        batch: list[tuple] = []
        for rec in combined.to_dict('records'):
            tn = rec.get('tracking_no') or ''
            if tn:
                if tn in existing_tracks:
                    rows_skipped += 1
                    continue
                existing_tracks.add(tn)
            else:
                pair = (rec.get('order_no') or '', rec.get('ship_time') or '')
                if pair in existing_pairs:
                    rows_skipped += 1
                    continue
                existing_pairs.add(pair)

            batch.append(tuple(
                rec.get(f) if rec.get(f) is not None else None
                for f in fields
            ))

        if batch:
            conn.executemany(
                f'INSERT INTO {SHIPMENT_TABLE_NAME} ({col_names}) VALUES ({placeholders})',
                batch,
            )
            rows_written = len(batch)
            conn.commit()

        _record_import_status(conn, len(staged_files), rows_written, rows_skipped)

    msg_parts = [
        f'导入完成：写入 {rows_written} 行，跳过重复 {rows_skipped} 行，'
        f'共处理 {len(staged_files)} 个文件。'
    ]
    if all_warnings:
        msg_parts.append('\n'.join(all_warnings))

    return {
        'success':       True,
        'message':       '\n'.join(msg_parts),
        'rows_written':  rows_written,
        'rows_skipped':  rows_skipped,
        'files_count':   len(staged_files),
    }


# ──────────────────────────────────────────────
# 查询
# ──────────────────────────────────────────────

def query_shipments(
    ship_time_start: str = '',
    ship_time_end: str = '',
    platform: str = '',
    shop_name: str = '',
    courier_name: str = '',
    tracking_no: str = '',
    order_no: str = '',
    ship_content_keyword: str = '',
    page: int = 1,
    page_size: int = 100,
) -> dict[str, Any]:
    """按条件分页查询底单记录，返回 {rows, total, page, page_size}。"""
    conditions: list[str] = []
    params: list[Any] = []

    if ship_time_start:
        conditions.append('ship_time >= ?')
        params.append(ship_time_start)
    if ship_time_end:
        end = ship_time_end
        if len(end) == 10:
            end += ' 23:59:59'
        conditions.append('ship_time <= ?')
        params.append(end)
    if platform:
        conditions.append('platform LIKE ?')
        params.append(f'%{platform}%')
    if shop_name:
        conditions.append('shop_name LIKE ?')
        params.append(f'%{shop_name}%')
    if courier_name:
        conditions.append('courier_name LIKE ?')
        params.append(f'%{courier_name}%')
    if tracking_no:
        conditions.append('tracking_no LIKE ?')
        params.append(f'%{tracking_no}%')
    if order_no:
        conditions.append('order_no LIKE ?')
        params.append(f'%{order_no}%')
    if ship_content_keyword:
        conditions.append('ship_content LIKE ?')
        params.append(f'%{ship_content_keyword}%')

    where = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''
    display_fields = [en for _, en in SHIPMENT_DISPLAY_COLUMNS]
    col_sql = ', '.join(display_fields)

    offset = (page - 1) * page_size

    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_shipment_table(conn)

        total = conn.execute(
            f'SELECT COUNT(*) FROM {SHIPMENT_TABLE_NAME} {where}', params
        ).fetchone()[0]

        rows_raw = conn.execute(
            f'SELECT {col_sql} FROM {SHIPMENT_TABLE_NAME} {where} '
            f'ORDER BY ship_time DESC LIMIT ? OFFSET ?',
            params + [page_size, offset],
        ).fetchall()

    rows = [dict(r) for r in rows_raw]
    return {
        'rows':      rows,
        'total':     total,
        'page':      page,
        'page_size': page_size,
        'headers':   [cn for cn, _ in SHIPMENT_DISPLAY_COLUMNS],
        'fields':    display_fields,
    }


# ──────────────────────────────────────────────
# 导出
# ──────────────────────────────────────────────

def export_shipments_to_excel(
    ship_time_start: str = '',
    ship_time_end: str = '',
    platform: str = '',
    shop_name: str = '',
    courier_name: str = '',
    tracking_no: str = '',
    order_no: str = '',
    ship_content_keyword: str = '',
) -> tuple[BytesIO, str]:
    """按条件导出底单记录为 Excel，返回 (BytesIO, 下载文件名)。"""
    conditions: list[str] = []
    params: list[Any] = []

    if ship_time_start:
        conditions.append('ship_time >= ?')
        params.append(ship_time_start)
    if ship_time_end:
        end = ship_time_end
        if len(end) == 10:
            end += ' 23:59:59'
        conditions.append('ship_time <= ?')
        params.append(end)
    if platform:
        conditions.append('platform LIKE ?')
        params.append(f'%{platform}%')
    if shop_name:
        conditions.append('shop_name LIKE ?')
        params.append(f'%{shop_name}%')
    if courier_name:
        conditions.append('courier_name LIKE ?')
        params.append(f'%{courier_name}%')
    if tracking_no:
        conditions.append('tracking_no LIKE ?')
        params.append(f'%{tracking_no}%')
    if order_no:
        conditions.append('order_no LIKE ?')
        params.append(f'%{order_no}%')
    if ship_content_keyword:
        conditions.append('ship_content LIKE ?')
        params.append(f'%{ship_content_keyword}%')

    where = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''
    export_fields = [en for _, en in SHIPMENT_EXPORT_COLUMNS]
    export_headers = [cn for cn, _ in SHIPMENT_EXPORT_COLUMNS]
    col_sql = ', '.join(export_fields)

    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_shipment_table(conn)
        rows_raw = conn.execute(
            f'SELECT {col_sql} FROM {SHIPMENT_TABLE_NAME} {where} ORDER BY ship_time DESC',
            params,
        ).fetchall()

    rows = [dict(r) for r in rows_raw]

    wb = Workbook()
    ws = wb.active
    ws.title = '底单记录'

    header_fill = PatternFill('solid', fgColor='E9EEF4')
    header_font = Font(bold=True)

    for col_idx, header in enumerate(export_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, rec in enumerate(rows, start=2):
        for col_idx, field in enumerate(export_fields, start=1):
            ws.cell(row=row_idx, column=col_idx, value=rec.get(field))

    for col_cells in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col_cells), default=8)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'底单记录导出_{ts}.xlsx'
    return buf, filename


# ──────────────────────────────────────────────
# 计费规则
# ──────────────────────────────────────────────

def _ensure_pricing_rules_table(conn: sqlite3.Connection) -> None:
    """建表（首次运行），并在空表时写入3条默认规则。"""
    conn.execute(f'''
        CREATE TABLE IF NOT EXISTS {PRICING_RULES_TABLE_NAME} (
            id                   INTEGER PRIMARY KEY AUTOINCREMENT,
            carrier_name         TEXT    NOT NULL,
            base_weight          REAL    NOT NULL DEFAULT 1.2,
            base_weight_op       TEXT    NOT NULL DEFAULT 'gt',
            base_fee             REAL    NOT NULL DEFAULT 0,
            rate_jzshw           REAL    NOT NULL DEFAULT 0,
            rate_other           REAL    NOT NULL DEFAULT 0,
            weight_tolerance_pct REAL    NOT NULL DEFAULT 0.20,
            effective_date       TEXT    NOT NULL DEFAULT '',
            is_active            INTEGER NOT NULL DEFAULT 1,
            note                 TEXT,
            created_at           TEXT    NOT NULL DEFAULT '',
            updated_at           TEXT    NOT NULL DEFAULT ''
        )
    ''')
    # 运行时迁移：为旧表补加 weight_tolerance_pct 列
    try:
        conn.execute(
            f'ALTER TABLE {PRICING_RULES_TABLE_NAME} '
            f'ADD COLUMN weight_tolerance_pct REAL NOT NULL DEFAULT 0.20'
        )
    except Exception:
        pass  # 列已存在则忽略
    conn.commit()

    # 空表时写入默认3条
    count = conn.execute(
        f'SELECT COUNT(*) FROM {PRICING_RULES_TABLE_NAME}'
    ).fetchone()[0]
    if count == 0:
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for row in PRICING_RULES_DEFAULTS:
            conn.execute(f'''
                INSERT INTO {PRICING_RULES_TABLE_NAME}
                    (carrier_name, base_weight, base_weight_op, base_fee,
                     rate_jzshw, rate_other, weight_tolerance_pct,
                     effective_date, is_active, note, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                row['carrier_name'], row['base_weight'], row['base_weight_op'],
                row['base_fee'], row['rate_jzshw'], row['rate_other'],
                row.get('weight_tolerance_pct', 0.20),
                row['effective_date'], row['is_active'], row['note'], now, now,
            ))
        conn.commit()


def get_pricing_rules() -> list[dict]:
    """返回所有计费规则，按 carrier_name / effective_date 排序。"""
    with sqlite3.connect(_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_pricing_rules_table(conn)
        rows = conn.execute(f'''
            SELECT id, carrier_name, base_weight, base_weight_op,
                   base_fee, rate_jzshw, rate_other,
                   weight_tolerance_pct,
                   effective_date, is_active, note,
                   created_at, updated_at
            FROM {PRICING_RULES_TABLE_NAME}
            ORDER BY carrier_name, effective_date DESC
        ''').fetchall()
    return [dict(r) for r in rows]


def save_pricing_rule(data: dict) -> tuple[bool, str, int | None]:
    """
    新增或更新一条计费规则。
    data 含 id（有则 UPDATE，无则 INSERT）。
    返回 (ok, error_msg, rule_id)。
    """
    # 字段校验
    carrier = (data.get('carrier_name') or '').strip()
    if not carrier:
        return False, '快递公司不能为空', None

    try:
        base_weight = float(data.get('base_weight', 1.2))
        if base_weight <= 0:
            raise ValueError
    except (TypeError, ValueError):
        return False, '底重必须为正数', None

    base_weight_op = data.get('base_weight_op', 'gt')
    if base_weight_op not in BASE_WEIGHT_OP_OPTIONS:
        return False, f'超出条件无效，允许值：{list(BASE_WEIGHT_OP_OPTIONS.keys())}', None

    try:
        base_fee             = float(data.get('base_fee', 0))
        rate_jzshw           = float(data.get('rate_jzshw', 0))
        rate_other           = float(data.get('rate_other', 0))
        weight_tolerance_pct = float(data.get('weight_tolerance_pct', 0.20))
        if not (0 < weight_tolerance_pct < 10):
            raise ValueError('允差百分比应在 0~1000% 之间')
    except (TypeError, ValueError) as e:
        return False, f'收费/费率/允差必须为数字：{e}', None

    effective_date = (data.get('effective_date') or '').strip()
    if not effective_date:
        return False, '起效日期不能为空', None

    is_active = 1 if data.get('is_active') in (1, True, '1', 'true', 'True') else 0
    note = (data.get('note') or '').strip() or None
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    rule_id = data.get('id')
    try:
        rule_id = int(rule_id) if rule_id is not None else None
    except (TypeError, ValueError):
        rule_id = None

    with sqlite3.connect(_db_path()) as conn:
        _ensure_pricing_rules_table(conn)

        if rule_id:
            # 更新
            exists = conn.execute(
                f'SELECT 1 FROM {PRICING_RULES_TABLE_NAME} WHERE id = ?', (rule_id,)
            ).fetchone()
            if not exists:
                return False, f'规则 id={rule_id} 不存在', None
            conn.execute(f'''
                UPDATE {PRICING_RULES_TABLE_NAME}
                SET carrier_name=?, base_weight=?, base_weight_op=?,
                    base_fee=?, rate_jzshw=?, rate_other=?,
                    weight_tolerance_pct=?,
                    effective_date=?, is_active=?, note=?, updated_at=?
                WHERE id=?
            ''', (
                carrier, base_weight, base_weight_op,
                base_fee, rate_jzshw, rate_other,
                weight_tolerance_pct,
                effective_date, is_active, note, now, rule_id,
            ))
            conn.commit()
            return True, '', rule_id
        else:
            # 新增
            cursor = conn.execute(f'''
                INSERT INTO {PRICING_RULES_TABLE_NAME}
                    (carrier_name, base_weight, base_weight_op, base_fee,
                     rate_jzshw, rate_other, weight_tolerance_pct,
                     effective_date, is_active, note, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                carrier, base_weight, base_weight_op,
                base_fee, rate_jzshw, rate_other,
                weight_tolerance_pct,
                effective_date, is_active, note, now, now,
            ))
            conn.commit()
            return True, '', cursor.lastrowid


def delete_pricing_rule(rule_id: int) -> tuple[bool, str]:
    """删除一条计费规则。"""
    try:
        rule_id = int(rule_id)
    except (TypeError, ValueError):
        return False, '规则 id 无效'

    with sqlite3.connect(_db_path()) as conn:
        _ensure_pricing_rules_table(conn)
        result = conn.execute(
            f'DELETE FROM {PRICING_RULES_TABLE_NAME} WHERE id = ?', (rule_id,)
        )
        conn.commit()
        if result.rowcount == 0:
            return False, f'规则 id={rule_id} 不存在'
    return True, ''
