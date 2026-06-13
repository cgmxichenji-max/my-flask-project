"""
快递费账单服务层
- 文件解析（中通/韵达/申通，列名自动匹配）
- 文件存储与管理
- OPT_02：从底单记录填入发货内容
- OPT_03：费用标准计算与核对
- OPT_07：异常行费用写回
- 正式入库 / 丢弃未核查数据
"""

import io
import math
import os
import re
import unicodedata
import uuid
import zipfile
from datetime import datetime

import pandas as pd

from auth.services import get_db_connection
from .weight_estimate import WeightEstimator, parse_ship_key_to_rows
from .table_schemas import (
    BILL_COLUMN_ALIASES,
    BILL_FILES_TABLE_NAME,
    BILL_REQUIRED_FIELDS,
    BILLS_DISPLAY_COLUMNS,
    BILLS_TABLE_NAME,
    PRICING_RULES_TABLE_NAME,
    SHIPMENT_TABLE_NAME,
)

# ── 存储目录 ──────────────────────────────────────────────────────
_BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STORAGE_DIR = os.path.join(_BASE_DIR, 'data', 'courier_bills')
os.makedirs(STORAGE_DIR, exist_ok=True)


# ── 通用工具 ──────────────────────────────────────────────────────
def _now():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def _safe_remove(path):
    """安全删除文件，忽略不存在/权限等异常。"""
    if not path:
        return
    try:
        if os.path.exists(path):
            os.remove(path)
    except OSError:
        pass


def _norm(s):
    return unicodedata.normalize('NFKC', str(s or '')).strip()


def _try_float(v):
    """Return (ok: bool, value: float)."""
    try:
        return True, float(str(v or '').replace(',', '').strip())
    except (ValueError, TypeError):
        return False, 0.0


# ── 省份判断 ──────────────────────────────────────────────────────
_JZSHW = frozenset(['江苏', '浙江', '上海', '安徽'])
_XJXZ_KW = ['新疆', '西藏']


def _normalize_province(prov):
    p = _norm(prov)
    p = re.sub(r'(省|市|自治区|回族自治区|维吾尔自治区|壮族自治区)$', '', p)
    return p.strip()


def _is_jzshw(prov):
    return _normalize_province(prov) in _JZSHW


def _is_xinjiang_tibet(prov):
    p = _normalize_province(prov)
    return any(kw in p for kw in _XJXZ_KW)


# ── 建表 ──────────────────────────────────────────────────────────
def _ensure_bill_tables(conn):
    conn.executescript(f"""
    CREATE TABLE IF NOT EXISTS {BILL_FILES_TABLE_NAME} (
        id               INTEGER PRIMARY KEY AUTOINCREMENT,
        carrier_name     TEXT NOT NULL,
        bill_year_month  TEXT NOT NULL,
        file_seq         INTEGER NOT NULL DEFAULT 1,
        original_filename TEXT NOT NULL,
        stored_filename  TEXT NOT NULL,
        stored_path      TEXT NOT NULL,
        row_count        INTEGER DEFAULT 0,
        total_fee        REAL DEFAULT 0,
        status           TEXT NOT NULL DEFAULT 'uploaded',
        uploaded_at      TEXT NOT NULL DEFAULT '',
        created_at       TEXT NOT NULL DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS {BILLS_TABLE_NAME} (
        id               INTEGER PRIMARY KEY AUTOINCREMENT,
        file_id          INTEGER NOT NULL,
        carrier_name     TEXT NOT NULL,
        bill_year_month  TEXT NOT NULL,
        tracking_no      TEXT NOT NULL,
        bill_date        TEXT DEFAULT '',
        settle_object    TEXT DEFAULT '',
        dest_province    TEXT DEFAULT '',
        settle_weight    REAL DEFAULT 0,
        actual_fee       REAL DEFAULT 0,
        sender           TEXT DEFAULT '',
        ship_content     TEXT DEFAULT '',
        ship_content_key TEXT DEFAULT '',
        std_fee          REAL,
        fee_diff         REAL,
        check_result     TEXT DEFAULT '',
        min_weight       REAL,
        max_weight       REAL,
        weight_in_range  TEXT DEFAULT '',
        est_max_fee      REAL,
        is_corrected     INTEGER DEFAULT 0,
        corrected_fee    REAL,
        is_verified      INTEGER DEFAULT 0,
        created_at       TEXT NOT NULL DEFAULT ''
    );
    CREATE INDEX IF NOT EXISTS idx_cfb_ym    ON {BILLS_TABLE_NAME}(bill_year_month);
    CREATE INDEX IF NOT EXISTS idx_cfb_trk   ON {BILLS_TABLE_NAME}(tracking_no);
    CREATE INDEX IF NOT EXISTS idx_cfb_fid   ON {BILLS_TABLE_NAME}(file_id);
    CREATE INDEX IF NOT EXISTS idx_cfbf_ym   ON {BILL_FILES_TABLE_NAME}(bill_year_month);
    """)
    # 运行时迁移：为旧表补加 ship_content_key 列（新建表已含此列，ALTER 仅针对旧库）
    try:
        conn.execute(
            f'ALTER TABLE {BILLS_TABLE_NAME} ADD COLUMN ship_content_key TEXT DEFAULT ""'
        )
    except Exception:
        pass  # 列已存在，忽略
    conn.commit()


# ── 列名匹配 ──────────────────────────────────────────────────────
def _find_col(headers, aliases):
    """Return 0-based index of first alias match in headers; -1 if not found."""
    norm_h = [_norm(h) for h in headers]
    for alias in aliases:
        a = _norm(alias)
        for i, h in enumerate(norm_h):
            if a == h:
                return i
    return -1


def _has_aggregate_col(headers):
    """True if ANY header cell contains '合计'."""
    return any('合计' in _norm(str(h)) for h in headers)


def _looks_like_header(row):
    """True if row contains at least one known column alias."""
    all_aliases = [a for aliases in BILL_COLUMN_ALIASES.values() for a in aliases]
    norm_row = {_norm(str(v)) for v in row}
    return any(_norm(a) in norm_row for a in all_aliases)


def _find_header_row(df):
    """
    扫描前 10 行，找最后一个「包含已知列名别名 且 不含「合计」列」的行。
    返回 (header_row_index, headers_list)，找不到返回 (-1, [])。
    """
    best_idx, best_headers = -1, []
    for i in range(min(10, len(df))):
        row = [str(v) if str(v) != 'nan' else '' for v in df.iloc[i]]
        if _looks_like_header(row) and not _has_aggregate_col(row):
            best_idx, best_headers = i, row
    return best_idx, best_headers


# ── 脏行判断 ──────────────────────────────────────────────────────
def _is_dirty_row(tracking, weight_s, fee_s):
    t = _norm(str(tracking or ''))
    if len(t) < 6:
        return True
    ok_w, _ = _try_float(weight_s)
    ok_f, _ = _try_float(fee_s)
    return not (ok_w and ok_f)


# ── OPT_04：发货内容解析 → 排序 Key（严格按 VBA SPX_PC_ParseShipText） ──
_RE_MULTI_SEMI   = re.compile(r';{2,}')
_RE_LEADING_NUM  = re.compile(r'^\d+\.')
_RE_STAR_QTY_END = re.compile(r'\*(\d+)\s*(?:瓶|盒|支|个|件)?\s*$')
_RE_BRACKET_QTY  = re.compile(r'\[(\d+)\]\s*$')
_RE_KEEP_ALNUM   = re.compile(r'[^A-Za-z0-9]')


def _normalize_ship_text(raw: str) -> str:
    """VBA SPX_PC_NormalizeShipText：统一分隔符、括号、引号，换行→分号。"""
    s = str(raw)
    s = s.replace('【', '[').replace('】', ']')
    s = s.replace('；', ';').replace('：', ':')
    s = s.replace('"', '').replace('“', '').replace('”', '')
    s = s.replace('\r\n', ';').replace('\n', ';').replace('\r', ';')
    s = re.sub(r'\]\s+', '];', s)       # "] " → "];"
    s = _RE_MULTI_SEMI.sub(';', s)      # 连续分号压缩
    return s.strip(';')


def _apply_patches_before_normalize(raw: str) -> str:
    """
    VBA SPX_PC_ApplyPatches_BeforeNormalize。
    VBA 原始代码无特殊补丁（ParseCombo 能自然处理各种组合格式），直接返回原文。
    """
    return raw


# VBA SPX_PC_CodePattern: 匹配产品编码（含可选 *数量）
_RE_CODE_PATTERN = re.compile(r'([A-Za-z]?\d+[A-Za-z]?)(?:\*(\d+))?')
# 策略1：编码后紧跟单位词（ml/g/kg 等）
_RE_CODE_WITH_UNIT = re.compile(
    r'([A-Za-z]*\d{3}[A-Za-z]*)(?=\s*\d+\s*(?:ml|g|kg|l|oz|pcs|片|袋|支)\b)'
    r'(?:\s*\*(\d+))?',
    re.IGNORECASE,
)


def _parse_combo_seg(s: str, buy_qty: int, agg: dict) -> None:
    """
    VBA SPX_PC_ParseCombo_ToAgg：
    取 "组合" 前的 main_part，用正则匹配所有编码+数量，乘以 buy_qty 累加。
    编码必须包含3位数字（VBA过滤规则）。
    """
    main_part = s.split('组合')[0]
    main_part = main_part.replace('：', ':')
    main_part = re.sub(r'\s*\*\s*', '*', main_part)   # 清理 * 两边空格
    main_part = re.sub(r'\s*:\s*', ':', main_part)    # 清理 : 两边空格
    main_part = re.sub(r':{2,}', ':', main_part)      # 合并多余冒号

    # 策略1：编码后有单位词（如 "A161 10ml"）
    matches = list(_RE_CODE_WITH_UNIT.finditer(main_part))
    if not matches:
        # 策略2 回退：通用 CodePattern
        matches = list(_RE_CODE_PATTERN.finditer(main_part))

    for m in matches:
        code = _RE_KEEP_ALNUM.sub('', m.group(1)).upper()
        if not code:
            continue
        # VBA 过滤：编码必须包含3位数字
        if not re.search(r'\d{3}', code):
            continue
        qty_str = m.group(2)
        qty = int(qty_str) if qty_str else 1
        if qty <= 0:
            qty = 1
        agg[code] = agg.get(code, 0) + qty * buy_qty


def _parse_ship_seg(seg: str, agg: dict) -> None:
    """
    VBA SPX_PC_ParseOneSegment_ToAgg：
    - 取末尾 [n] 作为 buyQty 乘数
    - 含"组合" → 走 _parse_combo_seg
    - 普通段：冒号前为编码，末尾 *qty 为数量，乘以 buyQty
    """
    seg = _RE_LEADING_NUM.sub('', seg)   # 去 "1." "2." 等前缀
    seg = seg.replace('：', ':')

    # 取 [n] 作为 buyQty（VBA TakeBracketQty_AndRemove）
    buy_qty = 1
    m_bq = _RE_BRACKET_QTY.search(seg)
    if m_bq:
        buy_qty = max(1, int(m_bq.group(1)))
        seg = seg[:m_bq.start()].rstrip()

    # 组合段：走 ParseCombo（注意：组合段不删空格，否则 "140 50ml" → "14050ml"）
    if '组合' in seg:
        _parse_combo_seg(seg, buy_qty, agg)
        return

    # 普通段：删空格，找冒号
    seg = seg.replace(' ', '')
    pos = seg.find(':')
    if pos < 0:
        return  # VBA: 缺冒号直接跳过

    code = _RE_KEEP_ALNUM.sub('', seg[:pos]).upper()
    if not code:
        return

    m = _RE_STAR_QTY_END.search(seg)
    qty = int(m.group(1)) if m else 1
    if qty <= 0:
        qty = 1

    agg[code] = agg.get(code, 0) + qty * buy_qty


def parse_ship_content_to_key(ship_content: str) -> str:
    """
    OPT_04 入口：发货内容原文 → 排序后聚合 Key。
    例：'140:*1;113B:*2' → '113B*2;140*1'
    """
    if not ship_content or not ship_content.strip():
        return ''
    raw  = _apply_patches_before_normalize(ship_content)
    norm = _normalize_ship_text(raw)
    norm = _RE_MULTI_SEMI.sub(';', norm).strip(';')
    if not norm:
        return ''
    agg: dict = {}
    for seg in norm.split(';'):
        seg = seg.strip()
        if seg:
            _parse_ship_seg(seg, agg)
    if not agg:
        return ''
    return ';'.join(f'{code}*{agg[code]}' for code in sorted(agg))


# ── OPT_05：重量核查（严格按 VBA OPT_CalcWeightFromKey + OPT_05_CheckWeightTolerance） ─
_DEFAULT_WEIGHT_TOL = 0.20   # 默认重量允差 20%（由计费规则中 weight_tolerance_pct 覆盖）


def _calc_weight_from_key(ship_key: str, weight_dict: dict) -> tuple:
    """
    VBA OPT_CalcWeightFromKey：从 aggKey 计算货物总重量。
    返回 (ok: bool, total_weight: float, bad_codes: str)
    ok=False 表示有编码不在重量表或总重量为0。
    """
    if not ship_key:
        return False, 0.0, ''
    total = 0.0
    bad   = []
    for part in ship_key.split(';'):
        part = part.strip()
        if not part:
            continue
        arr  = part.split('*')
        code = _RE_KEEP_ALNUM.sub('', arr[0]).upper()
        try:
            qty = float(arr[1]) if len(arr) > 1 else 1.0
        except (ValueError, IndexError):
            qty = 1.0
        if qty <= 0:
            qty = 1.0
        if code in weight_dict:
            total += weight_dict[code] * qty
        else:
            bad.append(code)
    ok = (len(bad) == 0 and total > 0)
    return ok, total, ','.join(bad)


# ── 解析单个账单文件 ──────────────────────────────────────────────
def parse_bill_file(filepath, carrier_name):
    """
    解析快递账单 Excel 文件。
    返回 (rows: list[dict], error: str, total_fee: float)
    error 为空字符串表示成功。
    """
    try:
        df = pd.read_excel(filepath, header=None, dtype=str)
    except Exception as e:
        return [], f'文件读取失败：{e}', 0.0

    if df.empty:
        return [], '文件为空', 0.0

    hdr_idx, headers = _find_header_row(df)
    if hdr_idx < 0:
        return [], (
            '无法自动识别标题行（已扫描前10行）。'
            '韵达账单若第1行含"合计XX元"，该行会自动跳过并寻找下一个有效标题行；'
            '若仍找不到，请检查文件后重新上传。'
        ), 0.0

    # 排除含"合计"的列
    clean_headers = ['' if '合计' in _norm(h) else h for h in headers]

    # 建立字段→列索引映射
    col = {}
    for field, aliases in BILL_COLUMN_ALIASES.items():
        col[field] = _find_col(clean_headers, aliases)

    # 校验必需列
    missing_labels = {
        'tracking_no':   '运单号/单号',
        'settle_weight': '结算重量',
        'actual_fee':    '结算费用/运费/应收金额/应收费用',
    }
    missing = [missing_labels[f] for f in BILL_REQUIRED_FIELDS if col.get(f, -1) < 0]
    if missing:
        return [], f'缺少必需列：{"、".join(missing)}，请检查文件后重新上传。', 0.0

    # 解析数据行
    rows = []
    total_fee = 0.0

    def get_val(row_series, field):
        idx = col.get(field, -1)
        if idx < 0 or idx >= len(row_series):
            return ''
        v = row_series.iloc[idx]
        return '' if str(v) == 'nan' else str(v).strip()

    for i in range(hdr_idx + 1, len(df)):
        r = df.iloc[i]
        tracking  = get_val(r, 'tracking_no')
        weight_s  = get_val(r, 'settle_weight')
        fee_s     = get_val(r, 'actual_fee')

        if _is_dirty_row(tracking, weight_s, fee_s):
            continue

        _, weight = _try_float(weight_s)
        _, fee    = _try_float(fee_s)
        total_fee += fee

        rows.append({
            'tracking_no':   _norm(tracking),
            'bill_date':     get_val(r, 'bill_date'),
            'settle_object': get_val(r, 'settle_object'),
            'dest_province': get_val(r, 'dest_province'),
            'settle_weight': weight,
            'actual_fee':    fee,
            'sender':        get_val(r, 'sender'),
        })

    if not rows:
        return [], '文件中未找到有效数据行（所有行均被判定为汇总行或格式无效）', 0.0

    return rows, '', round(total_fee, 2)


# ── 月份列表 ──────────────────────────────────────────────────────
def get_bill_months():
    """返回已有账单的年月列表（倒序）。"""
    with get_db_connection() as conn:
        _ensure_bill_tables(conn)
        rows = conn.execute(
            f'SELECT DISTINCT bill_year_month FROM {BILL_FILES_TABLE_NAME} '
            f'ORDER BY bill_year_month DESC'
        ).fetchall()
    return [r['bill_year_month'] for r in rows]


# ── 查重 ──────────────────────────────────────────────────────────
def check_file_exists(carrier_name, year_month):
    """
    检查同快递公司+年月是否已有上传文件。
    返回 {'exists': bool, 'files': [{id, stored_filename, file_seq}, ...]}
    """
    with get_db_connection() as conn:
        _ensure_bill_tables(conn)
        rows = conn.execute(
            f'SELECT id, stored_filename, file_seq FROM {BILL_FILES_TABLE_NAME} '
            f'WHERE carrier_name=? AND bill_year_month=? ORDER BY file_seq',
            (carrier_name, year_month),
        ).fetchall()
    return {'exists': bool(rows), 'files': [dict(r) for r in rows]}


# ── 导入账单文件 ──────────────────────────────────────────────────
def import_bill_file(file_storage, carrier_name, year_month, action='add'):
    """
    保存并解析快递账单文件。

    action:
        'add'     → 保留已有文件，新文件取下一个序号
        'replace' → 删除该 carrier+year_month 所有已有文件，再以序号1导入

    返回 {'success': bool, 'message': str, 'file_id': int, 'row_count': int,
           'stored_filename': str, 'total_fee': float}
    """
    os.makedirs(STORAGE_DIR, exist_ok=True)
    original_filename = file_storage.filename
    ext = os.path.splitext(original_filename)[1].lower() or '.xlsx'

    # ① 先把上传内容落到临时文件并校验解析。
    #    任何失败（网络中断/文件损坏/格式不符）都在此返回，绝不触碰已有数据。
    tmp_path = os.path.join(STORAGE_DIR, f'.uploading_{uuid.uuid4().hex}{ext}')
    try:
        file_storage.save(tmp_path)
    except Exception as e:
        _safe_remove(tmp_path)
        return {'success': False, 'message': f'文件保存失败（可能上传中断）：{e}'}

    rows, error, total_fee = parse_bill_file(tmp_path, carrier_name)
    if error:
        _safe_remove(tmp_path)
        return {'success': False, 'message': error}

    # ② 校验通过后才动数据库与旧文件。
    try:
        with get_db_connection() as conn:
            _ensure_bill_tables(conn)
            now = _now()

            old_paths = []
            if action == 'replace':
                old = conn.execute(
                    f'SELECT id, stored_path FROM {BILL_FILES_TABLE_NAME} '
                    f'WHERE carrier_name=? AND bill_year_month=?',
                    (carrier_name, year_month),
                ).fetchall()
                for f in old:
                    conn.execute(f'DELETE FROM {BILLS_TABLE_NAME} WHERE file_id=?', (f['id'],))
                    conn.execute(f'DELETE FROM {BILL_FILES_TABLE_NAME} WHERE id=?', (f['id'],))
                    old_paths.append(f['stored_path'])
                file_seq = 1
            else:
                row = conn.execute(
                    f'SELECT MAX(file_seq) AS m FROM {BILL_FILES_TABLE_NAME} '
                    f'WHERE carrier_name=? AND bill_year_month=?',
                    (carrier_name, year_month),
                ).fetchone()
                file_seq = (row['m'] or 0) + 1

            seq_str = f'{file_seq:02d}'
            stored_filename = f'{year_month}_{carrier_name}_{seq_str}{ext}'
            stored_path = os.path.join(STORAGE_DIR, stored_filename)

            # 原子落位：临时文件 → 正式文件
            os.replace(tmp_path, stored_path)
            tmp_path = None

            # 写文件台账
            cursor = conn.execute(
                f'INSERT INTO {BILL_FILES_TABLE_NAME} '
                f'(carrier_name, bill_year_month, file_seq, original_filename, stored_filename, '
                f' stored_path, row_count, total_fee, status, uploaded_at, created_at) '
                f'VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                (carrier_name, year_month, file_seq, original_filename, stored_filename,
                 stored_path, len(rows), total_fee, 'uploaded', now, now),
            )
            file_id = cursor.lastrowid

            # 批量写账单明细
            conn.executemany(
                f'INSERT INTO {BILLS_TABLE_NAME} '
                f'(file_id, carrier_name, bill_year_month, tracking_no, bill_date, '
                f' settle_object, dest_province, settle_weight, actual_fee, sender, created_at) '
                f'VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                [
                    (file_id, carrier_name, year_month,
                     r['tracking_no'], r['bill_date'], r['settle_object'],
                     r['dest_province'], r['settle_weight'], r['actual_fee'],
                     r['sender'], now)
                    for r in rows
                ],
            )
            conn.commit()
    finally:
        # 异常时清理未落位的临时文件（DB 由 with 块自动回滚，旧数据不受影响）
        if tmp_path:
            _safe_remove(tmp_path)

    # ③ DB 提交成功后，再删除被替换掉的旧物理文件（排除刚落位的新文件）
    for p in old_paths:
        if p and p != stored_path:
            _safe_remove(p)

    return {
        'success':         True,
        'message':         f'成功导入 {len(rows)} 行（{carrier_name} {year_month} 第{file_seq}份）',
        'file_id':         file_id,
        'row_count':       len(rows),
        'total_fee':       total_fee,
        'stored_filename': stored_filename,
    }


# ── 文件列表 ──────────────────────────────────────────────────────
def get_bill_files(year_month=None):
    """返回账单文件列表，可按年月过滤。"""
    with get_db_connection() as conn:
        _ensure_bill_tables(conn)
        if year_month:
            rows = conn.execute(
                f'SELECT * FROM {BILL_FILES_TABLE_NAME} '
                f'WHERE bill_year_month=? ORDER BY carrier_name, file_seq',
                (year_month,),
            ).fetchall()
        else:
            rows = conn.execute(
                f'SELECT * FROM {BILL_FILES_TABLE_NAME} '
                f'ORDER BY bill_year_month DESC, carrier_name, file_seq'
            ).fetchall()
    return [dict(r) for r in rows]


def delete_bill_file(file_id):
    """删除文件台账记录及其关联账单行。返回 (ok, message)。"""
    with get_db_connection() as conn:
        _ensure_bill_tables(conn)
        f = conn.execute(
            f'SELECT stored_path FROM {BILL_FILES_TABLE_NAME} WHERE id=?', (file_id,)
        ).fetchone()
        if not f:
            return False, '文件记录不存在'
        conn.execute(f'DELETE FROM {BILLS_TABLE_NAME} WHERE file_id=?', (file_id,))
        conn.execute(f'DELETE FROM {BILL_FILES_TABLE_NAME} WHERE id=?', (file_id,))
        conn.commit()
        try:
            if os.path.exists(f['stored_path']):
                os.remove(f['stored_path'])
        except OSError:
            pass
    return True, '已删除'


def get_bill_file_meta(file_id):
    """返回文件台账元数据 dict，找不到返回 None。"""
    with get_db_connection() as conn:
        _ensure_bill_tables(conn)
        row = conn.execute(
            f'SELECT * FROM {BILL_FILES_TABLE_NAME} WHERE id=?', (file_id,)
        ).fetchone()
    return dict(row) if row else None


# ── 打包下载原始文件 ──────────────────────────────────────────────
def generate_raw_zip(file_ids):
    """将指定 file_id 的原始存储文件打包为 ZIP，返回 (BytesIO, filename)。"""
    with get_db_connection() as conn:
        _ensure_bill_tables(conn)
        if file_ids:
            ph = ','.join('?' * len(file_ids))
            records = conn.execute(
                f'SELECT * FROM {BILL_FILES_TABLE_NAME} WHERE id IN ({ph})',
                list(file_ids),
            ).fetchall()
        else:
            records = []

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for r in records:
            p = r['stored_path']
            if os.path.exists(p):
                zf.write(p, r['stored_filename'])
    buf.seek(0)
    return buf, '快递账单原始文件.zip'


# ── 计算（OPT_02 + OPT_03） ───────────────────────────────────────
def _get_pricing_rule(carrier_name, conn):
    """取该快递公司最新启用的计费规则。"""
    row = conn.execute(
        f'SELECT * FROM {PRICING_RULES_TABLE_NAME} '
        f'WHERE carrier_name=? AND is_active=1 '
        f'ORDER BY effective_date DESC LIMIT 1',
        (carrier_name,),
    ).fetchone()
    return dict(row) if row else None


def _calc_std_fee(weight, province, rule):
    """
    按计费规则计算标准费用。
    返回 float，新疆/西藏返回 None（不适用），规则缺失返回 None。
    """
    if _is_xinjiang_tibet(province):
        return None

    base_w   = rule['base_weight']
    op       = rule['base_weight_op']   # 'gt' | 'gte'
    base_fee = rule['base_fee']
    rate     = rule['rate_jzshw'] if _is_jzshw(province) else rule['rate_other']

    # 超出部分（向上取整）
    if op == 'gte':
        over = max(0.0, weight - base_w)
    else:  # 'gt'
        over = max(0.0, weight - base_w)

    over_kg = math.ceil(over) if over > 1e-9 else 0
    return round(base_fee + over_kg * rate, 2)


def run_calculation(year_month):
    """
    对指定年月所有未核查行执行：
      OPT_02：从底单记录填入 ship_content（仅空行），同时调用 OPT_04 生成 ship_content_key
      OPT_03：计算 std_fee / fee_diff / check_result（含最低收费忽略判断）
      OPT_05：重量核查（min_weight / max_weight / weight_in_range / est_max_fee）
    返回汇总 dict。
    """
    with get_db_connection() as conn:
        _ensure_bill_tables(conn)

        # OPT_02：构建 tracking_no → ship_content 字典（底单记录）
        try:
            ship_rows = conn.execute(
                f'SELECT tracking_no, ship_content FROM {SHIPMENT_TABLE_NAME} '
                f'WHERE tracking_no IS NOT NULL AND tracking_no != "" '
                f'AND ship_content IS NOT NULL AND ship_content != ""'
            ).fetchall()
            ship_dict = {r['tracking_no']: r['ship_content'] for r in ship_rows}
        except Exception:
            ship_dict = {}

        # 重量预估器：按页头打印逻辑（箱型匹配→气泡袋匹配→货物+包装）预估整件重量
        try:
            estimator = WeightEstimator(conn)
            if not estimator.goods_weights:
                estimator = None
        except Exception:
            estimator = None

        # 取本月未核查行（含 ship_content 以便 OPT_02 仅填充空行）
        bills = conn.execute(
            f'SELECT id, carrier_name, tracking_no, dest_province, '
            f'       settle_weight, actual_fee, ship_content '
            f'FROM {BILLS_TABLE_NAME} '
            f'WHERE bill_year_month=? AND is_verified=0',
            (year_month,),
        ).fetchall()

        if not bills:
            return {
                'success':   True,
                'message':   f'{year_month} 无未核查数据，请先上传账单文件。',
                'processed': 0,
            }

        # 缓存计费规则
        rule_cache = {}
        for b in bills:
            c = b['carrier_name']
            if c not in rule_cache:
                rule_cache[c] = _get_pricing_rule(c, conn)

        price_err = ok_count = no_rule = invalid = weight_err = 0

        for b in bills:
            bid      = b['id']
            carrier  = b['carrier_name']
            tracking = b['tracking_no']
            province = b['dest_province'] or ''
            weight   = b['settle_weight'] or 0.0
            actual   = b['actual_fee'] or 0.0
            is_zto   = ('中通' in carrier)
            rule     = rule_cache.get(carrier)

            # OPT_02：仅当原行 ship_content 为空时才从底单填入
            old_content = b['ship_content'] or ''
            content = old_content if old_content.strip() else ship_dict.get(tracking, '')

            # OPT_04：解析排序 Key
            ship_key = parse_ship_content_to_key(content) if content.strip() else ''

            # OPT_05：重量核查（先于 OPT_03 执行）
            # 预估重量 = 货物重量 + 箱型重量 + 气泡袋重量（完全按页头打印推荐逻辑）。
            #   min_weight 列存「预估重量」，max_weight 列存「允差上限 = 预估 ×(1+允差%)」。
            # 规则（按用户确认）：
            #   1) 能算出重量 → 结算 ≤ 允差上限 记「是」，否则「否」；
            #   2) 算不出重量（货物缺重量/尺寸数据）→ 假设对方正确，记「是」，不预估；
            #   3) 重量「否」且「实收 > 起步费」才算真·重量异常（只收起步费的无论对错都这价，不较真）。
            tol_pct = float(rule['weight_tolerance_pct']) if (rule and rule.get('weight_tolerance_pct') is not None) else _DEFAULT_WEIGHT_TOL
            base_fee = 5.0 if is_zto else 4.5
            min_w = max_w = None      # min_w=预估重量, max_w=允差上限
            wt_in_range = ''
            est_fee = None
            weight_is_anomaly = False

            est_ok = False
            est_weight = 0.0
            if ship_key and estimator is not None:
                est_ok, est_weight, _detail = estimator.estimate(
                    parse_ship_key_to_rows(ship_key)
                )

            if not est_ok:
                # 算不出重量 → 假设对方正确
                wt_in_range = '是'
            else:
                min_w = round(est_weight, 2)                 # 预估整件重量
                max_w = round(est_weight * (1.0 + tol_pct), 2)  # 允差上限
                if weight <= max_w:
                    wt_in_range = '是'
                else:
                    wt_in_range = '否'
                    # 预估应收费用（按预估重量算对方应收）
                    if rule and not _is_xinjiang_tibet(province):
                        est_fee = _calc_std_fee(min_w, province, rule)
                        if est_fee is not None:
                            est_fee = round(est_fee, 2)
                    # 真异常过滤：仅当实收 > 起步费 时才算（只收起步费不较真）
                    if actual > base_fee + 0.0001:
                        weight_is_anomaly = True

            # OPT_03：费用核查（真·重量异常行不再判断价格异常）
            if weight_is_anomaly:
                weight_err += 1
                if rule and weight > 0:
                    std_fee = _calc_std_fee(weight, province, rule)
                    fee_diff = round(actual - std_fee, 2) if std_fee is not None else None
                else:
                    std_fee = fee_diff = None
                check_result = '重量异常'
            elif not rule:
                check_result = '无计费规则'
                std_fee = fee_diff = None
                no_rule += 1
            elif weight <= 0:
                check_result = '重量无效'
                std_fee = fee_diff = None
                invalid += 1
            else:
                std_fee = _calc_std_fee(weight, province, rule)
                if std_fee is None:
                    check_result = '不适用(新疆/西藏)'
                    fee_diff = None
                    ok_count += 1
                else:
                    fee_diff = round(actual - std_fee, 2)
                    if abs(actual - base_fee) < 0.0001:
                        check_result = '最低收费忽略'
                        ok_count += 1
                    elif abs(fee_diff) <= 0.01:
                        check_result = '正确'
                        ok_count += 1
                    else:
                        check_result = '不正确'
                        price_err += 1

            conn.execute(
                f'UPDATE {BILLS_TABLE_NAME} '
                f'SET ship_content=?, ship_content_key=?, '
                f'    std_fee=?, fee_diff=?, check_result=?, '
                f'    min_weight=?, max_weight=?, weight_in_range=?, est_max_fee=? '
                f'WHERE id=?',
                (content, ship_key,
                 std_fee, fee_diff, check_result,
                 min_w, max_w, wt_in_range, est_fee,
                 bid),
            )

        conn.commit()

    return {
        'success':      True,
        'message':      (
            f'计算完成：共 {len(bills)} 行 | '
            f'正确/忽略 {ok_count} | 价格异常 {price_err} | '
            f'重量异常 {weight_err} | 无规则 {no_rule} | 无效 {invalid}'
        ),
        'processed':    len(bills),
        'price_error':  price_err,
        'weight_error': weight_err,
        'ok':           ok_count,
        'no_rule':      no_rule,
        'invalid':      invalid,
    }


# ── 横表查询 ──────────────────────────────────────────────────────
def get_bill_anomaly_counts(year_month):
    """返回各筛选 tab 计数（未核查各类 + 已核查总数）。"""
    with get_db_connection() as conn:
        _ensure_bill_tables(conn)
        total = conn.execute(
            f'SELECT COUNT(*) FROM {BILLS_TABLE_NAME} '
            f'WHERE bill_year_month=? AND is_verified=0', (year_month,)
        ).fetchone()[0]
        price_err = conn.execute(
            f"SELECT COUNT(*) FROM {BILLS_TABLE_NAME} "
            f"WHERE bill_year_month=? AND is_verified=0 AND check_result='不正确'",
            (year_month,)
        ).fetchone()[0]
        empty_content = conn.execute(
            f"SELECT COUNT(*) FROM {BILLS_TABLE_NAME} "
            f"WHERE bill_year_month=? AND is_verified=0 "
            f"AND (ship_content IS NULL OR ship_content='')",
            (year_month,)
        ).fetchone()[0]
        weight_err = conn.execute(
            f"SELECT COUNT(*) FROM {BILLS_TABLE_NAME} "
            f"WHERE bill_year_month=? AND is_verified=0 "
            f"AND check_result='重量异常'",
            (year_month,)
        ).fetchone()[0]
        verified = conn.execute(
            f'SELECT COUNT(*) FROM {BILLS_TABLE_NAME} '
            f'WHERE bill_year_month=? AND is_verified=1', (year_month,)
        ).fetchone()[0]
    return {
        'total':         total,
        'price_error':   price_err,
        'empty_content': empty_content,
        'weight_error':  weight_err,
        'verified':      verified,
    }


_VALID_SORT_COLS = {col for col, _, _ in BILLS_DISPLAY_COLUMNS}


def query_bills(year_month, filter_type='all', search='',
                page=1, page_size=100, sort_col='bill_date', sort_dir='ASC'):
    """横表分页查询，支持筛选/多关键词搜索/排序。"""
    if sort_col not in _VALID_SORT_COLS:
        sort_col = 'bill_date'
    sort_dir = 'DESC' if str(sort_dir).upper() == 'DESC' else 'ASC'

    if filter_type == 'verified':
        # 已核查数据：is_verified=1
        conditions = ['bill_year_month=?', 'is_verified=1']
    else:
        # 其余均为未核查数据
        conditions = ['bill_year_month=?', 'is_verified=0']
        if filter_type == 'price_error':
            conditions.append("check_result='不正确'")
        elif filter_type == 'empty_content':
            conditions.append("(ship_content IS NULL OR ship_content='')")
        elif filter_type == 'weight_error':
            conditions.append("check_result='重量异常'")
    params = [year_month]

    if search.strip():
        for kw in search.strip().split():
            conditions.append(
                '(tracking_no LIKE ? OR carrier_name LIKE ? OR settle_object LIKE ? '
                'OR dest_province LIKE ? OR ship_content LIKE ? OR check_result LIKE ?)'
            )
            like = f'%{kw}%'
            params.extend([like] * 6)

    where = ' AND '.join(conditions)

    with get_db_connection() as conn:
        _ensure_bill_tables(conn)
        total = conn.execute(
            f'SELECT COUNT(*) FROM {BILLS_TABLE_NAME} WHERE {where}', params
        ).fetchone()[0]
        offset = (page - 1) * page_size
        rows = conn.execute(
            f'SELECT * FROM {BILLS_TABLE_NAME} WHERE {where} '
            f'ORDER BY {sort_col} {sort_dir} LIMIT ? OFFSET ?',
            params + [page_size, offset],
        ).fetchall()

    fields = [col for col, _, _ in BILLS_DISPLAY_COLUMNS] + ['id']
    return {
        'total':     total,
        'page':      page,
        'page_size': page_size,
        'rows':      [dict(r) for r in rows],
        'fields':    fields,
    }


# ── OPT_07：费用写回 ──────────────────────────────────────────────
def _is_correction_candidate(row):
    """需要按用户勾选状态决定是否修正的异常行。"""
    if row['check_result'] in ('不正确', '重量异常'):
        return True
    return not str(row['ship_content'] or '').strip()


def _correction_amount_for_row(row):
    """
    返回异常行应写回的正确费用。
    价格异常/空白异常按标准费用，重量异常优先按预估应收费用。
    """
    if row['check_result'] == '重量异常':
        if row['est_max_fee'] is not None:
            return float(row['est_max_fee'])
        raise ValueError(f'快递单号 {row["tracking_no"]} 为重量异常，但缺少预估应收费用，无法自动修正')
    if row['std_fee'] is not None:
        return float(row['std_fee'])
    raise ValueError(f'快递单号 {row["tracking_no"]} 缺少标准费用，无法自动修正')


def _apply_unchecked_corrections(conn, year_month, ok_row_ids):
    """
    勾选行视为正确：清除修正标记并保留原始费用。
    未勾选异常行视为需要修正：写入 corrected_fee，并标记 is_corrected=1。
    """
    ok_set = set(int(i) for i in (ok_row_ids or []))

    rows = conn.execute(
        f"SELECT id, tracking_no, actual_fee, std_fee, check_result, "
        f"       ship_content, est_max_fee "
        f"FROM {BILLS_TABLE_NAME} "
        f"WHERE bill_year_month=? AND is_verified=0 "
        f"AND (check_result IN ('不正确', '重量异常') "
        f"     OR ship_content IS NULL OR ship_content='' "
        f"     OR is_corrected=1)",
        (year_month,),
    ).fetchall()

    corrected = 0
    ok_count = 0
    for r in rows:
        if not _is_correction_candidate(r):
            continue
        if r['id'] in ok_set:
            conn.execute(
                f'UPDATE {BILLS_TABLE_NAME} '
                f'SET is_corrected=0, corrected_fee=NULL, fee_diff=0, check_result=? '
                f'WHERE id=?',
                ('正确', r['id']),
            )
            ok_count += 1
            continue

        corrected_fee = round(_correction_amount_for_row(r), 2)
        conn.execute(
            f'UPDATE {BILLS_TABLE_NAME} SET is_corrected=1, corrected_fee=? WHERE id=?',
            (corrected_fee, r['id']),
        )
        corrected += 1

    return corrected, ok_count


def run_opt07(year_month, ok_row_ids):
    """
    对本月未核查异常行按用户勾选状态写回费用：
    - 勾选行视为正确，不修正；
    - 未勾选价格/空白/重量异常行写入 corrected_fee。
    返回 {'success', 'message', 'corrected_count'}
    """
    with get_db_connection() as conn:
        _ensure_bill_tables(conn)
        corrected, ok_count = _apply_unchecked_corrections(conn, year_month, ok_row_ids)
        conn.commit()

    return {
        'success':         True,
        'message':         f'OPT_07 完成：已修正 {corrected} 行费用，保留勾选正确行 {ok_count} 行',
        'corrected_count': corrected,
        'ok_count':        ok_count,
    }


# ── 生成修正文件 ZIP（openpyxl 版：黄色标注 + 修正说明列） ───────
def _apply_corrections_openpyxl(stored_path, corrections, original_fees):
    """
    用 openpyxl 将修正内容写回文件：
    - 整行黄色填充（FFFFC000）
    - 更新费用列值为 corrected_fee
    - 末尾新增"修正说明"列：原价 X.XX → 修正: Y.YY (OPT_07价格修正)

    corrections:   {tracking_no_str: corrected_fee_float}
    original_fees: {tracking_no_str: original_fee_float}
    返回 BytesIO (xlsx 格式)。
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    YELLOW = PatternFill(start_color='FFFFC000', end_color='FFFFC000', fill_type='solid')

    empty_buf = io.BytesIO()
    try:
        wb = load_workbook(stored_path)
    except Exception:
        empty_buf.seek(0)
        return empty_buf

    ws = wb.active

    # 找标题行（openpyxl 行列均 1-based）
    hdr_row_idx = -1
    tracking_col = -1   # 1-based
    fee_col      = -1   # 1-based

    for ri in range(1, min(11, ws.max_row + 1)):
        row_vals = [str(ws.cell(ri, ci).value or '')
                    for ci in range(1, ws.max_column + 2)]
        if _looks_like_header(row_vals) and not _has_aggregate_col(row_vals):
            hdr_row_idx = ri
            clean = ['' if '合计' in _norm(h) else h for h in row_vals]
            t = _find_col(clean, BILL_COLUMN_ALIASES['tracking_no'])   # 0-based
            f = _find_col(clean, BILL_COLUMN_ALIASES['actual_fee'])    # 0-based
            if t >= 0:
                tracking_col = t + 1
            if f >= 0:
                fee_col = f + 1

    if hdr_row_idx < 0 or tracking_col < 0 or fee_col < 0:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # 末尾追加"修正说明"列标题
    note_col = ws.max_column + 1
    ws.cell(hdr_row_idx, note_col).value = '修正说明'

    for ri in range(hdr_row_idx + 1, ws.max_row + 1):
        t_val = _norm(str(ws.cell(ri, tracking_col).value or ''))
        if t_val not in corrections:
            continue

        new_fee  = corrections[t_val]
        orig_fee = original_fees.get(t_val, '?')

        # 更新费用列
        ws.cell(ri, fee_col).value = new_fee

        # 整行黄色填充
        for ci in range(1, note_col + 1):
            ws.cell(ri, ci).fill = YELLOW

        # 修正说明
        ws.cell(ri, note_col).value = (
            f'原价: {orig_fee} → 修正: {new_fee} (OPT_07价格修正)'
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_corrected_zip(year_month, ok_row_ids=None):
    """
    为本月未核查批次中有修正行的账单文件生成修正副本并打包 ZIP。
    - 只处理 is_verified=0 行对应的文件（当前批次，不含已核查旧批次）
    - 修正行：黄色填充 + 更新费用 + 修正说明列
    - 文件命名含修正后总金额
    返回 (BytesIO, zip_filename)。
    """
    with get_db_connection() as conn:
        _ensure_bill_tables(conn)
        _apply_unchecked_corrections(conn, year_month, ok_row_ids)
        conn.commit()

        # 仅取当前未核查批次的 file_id
        unverified_fids = {
            r[0] for r in conn.execute(
                f'SELECT DISTINCT file_id FROM {BILLS_TABLE_NAME} '
                f'WHERE bill_year_month=? AND is_verified=0',
                (year_month,),
            ).fetchall()
        }
        if not unverified_fids:
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, 'w'):
                pass
            buf.seek(0)
            return buf, f'快递账单修正_{year_month}.zip'

        ph = ','.join('?' * len(unverified_fids))
        file_records = conn.execute(
            f'SELECT * FROM {BILL_FILES_TABLE_NAME} WHERE id IN ({ph})',
            list(unverified_fids),
        ).fetchall()

        # 按 file_id 聚合修正数据（含原始费用用于注释）
        corr_rows = conn.execute(
            f'SELECT file_id, tracking_no, corrected_fee, actual_fee FROM {BILLS_TABLE_NAME} '
            f'WHERE bill_year_month=? AND is_verified=0 AND is_corrected=1',
            (year_month,),
        ).fetchall()

    corrections_by_fid: dict = {}
    originals_by_fid:   dict = {}
    for r in corr_rows:
        corrections_by_fid.setdefault(r['file_id'], {})[r['tracking_no']] = r['corrected_fee']
        originals_by_fid.setdefault(r['file_id'], {})[r['tracking_no']]   = r['actual_fee']

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for f in file_records:
            fid  = f['id']
            path = f['stored_path']
            if not os.path.exists(path):
                continue

            corrections = corrections_by_fid.get(fid, {})
            if not corrections:
                continue   # 无修正行，跳过

            originals = originals_by_fid.get(fid, {})

            # 修正后总金额（仅未核查行）
            with get_db_connection() as conn2:
                row = conn2.execute(
                    f'SELECT SUM(CASE WHEN is_corrected=1 THEN corrected_fee '
                    f'              ELSE actual_fee END) '
                    f'FROM {BILLS_TABLE_NAME} WHERE file_id=? AND is_verified=0',
                    (fid,),
                ).fetchone()
            total = round(row[0] or 0.0, 2)

            file_buf = _apply_corrections_openpyxl(path, corrections, originals)
            seq_str  = f'_{f["file_seq"]:02d}' if f['file_seq'] > 1 else ''
            out_name = (
                f'{f["carrier_name"]}_{year_month}'
                f'{seq_str}_{total:.0f}元_已修正.xlsx'
            )
            zf.writestr(out_name, file_buf.getvalue())

    buf.seek(0)
    return buf, f'快递账单修正_{year_month}.zip'


_SUMMARY_HEADERS = ['快递公司', '业务时间', '快递单号', '结算对象', '目的省份', '结算重量', '结算费用', '寄件人', '发货内容']
_SUMMARY_SHEETS = ('申通抖音', '申通微信澳柯', '韵达抖音', '韵达微信澳柯')
_CN_DIGITS = '〇一二三四五六七八九'
_CN_MONTHS = ('', '一月', '二月', '三月', '四月', '五月', '六月', '七月', '八月', '九月', '十月', '十一月', '十二月')


def _year_month_cn(year_month):
    """202605 -> 二〇二六年五月。"""
    ym = str(year_month or '').strip()
    if not re.fullmatch(r'\d{6}', ym):
        return ym
    year = ''.join(_CN_DIGITS[int(ch)] for ch in ym[:4])
    month = int(ym[4:])
    month_text = _CN_MONTHS[month] if 1 <= month <= 12 else f'{month}月'
    return f'{year}年{month_text}'


def _excel_datetime(value):
    text = str(value or '').strip()
    if not text:
        return ''
    parsed = pd.to_datetime(text, errors='coerce')
    if pd.isna(parsed):
        return text
    return parsed.to_pydatetime()


def _summary_amount(row):
    if row['is_corrected'] and row['corrected_fee'] is not None:
        return float(row['corrected_fee'] or 0)
    return float(row['actual_fee'] or 0)


def _summary_bucket(carrier_name, sender):
    carrier = str(carrier_name or '')
    sender = str(sender or '').strip()
    if '申通' in carrier:
        return '申通微信澳柯' if sender == '澳柯微信' else '申通抖音'
    if '韵达' in carrier:
        return '韵达微信澳柯' if '澳' in sender else '韵达抖音'
    return ''


def _apply_summary_sheet_style(ws):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    red_fill = PatternFill('solid', fgColor='C8323C')
    light_fill = PatternFill('solid', fgColor='FADADE')
    white_font = Font(color='FFFFFF', bold=True)
    bold_font = Font(bold=True)

    ws.merge_cells('A2:A3')
    ws.merge_cells('A4:A5')
    ws.column_dimensions['A'].width = 18.7166666666667
    ws.column_dimensions['B'].width = 17.775
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 20.9666666666667
    ws.row_dimensions[1].height = 24

    for row in ws.iter_rows(min_row=1, max_row=15, min_col=1, max_col=4):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for cell in ws[1]:
        cell.fill = red_fill
        cell.font = white_font
    for row_idx in (2, 4, 7, 11, 15):
        for col_idx in range(1, 5):
            ws.cell(row_idx, col_idx).fill = light_fill
    for cell in ws['A'][1:15]:
        cell.fill = red_fill
        cell.font = white_font

    for row_idx in (7, 8, 12, 15):
        ws.cell(row_idx, 2).font = bold_font
        ws.cell(row_idx, 3).font = bold_font
    for row_idx in range(1, 16):
        ws.cell(row_idx, 3).number_format = '#,##0.00'
        ws.cell(row_idx, 3).alignment = Alignment(horizontal='right', vertical='center')


def _write_summary_sheet(ws, year_month, sums):
    ws.title = '汇总表'
    ws['A1'] = _year_month_cn(year_month)
    ws['B1'] = '店 铺'
    ws['C1'] = '运费金额'

    ws['A2'] = '抖音老店'
    ws['B2'] = '申 通'
    ws['C2'] = round(sums['申通抖音'], 2)
    ws['B3'] = '韵 达 '
    ws['C3'] = round(sums['韵达抖音'], 2)

    ws['A4'] = '视频号澳柯'
    ws['B4'] = '申 通'
    ws['C4'] = round(sums['申通微信澳柯'], 2)
    ws['B5'] = '韵 达 '
    ws['C5'] = round(sums['韵达微信澳柯'], 2)

    ws['B6'] = '理 赔'
    ws['C6'] = 0
    ws['D6'] = ' '
    ws['B7'] = '小 计'
    ws['C7'] = '=SUM(C2:C6)'
    ws['B8'] = '含税价'
    ws['C8'] = '=C7*1.06'

    ws['B10'] = 'Eric 垫付费用'
    ws['C10'] = 0
    ws['D10'] = ' '
    ws['B11'] = 'Eric 垫付运费'
    ws['C11'] = 0
    ws['D11'] = ' '
    ws['B12'] = '含税价'
    ws['C12'] = '=(N(C10)+N(C11))*1.06'

    ws['B15'] = '合计含税价'
    ws['C15'] = '=C8+C12'
    ws['D15'] = ' '
    _apply_summary_sheet_style(ws)


def _write_detail_sheet(ws, rows):
    from openpyxl.styles import Font

    ws.append(_SUMMARY_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        fee = _summary_amount(row)
        ws.append([
            row['carrier_name'] or '',
            _excel_datetime(row['bill_date']),
            str(row['tracking_no'] or ''),
            row['settle_object'] or '',
            row['dest_province'] or '',
            row['settle_weight'] or 0,
            round(fee, 2),
            row['sender'] or '',
            row['ship_content_key'] or row['ship_content'] or '',
        ])

    for cell in ws['B'][1:]:
        cell.number_format = 'yyyy-mm-dd'
    for cell in ws['C']:
        cell.number_format = '@'
    for cell in ws['F']:
        cell.number_format = '0.00'
    for cell in ws['G']:
        cell.number_format = '0.00'
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 32)
    ws.freeze_panes = 'A2'


def generate_summary_workbook(year_month):
    """按 VBA 月度汇总口径，基于当前年月已核查账单生成快递费汇总表。"""
    ym = str(year_month or '').strip()
    if not re.fullmatch(r'\d{6}', ym):
        raise ValueError('请指定正确的年月')

    with get_db_connection() as conn:
        _ensure_bill_tables(conn)
        rows = conn.execute(
            f'''
            SELECT carrier_name, bill_date, tracking_no, settle_object, dest_province,
                   settle_weight, actual_fee, sender, ship_content, ship_content_key,
                   is_corrected, corrected_fee
            FROM {BILLS_TABLE_NAME}
            WHERE bill_year_month=? AND is_verified=1
            ORDER BY carrier_name, bill_date, tracking_no
            ''',
            (ym,),
        ).fetchall()

    if not rows:
        raise ValueError(f'{ym} 没有已核查入库数据，请先正式入库后再生成汇总表')

    grouped = {name: [] for name in _SUMMARY_SHEETS}
    sums = {name: 0.0 for name in _SUMMARY_SHEETS}
    for row in rows:
        bucket = _summary_bucket(row['carrier_name'], row['sender'])
        if not bucket:
            continue
        grouped[bucket].append(row)
        sums[bucket] += _summary_amount(row)

    from openpyxl import Workbook

    wb = Workbook()
    _write_summary_sheet(wb.active, ym, sums)
    for sheet_name in _SUMMARY_SHEETS:
        ws = wb.create_sheet(sheet_name)
        _write_detail_sheet(ws, grouped[sheet_name])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, f'快递费汇总{ym}月.xlsx'


def generate_summary_zip(year_month):
    """将月度快递费汇总 Excel 放入 ZIP 下载。"""
    workbook_buf, workbook_name = generate_summary_workbook(year_month)
    ym = str(year_month or '').strip()
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(workbook_name, workbook_buf.getvalue())
    buf.seek(0)
    return buf, f'快递费汇总{ym}月.zip'


# ── 正式入库 / 丢弃 ───────────────────────────────────────────────
def save_verified(year_month, ok_row_ids=None):
    """
    将本月所有未核查行标记为已核查（is_verified=1）。
    同时更新文件台账 status='verified'。
    """
    with get_db_connection() as conn:
        _ensure_bill_tables(conn)
        _apply_unchecked_corrections(conn, year_month, ok_row_ids)
        conn.execute(
            f'UPDATE {BILLS_TABLE_NAME} '
            f'SET actual_fee=corrected_fee, fee_diff=0, check_result=? '
            f'WHERE bill_year_month=? AND is_verified=0 '
            f'AND is_corrected=1 AND corrected_fee IS NOT NULL',
            ('正确', year_month),
        )
        cur = conn.execute(
            f'UPDATE {BILLS_TABLE_NAME} SET is_verified=1 '
            f'WHERE bill_year_month=? AND is_verified=0',
            (year_month,),
        )
        count = cur.rowcount
        conn.execute(
            f"UPDATE {BILL_FILES_TABLE_NAME} SET status='verified' WHERE bill_year_month=?",
            (year_month,),
        )
        conn.commit()
    return {'success': True, 'message': f'{count} 行数据已正式入库（标记已核查）', 'count': count}


def discard_unverified(year_month):
    """
    删除本月所有未核查行，以及相应空文件台账记录并清理文件。
    """
    with get_db_connection() as conn:
        _ensure_bill_tables(conn)
        file_ids = [
            r['id'] for r in conn.execute(
                f'SELECT id FROM {BILL_FILES_TABLE_NAME} WHERE bill_year_month=?',
                (year_month,)
            ).fetchall()
        ]

        cur = conn.execute(
            f'DELETE FROM {BILLS_TABLE_NAME} WHERE bill_year_month=? AND is_verified=0',
            (year_month,),
        )
        rows_deleted = cur.rowcount

        for fid in file_ids:
            remaining = conn.execute(
                f'SELECT COUNT(*) FROM {BILLS_TABLE_NAME} WHERE file_id=?', (fid,)
            ).fetchone()[0]
            if remaining == 0:
                f_row = conn.execute(
                    f'SELECT stored_path FROM {BILL_FILES_TABLE_NAME} WHERE id=?', (fid,)
                ).fetchone()
                if f_row:
                    try:
                        if os.path.exists(f_row['stored_path']):
                            os.remove(f_row['stored_path'])
                    except OSError:
                        pass
                conn.execute(f'DELETE FROM {BILL_FILES_TABLE_NAME} WHERE id=?', (fid,))

        conn.commit()
    return {'success': True, 'message': f'已丢弃 {rows_deleted} 行未核查数据', 'count': rows_deleted}
