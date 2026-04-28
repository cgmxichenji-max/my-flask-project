from flask import Blueprint, render_template, request, redirect, url_for, current_app
import json
import os
import re
import shutil
import sqlite3
import uuid
from pathlib import Path

from openpyxl import load_workbook
from flask import send_file
from werkzeug.utils import secure_filename

from auth.decorators import module_required
from invoicing.pdf_parser import parse_pdf, suggest_is_usable


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def get_database_path():
    db_path = current_app.config.get('DATABASE_PATH')
    if db_path:
        return db_path
    return os.path.join(BASE_DIR, 'data', 'main.db')


def get_db_connection():
    conn = sqlite3.connect(get_database_path())
    conn.row_factory = sqlite3.Row
    return conn


PENDING_DIR_NAME = 'invoice_pdfs_pending'
ARCHIVE_DIR_NAME = 'invoice_pdfs'
UNMATCHED_ENTITY_FOLDER = '_unmatched_'


def _data_dir():
    p = Path(BASE_DIR) / 'data'
    p.mkdir(parents=True, exist_ok=True)
    return p


def _pending_dir():
    p = _data_dir() / PENDING_DIR_NAME
    p.mkdir(parents=True, exist_ok=True)
    return p


def _archive_dir():
    p = _data_dir() / ARCHIVE_DIR_NAME
    p.mkdir(parents=True, exist_ok=True)
    return p


def _safe_folder_name(name):
    if not name:
        return UNMATCHED_ENTITY_FOLDER
    cleaned = re.sub(r'[\\/:*?"<>|]', '_', str(name).strip())
    return cleaned or UNMATCHED_ENTITY_FOLDER


def _match_customer_id(conn, raw_name):
    name = (raw_name or '').strip()
    if not name:
        return None
    row = conn.execute("SELECT id FROM customer WHERE short_name = ?", (name,)).fetchone()
    if row:
        return row['id']
    row = conn.execute("SELECT id FROM customer WHERE full_name = ?", (name,)).fetchone()
    if row:
        return row['id']
    alias_rows = conn.execute(
        "SELECT customer_id FROM customer_alias WHERE alias = ? ORDER BY id",
        (name,),
    ).fetchall()
    if len(alias_rows) == 1:
        return alias_rows[0]['customer_id']
    return None


def _match_entity_id(conn, buyer_name):
    if not buyer_name:
        return None
    rows = conn.execute("SELECT id, name FROM billing_entity").fetchall()
    for r in rows:
        ename = (r['name'] or '').strip()
        if ename and ename in buyer_name:
            return r['id']
    return None


def _parse_alias_match_value(raw_value):
    value = (raw_value or '').strip()
    if not value:
        return None, None
    if '::' not in value:
        return None, value
    customer_id_text, alias_name = value.split('::', 1)
    alias_name = alias_name.strip() or None
    try:
        customer_id = int(customer_id_text)
    except ValueError:
        customer_id = None
    return customer_id, alias_name


invoicing_bp = Blueprint('invoicing', __name__, template_folder='../templates')


CUSTOMER_HEADERS = ('达人', '客户', '客户简称', '带货账号昵称', '团长', '账号昵称')
AMOUNT_HEADERS = ('应开金额', '带货费用', '佣金', '求和项:带货费用', '金额')
PLATFORM_HEADERS = ('平台', '平台名称', '店铺', '店铺名称')
PERIOD_HEADERS = ('期间', '账期', '周期')
ENTITY_HEADERS = ('开票主体', '主体')
PERIOD_START_HEADERS = ('period_start', '账期起点', '期间起点', '开始日期')
PERIOD_END_HEADERS = ('period_end', '账期终点', '期间终点', '结束日期')


def normalize_header(value):
    return str(value or '').strip().replace(' ', '').replace('\n', '')


def find_header(headers, candidates):
    normalized_headers = {normalize_header(header): header for header in headers}
    for candidate in candidates:
        header = normalized_headers.get(normalize_header(candidate))
        if header:
            return header
    return None


def cell_text(value):
    if value is None:
        return ''
    return str(value).strip()


def parse_amount(value):
    if value is None or value == '':
        return None
    text = str(value).strip().replace(',', '').replace('¥', '').replace('￥', '')
    try:
        return float(text)
    except ValueError:
        return None


def find_or_create_customer(conn, raw_name, platform):
    name = (raw_name or '').strip()
    platform_name = (platform or '').strip()
    if not name:
        return None, False

    row = conn.execute(
        "SELECT id FROM customer WHERE short_name = ? AND platform = ?",
        (name, platform_name),
    ).fetchone()
    if row:
        return row['id'], False

    row = conn.execute(
        "SELECT id FROM customer WHERE full_name = ?",
        (name,),
    ).fetchone()
    if row:
        return row['id'], False

    alias_rows = conn.execute(
        "SELECT customer_id FROM customer_alias WHERE alias = ? ORDER BY id",
        (name,),
    ).fetchall()
    if len(alias_rows) == 1:
        return alias_rows[0]['customer_id'], False
    if len(alias_rows) > 1:
        return None, False

    cursor = conn.execute(
        "INSERT INTO customer (short_name, platform) VALUES (?, ?)",
        (name, platform_name),
    )
    return cursor.lastrowid, True


def get_entity_id_by_name(conn, entity_name):
    if not entity_name:
        return None
    row = conn.execute(
        "SELECT id FROM billing_entity WHERE name = ?",
        (entity_name,),
    ).fetchone()
    return row['id'] if row else None


# ===== 模块首页 =====

@invoicing_bp.route('/')
@module_required('invoicing')
def index():
    return render_template('invoicing_index.html')


# ===== 应开金额导入与查看 =====

@invoicing_bp.route('/expected-amounts')
@module_required('invoicing')
def expected_amounts():
    with get_db_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                e.id,
                c.short_name AS customer_name,
                b.name AS entity_name,
                e.platform,
                e.period,
                e.period_start,
                e.period_end,
                e.amount,
                e.created_at
            FROM expected_amount e
            LEFT JOIN customer c ON c.id = e.customer_id
            LEFT JOIN billing_entity b ON b.id = e.entity_id
            ORDER BY e.id DESC
            """
        ).fetchall()
        entities = conn.execute(
            "SELECT id, name FROM billing_entity ORDER BY id"
        ).fetchall()
    return render_template(
        'invoicing_expected_amounts.html',
        rows=rows,
        entities=entities,
        result=None,
    )


@invoicing_bp.route('/expected-amounts/import', methods=['POST'])
@module_required('invoicing')
def import_expected_amounts():
    default_entity_id = request.form.get('entity_id')
    default_platform = (request.form.get('platform') or '').strip()
    default_period = (request.form.get('period') or '').strip()
    default_period_start = (request.form.get('period_start') or '').strip() or None
    default_period_end = (request.form.get('period_end') or '').strip() or None
    sheet_name = (request.form.get('sheet_name') or '').strip()
    upload_file = request.files.get('file')

    result = {
        'success': False,
        'message': '',
        'imported_count': 0,
        'created_customer_count': 0,
        'duplicate_skipped_count': 0,
        'skipped_rows': [],
        'duplicate_rows': [],
    }

    if not upload_file or not upload_file.filename:
        result['message'] = '请选择 Excel 文件'
    elif not upload_file.filename.lower().endswith(('.xlsx', '.xlsm')):
        result['message'] = '仅支持 .xlsx / .xlsm 文件'
    elif not default_entity_id:
        result['message'] = '请选择默认归属'
    elif not default_platform:
        result['message'] = '请输入默认店铺/平台'
    elif not default_period:
        result['message'] = '请输入默认期间'
    else:
        try:
            workbook = load_workbook(upload_file, data_only=True)
            worksheet = workbook[sheet_name] if sheet_name else workbook.active
            rows_iter = worksheet.iter_rows(values_only=True)
            raw_headers = next(rows_iter, None)
            if not raw_headers:
                raise ValueError('Excel 表头为空')

            headers = [cell_text(header) for header in raw_headers]
            customer_header = find_header(headers, CUSTOMER_HEADERS)
            amount_header = find_header(headers, AMOUNT_HEADERS)
            platform_header = find_header(headers, PLATFORM_HEADERS)
            period_header = find_header(headers, PERIOD_HEADERS)
            entity_header = find_header(headers, ENTITY_HEADERS)
            period_start_header = find_header(headers, PERIOD_START_HEADERS)
            period_end_header = find_header(headers, PERIOD_END_HEADERS)

            if not customer_header or not amount_header:
                raise ValueError('Excel 缺少达人/客户列或应开金额列')

            with get_db_connection() as conn:
                for row_number, raw_row in enumerate(rows_iter, start=2):
                    row = {
                        headers[index]: raw_row[index] if index < len(raw_row) else None
                        for index in range(len(headers))
                        if headers[index]
                    }
                    customer_name = cell_text(row.get(customer_header))
                    amount = parse_amount(row.get(amount_header))
                    platform = cell_text(row.get(platform_header)) if platform_header else default_platform
                    period = cell_text(row.get(period_header)) if period_header else default_period
                    entity_name = cell_text(row.get(entity_header)) if entity_header else ''
                    entity_id = get_entity_id_by_name(conn, entity_name) if entity_name else int(default_entity_id)
                    period_start = cell_text(row.get(period_start_header)) if period_start_header else default_period_start
                    period_end = cell_text(row.get(period_end_header)) if period_end_header else default_period_end

                    if not customer_name and amount is None:
                        continue
                    if not customer_name or amount is None or not platform or not period or not entity_id:
                        result['skipped_rows'].append({
                            'row_number': row_number,
                            'reason': '缺少达人、金额、店铺/平台、期间或归属',
                            'customer_name': customer_name,
                            'amount': cell_text(row.get(amount_header)),
                            'platform': platform,
                            'period': period,
                            'entity_name': entity_name,
                        })
                        continue

                    customer_id, created_customer = find_or_create_customer(conn, customer_name, platform)
                    if not customer_id:
                        result['skipped_rows'].append({
                            'row_number': row_number,
                            'reason': '客户别名对应多个客户，无法自动判断',
                            'customer_name': customer_name,
                            'amount': cell_text(row.get(amount_header)),
                            'platform': platform,
                            'period': period,
                            'entity_name': entity_name,
                        })
                        continue
                    if created_customer:
                        result['created_customer_count'] += 1

                    existing = conn.execute(
                        """
                        SELECT id FROM expected_amount
                        WHERE customer_id = ?
                          AND entity_id = ?
                          AND platform = ?
                          AND period = ?
                          AND amount = ?
                        """,
                        (customer_id, entity_id, platform, period, amount),
                    ).fetchone()
                    if existing:
                        result['duplicate_skipped_count'] += 1
                        result['duplicate_rows'].append({
                            'row_number': row_number,
                            'customer_name': customer_name,
                            'amount': amount,
                            'platform': platform,
                            'period': period,
                            'entity_name': entity_name,
                        })
                        continue

                    conn.execute(
                        """
                        INSERT INTO expected_amount (
                            customer_id, entity_id, platform, period, amount,
                            period_start, period_end
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            customer_id,
                            entity_id,
                            platform,
                            period,
                            amount,
                            period_start or None,
                            period_end or None,
                        ),
                    )
                    result['imported_count'] += 1

                conn.commit()

            result['success'] = True
            result['message'] = (
                f"成功导入 {result['imported_count']} 条应开金额记录；"
                f"已存在跳过 {result['duplicate_skipped_count']} 条"
            )
        except Exception as exc:
            result['message'] = f'导入失败：{str(exc)}'

    with get_db_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                e.id,
                c.short_name AS customer_name,
                b.name AS entity_name,
                e.platform,
                e.period,
                e.period_start,
                e.period_end,
                e.amount,
                e.created_at
            FROM expected_amount e
            LEFT JOIN customer c ON c.id = e.customer_id
            LEFT JOIN billing_entity b ON b.id = e.entity_id
            ORDER BY e.id DESC
            """
        ).fetchall()
        entities = conn.execute(
            "SELECT id, name FROM billing_entity ORDER BY id"
        ).fetchall()

    return render_template(
        'invoicing_expected_amounts.html',
        rows=rows,
        entities=entities,
        result=result,
    )


# ===== 开票主体 CRUD =====

@invoicing_bp.route('/billing-entities')
@module_required('invoicing')
def billing_entities():
    with get_db_connection() as conn:
        rows = conn.execute(
            "SELECT id, name, created_at, updated_at FROM billing_entity ORDER BY id"
        ).fetchall()
    return render_template('invoicing_billing_entities.html', rows=rows)


@invoicing_bp.route('/billing-entities/create', methods=['POST'])
@module_required('invoicing')
def create_billing_entity():
    name = (request.form.get('name') or '').strip()
    if name:
        with get_db_connection() as conn:
            conn.execute("INSERT INTO billing_entity (name) VALUES (?)", (name,))
            conn.commit()
    return redirect(url_for('invoicing.billing_entities'))


@invoicing_bp.route('/billing-entities/<int:entity_id>/update', methods=['POST'])
@module_required('invoicing')
def update_billing_entity(entity_id):
    name = (request.form.get('name') or '').strip()
    if name:
        with get_db_connection() as conn:
            conn.execute(
                "UPDATE billing_entity SET name = ?, updated_at = datetime('now', 'localtime') WHERE id = ?",
                (name, entity_id),
            )
            conn.commit()
    return redirect(url_for('invoicing.billing_entities'))


@invoicing_bp.route('/billing-entities/<int:entity_id>/delete', methods=['POST'])
@module_required('invoicing')
def delete_billing_entity(entity_id):
    with get_db_connection() as conn:
        conn.execute("DELETE FROM billing_entity WHERE id = ?", (entity_id,))
        conn.commit()
    return redirect(url_for('invoicing.billing_entities'))


# ===== 达人/团长昵称 CRUD =====

@invoicing_bp.route('/customers')
@module_required('invoicing')
def customers():
    with get_db_connection() as conn:
        rows = conn.execute(
            """
            SELECT
                c.short_name,
                COALESCE(a.alias_list, '') AS alias_list,
                COALESCE(a.alias_items, '') AS alias_items,
                COALESCE(SUM(CASE WHEN e.platform = '澳柯' THEN e.amount ELSE 0 END), 0) AS amount_aoke,
                COALESCE(SUM(CASE WHEN e.platform = '香娜露儿' THEN e.amount ELSE 0 END), 0) AS amount_xiangnalu,
                COALESCE(SUM(CASE WHEN e.platform = '快手' THEN e.amount ELSE 0 END), 0) AS amount_kuaishou,
                COALESCE(SUM(CASE WHEN e.platform = '幕莲蔓' THEN e.amount ELSE 0 END), 0) AS amount_mulianman
            FROM customer c
            JOIN expected_amount e ON e.customer_id = c.id AND e.amount <> 0
            LEFT JOIN (
                SELECT
                    short_name,
                    GROUP_CONCAT(alias, ' / ') AS alias_list,
                    GROUP_CONCAT(alias, '|||') AS alias_items
                FROM (
                    SELECT DISTINCT c2.short_name, ca.alias
                    FROM customer_alias ca
                    JOIN customer c2 ON c2.id = ca.customer_id
                    ORDER BY ca.id
                )
                GROUP BY short_name
            ) a ON a.short_name = c.short_name
            GROUP BY c.short_name
            ORDER BY c.short_name
            """
        ).fetchall()
        alias_rows = conn.execute(
            """
            SELECT
                ca.alias,
                GROUP_CONCAT(DISTINCT c.short_name) AS nickname_list,
                COUNT(DISTINCT c.short_name) AS nickname_count,
                COALESCE(SUM(CASE WHEN e.platform = '澳柯' THEN e.amount ELSE 0 END), 0) AS amount_aoke,
                COALESCE(SUM(CASE WHEN e.platform = '香娜露儿' THEN e.amount ELSE 0 END), 0) AS amount_xiangnalu,
                COALESCE(SUM(CASE WHEN e.platform = '快手' THEN e.amount ELSE 0 END), 0) AS amount_kuaishou,
                COALESCE(SUM(CASE WHEN e.platform = '幕莲蔓' THEN e.amount ELSE 0 END), 0) AS amount_mulianman
            FROM customer_alias ca
            JOIN customer c ON c.id = ca.customer_id
            LEFT JOIN expected_amount e ON e.customer_id = c.id
            GROUP BY ca.alias
            ORDER BY ca.alias
            """
        ).fetchall()
    view = request.args.get('view') if request.args.get('view') in ('nickname', 'alias') else 'nickname'
    nickname_totals = {
        'amount_aoke': sum((r['amount_aoke'] or 0) for r in rows),
        'amount_xiangnalu': sum((r['amount_xiangnalu'] or 0) for r in rows),
        'amount_kuaishou': sum((r['amount_kuaishou'] or 0) for r in rows),
        'amount_mulianman': sum((r['amount_mulianman'] or 0) for r in rows),
    }
    alias_totals = {
        'amount_aoke': sum((r['amount_aoke'] or 0) for r in alias_rows),
        'amount_xiangnalu': sum((r['amount_xiangnalu'] or 0) for r in alias_rows),
        'amount_kuaishou': sum((r['amount_kuaishou'] or 0) for r in alias_rows),
        'amount_mulianman': sum((r['amount_mulianman'] or 0) for r in alias_rows),
    }
    return render_template(
        'invoicing_customers.html',
        rows=rows,
        alias_rows=alias_rows,
        view=view,
        nickname_totals=nickname_totals,
        alias_totals=alias_totals,
    )


@invoicing_bp.route('/customers/create', methods=['POST'])
@module_required('invoicing')
def create_customer():
    short_name = (request.form.get('short_name') or '').strip()
    full_name = (request.form.get('full_name') or '').strip() or None
    platform = (request.form.get('platform') or '').strip()
    if short_name:
        with get_db_connection() as conn:
            conn.execute(
                "INSERT INTO customer (short_name, full_name, platform) VALUES (?, ?, ?)",
                (short_name, full_name, platform),
            )
            conn.commit()
    return redirect(url_for('invoicing.customers'))


@invoicing_bp.route('/customers/bulk-alias', methods=['POST'])
@module_required('invoicing')
def bulk_create_alias():
    alias = (request.form.get('alias') or '').strip()
    nicknames = request.form.getlist('nicknames')
    customer_ids = request.form.getlist('customer_ids')
    added_count = 0
    skipped_count = 0

    if alias and (nicknames or customer_ids):
        with get_db_connection() as conn:
            target_ids = []
            if nicknames:
                for nickname in nicknames:
                    rows = conn.execute(
                        "SELECT id FROM customer WHERE short_name = ? ORDER BY platform, id",
                        (nickname,),
                    ).fetchall()
                    target_ids.extend([r['id'] for r in rows])
            else:
                target_ids = customer_ids

            for raw_id in target_ids:
                try:
                    customer_id = int(raw_id)
                except (TypeError, ValueError):
                    skipped_count += 1
                    continue

                exists = conn.execute(
                    "SELECT id FROM customer_alias WHERE customer_id = ? AND alias = ?",
                    (customer_id, alias),
                ).fetchone()
                if exists:
                    skipped_count += 1
                    continue

                conn.execute(
                    "INSERT INTO customer_alias (customer_id, alias) VALUES (?, ?)",
                    (customer_id, alias),
                )
                added_count += 1
            conn.commit()

    return redirect(url_for(
        'invoicing.customers',
        bulk_added=added_count,
        bulk_skipped=skipped_count,
        bulk_alias=alias,
    ))


@invoicing_bp.route('/customers/alias/delete-for-nickname', methods=['POST'])
@module_required('invoicing')
def delete_alias_for_nickname():
    nickname = (request.form.get('nickname') or '').strip()
    alias = (request.form.get('alias') or '').strip()
    if nickname and alias:
        with get_db_connection() as conn:
            conn.execute(
                """
                DELETE FROM customer_alias
                WHERE alias = ?
                  AND customer_id IN (
                      SELECT id FROM customer WHERE short_name = ?
                  )
                """,
                (alias, nickname),
            )
            conn.commit()
    return redirect(url_for('invoicing.customers'))


@invoicing_bp.route('/customers/alias/rename', methods=['POST'])
@module_required('invoicing')
def rename_alias():
    old_alias = (request.form.get('old_alias') or '').strip()
    new_alias = (request.form.get('new_alias') or '').strip()
    if old_alias and new_alias and old_alias != new_alias:
        with get_db_connection() as conn:
            rows = conn.execute(
                "SELECT id, customer_id FROM customer_alias WHERE alias = ? ORDER BY id",
                (old_alias,),
            ).fetchall()
            for row in rows:
                exists = conn.execute(
                    "SELECT id FROM customer_alias WHERE customer_id = ? AND alias = ?",
                    (row['customer_id'], new_alias),
                ).fetchone()
                if exists:
                    conn.execute("DELETE FROM customer_alias WHERE id = ?", (row['id'],))
                else:
                    conn.execute("UPDATE customer_alias SET alias = ? WHERE id = ?", (new_alias, row['id']))
            conn.commit()
    return redirect(url_for('invoicing.customers', view='alias'))


@invoicing_bp.route('/customers/alias/delete', methods=['POST'])
@module_required('invoicing')
def delete_alias_group():
    alias = (request.form.get('alias') or '').strip()
    if alias:
        with get_db_connection() as conn:
            conn.execute("DELETE FROM customer_alias WHERE alias = ?", (alias,))
            conn.commit()
    return redirect(url_for('invoicing.customers', view='alias'))


@invoicing_bp.route('/customers/<int:customer_id>')
@module_required('invoicing')
def customer_detail(customer_id):
    with get_db_connection() as conn:
        customer = conn.execute(
            "SELECT id, short_name, full_name, created_at, updated_at FROM customer WHERE id = ?",
            (customer_id,),
        ).fetchone()
        if not customer:
            return redirect(url_for('invoicing.customers'))
        aliases = conn.execute(
            "SELECT id, alias, created_at FROM customer_alias WHERE customer_id = ? ORDER BY id",
            (customer_id,),
        ).fetchall()
    return render_template(
        'invoicing_customer_detail.html',
        customer=customer,
        aliases=aliases,
    )


@invoicing_bp.route('/customers/<int:customer_id>/update', methods=['POST'])
@module_required('invoicing')
def update_customer(customer_id):
    short_name = (request.form.get('short_name') or '').strip()
    full_name = (request.form.get('full_name') or '').strip() or None
    if short_name:
        with get_db_connection() as conn:
            conn.execute(
                "UPDATE customer SET short_name = ?, full_name = ?, updated_at = datetime('now', 'localtime') WHERE id = ?",
                (short_name, full_name, customer_id),
            )
            conn.commit()
    return redirect(url_for('invoicing.customer_detail', customer_id=customer_id))


@invoicing_bp.route('/customers/<int:customer_id>/delete', methods=['POST'])
@module_required('invoicing')
def delete_customer(customer_id):
    with get_db_connection() as conn:
        # 外键未启用，需手动级联删除别名
        conn.execute("DELETE FROM customer_alias WHERE customer_id = ?", (customer_id,))
        conn.execute("DELETE FROM customer WHERE id = ?", (customer_id,))
        conn.commit()
    return redirect(url_for('invoicing.customers'))


# ===== 客户别名 =====

@invoicing_bp.route('/customers/<int:customer_id>/aliases/create', methods=['POST'])
@module_required('invoicing')
def create_alias(customer_id):
    alias = (request.form.get('alias') or '').strip()
    if alias:
        with get_db_connection() as conn:
            exists = conn.execute(
                "SELECT id FROM customer_alias WHERE customer_id = ? AND alias = ?",
                (customer_id, alias),
            ).fetchone()
            if not exists:
                conn.execute(
                    "INSERT INTO customer_alias (customer_id, alias) VALUES (?, ?)",
                    (customer_id, alias),
                )
                conn.commit()
    return redirect(url_for('invoicing.customer_detail', customer_id=customer_id))


@invoicing_bp.route('/aliases/<int:alias_id>/delete', methods=['POST'])
@module_required('invoicing')
def delete_alias(alias_id):
    with get_db_connection() as conn:
        row = conn.execute(
            "SELECT customer_id FROM customer_alias WHERE id = ?",
            (alias_id,),
        ).fetchone()
        if row is None:
            return redirect(url_for('invoicing.customers'))
        customer_id = row['customer_id']
        conn.execute("DELETE FROM customer_alias WHERE id = ?", (alias_id,))
        conn.commit()
    return redirect(url_for('invoicing.customer_detail', customer_id=customer_id))


# ===== 发票 上传 + 复核 + 列表 =====

@invoicing_bp.route('/invoices')
@module_required('invoicing')
def invoices_list():
    only_unmatched = request.args.get('filter') == 'unmatched'
    usable_filter = (request.args.get('usable') or '').strip()
    with get_db_connection() as conn:
        sql = """
            SELECT i.id, i.invoice_number, i.invoice_date, i.amount,
                   i.invoice_type, i.tax_rate,
                   i.seller_name, i.buyer_name, i.project_name,
                   i.alias_name,
                   i.pdf_remark, i.is_usable, i.customer_id, i.entity_id,
                   i.pdf_file_path, i.qr_content, i.created_at,
                   c.short_name AS customer_short_name
              FROM invoice i
              LEFT JOIN customer c ON c.id = i.customer_id
        """
        where_parts = []
        params = []
        if only_unmatched:
            where_parts.append("i.customer_id IS NULL AND (i.alias_name IS NULL OR i.alias_name = '')")
        if usable_filter in ('0', '1'):
            where_parts.append("i.is_usable = ?")
            params.append(int(usable_filter))
        if where_parts:
            sql += " WHERE " + " AND ".join(where_parts)
        sql += " ORDER BY i.id DESC"
        rows = conn.execute(sql, params).fetchall()
    return render_template(
        'invoicing_invoices.html',
        rows=rows,
        only_unmatched=only_unmatched,
        usable_filter=usable_filter,
    )


@invoicing_bp.route('/invoices/upload', methods=['GET'])
@module_required('invoicing')
def invoices_upload_form():
    return render_template('invoicing_invoices_upload.html', error=None)


@invoicing_bp.route('/invoices/upload', methods=['POST'])
@module_required('invoicing')
def invoices_upload_submit():
    pdf_file = request.files.get('pdf_file')
    if not pdf_file or not pdf_file.filename:
        return render_template('invoicing_invoices_upload.html', error='请选择 PDF 文件')
    if not pdf_file.filename.lower().endswith('.pdf'):
        return render_template('invoicing_invoices_upload.html', error='仅支持 .pdf 文件')

    pending_id = uuid.uuid4().hex
    pending_pdf = _pending_dir() / f'{pending_id}.pdf'
    pdf_file.save(str(pending_pdf))

    try:
        parsed = parse_pdf(str(pending_pdf))
    except Exception as exc:
        pending_pdf.unlink(missing_ok=True)
        return render_template('invoicing_invoices_upload.html', error=f'解析失败：{exc}')

    with get_db_connection() as conn:
        suggested_customer_id = _match_customer_id(conn, parsed.get('seller_name'))
        suggested_entity_id = _match_entity_id(conn, parsed.get('buyer_name'))
        duplicate_id = None
        invn = (parsed.get('invoice_number') or '').strip()
        if invn:
            row = conn.execute("SELECT id FROM invoice WHERE invoice_number = ?", (invn,)).fetchone()
            if row:
                duplicate_id = row['id']

    parsed['suggested_customer_id'] = suggested_customer_id
    parsed['suggested_entity_id'] = suggested_entity_id
    parsed['suggested_is_usable'] = suggest_is_usable(parsed.get('project_name'), parsed.get('pdf_remark'))
    parsed['duplicate_existing_invoice_id'] = duplicate_id

    json_path = _pending_dir() / f'{pending_id}.json'
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(parsed, f, ensure_ascii=False, indent=2)

    return redirect(url_for('invoicing.invoices_review', pending_id=pending_id))


@invoicing_bp.route('/invoices/review/<pending_id>')
@module_required('invoicing')
def invoices_review(pending_id):
    pdf_path = _pending_dir() / f'{pending_id}.pdf'
    json_path = _pending_dir() / f'{pending_id}.json'
    if not pdf_path.exists() or not json_path.exists():
        return redirect(url_for('invoicing.invoices_list'))
    with open(json_path, encoding='utf-8') as f:
        parsed = json.load(f)
    current_amount = parsed.get('amount') or 0
    with get_db_connection() as conn:
        customers = conn.execute(
            """
            SELECT
                c.id,
                c.short_name,
                COALESCE(e.platform, c.platform) AS platform,
                COALESCE(e.period, '') AS period,
                COALESCE(SUM(e.amount), 0) AS expected_total,
                COALESCE((
                    SELECT SUM(i.amount)
                    FROM invoice i
                    WHERE i.is_usable = 1
                      AND i.customer_id = c.id
                      AND (i.alias_name IS NULL OR i.alias_name = '')
                ), 0) AS invoiced_total,
                COALESCE(SUM(e.amount), 0) - COALESCE((
                    SELECT SUM(i.amount)
                    FROM invoice i
                    WHERE i.is_usable = 1
                      AND i.customer_id = c.id
                      AND (i.alias_name IS NULL OR i.alias_name = '')
                ), 0) AS remaining_total,
                COALESCE(a.alias_list, '') AS alias_list
            FROM customer c
            JOIN expected_amount e ON e.customer_id = c.id AND e.amount <> 0
            LEFT JOIN (
                SELECT customer_id, GROUP_CONCAT(alias, ' / ') AS alias_list
                FROM (
                    SELECT customer_id, alias
                    FROM customer_alias
                    ORDER BY id
                )
                GROUP BY customer_id
            ) a ON a.customer_id = c.id
            GROUP BY c.id, COALESCE(e.platform, c.platform), COALESCE(e.period, '')
            ORDER BY
                COALESCE(e.platform, c.platform),
                CASE WHEN COALESCE(a.alias_list, '') <> '' THEN 0 ELSE 1 END,
                (COALESCE(SUM(e.amount), 0) - ?) ASC,
                COALESCE(e.period, '') DESC,
                c.short_name
            """,
            (current_amount,),
        ).fetchall()
        entities = conn.execute(
            "SELECT id, name FROM billing_entity ORDER BY name"
        ).fetchall()
        aliases = conn.execute(
            """
            SELECT
                ca.alias,
                c.id AS customer_id,
                c.short_name,
                COALESCE(e.platform, c.platform) AS platform,
                COALESCE(e.period, '') AS period,
                COALESCE(SUM(e.amount), 0) AS expected_total,
                COALESCE((
                    SELECT SUM(i.amount)
                    FROM invoice i
                    WHERE i.is_usable = 1
                      AND i.alias_name = ca.alias
                      AND i.customer_id = c.id
                ), 0) AS invoiced_total,
                COALESCE(SUM(e.amount), 0) - COALESCE((
                    SELECT SUM(i.amount)
                    FROM invoice i
                    WHERE i.is_usable = 1
                      AND i.alias_name = ca.alias
                      AND i.customer_id = c.id
                ), 0) AS remaining_total,
                COUNT(*) AS expected_count
            FROM customer_alias ca
            JOIN customer c ON c.id = ca.customer_id
            JOIN expected_amount e ON e.customer_id = c.id AND e.amount <> 0
            GROUP BY ca.alias, c.id, c.short_name, COALESCE(e.platform, c.platform), COALESCE(e.period, '')
            ORDER BY
                COALESCE(e.platform, c.platform),
                (COALESCE(SUM(e.amount), 0) - ?) ASC,
                COALESCE(e.period, '') DESC,
                c.short_name,
                ca.alias
            """,
            (current_amount,),
        ).fetchall()
    return render_template(
        'invoicing_invoices_review.html',
        pending_id=pending_id,
        parsed=parsed,
        customers=customers,
        entities=entities,
        aliases=aliases,
        duplicate_warning=request.args.get('duplicate') == '1',
    )


@invoicing_bp.route('/invoices/review/<pending_id>/confirm', methods=['POST'])
@module_required('invoicing')
def invoices_review_confirm(pending_id):
    pdf_path = _pending_dir() / f'{pending_id}.pdf'
    json_path = _pending_dir() / f'{pending_id}.json'
    if not pdf_path.exists():
        return redirect(url_for('invoicing.invoices_list'))

    invoice_number = (request.form.get('invoice_number') or '').strip()
    invoice_date = (request.form.get('invoice_date') or '').strip() or None
    invoice_type = (request.form.get('invoice_type') or '').strip() or None
    tax_rate = (request.form.get('tax_rate') or '').strip() or None
    amount_str = (request.form.get('amount') or '').strip()
    seller_name = (request.form.get('seller_name') or '').strip() or None
    buyer_name = (request.form.get('buyer_name') or '').strip() or None
    project_name = (request.form.get('project_name') or '').strip() or None
    pdf_remark = (request.form.get('pdf_remark') or '').strip() or None
    qr_content = (request.form.get('qr_content') or '').strip() or None
    customer_id_raw = (request.form.get('customer_id') or '').strip()
    alias_customer_id, alias_name = _parse_alias_match_value(request.form.get('alias_name'))
    entity_id_raw = (request.form.get('entity_id') or '').strip()
    is_usable = 1 if (request.form.get('is_usable') == '1') else 0

    try:
        amount = float(amount_str) if amount_str else None
    except ValueError:
        amount = None

    customer_id = int(customer_id_raw) if customer_id_raw else None
    if alias_name:
        customer_id = alias_customer_id
    entity_id = int(entity_id_raw) if entity_id_raw else None

    if not invoice_number:
        return redirect(url_for('invoicing.invoices_review', pending_id=pending_id))

    with get_db_connection() as conn:
        if conn.execute("SELECT id FROM invoice WHERE invoice_number = ?", (invoice_number,)).fetchone():
            return redirect(url_for('invoicing.invoices_review', pending_id=pending_id) + '?duplicate=1')

        if entity_id:
            er = conn.execute("SELECT name FROM billing_entity WHERE id = ?", (entity_id,)).fetchone()
            entity_folder = _safe_folder_name(er['name'] if er else '')
        else:
            entity_folder = UNMATCHED_ENTITY_FOLDER

        year_folder = (invoice_date or '')[:4] or 'unknown_year'
        target_dir = _archive_dir() / entity_folder / year_folder
        target_dir.mkdir(parents=True, exist_ok=True)

        final_name = f'{invoice_number}.pdf'
        final_path = target_dir / final_name
        if final_path.exists():
            final_name = f'{invoice_number}_{uuid.uuid4().hex[:8]}.pdf'
            final_path = target_dir / final_name

        shutil.move(str(pdf_path), str(final_path))
        if json_path.exists():
            json_path.unlink(missing_ok=True)

        relative_path = str(final_path.relative_to(Path(BASE_DIR)))

        conn.execute("""
            INSERT INTO invoice (
                invoice_number, invoice_date, customer_id, entity_id,
                alias_name,
                amount, invoice_type, tax_rate, seller_name, buyer_name,
                pdf_file_path, qr_content, manual_confirmed,
                project_name, pdf_remark, is_usable
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?, ?)
        """, (
            invoice_number, invoice_date, customer_id, entity_id,
            alias_name,
            amount, invoice_type, tax_rate, seller_name, buyer_name,
            relative_path, qr_content,
            project_name, pdf_remark, is_usable,
        ))
        conn.commit()

    return redirect(url_for('invoicing.invoices_list'))


@invoicing_bp.route('/invoices/review/<pending_id>/discard', methods=['POST'])
@module_required('invoicing')
def invoices_review_discard(pending_id):
    pdf_path = _pending_dir() / f'{pending_id}.pdf'
    json_path = _pending_dir() / f'{pending_id}.json'
    pdf_path.unlink(missing_ok=True)
    json_path.unlink(missing_ok=True)
    return redirect(url_for('invoicing.invoices_upload_form'))


@invoicing_bp.route('/invoices/<int:invoice_id>/pdf')
@module_required('invoicing')
def invoice_pdf(invoice_id):
    with get_db_connection() as conn:
        row = conn.execute("SELECT pdf_file_path FROM invoice WHERE id = ?", (invoice_id,)).fetchone()
    if not row or not row['pdf_file_path']:
        return '', 404
    abs_path = Path(BASE_DIR) / row['pdf_file_path']
    if not abs_path.exists():
        return '', 404
    return send_file(str(abs_path), mimetype='application/pdf')


@invoicing_bp.route('/invoices/pending/<pending_id>/pdf')
@module_required('invoicing')
def invoice_pdf_pending(pending_id):
    pdf_path = _pending_dir() / f'{pending_id}.pdf'
    if not pdf_path.exists():
        return '', 404
    return send_file(str(pdf_path), mimetype='application/pdf')


@invoicing_bp.route('/invoices/<int:invoice_id>/match', methods=['GET', 'POST'])
@module_required('invoicing')
def invoice_match(invoice_id):
    if request.method == 'GET':
        with get_db_connection() as conn:
            invoice = conn.execute(
                """
                SELECT i.id, i.invoice_number, i.invoice_date, i.amount,
                       i.invoice_type, i.tax_rate, i.seller_name, i.buyer_name,
                       i.project_name, i.customer_id, i.alias_name, i.is_usable,
                       c.short_name AS customer_short_name
                FROM invoice i
                LEFT JOIN customer c ON c.id = i.customer_id
                WHERE i.id = ?
                """,
                (invoice_id,),
            ).fetchone()
            if not invoice:
                return redirect(url_for('invoicing.invoices_list'))
            current_amount = invoice['amount'] or 0
            customers = conn.execute(
                """
                SELECT
                    c.id,
                    c.short_name,
                    COALESCE(e.platform, c.platform) AS platform,
                    COALESCE(e.period, '') AS period,
                    COALESCE(SUM(e.amount), 0) AS expected_total,
                    COALESCE((
                        SELECT SUM(i.amount)
                        FROM invoice i
                        WHERE i.is_usable = 1
                          AND i.customer_id = c.id
                          AND (i.alias_name IS NULL OR i.alias_name = '')
                    ), 0) AS invoiced_total,
                    COALESCE(SUM(e.amount), 0) - COALESCE((
                        SELECT SUM(i.amount)
                        FROM invoice i
                        WHERE i.is_usable = 1
                          AND i.customer_id = c.id
                          AND (i.alias_name IS NULL OR i.alias_name = '')
                    ), 0) AS remaining_total,
                    COALESCE(a.alias_list, '') AS alias_list
                FROM customer c
                JOIN expected_amount e ON e.customer_id = c.id AND e.amount <> 0
                LEFT JOIN (
                    SELECT customer_id, GROUP_CONCAT(alias, ' / ') AS alias_list
                    FROM (
                        SELECT customer_id, alias
                        FROM customer_alias
                        ORDER BY id
                    )
                    GROUP BY customer_id
                ) a ON a.customer_id = c.id
                GROUP BY c.id, COALESCE(e.platform, c.platform), COALESCE(e.period, '')
                ORDER BY
                    COALESCE(e.platform, c.platform),
                    CASE WHEN COALESCE(a.alias_list, '') <> '' THEN 0 ELSE 1 END,
                    (COALESCE(SUM(e.amount), 0) - ?) ASC,
                    COALESCE(e.period, '') DESC,
                    c.short_name
                """,
                (current_amount,),
            ).fetchall()
            aliases = conn.execute(
                """
                SELECT
                    ca.alias,
                    c.id AS customer_id,
                    c.short_name,
                    COALESCE(e.platform, c.platform) AS platform,
                    COALESCE(e.period, '') AS period,
                    COALESCE(SUM(e.amount), 0) AS expected_total,
                    COALESCE((
                        SELECT SUM(i.amount)
                        FROM invoice i
                        WHERE i.is_usable = 1
                          AND i.alias_name = ca.alias
                          AND i.customer_id = c.id
                    ), 0) AS invoiced_total,
                    COALESCE(SUM(e.amount), 0) - COALESCE((
                        SELECT SUM(i.amount)
                        FROM invoice i
                        WHERE i.is_usable = 1
                          AND i.alias_name = ca.alias
                          AND i.customer_id = c.id
                    ), 0) AS remaining_total
                FROM customer_alias ca
                JOIN customer c ON c.id = ca.customer_id
                JOIN expected_amount e ON e.customer_id = c.id AND e.amount <> 0
                GROUP BY ca.alias, c.id, c.short_name, COALESCE(e.platform, c.platform), COALESCE(e.period, '')
                ORDER BY
                    COALESCE(e.platform, c.platform),
                    (COALESCE(SUM(e.amount), 0) - ?) ASC,
                    COALESCE(e.period, '') DESC,
                    c.short_name,
                    ca.alias
                """,
                (current_amount,),
            ).fetchall()
            entities = conn.execute(
                "SELECT id, name FROM billing_entity ORDER BY name"
            ).fetchall()
        return render_template(
            'invoicing_invoice_match.html',
            invoice=invoice,
            customers=customers,
            aliases=aliases,
            entities=entities,
        )

    customer_id_raw = (request.form.get('customer_id') or '').strip()
    customer_id = int(customer_id_raw) if customer_id_raw else None
    alias_customer_id, alias_name = _parse_alias_match_value(request.form.get('alias_name'))
    if alias_name:
        customer_id = alias_customer_id
    is_usable_raw = request.form.get('is_usable')
    next_url = (request.form.get('next') or '').strip()
    with get_db_connection() as conn:
        if is_usable_raw is not None:
            is_usable = 1 if is_usable_raw == '1' else 0
            conn.execute(
                "UPDATE invoice SET customer_id = ?, alias_name = ?, is_usable = ?, updated_at = datetime('now', 'localtime') WHERE id = ?",
                (customer_id, alias_name, is_usable, invoice_id),
            )
        else:
            conn.execute(
                "UPDATE invoice SET customer_id = ?, alias_name = ?, updated_at = datetime('now', 'localtime') WHERE id = ?",
                (customer_id, alias_name, invoice_id),
            )
        conn.commit()
    return redirect(next_url or url_for('invoicing.invoices_list'))


@invoicing_bp.route('/invoices/<int:invoice_id>/delete', methods=['POST'])
@module_required('invoicing')
def invoice_delete(invoice_id):
    with get_db_connection() as conn:
        row = conn.execute("SELECT pdf_file_path FROM invoice WHERE id = ?", (invoice_id,)).fetchone()
        if row and row['pdf_file_path']:
            try:
                (Path(BASE_DIR) / row['pdf_file_path']).unlink(missing_ok=True)
            except Exception:
                pass
        conn.execute("DELETE FROM invoice WHERE id = ?", (invoice_id,))
        conn.commit()
    return redirect(url_for('invoicing.invoices_list'))


# ===== 应开 vs 已开核对 =====

@module_required('invoicing')
def old_reconciliation():
    start_date = (request.args.get('start_date') or '').strip() or None
    end_date = (request.args.get('end_date') or '').strip() or None
    platform_keys = ('澳柯', '香娜露儿', '快手', '幕莲蔓')

    expected_where = "1=1"
    expected_params = []
    if end_date:
        expected_where += " AND (period_start IS NULL OR period_start <= ?)"
        expected_params.append(end_date)
    if start_date:
        expected_where += " AND (period_end IS NULL OR period_end >= ?)"
        expected_params.append(start_date)

    invoiced_where = "is_usable = 1"
    invoiced_params = []
    if start_date:
        invoiced_where += " AND invoice_date >= ?"
        invoiced_params.append(start_date)
    if end_date:
        invoiced_where += " AND invoice_date <= ?"
        invoiced_params.append(end_date)

    expected_sql = f"""
        SELECT platform,
               COALESCE(SUM(amount), 0) AS expected_total,
               COUNT(*) AS expected_count
        FROM expected_amount
        WHERE {expected_where}
        GROUP BY platform
    """

    invoiced_sql = f"""
        WITH alias_platform AS (
            SELECT alias,
                   CASE WHEN COUNT(DISTINCT platform) = 1 THEN MIN(platform) ELSE NULL END AS platform
            FROM (
                SELECT ca.alias, COALESCE(e.platform, c.platform) AS platform
                FROM customer_alias ca
                JOIN customer c ON c.id = ca.customer_id
                LEFT JOIN expected_amount e ON e.customer_id = c.id
                WHERE COALESCE(e.platform, c.platform, '') <> ''
            )
            GROUP BY alias
        ),
        invoice_with_platform AS (
	            SELECT i.id, i.amount,
	                   COALESCE(
	                       NULLIF(i.platform, ''),
	                       c.platform,
	                       ap.platform,
	                       CASE
	                           WHEN i.buyer_name LIKE '%澳柯%' THEN '澳柯'
	                           WHEN i.buyer_name LIKE '%香娜露儿%' THEN '香娜露儿'
	                           WHEN i.buyer_name LIKE '%快手%' THEN '快手'
	                           WHEN i.buyer_name LIKE '%幕莲蔓%' THEN '幕莲蔓'
	                       END
	                   ) AS platform
            FROM invoice i
            LEFT JOIN customer c ON c.id = i.customer_id
            LEFT JOIN alias_platform ap ON ap.alias = i.alias_name
            WHERE {invoiced_where}
        )
        SELECT platform,
               COALESCE(SUM(amount), 0) AS invoiced_total,
               COUNT(*) AS invoiced_count
        FROM invoice_with_platform
        WHERE platform IS NOT NULL AND platform <> ''
        GROUP BY platform
    """

    expected_detail_sql = f"""
        WITH customer_alias_one AS (
            SELECT customer_id, MIN(alias) AS alias
            FROM customer_alias
            GROUP BY customer_id
        )
        SELECT
            e.platform,
            CASE WHEN ca.alias IS NOT NULL AND ca.alias <> '' THEN 'alias:' || ca.alias ELSE 'customer:' || c.id END AS group_key,
            COALESCE(ca.alias, '') AS alias_name,
            CASE
                WHEN ca.alias IS NOT NULL AND ca.alias <> '' THEN GROUP_CONCAT(DISTINCT c.short_name)
                ELSE c.short_name
            END AS nickname_list,
            COALESCE(SUM(e.amount), 0) AS expected_total,
            COUNT(*) AS expected_count,
            GROUP_CONCAT(COALESCE(e.period, '未设期间') || '::' || printf('%.2f', e.amount), '|||') AS expected_items
        FROM expected_amount e
        JOIN customer c ON c.id = e.customer_id
        LEFT JOIN customer_alias_one ca ON ca.customer_id = c.id
        WHERE {expected_where}
        GROUP BY e.platform, group_key, alias_name
    """

    invoiced_detail_sql = f"""
        WITH customer_alias_one AS (
            SELECT customer_id, MIN(alias) AS alias
            FROM customer_alias
            GROUP BY customer_id
        ),
        alias_platform AS (
            SELECT alias,
                   CASE WHEN COUNT(DISTINCT platform) = 1 THEN MIN(platform) ELSE NULL END AS platform
            FROM (
                SELECT ca.alias, COALESCE(e.platform, c.platform) AS platform
                FROM customer_alias ca
                JOIN customer c ON c.id = ca.customer_id
                LEFT JOIN expected_amount e ON e.customer_id = c.id
                WHERE COALESCE(e.platform, c.platform, '') <> ''
            )
            GROUP BY alias
        ),
        invoice_base AS (
            SELECT
                i.id,
                i.invoice_number,
                i.amount,
	                COALESCE(
	                    NULLIF(i.platform, ''),
	                    c.platform,
	                    ap.platform,
	                    CASE
	                        WHEN i.buyer_name LIKE '%澳柯%' THEN '澳柯'
	                        WHEN i.buyer_name LIKE '%香娜露儿%' THEN '香娜露儿'
	                        WHEN i.buyer_name LIKE '%快手%' THEN '快手'
	                        WHEN i.buyer_name LIKE '%幕莲蔓%' THEN '幕莲蔓'
	                    END
	                ) AS platform,
                COALESCE(NULLIF(i.alias_name, ''), ca.alias, '') AS alias_name,
                c.id AS customer_id,
                c.short_name AS short_name
            FROM invoice i
            LEFT JOIN customer c ON c.id = i.customer_id
            LEFT JOIN customer_alias_one ca ON ca.customer_id = c.id
            LEFT JOIN alias_platform ap ON ap.alias = i.alias_name
            WHERE {invoiced_where}
        )
        SELECT
            b.platform,
            CASE
                WHEN b.alias_name <> '' THEN 'alias:' || b.alias_name
                WHEN b.customer_id IS NULL THEN 'unmatched'
                ELSE 'customer:' || b.customer_id
            END AS group_key,
            b.alias_name,
            CASE
                WHEN b.alias_name <> '' THEN (
                    SELECT GROUP_CONCAT(DISTINCT c2.short_name)
                    FROM customer_alias ca2
                    JOIN customer c2 ON c2.id = ca2.customer_id
                    WHERE ca2.alias = b.alias_name
	                      AND (b.platform IS NULL OR c2.platform = b.platform)
	                )
                WHEN b.customer_id IS NULL THEN '未匹配 ' || printf('%.2f', COALESCE(SUM(b.amount), 0))
                ELSE COALESCE(MAX(b.short_name), '')
            END AS nickname_list,
            COALESCE(SUM(b.amount), 0) AS invoiced_total,
            COUNT(*) AS invoiced_count,
            GROUP_CONCAT(b.id || '::' || b.invoice_number || '::' || printf('%.2f', b.amount), '|||') AS invoice_items
        FROM invoice_base b
        WHERE b.platform IS NOT NULL AND b.platform <> ''
        GROUP BY b.platform, group_key, b.alias_name
    """

    unmatched_sql = """
        SELECT COALESCE(SUM(amount), 0) AS total, COUNT(*) AS cnt
        FROM invoice
        WHERE is_usable = 1
          AND customer_id IS NULL
          AND (alias_name IS NULL OR alias_name = '')
    """
    unmatched_params = []
    if start_date:
        unmatched_sql += " AND invoice_date >= ?"
        unmatched_params.append(start_date)
    if end_date:
        unmatched_sql += " AND invoice_date <= ?"
        unmatched_params.append(end_date)

    unassigned_sql = f"""
        WITH alias_platform AS (
            SELECT alias,
                   CASE WHEN COUNT(DISTINCT platform) = 1 THEN MIN(platform) ELSE NULL END AS platform
            FROM (
                SELECT ca.alias, COALESCE(e.platform, c.platform) AS platform
                FROM customer_alias ca
                JOIN customer c ON c.id = ca.customer_id
                LEFT JOIN expected_amount e ON e.customer_id = c.id
                WHERE COALESCE(e.platform, c.platform, '') <> ''
            )
            GROUP BY alias
        ),
        invoice_with_platform AS (
	            SELECT i.id, i.amount,
	                   COALESCE(
	                       NULLIF(i.platform, ''),
	                       c.platform,
	                       ap.platform,
	                       CASE
	                           WHEN i.buyer_name LIKE '%澳柯%' THEN '澳柯'
	                           WHEN i.buyer_name LIKE '%香娜露儿%' THEN '香娜露儿'
	                           WHEN i.buyer_name LIKE '%快手%' THEN '快手'
	                           WHEN i.buyer_name LIKE '%幕莲蔓%' THEN '幕莲蔓'
	                       END
	                   ) AS platform
            FROM invoice i
            LEFT JOIN customer c ON c.id = i.customer_id
            LEFT JOIN alias_platform ap ON ap.alias = i.alias_name
            WHERE {invoiced_where}
              AND NOT (i.customer_id IS NULL AND (i.alias_name IS NULL OR i.alias_name = ''))
        )
        SELECT COALESCE(SUM(amount), 0) AS total, COUNT(*) AS cnt
        FROM invoice_with_platform
        WHERE platform IS NULL OR platform = ''
    """

    transfer_sql = f"""
        WITH alias_platform AS (
            SELECT alias,
                   CASE WHEN COUNT(DISTINCT platform) = 1 THEN MIN(platform) ELSE NULL END AS platform
            FROM (
                SELECT ca.alias, COALESCE(e.platform, c.platform) AS platform
                FROM customer_alias ca
                JOIN customer c ON c.id = ca.customer_id
                LEFT JOIN expected_amount e ON e.customer_id = c.id
                WHERE COALESCE(e.platform, c.platform, '') <> ''
            )
            GROUP BY alias
        ),
        invoice_base AS (
            SELECT
                i.id,
                i.invoice_number,
                i.amount,
                COALESCE(NULLIF(i.platform, ''), c.platform, ap.platform) AS source_platform,
                CASE
                    WHEN i.buyer_name LIKE '%澳柯%' THEN '澳柯'
                    WHEN i.buyer_name LIKE '%香娜露儿%' THEN '香娜露儿'
                    WHEN i.buyer_name LIKE '%快手%' THEN '快手'
                    WHEN i.buyer_name LIKE '%幕莲蔓%' THEN '幕莲蔓'
                END AS billing_platform,
                COALESCE(NULLIF(i.alias_name, ''), ca.alias, '') AS alias_name,
                c.short_name AS short_name
            FROM invoice i
            LEFT JOIN customer c ON c.id = i.customer_id
            LEFT JOIN (
                SELECT customer_id, MIN(alias) AS alias
                FROM customer_alias
                GROUP BY customer_id
            ) ca ON ca.customer_id = c.id
            LEFT JOIN alias_platform ap ON ap.alias = i.alias_name
            WHERE {invoiced_where}
        )
        SELECT
            billing_platform AS platform,
            source_platform,
            alias_name,
            short_name,
            COALESCE(SUM(amount), 0) AS invoiced_total,
            COUNT(*) AS invoiced_count,
            GROUP_CONCAT(id || '::' || invoice_number || '::' || printf('%.2f', amount), '|||') AS invoice_items
        FROM invoice_base
        WHERE billing_platform IS NOT NULL
          AND source_platform IS NOT NULL
          AND billing_platform <> source_platform
        GROUP BY billing_platform, source_platform, alias_name, short_name
    """

    with get_db_connection() as conn:
        expected_rows = conn.execute(expected_sql, expected_params).fetchall()
        invoiced_rows = conn.execute(invoiced_sql, invoiced_params).fetchall()
        expected_detail_rows = conn.execute(expected_detail_sql, expected_params).fetchall()
        invoiced_detail_rows = conn.execute(invoiced_detail_sql, invoiced_params).fetchall()
        transfer_rows = conn.execute(transfer_sql, invoiced_params).fetchall()
        unmatched = conn.execute(unmatched_sql, unmatched_params).fetchone()
        unassigned = conn.execute(unassigned_sql, invoiced_params).fetchone()

    expected_map = {r['platform']: r for r in expected_rows}
    invoiced_map = {r['platform']: r for r in invoiced_rows}
    detail_map = {}
    for r in expected_detail_rows:
        key = (r['platform'], r['group_key'])
        detail_map.setdefault(key, {
            'platform': r['platform'],
            'alias_name': r['alias_name'] or '',
            'nickname_list': r['nickname_list'] or '',
            'expected_total': 0,
            'expected_count': 0,
            'expected_items': '',
            'invoiced_total': 0,
            'invoiced_count': 0,
            'invoice_items': '',
        })
        detail_map[key]['expected_total'] += r['expected_total'] or 0
        detail_map[key]['expected_count'] += r['expected_count'] or 0
        if r['expected_items']:
            existing_items = detail_map[key].get('expected_items') or ''
            detail_map[key]['expected_items'] = (
                existing_items + '|||' + r['expected_items']
                if existing_items else r['expected_items']
            )
    for r in invoiced_detail_rows:
        key = (r['platform'], r['group_key'])
        detail_map.setdefault(key, {
            'platform': r['platform'],
            'alias_name': r['alias_name'] or '',
            'nickname_list': r['nickname_list'] or '',
            'expected_total': 0,
            'expected_count': 0,
            'expected_items': '',
            'invoiced_total': 0,
            'invoiced_count': 0,
            'invoice_items': '',
        })
        if r['alias_name'] and not detail_map[key]['alias_name']:
            detail_map[key]['alias_name'] = r['alias_name']
        if r['nickname_list'] and not detail_map[key]['nickname_list']:
            detail_map[key]['nickname_list'] = r['nickname_list']
        detail_map[key]['invoiced_total'] += r['invoiced_total'] or 0
        detail_map[key]['invoiced_count'] += r['invoiced_count'] or 0
        if r['invoice_items']:
            existing_items = detail_map[key].get('invoice_items') or ''
            detail_map[key]['invoice_items'] = (
                existing_items + '|||' + r['invoice_items']
                if existing_items else r['invoice_items']
            )

    transfer_totals = {}
    for r in transfer_rows:
        platform = r['platform']
        source_platform = r['source_platform'] or '未知来源'
        label_name = r['alias_name'] or r['short_name'] or '未命名'
        key = (platform, f"transfer:{source_platform}:{label_name}")
        detail_map.setdefault(key, {
            'platform': platform,
            'alias_name': '转移开票',
            'nickname_list': f"来源：{source_platform} / {label_name}",
            'expected_total': 0,
            'expected_count': 0,
            'expected_items': '',
            'invoiced_total': 0,
            'invoiced_count': 0,
            'invoice_items': '',
        })
        detail_map[key]['invoiced_total'] += r['invoiced_total'] or 0
        detail_map[key]['invoiced_count'] += r['invoiced_count'] or 0
        if r['invoice_items']:
            existing_items = detail_map[key].get('invoice_items') or ''
            detail_map[key]['invoice_items'] = (
                existing_items + '|||' + r['invoice_items']
                if existing_items else r['invoice_items']
            )
        transfer_totals.setdefault(platform, {'total': 0, 'count': 0})
        transfer_totals[platform]['total'] += r['invoiced_total'] or 0
        transfer_totals[platform]['count'] += r['invoiced_count'] or 0

    details_by_platform = {platform: [] for platform in platform_keys}
    for item in detail_map.values():
        platform = item['platform']
        if platform not in details_by_platform:
            continue
        item['diff'] = (item['expected_total'] or 0) - (item['invoiced_total'] or 0)
        details_by_platform[platform].append(item)
    for platform in platform_keys:
        details_by_platform[platform].sort(
            key=lambda item: (
                0 if item['alias_name'] else 1,
                -(item['expected_total'] or 0),
                item['alias_name'] or item['nickname_list'] or '',
            )
        )

    rows = []
    for platform in platform_keys:
        expected_row = expected_map.get(platform)
        invoiced_row = invoiced_map.get(platform)
        expected_total = (expected_row['expected_total'] if expected_row else 0) or 0
        expected_count = (expected_row['expected_count'] if expected_row else 0) or 0
        invoiced_total = (invoiced_row['invoiced_total'] if invoiced_row else 0) or 0
        invoiced_count = (invoiced_row['invoiced_count'] if invoiced_row else 0) or 0
        if platform in transfer_totals:
            invoiced_total += transfer_totals[platform]['total']
            invoiced_count += transfer_totals[platform]['count']
        rows.append({
            'platform': platform,
            'expected_total': expected_total,
            'expected_count': expected_count,
            'invoiced_total': invoiced_total,
            'invoiced_count': invoiced_count,
            'diff': expected_total - invoiced_total,
            'details': details_by_platform.get(platform, []),
        })

    total_expected = sum((r['expected_total'] or 0) for r in rows)
    total_invoiced = sum((r['invoiced_total'] or 0) for r in rows)
    total_diff = total_expected - total_invoiced

    return render_template(
        'invoicing_reconciliation.html',
        rows=rows,
        start_date=start_date or '',
        end_date=end_date or '',
        unmatched_total=unmatched['total'] or 0,
        unmatched_count=unmatched['cnt'] or 0,
        unassigned_total=unassigned['total'] or 0,
        unassigned_count=unassigned['cnt'] or 0,
        total_expected=total_expected,
        total_invoiced=total_invoiced,
        total_diff=total_diff,
    )


@invoicing_bp.route('/reconciliation')
@module_required('invoicing')
def reconciliation():
    start_date = (request.args.get('start_date') or '').strip() or None
    end_date = (request.args.get('end_date') or '').strip() or None
    platform_keys = ('澳柯', '香娜露儿', '快手', '幕莲蔓')

    def _buyer_platform(buyer_name):
        text = buyer_name or ''
        for key in platform_keys:
            if key in text:
                return key
        return None

    def _append_item(existing, item):
        return f"{existing}|||{item}" if existing else item

    def _new_summary():
        return {
            'expected_total': 0,
            'expected_count': 0,
            'belong_total': 0,
            'belong_count': 0,
            'transfer_in_total': 0,
            'transfer_in_count': 0,
            'transfer_out_total': 0,
            'transfer_out_count': 0,
        }

    def _new_detail(platform, alias_name='', nickname_list=''):
        return {
            'platform': platform,
            'alias_name': alias_name or '',
            'nickname_list': nickname_list or '',
            'expected_total': 0,
            'expected_count': 0,
            'expected_items': '',
            'belong_total': 0,
            'belong_count': 0,
            'belong_items': '',
            'transfer_in_total': 0,
            'transfer_in_count': 0,
            'transfer_in_items': '',
            'transfer_out_total': 0,
            'transfer_out_count': 0,
            'transfer_out_items': '',
        }

    expected_where = "1=1"
    expected_params = []
    if end_date:
        expected_where += " AND (period_start IS NULL OR period_start <= ?)"
        expected_params.append(end_date)
    if start_date:
        expected_where += " AND (period_end IS NULL OR period_end >= ?)"
        expected_params.append(start_date)

    invoice_where = "i.is_usable = 1"
    invoice_params = []
    if start_date:
        invoice_where += " AND i.invoice_date >= ?"
        invoice_params.append(start_date)
    if end_date:
        invoice_where += " AND i.invoice_date <= ?"
        invoice_params.append(end_date)

    expected_sql = f"""
        WITH customer_alias_one AS (
            SELECT customer_id, MIN(alias) AS alias
            FROM customer_alias
            GROUP BY customer_id
        )
        SELECT e.platform,
               e.period,
               e.amount,
               c.id AS customer_id,
               c.short_name,
               COALESCE(ca.alias, '') AS alias_name
        FROM expected_amount e
        JOIN customer c ON c.id = e.customer_id
        LEFT JOIN customer_alias_one ca ON ca.customer_id = c.id
        WHERE {expected_where}
    """

    invoice_sql = f"""
        WITH customer_alias_one AS (
            SELECT customer_id, MIN(alias) AS alias
            FROM customer_alias
            GROUP BY customer_id
        )
        SELECT i.id,
               i.invoice_number,
               i.amount,
               i.platform AS invoice_platform,
               i.buyer_name,
               i.customer_id,
               COALESCE(NULLIF(i.alias_name, ''), ca.alias, '') AS alias_name,
               c.short_name,
               c.platform AS customer_platform
        FROM invoice i
        LEFT JOIN customer c ON c.id = i.customer_id
        LEFT JOIN customer_alias_one ca ON ca.customer_id = c.id
        WHERE {invoice_where}
    """

    alias_platform_sql = """
        SELECT ca.alias, COALESCE(e.platform, c.platform) AS platform
        FROM customer_alias ca
        JOIN customer c ON c.id = ca.customer_id
        LEFT JOIN expected_amount e ON e.customer_id = c.id
        WHERE COALESCE(e.platform, c.platform, '') <> ''
    """
    alias_members_sql = """
        SELECT ca.alias, c.platform, c.short_name
        FROM customer_alias ca
        JOIN customer c ON c.id = ca.customer_id
        ORDER BY c.short_name
    """
    unmatched_sql = """
        SELECT COALESCE(SUM(amount), 0) AS total, COUNT(*) AS cnt
        FROM invoice
        WHERE is_usable = 1
          AND customer_id IS NULL
          AND (alias_name IS NULL OR alias_name = '')
    """
    unmatched_params = []
    if start_date:
        unmatched_sql += " AND invoice_date >= ?"
        unmatched_params.append(start_date)
    if end_date:
        unmatched_sql += " AND invoice_date <= ?"
        unmatched_params.append(end_date)

    with get_db_connection() as conn:
        expected_rows = conn.execute(expected_sql, expected_params).fetchall()
        invoice_rows = conn.execute(invoice_sql, invoice_params).fetchall()
        alias_platform_rows = conn.execute(alias_platform_sql).fetchall()
        alias_member_rows = conn.execute(alias_members_sql).fetchall()
        unmatched = conn.execute(unmatched_sql, unmatched_params).fetchone()

    alias_platforms = {}
    for r in alias_platform_rows:
        alias_platforms.setdefault(r['alias'], set()).add(r['platform'])
    alias_platform_map = {
        alias: next(iter(platforms))
        for alias, platforms in alias_platforms.items()
        if len(platforms) == 1
    }

    alias_members = {}
    alias_members_any = {}
    for r in alias_member_rows:
        alias = r['alias']
        platform = r['platform'] or ''
        alias_members.setdefault((alias, platform), []).append(r['short_name'])
        alias_members_any.setdefault(alias, []).append(r['short_name'])

    summaries = {platform: _new_summary() for platform in platform_keys}
    detail_map = {}

    def _alias_nicknames(alias_name, platform):
        names = alias_members.get((alias_name, platform)) or alias_members_any.get(alias_name) or []
        return '、'.join(dict.fromkeys(names))

    def _expected_group(row):
        alias_name = row['alias_name'] or ''
        if alias_name:
            return f"alias:{alias_name}", alias_name, _alias_nicknames(alias_name, row['platform'])
        return f"customer:{row['customer_id']}", '', row['short_name'] or ''

    def _invoice_group(row, source_platform):
        alias_name = row['alias_name'] or ''
        if alias_name:
            return f"alias:{alias_name}", alias_name, _alias_nicknames(alias_name, source_platform)
        if row['customer_id']:
            return f"customer:{row['customer_id']}", '', row['short_name'] or ''
        return 'unmatched', '', '未匹配发票'

    for r in expected_rows:
        platform = r['platform']
        if platform not in summaries:
            continue
        summaries[platform]['expected_total'] += r['amount'] or 0
        summaries[platform]['expected_count'] += 1
        group_key, alias_name, nickname_list = _expected_group(r)
        key = (platform, group_key)
        detail = detail_map.setdefault(key, _new_detail(platform, alias_name, nickname_list))
        if alias_name and not detail['alias_name']:
            detail['alias_name'] = alias_name
        if nickname_list and not detail['nickname_list']:
            detail['nickname_list'] = nickname_list
        detail['expected_total'] += r['amount'] or 0
        detail['expected_count'] += 1
        period_label = r['period'] or '未设期间'
        detail['expected_items'] = _append_item(
            detail['expected_items'],
            f"{period_label}::{(r['amount'] or 0):.2f}",
        )

    for r in invoice_rows:
        amount = r['amount'] or 0
        alias_name = r['alias_name'] or ''
        source_platform = (
            (r['invoice_platform'] or '').strip()
            or (r['customer_platform'] or '').strip()
            or alias_platform_map.get(alias_name)
        )
        billing_platform = _buyer_platform(r['buyer_name'])
        group_key, row_alias, nickname_list = _invoice_group(r, source_platform)
        base_item = f"{r['id']}::{r['invoice_number']}::{amount:.2f}"

        is_transfer_out = (
            source_platform in summaries
            and billing_platform
            and billing_platform != source_platform
        )

        if source_platform in summaries:
            if is_transfer_out:
                summaries[source_platform]['transfer_out_total'] += amount
                summaries[source_platform]['transfer_out_count'] += 1
            else:
                summaries[source_platform]['belong_total'] += amount
                summaries[source_platform]['belong_count'] += 1
            key = (source_platform, group_key)
            detail = detail_map.setdefault(
                key, _new_detail(source_platform, row_alias, nickname_list)
            )
            if row_alias and not detail['alias_name']:
                detail['alias_name'] = row_alias
            if nickname_list and not detail['nickname_list']:
                detail['nickname_list'] = nickname_list
            note = f"开票至{billing_platform}" if billing_platform and billing_platform != source_platform else ''
            if is_transfer_out:
                detail['transfer_out_total'] += amount
                detail['transfer_out_count'] += 1
                detail['transfer_out_items'] = _append_item(
                    detail['transfer_out_items'],
                    f"{base_item}::转移至{billing_platform}",
                )
            else:
                detail['belong_total'] += amount
                detail['belong_count'] += 1
                detail['belong_items'] = _append_item(
                    detail['belong_items'],
                    f"{base_item}::{note}",
                )

        if billing_platform in summaries and billing_platform != source_platform:
            summaries[billing_platform]['transfer_in_total'] += amount
            summaries[billing_platform]['transfer_in_count'] += 1
            if source_platform:
                transfer_key = f"transfer_in:{source_platform}:{group_key}"
                label = nickname_list or row_alias or r['short_name'] or '未命名'
                transfer_label = f"来源：{source_platform} / {label}"
                note = f"来源{source_platform}"
                transfer_alias = ''
            elif row_alias:
                transfer_key = f"alias:{row_alias}"
                transfer_label = _alias_nicknames(row_alias, billing_platform) or row_alias
                note = '按购买方归入'
                transfer_alias = row_alias
            else:
                transfer_key = 'billing_unmatched'
                transfer_label = f"未匹配 {amount:.2f}"
                note = '未匹配归属'
                transfer_alias = ''
            key = (billing_platform, transfer_key)
            detail = detail_map.setdefault(
                key, _new_detail(billing_platform, transfer_alias, transfer_label)
            )
            if transfer_alias and not detail['alias_name']:
                detail['alias_name'] = transfer_alias
            if transfer_key == 'billing_unmatched':
                detail['nickname_list'] = f"未匹配 {detail['transfer_in_total'] + amount:.2f}"
            elif transfer_label and not detail['nickname_list']:
                detail['nickname_list'] = transfer_label
            detail['transfer_in_total'] += amount
            detail['transfer_in_count'] += 1
            detail['transfer_in_items'] = _append_item(
                detail['transfer_in_items'],
                f"{base_item}::{note}",
            )

    details_by_platform = {platform: [] for platform in platform_keys}
    for item in detail_map.values():
        platform = item['platform']
        if platform not in details_by_platform:
            continue
        item['invoiced_total'] = (item['belong_total'] or 0) + (item['transfer_in_total'] or 0)
        item['invoiced_count'] = (item['belong_count'] or 0) + (item['transfer_in_count'] or 0)
        item['diff'] = (
            (item['expected_total'] or 0)
            - (item['invoiced_total'] or 0)
            - (item['transfer_out_total'] or 0)
        )
        details_by_platform[platform].append(item)
    for platform in platform_keys:
        details_by_platform[platform].sort(
            key=lambda item: (
                0 if item['alias_name'] else 1,
                0 if item['expected_total'] else 1,
                -(item['expected_total'] or item['invoiced_total'] or 0),
                item['alias_name'] or item['nickname_list'] or '',
            )
        )

    rows = []
    for platform in platform_keys:
        s = summaries[platform]
        invoiced_total = (s['belong_total'] or 0) + (s['transfer_in_total'] or 0)
        invoiced_count = (s['belong_count'] or 0) + (s['transfer_in_count'] or 0)
        rows.append({
            'platform': platform,
            'expected_total': s['expected_total'],
            'expected_count': s['expected_count'],
            'belong_total': s['belong_total'],
            'belong_count': s['belong_count'],
            'transfer_in_total': s['transfer_in_total'],
            'transfer_in_count': s['transfer_in_count'],
            'transfer_out_total': s['transfer_out_total'],
            'transfer_out_count': s['transfer_out_count'],
            'invoiced_total': invoiced_total,
            'invoiced_count': invoiced_count,
            'settled_total': invoiced_total + (s['transfer_out_total'] or 0),
            'diff': s['expected_total'] - invoiced_total - (s['transfer_out_total'] or 0),
            'details': details_by_platform.get(platform, []),
        })

    total_expected = sum((r['expected_total'] or 0) for r in rows)
    total_invoiced = sum((r['invoiced_total'] or 0) for r in rows)
    total_diff = total_expected - total_invoiced

    return render_template(
        'invoicing_reconciliation.html',
        rows=rows,
        start_date=start_date or '',
        end_date=end_date or '',
        unmatched_total=unmatched['total'] or 0,
        unmatched_count=unmatched['cnt'] or 0,
        total_expected=total_expected,
        total_invoiced=total_invoiced,
        total_diff=total_diff,
    )
