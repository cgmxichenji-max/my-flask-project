import os
import json
import math
import base64
import csv
import hashlib
import http.cookiejar
import io
import re
import ssl
import secrets
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime

from flask import Blueprint, render_template, jsonify, request, redirect, url_for, current_app, g
import sqlite3
from dateutil import parser as date_parser
from cryptography.hazmat.primitives.asymmetric import ec, utils as ec_utils
from cryptography.hazmat.primitives import hashes

from auth.decorators import module_required

label_print_bp = Blueprint('label_print', __name__)
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
KDOCS_SOURCE_URL = 'https://kdocs.cn/l/cnaogtuBWmXW'
KDOCS_SOURCES = [
    {'key': 'kdocs_main', 'name': 'WPS主表', 'url': 'https://kdocs.cn/l/cnaogtuBWmXW'},
    {'key': 'kdocs_extra', 'name': 'WPS补充表', 'url': 'https://kdocs.cn/l/cqUIdRBp3yCl'},
]
KDOCS_SOURCE_BY_KEY = {item['key']: item for item in KDOCS_SOURCES}
KDOCS_COOKIE_PATH = os.path.join(BASE_DIR, 'data', 'kdocs_cookie.txt')
KDOCS_LOGIN_ACCOUNT = '香水梨'
KDOCS_LOGIN_PASSWORD = 'chenxi98'
KDOCS_QR_BASE = 'https://qr.wps.cn'
KDOCS_ACCOUNT_BASE = 'https://account.wps.cn'
KDOCS_QR_SESSIONS = {}
KDOCS_SOURCE_KEY = 'kdocs_main'
KDOCS_DBT_PAGE_SIZE = 500
KDOCS_CONTENT_FIELD_HINTS = ('箱唛', '规格*数量', '规格', '数量')

# ─────────────────────────────── 数据库工具 ───────────────────────────────

def get_db_connection():
    db_path = current_app.config.get('DATABASE_PATH', os.path.join(BASE_DIR, 'data', 'main.db'))
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def _ensure_column(conn, table_name, column_name, column_def):
    columns = [row['name'] for row in conn.execute(f'PRAGMA table_info({table_name})').fetchall()]
    if column_name not in columns:
        conn.execute(f'ALTER TABLE {table_name} ADD COLUMN {column_name} {column_def}')


def ensure_tables(conn):
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS label_products (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            code         TEXT    NOT NULL UNIQUE,
            short_name   TEXT    NOT NULL DEFAULT '',
            product_name TEXT    NOT NULL DEFAULT '',
            spec         TEXT    NOT NULL DEFAULT '',
            box_spec     INTEGER NOT NULL DEFAULT 0,
            created_at   TEXT    NOT NULL,
            updated_at   TEXT    NOT NULL
        );
        CREATE TABLE IF NOT EXISTS label_weights (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            code       TEXT    NOT NULL UNIQUE,
            weight     REAL    NOT NULL DEFAULT 0,
            updated_at TEXT    NOT NULL
        );
        CREATE TABLE IF NOT EXISTS label_sizes (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            code         TEXT    NOT NULL UNIQUE,
            length       INTEGER NOT NULL DEFAULT 0,
            width        INTEGER NOT NULL DEFAULT 0,
            height       INTEGER NOT NULL DEFAULT 0,
            is_irregular INTEGER NOT NULL DEFAULT 0,
            updated_at   TEXT    NOT NULL
        );
        CREATE TABLE IF NOT EXISTS label_packing_weights (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            pack_name  TEXT    NOT NULL UNIQUE,
            weight     REAL    NOT NULL DEFAULT 0,
            updated_at TEXT    NOT NULL
        );
        CREATE TABLE IF NOT EXISTS label_packing_sizes (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            pack_name  TEXT    NOT NULL UNIQUE,
            size       TEXT    NOT NULL DEFAULT '',
            updated_at TEXT    NOT NULL
        );
        CREATE TABLE IF NOT EXISTS label_print_history (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            total_tickets  INTEGER NOT NULL DEFAULT 0,
            total_qty      INTEGER NOT NULL DEFAULT 0,
            items_json     TEXT    NOT NULL DEFAULT '[]',
            printed_at     TEXT    NOT NULL
        );
        CREATE TABLE IF NOT EXISTS label_wps_records (
            id                      INTEGER PRIMARY KEY AUTOINCREMENT,
            source_key              TEXT    NOT NULL DEFAULT '',
            submit_date             TEXT    NOT NULL DEFAULT '',
            submit_time_text        TEXT    NOT NULL DEFAULT '',
            row_hash                TEXT    NOT NULL DEFAULT '',
            row_json                TEXT    NOT NULL DEFAULT '{}',
            content_text            TEXT    NOT NULL DEFAULT '',
            parsed_items_json       TEXT    NOT NULL DEFAULT '[]',
            parse_status            TEXT    NOT NULL DEFAULT 'unparsed',
            printed_count           INTEGER NOT NULL DEFAULT 0,
            last_printed_at         TEXT    NOT NULL DEFAULT '',
            last_printed_by_user_id INTEGER,
            last_printed_by_username TEXT   NOT NULL DEFAULT '',
            created_at              TEXT    NOT NULL,
            updated_at              TEXT    NOT NULL,
            UNIQUE(source_key, row_hash)
        );
        CREATE TABLE IF NOT EXISTS label_pack_presets (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            combo_key  TEXT    NOT NULL UNIQUE,
            box_name   TEXT    NOT NULL DEFAULT '',
            bag_name   TEXT    NOT NULL DEFAULT '',
            bag_qty    INTEGER NOT NULL DEFAULT 1,
            note       TEXT    NOT NULL DEFAULT '',
            updated_at TEXT    NOT NULL
        );
        CREATE TABLE IF NOT EXISTS label_pack_settings (
            key        TEXT    PRIMARY KEY,
            value      TEXT    NOT NULL DEFAULT '',
            label      TEXT    NOT NULL DEFAULT '',
            updated_at TEXT    NOT NULL
        );
    ''')
    _ensure_column(conn, 'label_pack_presets', 'bag_qty', 'INTEGER NOT NULL DEFAULT 1')
    _ensure_column(conn, 'label_print_history', 'wps_record_id', 'INTEGER')
    _ensure_column(conn, 'label_print_history', 'printed_by_user_id', 'INTEGER')
    _ensure_column(conn, 'label_print_history', 'printed_by_username', "TEXT NOT NULL DEFAULT ''")
    _ensure_column(conn, 'label_print_history', 'is_forced_reprint', 'INTEGER NOT NULL DEFAULT 0')
    _ensure_column(conn, 'label_print_history', 'parsed_items_json', "TEXT NOT NULL DEFAULT '[]'")
    conn.commit()


# ─────────────────────────────── 初始数据 ───────────────────────────────

INIT_WEIGHTS = [
    ('140',0.24),('184',0.13),('702',0.35),('161',0.23),
    ('A140',0.06),('113B',0.09),('P136',0.06),('139',0.13),
    ('A113',0.05),('163',0.12),('164',0.40),('103',0.11),
    ('705',0.12),('160',0.37),('155',0.17),('A160',0.04),
    ('A161',0.05),('P147',0.27),('112',0.37),('165',0.40),
    ('145',0.13),('156',0.18),
]

INIT_SIZES = [
    ('140',80,75,71,0),('184',180,62,35,1),('702',105,105,55,0),
    ('161',78,76,72,0),('A140',57,57,35,0),('113B',155,51,25,0),
    ('P136',135,48,21,0),('139',70,70,63,0),('A113',57,57,35,0),
    ('163',108,46,46,0),('164',191,55,55,0),('103',57,57,35,0),
    ('705',130,103,30,0),('160',190,133,48,0),('155',138,52,52,0),
    ('A160',135,48,23,0),('A161',57,57,35,0),('P147',197,100,62,0),
    ('112',190,133,48,0),('165',191,55,55,0),('145',152,53,40,0),
    ('156',185,115,20,0),
]

INIT_PACKING_SIZES = [
    ('5',    '290×170×190'),
    ('6',    '260×150×180'),
    ('6.5',  '260×150×90'),
    ('7',    '230×130×160'),
    ('7.5',  '230×130×80'),
    ('8',    '210×110×140'),
    ('8.5',  '210×110×70'),
    ('9',    '195×105×135'),
    ('9.5',  '195×105×68'),
    ('10',   '175×95×115'),
    ('10.5', '175×95×58'),
    ('11',   '145×85×105'),
    ('11.5', '145×85×53'),
    ('大泡',  '250×350'),
    ('中泡',  '200×300'),
    ('小泡',  '180×200'),
]


INIT_SETTINGS = [
    ('buffer_mm',        '0',    '每边缓冲 (mm)：加在货物每个尺寸上的安全余量，手量尺寸已有余量时可设为0'),
    ('fill_rate_single', '0.85', '单品填充率上限：单品精确计算时留出的余量比例（当前仅供扩展，精确计算不使用）'),
    ('fill_rate_multi',  '0.75', '混装填充率上限：多品体积估算时要求总体积 ≤ 箱容积 × 此值'),
    ('irregular_factor', '1.15', '异形货物体积系数：is_irregular=1 的货物体积乘以此系数，避免推荐过小箱型'),
    ('complex_threshold','10',   '不预览阈值：货物种类数超过此值时不给出推荐'),
    ('bag_girth_ratio',  '1.05', '气泡袋周长比例（余量参数）：货物截面周长 2*(a+b) ≤ 袋长边 × 此值；1.0=不留余量，>1.0=允许袋子轻微撑开，值越大推荐越小袋型；建议 0.95~1.1'),
]
SETTING_KEYS = [k for k, _, _ in INIT_SETTINGS]


def seed_aux_tables(conn):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if conn.execute('SELECT COUNT(*) FROM label_weights').fetchone()[0] == 0:
        conn.executemany(
            'INSERT OR IGNORE INTO label_weights (code,weight,updated_at) VALUES (?,?,?)',
            [(c, w, now) for c, w in INIT_WEIGHTS]
        )
    if conn.execute('SELECT COUNT(*) FROM label_sizes').fetchone()[0] == 0:
        conn.executemany(
            'INSERT OR IGNORE INTO label_sizes (code,length,width,height,is_irregular,updated_at) VALUES (?,?,?,?,?,?)',
            [(c,l,w,h,ir,now) for c,l,w,h,ir in INIT_SIZES]
        )
    # 包材尺寸：始终用 INSERT OR REPLACE 确保初始数据与当前 INIT_PACKING_SIZES 同步
    conn.executemany(
        'INSERT OR REPLACE INTO label_packing_sizes (pack_name,size,updated_at) VALUES (?,?,?)',
        [(n, s, now) for n, s in INIT_PACKING_SIZES]
    )
    # 参数设置：INSERT OR IGNORE，新增参数自动补入，不覆盖用户已修改的值；仅同步 label 字段
    for key, value, label in INIT_SETTINGS:
        conn.execute(
            'INSERT OR IGNORE INTO label_pack_settings (key,value,label,updated_at) VALUES (?,?,?,?)',
            (key, value, label, now)
        )
        conn.execute(
            'UPDATE label_pack_settings SET label=? WHERE key=?',
            (label, key)
        )
    # 迁移：bag_girth_ratio 旧默认值（0.9 / 0.95）→ 1.05（仅当用户未手动修改过时才升级）
    conn.execute(
        "UPDATE label_pack_settings SET value='1.05', updated_at=? WHERE key='bag_girth_ratio' AND value IN ('0.9','0.95')",
        (now,)
    )
    conn.commit()


def import_products_from_xlsx(conn):
    xlsx_path = os.path.join(BASE_DIR, 'temp', '打印页头.xlsx')
    if not os.path.exists(xlsx_path):
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb['Sheet2']
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[1] is None:
                continue
            conn.execute(
                'INSERT OR IGNORE INTO label_products '
                '(code,short_name,product_name,spec,box_spec,created_at,updated_at) VALUES (?,?,?,?,?,?,?)',
                (str(row[1]).strip(), str(row[2] or ''), str(row[0] or ''),
                 str(row[3] or ''), int(row[4] or 0), now, now)
            )
        conn.commit()
    except Exception as e:
        current_app.logger.warning(f'[label_print] xlsx import: {e}')


def _get_pack_names(conn):
    """从 pack_item 表获取包材名称列表"""
    try:
        return [r[0] for r in conn.execute('SELECT name FROM pack_item ORDER BY name').fetchall()]
    except Exception:
        return []


def _get_pack_stock(conn):
    """取 pack_stock_snapshot 中每个 spec 的最新库存，返回 {spec: qty} 字典"""
    try:
        rows = conn.execute('''
            SELECT spec, qty FROM pack_stock_snapshot
            WHERE (spec, stocktake_ts) IN (
                SELECT spec, MAX(stocktake_ts) FROM pack_stock_snapshot GROUP BY spec
            )
        ''').fetchall()
        return {r['spec']: (r['qty'] if r['qty'] is not None else 0) for r in rows}
    except Exception:
        return {}


def _load_all(conn):
    order = 'ORDER BY CAST(code AS INTEGER), code'
    products        = conn.execute(f'SELECT * FROM label_products {order}').fetchall()
    weights         = conn.execute(f'SELECT * FROM label_weights  {order}').fetchall()
    sizes           = conn.execute(f'SELECT * FROM label_sizes    {order}').fetchall()
    packing_weights = conn.execute('SELECT * FROM label_packing_weights ORDER BY pack_name').fetchall()
    packing_sizes   = conn.execute('SELECT * FROM label_packing_sizes ORDER BY id').fetchall()
    pack_presets    = conn.execute('SELECT * FROM label_pack_presets ORDER BY combo_key').fetchall()
    pack_settings   = conn.execute('SELECT * FROM label_pack_settings ORDER BY key').fetchall()
    history         = conn.execute(
        'SELECT * FROM label_print_history ORDER BY printed_at DESC LIMIT 100'
    ).fetchall()
    pack_names      = _get_pack_names(conn)
    pack_stock      = _get_pack_stock(conn)
    return products, weights, sizes, packing_weights, packing_sizes, pack_presets, pack_settings, history, pack_names, pack_stock


def _init_db(conn):
    ensure_tables(conn)
    if conn.execute('SELECT COUNT(*) FROM label_products').fetchone()[0] == 0:
        import_products_from_xlsx(conn)
    seed_aux_tables(conn)


# ─────────────────────────────── 金山表格解析 ───────────────────────────────

def _clean_cell(value):
    if value is None:
        return ''
    text = str(value).strip()
    if text.endswith('.0') and text[:-2].isdigit():
        return text[:-2]
    return text


def _parse_date_cell(value):
    text = _clean_cell(value)
    if not text:
        return None
    match = re.search(r'(\d{4})[./\-年](\d{1,2})[./\-月](\d{1,2})', text)
    if match:
        try:
            return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3))).date()
        except ValueError:
            return None
    try:
        return date_parser.parse(text, fuzzy=True).date()
    except Exception:
        return None


def _rows_from_delimited_text(text):
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters='\t,;')
        reader = csv.reader(io.StringIO(text), dialect)
    except Exception:
        reader = csv.reader(io.StringIO(text), delimiter='\t')
    rows = [[_clean_cell(c) for c in row] for row in reader]
    return [row for row in rows if any(c for c in row)]


def _rows_from_html(text):
    try:
        import pandas as pd
        tables = pd.read_html(io.StringIO(text))
    except Exception:
        return []
    rows = []
    for table in tables:
        rows.extend([[_clean_cell(c) for c in table.columns.tolist()]])
        rows.extend([[_clean_cell(c) for c in row] for row in table.values.tolist()])
    return [row for row in rows if any(c for c in row)]


def _rows_from_xlsx(content):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    rows = [[_clean_cell(c) for c in row] for row in ws.iter_rows(values_only=True)]
    return [row for row in rows if any(c for c in row)]


def _kdocs_link_id(source=None):
    source = source or KDOCS_SOURCES[0]
    parsed = urllib.parse.urlparse(source.get('url') or KDOCS_SOURCE_URL)
    return parsed.path.rstrip('/').split('/')[-1]


def _kdocs_json_request(url, payload, source=None):
    source = source or KDOCS_SOURCES[0]
    cookie = _get_kdocs_cookie()
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        'Accept': 'application/json, text/plain, */*',
        'Content-Type': 'application/json',
        'Origin': 'https://www.kdocs.cn',
        'Referer': source.get('url') or KDOCS_SOURCE_URL,
    }
    if cookie:
        headers['Cookie'] = cookie
    body = json.dumps(payload, ensure_ascii=False).encode('utf-8')
    req = urllib.request.Request(url, data=body, headers=headers, method='POST')
    try:
        ssl_context = ssl._create_unverified_context()
        opener = urllib.request.build_opener(
            urllib.request.ProxyHandler({}),
            urllib.request.HTTPSHandler(context=ssl_context),
        )
        with opener.open(req, timeout=30) as resp:
            text = resp.read().decode('utf-8', errors='ignore')
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode('utf-8', errors='ignore')[:300]
        raise ValueError(f'读取金山轻维表失败：HTTP {exc.code} {detail}')
    except urllib.error.URLError as exc:
        raise ValueError(f'读取金山轻维表失败：{exc}')
    try:
        data = json.loads(text or '{}')
    except Exception:
        raise ValueError('读取金山轻维表失败：接口返回格式不是JSON')
    result = str(data.get('result') or '').lower()
    if result and result not in ('ok', 'success'):
        message = data.get('msg') or data.get('message') or data.get('error') or result
        raise ValueError(f'读取金山轻维表失败：{message}')
    return data


def _kdocs_core_execute(command, param, source=None):
    url = f'https://www.kdocs.cn/api/v3/office/file/{_kdocs_link_id(source)}/core/execute'
    return _kdocs_json_request(url, {'command': command, 'param': param}, source=source)


def _kdocs_value_to_text(value):
    if value is None:
        return ''
    if isinstance(value, str):
        return _clean_cell(value)
    if isinstance(value, (int, float)):
        return _clean_cell(value)
    if isinstance(value, list):
        parts = []
        for item in value:
            text = _kdocs_value_to_text(item)
            if text:
                parts.append(text)
        return '，'.join(parts)
    if isinstance(value, dict):
        for key in ('value', 'text', 'name', 'nickName', 'fileName', 'id'):
            if value.get(key):
                return _clean_cell(value.get(key))
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return _clean_cell(value)


def _fetch_kdocs_dbt_rows(source=None):
    source = source or KDOCS_SOURCES[0]
    sheets_data = _kdocs_core_execute('http.db.listSheets', {'showVeryhidden': False}, source=source)
    sheets = (sheets_data.get('detail') or {}).get('sheets') or []
    if not sheets:
        raise ValueError('读取金山轻维表失败：没有找到数据表')

    sheet = next(
        (
            item for item in sheets
            if any(field.get('name') == '提交时间' for field in item.get('fields') or [])
            and any('箱唛' in str(field.get('name') or '') for field in item.get('fields') or [])
        ),
        sheets[0],
    )
    fields = sheet.get('fields') or []
    if not fields:
        raise ValueError('读取金山轻维表失败：没有找到字段')

    header = [_clean_cell(field.get('name') or field.get('id')) for field in fields]
    field_ids = [field.get('id') for field in fields]
    rows = [header]
    offset = '0'
    seen_offsets = set()

    while True:
        payload = {
            'sheetId': sheet.get('id'),
            'offset': offset,
            'pageSize': KDOCS_DBT_PAGE_SIZE,
            'preferId': True,
            'showRecordExtraInfo': False,
            'showFieldsInfo': False,
            'textValue': 'true',
        }
        records_data = _kdocs_core_execute('http.db.listRecords', payload, source=source)
        detail = records_data.get('detail') or {}
        records = detail.get('records') or []
        for record in records:
            values = record.get('fields') or {}
            rows.append([_kdocs_value_to_text(values.get(field_id)) for field_id in field_ids])

        next_offset = str(detail.get('offset') or '')
        if not records or not next_offset or next_offset == offset or next_offset in seen_offsets:
            break
        seen_offsets.add(offset)
        offset = next_offset

    return [row for row in rows if any(row)]


def _fetch_kdocs_rows(source=None):
    source = source or KDOCS_SOURCES[0]
    try:
        return _fetch_kdocs_dbt_rows(source=source)
    except ValueError as exc:
        if '登录' in str(exc) or 'permission' in str(exc).lower():
            raise

    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,text/csv,*/*',
    }
    cookie = _get_kdocs_cookie()
    if cookie:
        headers['Cookie'] = cookie
    req = urllib.request.Request(source.get('url') or KDOCS_SOURCE_URL, headers=headers)
    try:
        ssl_context = ssl._create_unverified_context()
        opener = urllib.request.build_opener(
            urllib.request.ProxyHandler({}),
            urllib.request.HTTPSHandler(context=ssl_context),
        )
        with opener.open(req, timeout=18) as resp:
            content = resp.read()
            content_type = resp.headers.get('Content-Type', '').lower()
            final_url = resp.geturl()
    except urllib.error.URLError as exc:
        raise ValueError(f'读取金山文档失败：{exc}')

    if 'spreadsheetml.sheet' in content_type or final_url.lower().endswith('.xlsx'):
        return _rows_from_xlsx(content)

    text = content.decode('utf-8', errors='ignore')
    if 'passport' in final_url or ('请输入账号' in text and '请输入密码' in text):
        raise ValueError('金山文档要求登录；请在页面保存已登录 Cookie，或在服务器设置 KDOCS_COOKIE 后再点解析。')
    if '<table' in text.lower():
        rows = _rows_from_html(text)
        if rows:
            return rows
    return _rows_from_delimited_text(text)


def _read_der_length(data, idx):
    length = data[idx]
    idx += 1
    if length < 128:
        return length, idx
    size = length & 0x7f
    return int.from_bytes(data[idx:idx + size], 'big'), idx + size


def _read_der_tlv(data, idx, expected_tag):
    tag = data[idx]
    idx += 1
    if tag != expected_tag:
        raise ValueError('金山登录密钥格式异常')
    length, idx = _read_der_length(data, idx)
    return data[idx:idx + length], idx + length


def _parse_rsa_public_key(pass_key):
    data = base64.b64decode(pass_key)
    seq, _ = _read_der_tlv(data, 0, 0x30)
    _, pos = _read_der_tlv(seq, 0, 0x30)
    bit_string, _ = _read_der_tlv(seq, pos, 0x03)
    rsa_seq, _ = _read_der_tlv(bit_string[1:], 0, 0x30)
    modulus, pos = _read_der_tlv(rsa_seq, 0, 0x02)
    exponent, _ = _read_der_tlv(rsa_seq, pos, 0x02)
    return int.from_bytes(modulus, 'big'), int.from_bytes(exponent, 'big'), len(modulus)


def _rsa_encrypt_password(password, pass_key):
    modulus, exponent, key_len = _parse_rsa_public_key(pass_key)
    message = password.encode('utf-8')
    padding_len = key_len - len(message) - 3
    if padding_len < 8:
        raise ValueError('金山登录密钥长度不足')
    padding = bytearray()
    while len(padding) < padding_len:
        chunk = secrets.token_bytes(1)
        if chunk != b'\x00':
            padding.extend(chunk)
    encoded = b'\x00\x02' + bytes(padding) + b'\x00' + message
    encrypted = pow(int.from_bytes(encoded, 'big'), exponent, modulus)
    return base64.b64encode(encrypted.to_bytes(key_len, 'big')).decode('ascii')


def _cookie_header_from_jar(cookie_jar):
    parts = []
    for cookie in cookie_jar:
        if cookie.value:
            parts.append(f'{cookie.name}={cookie.value}')
    return '; '.join(parts)


def _make_cookie(name, value, domain):
    return http.cookiejar.Cookie(
        version=0, name=name, value=value, port=None, port_specified=False,
        domain=domain, domain_specified=True, domain_initial_dot=False,
        path='/', path_specified=True, secure=True, expires=None,
        discard=True, comment=None, comment_url=None, rest={}, rfc2109=False
    )


def _open_json(opener, url, data=None, headers=None, method=None):
    body = None
    if data is not None:
        body = urllib.parse.urlencode(data).encode('utf-8')
    req = urllib.request.Request(url, data=body, headers=headers or {}, method=method)
    try:
        with opener.open(req, timeout=18) as resp:
            return json.loads(resp.read().decode('utf-8', errors='ignore') or '{}')
    except urllib.error.HTTPError as exc:
        text = exc.read().decode('utf-8', errors='ignore')
        try:
            return json.loads(text or '{}')
        except Exception:
            raise ValueError(f'金山登录请求失败：HTTP {exc.code}')


def _make_kdocs_opener():
    ssl_context = ssl._create_unverified_context()
    cookie_jar = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(
        urllib.request.ProxyHandler({}),
        urllib.request.HTTPSHandler(context=ssl_context),
        urllib.request.HTTPCookieProcessor(cookie_jar),
    )
    return opener, cookie_jar


def _open_text(opener, url, data=None, headers=None, method=None, timeout=18):
    body = None
    if data is not None:
        body = urllib.parse.urlencode(data).encode('utf-8')
    req = urllib.request.Request(url, data=body, headers=headers or {}, method=method)
    with opener.open(req, timeout=timeout) as resp:
        return resp.read().decode('utf-8', errors='ignore')


def _open_jsonp(opener, url, timeout=18):
    text = _open_text(opener, url, timeout=timeout)
    match = re.search(r'^[^(]*\((.*)\)\s*;?\s*$', text.strip(), re.S)
    payload = match.group(1) if match else text
    return json.loads(payload or '{}')


def _cleanup_kdocs_qr_sessions():
    now = datetime.now().timestamp()
    expired = [loginid for loginid, item in KDOCS_QR_SESSIONS.items() if item.get('expires_at', 0) < now]
    for loginid in expired:
        KDOCS_QR_SESSIONS.pop(loginid, None)


def _exchange_kdocs_authcode(opener, authcode):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        'Referer': 'https://account.wps.cn/',
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
        'Accept': 'application/json, text/javascript, */*; q=0.01',
    }
    result = _open_json(
        opener,
        f'{KDOCS_ACCOUNT_BASE}/api/session/exchange/login',
        data={'authcode': authcode},
        headers=headers,
        method='POST',
    )
    if result.get('result') not in (None, 'ok') and not result.get('userid'):
        raise ValueError(result.get('msg') or result.get('result') or '扫码登录换取 Cookie 失败')
    return result


def _open_json_body(opener, url, payload, headers=None, timeout=18):
    req = urllib.request.Request(
        url,
        data=json.dumps(payload, separators=(',', ':')).encode('utf-8'),
        headers=headers or {},
        method='POST',
    )
    try:
        with opener.open(req, timeout=timeout) as resp:
            return json.loads(resp.read().decode('utf-8', errors='ignore') or '{}')
    except urllib.error.HTTPError as exc:
        text = exc.read().decode('utf-8', errors='ignore')
        try:
            return json.loads(text or '{}')
        except Exception:
            raise ValueError(f'金山授权请求失败：HTTP {exc.code}')


def _b64url(data):
    return base64.urlsafe_b64encode(data).decode('ascii').rstrip('=')


def _make_kdocs_ec_key():
    private_key = ec.generate_private_key(ec.SECP256R1())
    public_numbers = private_key.public_key().public_numbers()
    jwk = {
        'key_ops': ['verify'],
        'ext': True,
        'kty': 'EC',
        'x': _b64url(public_numbers.x.to_bytes(32, 'big')),
        'y': _b64url(public_numbers.y.to_bytes(32, 'big')),
        'crv': 'P-256',
    }
    public_key = _b64url(json.dumps(jwk, separators=(',', ':')).encode('utf-8'))
    return private_key, public_key


def _sign_kdocs_login_data(private_key, text):
    signature_der = private_key.sign(text.encode('utf-8'), ec.ECDSA(hashes.SHA256()))
    r, s = ec_utils.decode_dss_signature(signature_der)
    return _b64url(r.to_bytes(32, 'big') + s.to_bytes(32, 'big'))


def _grant_kdocs_token(opener, kso_authcode, code_verifier, private_key, public_key):
    payload = {
        'grant_type': 'authorization_code',
        'code': kso_authcode,
        'code_verifier': code_verifier,
        'code_sign': _sign_kdocs_login_data(private_key, kso_authcode),
        'public_key': public_key,
        'is_append': False,
        'slv': 'ecdsa_itk',
    }
    result = _open_json_body(
        opener,
        f'{KDOCS_ACCOUNT_BASE}/passport/secure/api/grant_token',
        payload,
        headers={
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
            'Referer': 'https://account.wps.cn/wpspersonallogin',
            'Content-Type': 'application/json',
            'Accept': 'application/json, text/javascript, */*; q=0.01',
        },
        timeout=20,
    )
    if result.get('result') not in (None, 'ok') and not result.get('data'):
        raise ValueError(result.get('msg') or result.get('result') or '金山扫码授权失败')
    return result


def _message_needs_kdocs_login(message):
    return any(key in message for key in ('登录', '验证码', 'Cookie', '扫码', 'InvalidCaptcha', 'ErrNeedCaptcha'))


def _generate_kdocs_pkce():
    alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-._~'
    verifier_len = 64
    verifier = ''.join(secrets.choice(alphabet) for _ in range(verifier_len))
    digest = hashlib.sha256(verifier.encode('utf-8')).digest()
    challenge = base64.urlsafe_b64encode(digest).decode('ascii').rstrip('=')
    return verifier, challenge


def _login_kdocs_with_password():
    opener, cookie_jar = _make_kdocs_opener()
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        'Referer': 'https://account.wps.cn/v1/accountlogin?keeponline=true&loginpageiframe=true',
        'X-Requested-With': 'XMLHttpRequest',
        'Accept': 'application/json, text/javascript, */*; q=0.01',
    }
    opener.open(urllib.request.Request(headers['Referer'], headers=headers), timeout=18).read()

    passkey_url = 'https://account.wps.cn/api/v3/passkey?' + urllib.parse.urlencode({
        'account': KDOCS_LOGIN_ACCOUNT
    })
    passkey = _open_json(opener, passkey_url, headers=headers)
    if passkey.get('result') != 'ok':
        raise ValueError(f"金山登录取密钥失败：{passkey.get('msg') or passkey.get('result')}")

    csrf = ''.join(secrets.choice('ABCDEFGHJKMNPQRSTWXYZabcdefhijkmnprstwxyz2345678') for _ in range(32))
    cookie_jar.set_cookie(_make_cookie('csrf', csrf, 'account.wps.cn'))
    encrypted_password = _rsa_encrypt_password(KDOCS_LOGIN_PASSWORD, passkey['pass_key'])
    verify_headers = dict(headers)
    verify_headers.update({
        'X-CSRFToken': csrf,
        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    })
    result = _open_json(
        opener,
        'https://account.wps.cn/api/v3/account/safe_verify',
        data={
            'ssid': passkey.get('ssid', ''),
            'password': encrypted_password,
            'account': KDOCS_LOGIN_ACCOUNT,
            'from': 'v1/accountlogin',
            'source': 'web',
            'keeponline': '1',
        },
        headers=verify_headers,
        method='POST',
    )
    if result.get('result') != 'ok':
        msg = result.get('msg') or result.get('result') or '未知错误'
        if '验证码' in msg or result.get('result') in ('InvalidCaptcha', 'ErrNeedCaptcha'):
            msg += '；请在页面二维码扫码验证'
        raise ValueError(f'金山账号密码自动登录失败：{msg}')

    ssid = result.get('ssid') or passkey.get('ssid')
    if ssid:
        callback_url = 'https://account.wps.cn/v1/verifycallbackapp?' + urllib.parse.urlencode({
            'verifyresult': 'ok',
            'logintype': 'v1/accountlogin',
            'ssid': ssid,
        })
        opener.open(urllib.request.Request(callback_url, headers=headers), timeout=18).read()

    cookie = _cookie_header_from_jar(cookie_jar)
    if not cookie:
        raise ValueError('金山账号密码自动登录失败：未获得登录 Cookie')
    _save_kdocs_cookie(cookie)
    return cookie


def _get_kdocs_cookie():
    cookie = os.environ.get('KDOCS_COOKIE') or current_app.config.get('KDOCS_COOKIE')
    if cookie:
        return str(cookie).strip()
    try:
        with open(KDOCS_COOKIE_PATH, 'r', encoding='utf-8') as f:
            return f.read().strip()
    except FileNotFoundError:
        return _login_kdocs_with_password()


def _save_kdocs_cookie(cookie):
    cookie = (cookie or '').strip()
    if not cookie:
        raise ValueError('Cookie 不能为空')
    if '\n' in cookie or '\r' in cookie:
        raise ValueError('Cookie 请保持为一整行')
    os.makedirs(os.path.dirname(KDOCS_COOKIE_PATH), exist_ok=True)
    with open(KDOCS_COOKIE_PATH, 'w', encoding='utf-8') as f:
        f.write(cookie)
    try:
        os.chmod(KDOCS_COOKIE_PATH, 0o600)
    except Exception:
        pass


def _extract_today_submit_text(rows, target_date=None):
    target_date = target_date or datetime.now().date()
    if not rows:
        raise ValueError('没有可解析的数据')

    header_idx, submit_idx = _find_kdocs_header(rows)
    header = rows[header_idx] if header_idx < len(rows) else []
    target_idx = _find_kdocs_content_column(header)
    column_name = header[target_idx] if len(header) > target_idx and header[target_idx] else '货物清单列'
    values = []
    for row in rows[header_idx + 1:]:
        if len(row) <= max(submit_idx, target_idx):
            continue
        submitted_on = _parse_date_cell(row[submit_idx])
        if submitted_on == target_date:
            value = _clean_cell(row[target_idx])
            if value:
                values.append(value)

    return {
        'date': target_date.isoformat(),
        'column_name': column_name,
        'count': len(values),
        'values': values,
        'text': '\n'.join(values),
    }


def _find_kdocs_header(rows):
    header_idx = 0
    submit_idx = 1
    for idx, row in enumerate(rows[:20]):
        if len(row) > 1 and '提交时间' in row[1]:
            return idx, 1
        found = next((i for i, cell in enumerate(row) if '提交时间' in cell), None)
        if found is not None:
            return idx, found
    return header_idx, submit_idx


def _find_kdocs_content_column(header):
    for idx, cell in enumerate(header):
        text = _clean_cell(cell)
        if '箱唛' in text or ('规格' in text and '数量' in text):
            return idx
    for idx, cell in enumerate(header):
        text = _clean_cell(cell)
        if any(hint in text for hint in KDOCS_CONTENT_FIELD_HINTS):
            return idx
    return 5


def _today_kdocs_record_rows(rows, target_date=None, source=None):
    source = source or KDOCS_SOURCES[0]
    target_date = target_date or datetime.now().date()
    header_idx, submit_idx = _find_kdocs_header(rows)
    header = rows[header_idx] if header_idx < len(rows) else []
    target_idx = _find_kdocs_content_column(header)
    records = []
    for idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if len(row) <= max(submit_idx, target_idx):
            continue
        submitted_on = _parse_date_cell(row[submit_idx])
        if submitted_on != target_date:
            continue
        row_payload = {'header': header, 'row': row, 'row_number': idx}
        row_json = json.dumps(row_payload, ensure_ascii=False, sort_keys=True)
        records.append({
            'source_key': source.get('key') or KDOCS_SOURCE_KEY,
            'submit_date': target_date.isoformat(),
            'submit_time_text': _clean_cell(row[submit_idx]),
            'row_hash': hashlib.sha256(row_json.encode('utf-8')).hexdigest(),
            'row_json': json.dumps(row_payload, ensure_ascii=False),
            'content_text': _clean_cell(row[target_idx]),
        })
    return records


def _record_row_to_dict(row):
    data = dict(row)
    source = KDOCS_SOURCE_BY_KEY.get(data.get('source_key') or '')
    data['source_name'] = (source or {}).get('name') or data.get('source_key') or 'WPS'
    try:
        data['row_data'] = json.loads(data.get('row_json') or '{}')
    except Exception:
        data['row_data'] = {}
    try:
        data['parsed_items'] = json.loads(data.get('parsed_items_json') or '[]')
    except Exception:
        data['parsed_items'] = []
    data['is_printed'] = int(data.get('printed_count') or 0) > 0
    return data


def _list_today_wps_records(conn, target_date=None):
    target_date = target_date or datetime.now().date()
    source_keys = [source['key'] for source in KDOCS_SOURCES]
    placeholders = ','.join('?' for _ in source_keys)
    rows = conn.execute(
        f'''
        SELECT * FROM label_wps_records
        WHERE source_key IN ({placeholders}) AND submit_date=?
        ORDER BY submit_time_text, source_key, id
        ''',
        (*source_keys, target_date.isoformat()),
    ).fetchall()
    return [_record_row_to_dict(row) for row in rows]


def _upsert_today_wps_records(conn, records):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    inserted = 0
    updated = 0
    for record in records:
        existing = conn.execute(
            'SELECT id FROM label_wps_records WHERE source_key=? AND row_hash=?',
            (record['source_key'], record['row_hash']),
        ).fetchone()
        if existing:
            conn.execute(
                '''
                UPDATE label_wps_records
                SET submit_date=?, submit_time_text=?, row_json=?, content_text=?, updated_at=?
                WHERE id=?
                ''',
                (
                    record['submit_date'], record['submit_time_text'], record['row_json'],
                    record['content_text'], now, existing['id']
                ),
            )
            updated += 1
        else:
            conn.execute(
                '''
                INSERT INTO label_wps_records (
                    source_key, submit_date, submit_time_text, row_hash, row_json, content_text,
                    parsed_items_json, parse_status, created_at, updated_at
                ) VALUES (?,?,?,?,?,?,?,?,?,?)
                ''',
                (
                    record['source_key'], record['submit_date'], record['submit_time_text'],
                    record['row_hash'], record['row_json'], record['content_text'],
                    '[]', 'unparsed', now, now
                ),
            )
            inserted += 1
    return inserted, updated


CHINESE_NUMERAL_MAP = {
    '零': 0, '〇': 0, '一': 1, '二': 2, '两': 2, '俩': 2, '三': 3, '四': 4,
    '五': 5, '六': 6, '七': 7, '八': 8, '九': 9, '十': 10,
}


def _parse_chinese_int(text):
    text = _clean_cell(text)
    if not text:
        return None
    if text.isdigit():
        return int(text)
    if text in CHINESE_NUMERAL_MAP:
        return CHINESE_NUMERAL_MAP[text]
    if '十' in text:
        left, _, right = text.partition('十')
        tens = CHINESE_NUMERAL_MAP.get(left, 1) if left else 1
        ones = CHINESE_NUMERAL_MAP.get(right, 0) if right else 0
        return tens * 10 + ones
    return None


def _qty_from_match_text(text):
    value = _parse_chinese_int(text)
    return max(1, value) if value is not None else None


def _parse_qty_near_token(line, token, allow_line_fallback=False):
    escaped = re.escape(token)
    number_pattern = r'(\d+|[零〇一二两俩三四五六七八九十]{1,3})'
    patterns = [
        rf'{escaped}.*?[xX×*]\s*{number_pattern}',
        rf'{number_pattern}\s*(?:个|件|只|条|瓶|支|盒|套|包)?\s*{escaped}',
        rf'{escaped}\s*(?:[：:]\s*){number_pattern}\s*(?:个|件|只|条|瓶|支|盒|套|包)?',
        rf'{escaped}\s+{number_pattern}\s*(?:个|件|只|条|瓶|支|盒|套|包)?',
    ]
    for pattern in patterns:
        match = re.search(pattern, line, re.I)
        if match:
            qty = _qty_from_match_text(match.group(1))
            if qty is not None:
                return qty
    if allow_line_fallback:
        nums = re.findall(r'(?<![A-Za-z])\d+(?!\s*(?:ml|g|kg|克|毫升|升))', line, re.I)
        if len(nums) == 1 and nums[0].lower() == str(token).lower():
            return 1
        if nums:
            return max(1, int(nums[-1]))
    return 1


def _parse_qty_near_code(line, code):
    return _parse_qty_near_token(line, code, allow_line_fallback=True)


def _parse_qty_from_text(line):
    patterns = [
        r'[xX×*]\s*(\d+|[零〇一二两俩三四五六七八九十]{1,3})',
        r'(\d+|[零〇一二两俩三四五六七八九十]{1,3})\s*(?:个|件|只|条|瓶|支|盒|套|包)',
    ]
    for pattern in patterns:
        match = re.search(pattern, line, re.I)
        if match:
            qty = _qty_from_match_text(match.group(1))
            if qty is not None:
                return qty
    return 0


def _split_wps_content_segments(content_text):
    segments = []
    for line in re.split(r'[\r\n]+', content_text or ''):
        line = line.strip()
        if not line:
            continue
        for part in re.split(r'\t+| {2,}|\u3000{2,}', line):
            part = part.strip()
            if part in ('黑', '白'):
                continue
            if part and not re.fullmatch(r'[.\-—_·…。．、，,;；:：\s]+', part):
                segments.append(part)
    return segments


def _text_keywords(text):
    text = _clean_cell(text)
    if not text:
        return []
    keywords = [text]
    keywords.extend(re.findall(r'[\u4e00-\u9fff]{2,}', text))
    return keywords


def _product_keyword_candidates(product):
    candidates = []
    short_name = _clean_cell(product['short_name'])
    product_name = _clean_cell(product['product_name'])
    spec = _clean_cell(product['spec'])
    for text in (short_name, product_name):
        candidates.extend(_text_keywords(text))
    if short_name and spec:
        candidates.append(f'{short_name}{spec}')
    return sorted({item for item in candidates if len(item) >= 2}, key=len, reverse=True)


def _has_conflicting_spec(line, product):
    product_spec = _clean_cell(product['spec']).lower().replace(' ', '')
    if not product_spec:
        return False
    specs = [item.lower().replace(' ', '') for item in re.findall(r'\d+(?:\.\d+)?\s*(?:ml|g|kg|克|毫升|升)', line, re.I)]
    return bool(specs) and product_spec not in specs


def _match_product_by_keyword(line, products):
    line_lower = line.lower()
    for product in products:
        if _has_conflicting_spec(line, product):
            continue
        for keyword in _product_keyword_candidates(product):
            keyword_lower = keyword.lower()
            if keyword and keyword_lower in line_lower:
                return product, keyword
    return None, ''


def _parse_wps_content_items(conn, content_text):
    products = conn.execute(
        '''
        SELECT code,short_name,product_name,spec,box_spec
        FROM label_products
        ORDER BY LENGTH(short_name) DESC, LENGTH(product_name) DESC, LENGTH(code) DESC
        '''
    ).fetchall()
    items = []
    for line in _split_wps_content_segments(content_text):
        product = None
        matched_keyword = ''
        for prod in products:
            code = str(prod['code'])
            if re.search(rf'(?<![A-Za-z0-9]){re.escape(code)}(?![A-Za-z0-9])', line, re.I):
                product = prod
                matched_keyword = code
                break
        if not product:
            product, matched_keyword = _match_product_by_keyword(line, products)
        if product:
            qty = _parse_qty_near_code(line, product['code']) if matched_keyword == str(product['code']) else _parse_qty_near_token(line, matched_keyword)
            items.append({
                'kind': 'product',
                'code': product['code'],
                'label': product['short_name'] or product['code'],
                'pcs': qty,
                'raw': line,
                'status': 'matched' if matched_keyword == str(product['code']) else f'keyword:{matched_keyword}',
            })
        else:
            text_qty = _parse_qty_from_text(line)
            items.append({
                'kind': 'text',
                'code': '',
                'label': line,
                'pcs': text_qty,
                'raw': line,
                'status': 'text',
            })
    return items


# ─────────────────────────────── 通用渲染 ───────────────────────────────

def _render(active_tab, conn=None):
    close = conn is None
    if conn is None:
        conn = get_db_connection()
    _init_db(conn)
    products, weights, sizes, packing_weights, packing_sizes, pack_presets, pack_settings, history, pack_names, pack_stock = _load_all(conn)
    if close:
        conn.close()
    return render_template(
        'label_print.html', active_tab=active_tab,
        products=products, weights=weights, sizes=sizes,
        packing_weights=packing_weights, packing_sizes=packing_sizes,
        pack_presets=pack_presets, pack_settings=pack_settings,
        history=history, pack_names=pack_names, pack_stock=pack_stock
    )


# ─────────────────────────────── 页面路由 ───────────────────────────────

@label_print_bp.route('/', strict_slashes=False)
@module_required('label_print')
def index():
    return _render('print')


@label_print_bp.route('/products', strict_slashes=False)
@module_required('label_print')
def products_page():
    return _render('products')


@label_print_bp.route('/weights', strict_slashes=False)
@module_required('label_print')
def weights_page():
    return _render('weights')


@label_print_bp.route('/sizes', strict_slashes=False)
@module_required('label_print')
def sizes_page():
    return _render('sizes')


@label_print_bp.route('/packing_weights', strict_slashes=False)
@module_required('label_print')
def packing_weights_page():
    return _render('packing_weights')


@label_print_bp.route('/history', strict_slashes=False)
@module_required('label_print')
def history_page():
    return _render('history')


@label_print_bp.route('/pack_presets', strict_slashes=False)
@module_required('label_print')
def pack_presets_page():
    return _render('pack_presets')


@label_print_bp.route('/pack_settings', strict_slashes=False)
@module_required('label_print')
def pack_settings_page():
    return _render('pack_settings')


# ─────────────────────────────── 产品 CRUD ───────────────────────────────

@label_print_bp.route('/products/upsert', methods=['POST'])
@module_required('label_print')
def product_upsert():
    code  = request.form.get('code','').strip()
    sn    = request.form.get('short_name','').strip()
    pn    = request.form.get('product_name','').strip()
    sp    = request.form.get('spec','').strip()
    try:   bs = int(request.form.get('box_spec',0))
    except: bs = 0
    if not code or not sn:
        return redirect(url_for('label_print.products_page') + '?error=编号和简称不能为空')
    now  = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = get_db_connection()
    if conn.execute('SELECT id FROM label_products WHERE code=?',(code,)).fetchone():
        conn.execute('UPDATE label_products SET short_name=?,product_name=?,spec=?,box_spec=?,updated_at=? WHERE code=?',
                     (sn,pn,sp,bs,now,code)); msg='updated'
    else:
        conn.execute('INSERT INTO label_products (code,short_name,product_name,spec,box_spec,created_at,updated_at) VALUES (?,?,?,?,?,?,?)',
                     (code,sn,pn,sp,bs,now,now)); msg='added'
    conn.commit(); conn.close()
    return redirect(url_for('label_print.products_page') + f'?ok={msg}')


@label_print_bp.route('/products/<int:pid>/delete', methods=['POST'])
@module_required('label_print')
def product_delete(pid):
    conn = get_db_connection()
    conn.execute('DELETE FROM label_products WHERE id=?',(pid,))
    conn.commit(); conn.close()
    return redirect(url_for('label_print.products_page') + '?ok=1')


# ─────────────────────────────── 货物重量 CRUD ───────────────────────────────

@label_print_bp.route('/weights/upsert', methods=['POST'])
@module_required('label_print')
def weight_upsert():
    code = request.form.get('code','').strip()
    try:   wt = float(request.form.get('weight',0))
    except: wt = 0.0
    if not code:
        return redirect(url_for('label_print.weights_page') + '?error=编号不能为空')
    now  = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = get_db_connection()
    if conn.execute('SELECT id FROM label_weights WHERE code=?',(code,)).fetchone():
        conn.execute('UPDATE label_weights SET weight=?,updated_at=? WHERE code=?',(wt,now,code)); msg='updated'
    else:
        conn.execute('INSERT INTO label_weights (code,weight,updated_at) VALUES (?,?,?)',(code,wt,now)); msg='added'
    conn.commit(); conn.close()
    return redirect(url_for('label_print.weights_page') + f'?ok={msg}')


@label_print_bp.route('/weights/<int:wid>/delete', methods=['POST'])
@module_required('label_print')
def weight_delete(wid):
    conn = get_db_connection()
    conn.execute('DELETE FROM label_weights WHERE id=?',(wid,))
    conn.commit(); conn.close()
    return redirect(url_for('label_print.weights_page') + '?ok=1')


# ─────────────────────────────── 尺寸 CRUD ───────────────────────────────

@label_print_bp.route('/sizes/upsert', methods=['POST'])
@module_required('label_print')
def size_upsert():
    code = request.form.get('code','').strip()
    ir   = 1 if request.form.get('is_irregular') else 0
    def ti(k):
        try: return int(request.form.get(k,0))
        except: return 0
    if not code:
        return redirect(url_for('label_print.sizes_page') + '?error=编号不能为空')
    l,w,h = ti('length'),ti('width'),ti('height')
    now  = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = get_db_connection()
    if conn.execute('SELECT id FROM label_sizes WHERE code=?',(code,)).fetchone():
        conn.execute('UPDATE label_sizes SET length=?,width=?,height=?,is_irregular=?,updated_at=? WHERE code=?',
                     (l,w,h,ir,now,code)); msg='updated'
    else:
        conn.execute('INSERT INTO label_sizes (code,length,width,height,is_irregular,updated_at) VALUES (?,?,?,?,?,?)',
                     (code,l,w,h,ir,now)); msg='added'
    conn.commit(); conn.close()
    return redirect(url_for('label_print.sizes_page') + f'?ok={msg}')


@label_print_bp.route('/sizes/<int:sid>/delete', methods=['POST'])
@module_required('label_print')
def size_delete(sid):
    conn = get_db_connection()
    conn.execute('DELETE FROM label_sizes WHERE id=?',(sid,))
    conn.commit(); conn.close()
    return redirect(url_for('label_print.sizes_page') + '?ok=1')


# ─────────────────────────────── 包材重量 CRUD ───────────────────────────────

@label_print_bp.route('/packing_weights/upsert', methods=['POST'])
@module_required('label_print')
def packing_weight_upsert():
    pack_name = request.form.get('pack_name','').strip()
    try:   wt = float(request.form.get('weight',0))
    except: wt = 0.0
    if not pack_name:
        return redirect(url_for('label_print.packing_weights_page') + '?error=包材名称不能为空')
    now  = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = get_db_connection()
    if conn.execute('SELECT id FROM label_packing_weights WHERE pack_name=?',(pack_name,)).fetchone():
        conn.execute('UPDATE label_packing_weights SET weight=?,updated_at=? WHERE pack_name=?',
                     (wt,now,pack_name)); msg='updated'
    else:
        conn.execute('INSERT INTO label_packing_weights (pack_name,weight,updated_at) VALUES (?,?,?)',
                     (pack_name,wt,now)); msg='added'
    conn.commit(); conn.close()
    return redirect(url_for('label_print.packing_weights_page') + f'?ok={msg}')


@label_print_bp.route('/packing_weights/<int:pid>/delete', methods=['POST'])
@module_required('label_print')
def packing_weight_delete(pid):
    conn = get_db_connection()
    conn.execute('DELETE FROM label_packing_weights WHERE id=?',(pid,))
    conn.commit(); conn.close()
    return redirect(url_for('label_print.packing_weights_page') + '?ok=1')


# ─────────────────────────────── 包材尺寸 CRUD ───────────────────────────────

@label_print_bp.route('/packing_sizes', strict_slashes=False)
@module_required('label_print')
def packing_sizes_page():
    return _render('packing_sizes')


@label_print_bp.route('/packing_sizes/upsert', methods=['POST'])
@module_required('label_print')
def packing_size_upsert():
    pack_name = request.form.get('pack_name','').strip()
    size      = request.form.get('size','').strip()
    if not pack_name:
        return redirect(url_for('label_print.packing_sizes_page') + '?error=包材名称不能为空')
    now  = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = get_db_connection()
    if conn.execute('SELECT id FROM label_packing_sizes WHERE pack_name=?',(pack_name,)).fetchone():
        conn.execute('UPDATE label_packing_sizes SET size=?,updated_at=? WHERE pack_name=?',
                     (size,now,pack_name)); msg='updated'
    else:
        conn.execute('INSERT INTO label_packing_sizes (pack_name,size,updated_at) VALUES (?,?,?)',
                     (pack_name,size,now)); msg='added'
    conn.commit(); conn.close()
    return redirect(url_for('label_print.packing_sizes_page') + f'?ok={msg}')


@label_print_bp.route('/packing_sizes/<int:pid>/delete', methods=['POST'])
@module_required('label_print')
def packing_size_delete(pid):
    conn = get_db_connection()
    conn.execute('DELETE FROM label_packing_sizes WHERE id=?',(pid,))
    conn.commit(); conn.close()
    return redirect(url_for('label_print.packing_sizes_page') + '?ok=1')


# ─────────────────────────────── 打印历史 ───────────────────────────────

@label_print_bp.route('/history/<int:hid>/delete', methods=['POST'])
@module_required('label_print')
def history_delete(hid):
    conn = get_db_connection()
    conn.execute('DELETE FROM label_print_history WHERE id=?',(hid,))
    conn.commit(); conn.close()
    return redirect(url_for('label_print.history_page') + '?ok=1')


# ─────────────────────────────── 预设规则 CRUD ───────────────────────────────

@label_print_bp.route('/pack_presets/upsert', methods=['POST'])
@module_required('label_print')
def pack_preset_upsert():
    combo_key = request.form.get('combo_key', '').strip()
    box_name  = request.form.get('box_name', '').strip()
    bag_name  = request.form.get('bag_name', '').strip()
    try:
        bag_qty = int(request.form.get('bag_qty', 1) or 1)
    except Exception:
        bag_qty = 1
    if not bag_name:
        bag_qty = 0
    elif bag_qty < 1:
        bag_qty = 1
    note      = request.form.get('note', '').strip()
    if not combo_key:
        return redirect(url_for('label_print.pack_presets_page') + '?error=组合编号不能为空')
    now  = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = get_db_connection()
    ensure_tables(conn)
    if conn.execute('SELECT id FROM label_pack_presets WHERE combo_key=?', (combo_key,)).fetchone():
        conn.execute('UPDATE label_pack_presets SET box_name=?,bag_name=?,bag_qty=?,note=?,updated_at=? WHERE combo_key=?',
                     (box_name, bag_name, bag_qty, note, now, combo_key)); msg = 'updated'
    else:
        conn.execute('INSERT INTO label_pack_presets (combo_key,box_name,bag_name,bag_qty,note,updated_at) VALUES (?,?,?,?,?,?)',
                     (combo_key, box_name, bag_name, bag_qty, note, now)); msg = 'added'
    conn.commit(); conn.close()
    return redirect(url_for('label_print.pack_presets_page') + f'?ok={msg}')


@label_print_bp.route('/pack_presets/<int:pid>/delete', methods=['POST'])
@module_required('label_print')
def pack_preset_delete(pid):
    conn = get_db_connection()
    conn.execute('DELETE FROM label_pack_presets WHERE id=?', (pid,))
    conn.commit(); conn.close()
    return redirect(url_for('label_print.pack_presets_page') + '?ok=1')


# ─────────────────────────────── 参数设置 ───────────────────────────────

@label_print_bp.route('/pack_settings/update', methods=['POST'])
@module_required('label_print')
def pack_settings_update():
    now  = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = get_db_connection()
    ensure_tables(conn)
    for key in SETTING_KEYS:
        val = request.form.get(key, '').strip()
        if val:
            conn.execute('UPDATE label_pack_settings SET value=?,updated_at=? WHERE key=?', (val, now, key))
    conn.commit(); conn.close()
    return redirect(url_for('label_print.pack_settings_page') + '?ok=saved')


# ─────────────────────────────── API ───────────────────────────────

@label_print_bp.route('/api/products')
@module_required('label_print')
def api_products():
    conn = get_db_connection()
    ensure_tables(conn)
    rows = conn.execute(
        'SELECT code,short_name,spec,box_spec FROM label_products ORDER BY CAST(code AS INTEGER),code'
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@label_print_bp.route('/api/kdocs_today_text', methods=['POST'])
@module_required('label_print')
def api_kdocs_today_text():
    data = request.get_json(silent=True) or {}
    raw_text = (data.get('raw_text') or '').strip()
    try:
        rows = _rows_from_delimited_text(raw_text) if raw_text else _fetch_kdocs_rows()
        result = _extract_today_submit_text(rows)
        result['ok'] = True
        result['source'] = 'txt' if raw_text else 'kdocs'
        return jsonify(result)
    except ValueError as exc:
        message = str(exc)
        return jsonify({'ok': False, 'message': message, 'need_login': _message_needs_kdocs_login(message)}), 400
    except Exception as exc:
        current_app.logger.exception('[label_print] kdocs parse failed')
        return jsonify({'ok': False, 'message': f'解析失败：{exc}'}), 500


@label_print_bp.route('/api/wps_records/import_today', methods=['POST'])
@module_required('label_print')
def api_wps_records_import_today():
    try:
        records = []
        errors = []
        for source in KDOCS_SOURCES:
            try:
                rows = _fetch_kdocs_rows(source=source)
                records.extend(_today_kdocs_record_rows(rows, source=source))
            except ValueError as exc:
                errors.append({'source_key': source['key'], 'source_name': source['name'], 'message': str(exc)})
        if errors and not records:
            message = '；'.join(f"{item['source_name']}：{item['message']}" for item in errors)
            need_login = any(_message_needs_kdocs_login(item['message']) for item in errors)
            return jsonify({'ok': False, 'message': message, 'need_login': need_login}), 400
        conn = get_db_connection()
        ensure_tables(conn)
        inserted, updated = _upsert_today_wps_records(conn, records)
        conn.commit()
        result_records = _list_today_wps_records(conn)
        conn.close()
        return jsonify({
            'ok': True,
            'inserted': inserted,
            'updated': updated,
            'count': len(result_records),
            'records': result_records,
            'errors': errors,
        })
    except ValueError as exc:
        message = str(exc)
        return jsonify({'ok': False, 'message': message, 'need_login': _message_needs_kdocs_login(message)}), 400
    except Exception as exc:
        current_app.logger.exception('[label_print] import wps records failed')
        return jsonify({'ok': False, 'message': f'读取WPS今日记录失败：{exc}'}), 500


@label_print_bp.route('/api/wps_records/today')
@module_required('label_print')
def api_wps_records_today():
    conn = get_db_connection()
    ensure_tables(conn)
    records = _list_today_wps_records(conn)
    conn.close()
    return jsonify({'ok': True, 'records': records, 'count': len(records)})


@label_print_bp.route('/api/wps_records/<int:record_id>/parse', methods=['POST'])
@module_required('label_print')
def api_wps_record_parse(record_id):
    conn = get_db_connection()
    ensure_tables(conn)
    record = conn.execute('SELECT * FROM label_wps_records WHERE id=?', (record_id,)).fetchone()
    if not record:
        conn.close()
        return jsonify({'ok': False, 'message': 'WPS记录不存在'}), 404
    items = _parse_wps_content_items(conn, record['content_text'])
    status = 'parsed' if items else 'needs_review'
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute(
        'UPDATE label_wps_records SET parsed_items_json=?, parse_status=?, updated_at=? WHERE id=?',
        (json.dumps(items, ensure_ascii=False), status, now, record_id),
    )
    conn.commit()
    updated = conn.execute('SELECT * FROM label_wps_records WHERE id=?', (record_id,)).fetchone()
    conn.close()
    return jsonify({'ok': True, 'record': _record_row_to_dict(updated), 'items': items})


@label_print_bp.route('/api/wps_records/<int:record_id>/save_parse', methods=['POST'])
@module_required('label_print')
def api_wps_record_save_parse(record_id):
    data = request.get_json(silent=True) or {}
    items = data.get('items') or []
    if not isinstance(items, list):
        return jsonify({'ok': False, 'message': '解析结果格式不正确'}), 400
    conn = get_db_connection()
    ensure_tables(conn)
    record = conn.execute('SELECT * FROM label_wps_records WHERE id=?', (record_id,)).fetchone()
    if not record:
        conn.close()
        return jsonify({'ok': False, 'message': 'WPS记录不存在'}), 404
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute(
        'UPDATE label_wps_records SET parsed_items_json=?, parse_status=?, updated_at=? WHERE id=?',
        (json.dumps(items, ensure_ascii=False), 'parsed', now, record_id),
    )
    conn.commit()
    updated = conn.execute('SELECT * FROM label_wps_records WHERE id=?', (record_id,)).fetchone()
    conn.close()
    return jsonify({'ok': True, 'record': _record_row_to_dict(updated)})


@label_print_bp.route('/api/kdocs_qr/start', methods=['POST'])
@module_required('label_print')
def api_kdocs_qr_start():
    try:
        _cleanup_kdocs_qr_sessions()
        opener, cookie_jar = _make_kdocs_opener()
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
            'Accept': 'application/json, text/javascript, */*; q=0.01',
            'Referer': 'https://account.wps.cn/wpspersonallogin',
        }
        code_verifier, code_challenge = _generate_kdocs_pkce()
        private_key, public_key = _make_kdocs_ec_key()
        _open_text(opener, headers['Referer'], headers=headers, timeout=18)
        login = _open_jsonp(
            opener,
            f'{KDOCS_QR_BASE}/api/v3/login_qrcode?' + urllib.parse.urlencode({
                '_jsonp': 'quickGetQrcodeJsonpCallback',
                'code_challenge': code_challenge,
            }),
        )
        loginid = login.get('loginid')
        if not loginid:
            raise ValueError(login.get('msg') or login.get('result') or '生成二维码登录 ID 失败')

        qr_url = f'{KDOCS_QR_BASE}/api/v3/login_qrcode/url?' + urllib.parse.urlencode({
            'loginid': loginid,
        })
        qr_info = _open_json(opener, qr_url, headers=headers)
        image_url = qr_info.get('url')
        if qr_info.get('result') not in (None, 'ok') or not image_url:
            raise ValueError(qr_info.get('msg') or qr_info.get('result') or '生成二维码失败')

        KDOCS_QR_SESSIONS[loginid] = {
            'opener': opener,
            'cookie_jar': cookie_jar,
            'code_verifier': code_verifier,
            'private_key': private_key,
            'public_key': public_key,
            'state': 'scan',
            'expires_at': datetime.now().timestamp() + 300,
        }
        return jsonify({'ok': True, 'loginid': loginid, 'qr_url': image_url, 'state': 'scan', 'expires_in': 300})
    except Exception as exc:
        current_app.logger.exception('[label_print] start kdocs qr login failed')
        return jsonify({'ok': False, 'message': f'生成金山登录二维码失败：{exc}'}), 500


@label_print_bp.route('/api/kdocs_qr/poll', methods=['POST'])
@module_required('label_print')
def api_kdocs_qr_poll():
    data = request.get_json(silent=True) or {}
    loginid = (data.get('loginid') or '').strip()
    state = (data.get('state') or 'scan').strip() or 'scan'
    try:
        _cleanup_kdocs_qr_sessions()
        item = KDOCS_QR_SESSIONS.get(loginid)
        if not item:
            raise ValueError('二维码已过期，请重新点击解析生成')

        poll_url = f'{KDOCS_QR_BASE}/api/v3/login_qrcode/login?' + urllib.parse.urlencode({
            'loginid': loginid,
            'state': state,
            '_jsonp': 'callback',
        })
        result = _open_jsonp(item['opener'], poll_url, timeout=35)
        result_code = result.get('result')
        qr_state = result.get('state') or result_code

        if result_code not in (None, 'ok') and qr_state not in ('pending', 'scan', 'logined'):
            KDOCS_QR_SESSIONS.pop(loginid, None)
            raise ValueError(result.get('msg') or result_code or '二维码登录失败')

        if qr_state == 'pending':
            return jsonify({'ok': True, 'state': 'pending', 'next_state': state, 'message': '等待手机扫码'})
        if qr_state == 'scan':
            item['state'] = 'confirm'
            return jsonify({'ok': True, 'state': 'scan', 'next_state': 'confirm', 'message': '已扫码，请在手机上确认登录'})
        if qr_state == 'logined':
            authcode = result.get('authcode')
            kso_authcode = result.get('kso_authcode')
            if kso_authcode:
                _grant_kdocs_token(
                    item['opener'],
                    kso_authcode,
                    item.get('code_verifier') or '',
                    item.get('private_key'),
                    item.get('public_key') or '',
                )
            elif authcode:
                _exchange_kdocs_authcode(item['opener'], authcode)
            else:
                raise ValueError('扫码已确认，但金山未返回 authcode')
            cookie = _cookie_header_from_jar(item['cookie_jar'])
            if not cookie:
                raise ValueError('扫码登录成功但未获得 Cookie')
            _save_kdocs_cookie(cookie)
            KDOCS_QR_SESSIONS.pop(loginid, None)
            return jsonify({'ok': True, 'state': 'logined', 'saved': True, 'message': '登录验证完成，正在继续解析'})

        return jsonify({'ok': True, 'state': qr_state or 'pending', 'next_state': state, 'message': '等待验证'})
    except ValueError as exc:
        return jsonify({'ok': False, 'message': str(exc)}), 400
    except Exception as exc:
        current_app.logger.exception('[label_print] poll kdocs qr login failed')
        return jsonify({'ok': False, 'message': f'扫码验证失败：{exc}'}), 500


@label_print_bp.route('/api/kdocs_cookie', methods=['POST'])
@module_required('label_print')
def api_kdocs_cookie():
    data = request.get_json(silent=True) or {}
    try:
        _save_kdocs_cookie(data.get('cookie') or '')
        return jsonify({'ok': True})
    except ValueError as exc:
        return jsonify({'ok': False, 'message': str(exc)}), 400
    except Exception as exc:
        current_app.logger.exception('[label_print] save kdocs cookie failed')
        return jsonify({'ok': False, 'message': f'保存失败：{exc}'}), 500


@label_print_bp.route('/api/save_print', methods=['POST'])
@module_required('label_print')
def api_save_print():
    data = request.get_json(silent=True) or {}
    total_tickets = int(data.get('total_tickets', 0))
    total_qty     = int(data.get('total_qty', 0))
    items         = data.get('items', [])
    parsed_items  = data.get('parsed_items') or items
    wps_record_id = data.get('wps_record_id')
    is_forced     = 1 if data.get('is_forced_reprint') else 0
    printed_at    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    current_user  = getattr(g, 'current_user', None) or {}
    user_id       = current_user.get('id')
    username      = current_user.get('username') or ''
    conn = get_db_connection()
    ensure_tables(conn)
    conn.execute(
        '''
        INSERT INTO label_print_history (
            total_tickets,total_qty,items_json,printed_at,
            wps_record_id,printed_by_user_id,printed_by_username,is_forced_reprint,parsed_items_json
        ) VALUES (?,?,?,?,?,?,?,?,?)
        ''',
        (
            total_tickets, total_qty, json.dumps(items, ensure_ascii=False), printed_at,
            wps_record_id, user_id, username, is_forced, json.dumps(parsed_items, ensure_ascii=False)
        )
    )
    if wps_record_id:
        conn.execute(
            '''
            UPDATE label_wps_records
            SET printed_count=printed_count+1,
                last_printed_at=?,
                last_printed_by_user_id=?,
                last_printed_by_username=?,
                parsed_items_json=?,
                parse_status='parsed',
                updated_at=?
            WHERE id=?
            ''',
            (
                printed_at, user_id, username, json.dumps(parsed_items, ensure_ascii=False),
                printed_at, wps_record_id
            ),
        )
    conn.commit(); conn.close()
    return jsonify({'ok': True})
