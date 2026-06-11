from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any
import re
import sqlite3
import zipfile

import pandas as pd
from flask import current_app
from openpyxl.utils import get_column_letter

from common.excel_utils import check_columns_match, is_excel_filename, normalize_columns, normalize_header_text

from .table_schemas import (
    AFTER_SALES_COLUMN_MAPPING,
    AFTER_SALES_COLUMN_TYPES,
    AFTER_SALES_DEDUP_KEY_COLUMNS,
    AFTER_SALES_REQUIRED_COLUMNS,
    DATA_STATUS_CONFIG,
    DATA_STATUS_TABLE_NAME,
    EXPORT_TABLE_CONFIG,
    FUND_FLOW_COLUMN_MAPPING,
    FUND_FLOW_COLUMN_TYPES,
    FUND_FLOW_DEDUP_KEY_COLUMNS,
    FUND_FLOW_REQUIRED_COLUMNS,
    KUAISHOU_AFTER_SALES_TABLE_NAME,
    KUAISHOU_FUND_FLOW_TABLE_NAME,
    KUAISHOU_ORDER_TABLE_NAME,
    ORDER_COLUMN_MAPPING,
    ORDER_COLUMN_TYPES,
    ORDER_DEDUP_KEY_COLUMNS,
    ORDER_REQUIRED_COLUMNS,
)


TEXT_SOURCE_COLUMNS = {
    '订单号',
    '赠品订单号',
    '活动订单编号',
    '商品ID',
    'SKU编码',
    'CPS达人ID',
    '团长ID',
    '团长id',
    '快赚客ID',
    '授权推广者ID',
    '收货人电话',
    '快递单号',
    '服务门店ID',
    '商家ID',
    '达人ID',
    'MCN机构ID',
    '服务商ID',
    '结算商户号',
    '售后单号',
    '订单编号',
    '买家ID',
    '发货快递单号(如为西北订单，仅展示一段物流信息)',
    '退货物流单号',
    '商家退货联系人电话',
    '换货物流单号',
}

TABLE_IMPORT_CONFIG = {
    'orders': {
        'title': '订单',
        'source_table': KUAISHOU_ORDER_TABLE_NAME,
        'column_mapping': ORDER_COLUMN_MAPPING,
        'required_columns': ORDER_REQUIRED_COLUMNS,
        'column_types': ORDER_COLUMN_TYPES,
        'dedup_key_columns': ORDER_DEDUP_KEY_COLUMNS,
    },
    'fund_flows': {
        'title': '资金流水',
        'source_table': KUAISHOU_FUND_FLOW_TABLE_NAME,
        'column_mapping': FUND_FLOW_COLUMN_MAPPING,
        'required_columns': FUND_FLOW_REQUIRED_COLUMNS,
        'column_types': FUND_FLOW_COLUMN_TYPES,
        'dedup_key_columns': FUND_FLOW_DEDUP_KEY_COLUMNS,
    },
    'aftersales': {
        'title': '售后',
        'source_table': KUAISHOU_AFTER_SALES_TABLE_NAME,
        'column_mapping': AFTER_SALES_COLUMN_MAPPING,
        'required_columns': AFTER_SALES_REQUIRED_COLUMNS,
        'column_types': AFTER_SALES_COLUMN_TYPES,
        'dedup_key_columns': AFTER_SALES_DEDUP_KEY_COLUMNS,
    },
}


def _get_database_path() -> Path:
    db_path = current_app.config.get('DATABASE_PATH')
    if db_path:
        return Path(db_path)
    return Path(current_app.root_path) / 'data' / 'main.db'


def _get_upload_source_filename(file_obj: Any) -> str:
    return str(getattr(file_obj, 'filename', '') or '').strip()


def _read_upload_source_bytes(file_obj: Any) -> bytes:
    path = getattr(file_obj, 'path', None)
    if path:
        return Path(path).read_bytes()
    return file_obj.read()


def _reset_upload_source(file_obj: Any) -> None:
    stream = getattr(file_obj, 'stream', None)
    if stream is not None:
        stream.seek(0)


def _safe_password_from_filename(filename: str) -> str:
    stem = Path(filename).stem
    if len(stem) < 6:
        raise ValueError(f'订单表文件名不足 6 位，无法提取打开密码：{filename}')
    return stem[-6:]


def _excel_buffer_from_upload(file_bytes: bytes, filename: str) -> BytesIO:
    """读取 Excel；如文件加密，则使用文件名最后 6 位作为密码解密。"""
    raw_buffer = BytesIO(file_bytes)
    try:
        import msoffcrypto
        from msoffcrypto.exceptions import DecryptionError
    except ImportError:
        raw_buffer.seek(0)
        return raw_buffer

    try:
        office_file = msoffcrypto.OfficeFile(raw_buffer)
        if not office_file.is_encrypted():
            raw_buffer.seek(0)
            return raw_buffer

        password = _safe_password_from_filename(filename)
        decrypted = BytesIO()
        office_file.load_key(password=password)
        office_file.decrypt(decrypted)
        decrypted.seek(0)
        return decrypted
    except DecryptionError as exc:
        raise ValueError(f'文件已加密但密码校验失败，请确认文件名最后 6 位是否为打开密码：{filename}') from exc
    except Exception:
        raw_buffer.seek(0)
        return raw_buffer


def _ensure_data_status_seeded(conn: sqlite3.Connection) -> None:
    conn.execute(
        f'''
        CREATE TABLE IF NOT EXISTS {DATA_STATUS_TABLE_NAME} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            table_key TEXT NOT NULL UNIQUE,
            table_name TEXT NOT NULL,
            record_count INTEGER DEFAULT 0,
            min_date TEXT,
            max_date TEXT,
            last_import_time TEXT
        )
        '''
    )
    for table_key, cfg in DATA_STATUS_CONFIG.items():
        conn.execute(
            f'''
            INSERT OR IGNORE INTO {DATA_STATUS_TABLE_NAME}
            (table_key, table_name, record_count)
            VALUES (?, ?, 0)
            ''',
            (table_key, cfg['table_name']),
        )
    conn.commit()


def _build_create_table_sql(table_name: str, column_types: dict[str, str]) -> str:
    columns_sql = ',\n    '.join(f'{name} {col_type}' for name, col_type in column_types.items())
    return f'''CREATE TABLE IF NOT EXISTS {table_name} (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    {columns_sql}
);'''


def _get_existing_table_columns(conn: sqlite3.Connection, table_name: str) -> list[str]:
    rows = conn.execute(f'PRAGMA table_info({table_name})').fetchall()
    return [row[1] for row in rows]


def _ensure_table(conn: sqlite3.Connection, table_name: str, column_types: dict[str, str]) -> tuple[bool, str]:
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,),
    ).fetchone()
    if row is None:
        conn.execute(_build_create_table_sql(table_name, column_types))
        conn.commit()
        return True, f'数据表不存在，已自动创建：{table_name}'

    existing_columns = set(_get_existing_table_columns(conn, table_name))
    added_columns: list[str] = []
    for column_name, column_type in column_types.items():
        if column_name not in existing_columns:
            conn.execute(f'ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type}')
            added_columns.append(column_name)
    if added_columns:
        conn.commit()
        return False, f'数据表已存在：{table_name}，并已补齐字段：{", ".join(added_columns)}'
    return False, f'数据表已存在：{table_name}'


def ensure_all_tables() -> None:
    db_path = _get_database_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db_path) as conn:
        for cfg in TABLE_IMPORT_CONFIG.values():
            _ensure_table(conn, cfg['source_table'], cfg['column_types'])
        _ensure_data_status_seeded(conn)


def _update_data_status(table_key: str) -> str:
    cfg = DATA_STATUS_CONFIG.get(table_key)
    if not cfg:
        return ''

    db_path = _get_database_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db_path) as conn:
        _ensure_data_status_seeded(conn)
        conn.execute(
            f'''
            UPDATE {DATA_STATUS_TABLE_NAME}
            SET
                table_name = ?,
                record_count = (SELECT COUNT(*) FROM {cfg['source_table']}),
                min_date = (SELECT MIN({cfg['date_field']}) FROM {cfg['source_table']}),
                max_date = (SELECT MAX({cfg['date_field']}) FROM {cfg['source_table']}),
                last_import_time = datetime('now', 'localtime')
            WHERE table_key = ?
            ''',
            (cfg['table_name'], table_key),
        )
        conn.commit()
    return ''


def get_data_status_rows() -> list[dict[str, Any]]:
    ensure_all_tables()
    with sqlite3.connect(_get_database_path()) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            f'''
            SELECT table_key, table_name, record_count, min_date, max_date, last_import_time
            FROM {DATA_STATUS_TABLE_NAME}
            ORDER BY CASE table_key
                WHEN 'orders' THEN 1
                WHEN 'fund_flows' THEN 2
                WHEN 'aftersales' THEN 3
                ELSE 99
            END, id
            '''
        ).fetchall()
    return [dict(row) for row in rows]


def _clean_text_value(value: Any) -> Any:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return format(value, '.0f')
        return format(value, 'f').rstrip('0').rstrip('.')
    return str(value).strip()


def _clean_numeric_value(value: Any) -> Any:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(',', '').replace('¥', '').replace('￥', '').replace('%', '')
    text = re.sub(r'\s+', '', text)
    try:
        return float(text)
    except ValueError:
        return None


def _build_text_dtype_mapping(excel_buffer: BytesIO) -> dict[str, str]:
    header_df = pd.read_excel(excel_buffer, nrows=0)
    excel_buffer.seek(0)
    normalized_text_columns = {normalize_header_text(col) for col in TEXT_SOURCE_COLUMNS}
    return {
        raw_col: 'string'
        for raw_col in header_df.columns.tolist()
        if normalize_header_text(raw_col) in normalized_text_columns
    }


def _prepare_df_for_db(df: pd.DataFrame, column_types: dict[str, str]) -> pd.DataFrame:
    prepared_df = df.copy()
    db_columns = list(column_types.keys())
    for column_name in db_columns:
        if column_name not in prepared_df.columns:
            prepared_df[column_name] = None
    prepared_df = prepared_df[db_columns]

    for column_name, column_type in column_types.items():
        if column_type == 'REAL':
            prepared_df[column_name] = prepared_df[column_name].apply(_clean_numeric_value)
        elif column_type == 'INTEGER':
            prepared_df[column_name] = (
                prepared_df[column_name]
                .apply(_clean_numeric_value)
                .apply(lambda value: int(value) if value is not None else None)
            )
        else:
            prepared_df[column_name] = prepared_df[column_name].apply(_clean_text_value)
    return prepared_df


def _build_dedup_key(row: pd.Series, key_columns: list[str]) -> str | None:
    parts: list[str] = []
    has_non_empty_value = False
    for column_name in key_columns:
        value = _clean_text_value(row.get(column_name))
        value_text = '' if value is None else str(value).strip()
        if value_text:
            has_non_empty_value = True
            parts.append(value_text)
        else:
            parts.append('<EMPTY>')
    if not has_non_empty_value:
        return None
    return '||'.join(parts)


def _get_existing_keys(conn: sqlite3.Connection, table_name: str, key_columns: list[str]) -> set[str]:
    columns_sql = ', '.join(key_columns)
    rows = conn.execute(f'SELECT {columns_sql} FROM {table_name}').fetchall()
    keys: set[str] = set()
    for row in rows:
        row_dict = {column_name: row[index] for index, column_name in enumerate(key_columns)}
        key = _build_dedup_key(pd.Series(row_dict), key_columns)
        if key:
            keys.add(key)
    return keys


def _deduplicate_df(
    merged_df: pd.DataFrame,
    conn: sqlite3.Connection,
    table_name: str,
    key_columns: list[str],
) -> tuple[pd.DataFrame, int, int]:
    missing_key_columns = [col for col in key_columns if col not in merged_df.columns]
    if missing_key_columns:
        return merged_df, 0, 0

    working_df = merged_df.copy()
    working_df['_dedup_key'] = working_df.apply(lambda row: _build_dedup_key(row, key_columns), axis=1)

    valid_key_mask = working_df['_dedup_key'].notna() & (working_df['_dedup_key'].astype(str).str.strip() != '')
    valid_df = working_df[valid_key_mask].copy()
    unique_valid_df = valid_df.drop_duplicates(subset=['_dedup_key'], keep='first')
    batch_duplicate_count = int(len(valid_df) - len(unique_valid_df))

    no_key_df = working_df[~valid_key_mask].copy()
    combined_df = pd.concat([unique_valid_df, no_key_df], ignore_index=True)

    existing_keys = _get_existing_keys(conn, table_name, key_columns)
    if existing_keys:
        valid_after_batch_mask = combined_df['_dedup_key'].notna() & (
            combined_df['_dedup_key'].astype(str).str.strip() != ''
        )
        valid_after_batch_df = combined_df[valid_after_batch_mask].copy()
        no_key_after_batch_df = combined_df[~valid_after_batch_mask].copy()
        new_valid_df = valid_after_batch_df[~valid_after_batch_df['_dedup_key'].isin(existing_keys)].copy()
        db_duplicate_count = int(len(valid_after_batch_df) - len(new_valid_df))
        combined_df = pd.concat([new_valid_df, no_key_after_batch_df], ignore_index=True)
    else:
        db_duplicate_count = 0

    return combined_df.drop(columns=['_dedup_key'], errors='ignore'), batch_duplicate_count, db_duplicate_count


def _build_precheck_failed_response(
    title: str,
    file_summaries: list[dict[str, Any]],
    invalid_files: list[str],
    failed_files: list[dict[str, str]],
) -> dict[str, Any]:
    message_parts = [f'本次{title}导入预检未通过，未写入数据库']
    for summary in file_summaries:
        message_parts.extend([
            '',
            f"文件：{summary['filename']}",
            f"行数：{summary['row_count']}",
            f"列数：{summary['column_count']}",
            f"列名：{'，'.join(summary['columns'])}",
        ])
    if failed_files:
        message_parts.extend(['', '预检失败：'])
        message_parts.extend(f"- {item['filename']}（{item['error']}）" for item in failed_files)
    if invalid_files:
        message_parts.extend(['', f"无效文件：{'，'.join(invalid_files)}"])
    return {
        'success': False,
        'message': '\n'.join(message_parts),
        'file_count': len(file_summaries),
        'files': file_summaries,
        'invalid_files': invalid_files,
        'failed_files': failed_files,
        'precheck_failed': True,
        'written_rows': 0,
    }


def _write_table_to_db(table_key: str, dataframes: list[pd.DataFrame]) -> tuple[int, str, str]:
    cfg = TABLE_IMPORT_CONFIG[table_key]
    if not dataframes:
        return 0, f"没有可写入的{cfg['title']}数据", ''

    db_path = _get_database_path()
    db_path.parent.mkdir(parents=True, exist_ok=True)
    prepared_frames = [_prepare_df_for_db(df, cfg['column_types']) for df in dataframes]
    merged_df = pd.concat(prepared_frames, ignore_index=True)

    with sqlite3.connect(db_path) as conn:
        _ensure_table(conn, cfg['source_table'], cfg['column_types'])
        deduped_df, batch_duplicate_count, db_duplicate_count = _deduplicate_df(
            merged_df,
            conn,
            cfg['source_table'],
            cfg['dedup_key_columns'],
        )
        if deduped_df.empty:
            parts = [f"没有可写入的新{cfg['title']}数据"]
            if batch_duplicate_count:
                parts.append(f'本次文件内重复已跳过：{batch_duplicate_count} 条')
            if db_duplicate_count:
                parts.append(f'数据库中已存在记录已跳过：{db_duplicate_count} 条')
            return 0, '；'.join(parts), ''

        deduped_df.to_sql(cfg['source_table'], conn, if_exists='append', index=False)

    parts = [f"成功写入 {len(deduped_df)} 行{cfg['title']}数据"]
    if batch_duplicate_count:
        parts.append(f'本次文件内重复已跳过：{batch_duplicate_count} 条')
    if db_duplicate_count:
        parts.append(f'数据库中已存在记录已跳过：{db_duplicate_count} 条')
    status_warning = _update_data_status(table_key)
    return int(len(deduped_df)), '；'.join(parts), status_warning


def import_excel_files(files: list[Any], table_key: str) -> dict[str, Any]:
    cfg = TABLE_IMPORT_CONFIG[table_key]
    valid_files: list[Any] = []
    invalid_files: list[str] = []
    failed_files: list[dict[str, str]] = []
    file_summaries: list[dict[str, Any]] = []
    prepared_dataframes: list[pd.DataFrame] = []
    base_columns: list[str] | None = None
    base_filename: str | None = None
    has_structure_mismatch = False

    for file_obj in files:
        filename = _get_upload_source_filename(file_obj)
        if not filename:
            invalid_files.append('未命名文件')
            continue
        if not is_excel_filename(filename):
            invalid_files.append(filename)
            continue
        valid_files.append(file_obj)

    if not valid_files:
        return {
            'success': False,
            'message': '没有有效的 Excel 文件（.xlsx/.xls）',
            'file_count': 0,
            'files': [],
            'invalid_files': invalid_files,
            'failed_files': failed_files,
        }

    for file_obj in valid_files:
        filename = _get_upload_source_filename(file_obj)
        try:
            file_bytes = _read_upload_source_bytes(file_obj)
            excel_buffer = _excel_buffer_from_upload(file_bytes, filename)
            dtype_mapping = _build_text_dtype_mapping(excel_buffer)
            df = pd.read_excel(excel_buffer, dtype=dtype_mapping if dtype_mapping else None)
            df.columns = [normalize_header_text(col) for col in df.columns.tolist()]
            current_columns = normalize_columns(df.columns.tolist())

            normalized_required_columns = [normalize_header_text(col) for col in cfg['required_columns']]
            normalized_column_mapping = {
                normalize_header_text(chinese_name): english_name
                for chinese_name, english_name in cfg['column_mapping'].items()
            }
            missing_required = [col for col in normalized_required_columns if col not in current_columns]
            if missing_required:
                failed_files.append({
                    'filename': filename,
                    'error': f"缺少必需字段：{', '.join(missing_required)}",
                })
                continue

            df = df.rename(columns=normalized_column_mapping)
            summary = {
                'filename': filename,
                'row_count': int(len(df)),
                'column_count': int(len(df.columns)),
                'columns': normalize_columns(df.columns.tolist()),
            }

            if base_columns is None:
                base_columns = current_columns
                base_filename = filename
                file_summaries.append(summary)
                prepared_dataframes.append(df)
                continue

            columns_match, missing_columns, extra_columns = check_columns_match(base_columns, current_columns)
            if not columns_match:
                details: list[str] = []
                if missing_columns:
                    details.append(f"缺少列：{', '.join(missing_columns)}")
                if extra_columns:
                    details.append(f"多出列：{', '.join(extra_columns)}")
                if not details:
                    details.append('列名顺序不一致')
                failed_files.append({
                    'filename': filename,
                    'error': f"列结构不一致（基准文件：{base_filename}；{'；'.join(details)}）",
                })
                has_structure_mismatch = True
                file_summaries.append(summary)
                continue

            file_summaries.append(summary)
            prepared_dataframes.append(df)
        except Exception as exc:
            failed_files.append({'filename': filename, 'error': str(exc)})
        finally:
            _reset_upload_source(file_obj)

    if has_structure_mismatch or failed_files:
        return _build_precheck_failed_response(cfg['title'], file_summaries, invalid_files, failed_files)

    if not file_summaries:
        return {
            'success': False,
            'message': f"{cfg['title']}文件已接收，但读取失败",
            'file_count': 0,
            'files': [],
            'invalid_files': invalid_files,
            'failed_files': failed_files,
        }

    table_created, table_message = False, ''
    try:
        with sqlite3.connect(_get_database_path()) as conn:
            table_created, table_message = _ensure_table(conn, cfg['source_table'], cfg['column_types'])
        written_rows, write_message, status_warning = _write_table_to_db(table_key, prepared_dataframes)
    except Exception as exc:
        return {
            'success': False,
            'message': f'写入数据库失败：{exc}\n数据库路径：{_get_database_path()}',
            'file_count': len(file_summaries),
            'files': file_summaries,
            'invalid_files': invalid_files,
            'failed_files': failed_files,
        }

    message = f"成功读取 {len(file_summaries)} 个{cfg['title']}文件"
    for part in (table_message, write_message, status_warning):
        if part:
            message += f'\n{part}'

    return {
        'success': True,
        'message': message,
        'file_count': len(file_summaries),
        'files': file_summaries,
        'invalid_files': invalid_files,
        'failed_files': failed_files,
        'table_created': table_created,
        'table_message': table_message,
        'written_rows': written_rows,
        'write_message': write_message,
    }


def import_orders_files(files: list[Any]) -> dict[str, Any]:
    return import_excel_files(files, 'orders')


def import_fund_flow_files(files: list[Any]) -> dict[str, Any]:
    return import_excel_files(files, 'fund_flows')


def import_after_sales_files(files: list[Any]) -> dict[str, Any]:
    return import_excel_files(files, 'aftersales')


def _normalize_export_datetime_text(value: str | None, boundary: str = 'start') -> str | None:
    text = str(value or '').strip()
    if not text:
        return None
    normalized = text.replace('T', ' ').replace('/', '-')
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
        try:
            dt = datetime.strptime(normalized, fmt)
            if fmt == '%Y-%m-%d':
                return dt.strftime('%Y-%m-%d 23:59:59' if boundary == 'end' else '%Y-%m-%d 00:00:00')
            if fmt == '%Y-%m-%d %H:%M':
                return dt.strftime('%Y-%m-%d %H:%M:00')
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except ValueError:
            continue
    return normalized


def _is_numeric_column(column_name: str, column_types: dict[str, str]) -> bool:
    return str(column_types.get(column_name, '')).upper() in {'REAL', 'INTEGER', 'NUMERIC', 'FLOAT', 'DECIMAL'}


def _is_datetime_column(column_name: str) -> bool:
    lower_name = column_name.lower()
    return lower_name.endswith('_at') or lower_name.endswith('_time') or 'time' in lower_name


def _normalize_filter_operator(value: Any) -> str:
    return str(value or '').strip().lower()


def _normalize_filter_logic(value: Any) -> str:
    return 'OR' if str(value or '').strip().lower() == 'or' else 'AND'


def _build_filter_sql_parts(
    filter_conditions: list[dict[str, Any]],
    allowed_fields: set[str],
    column_types: dict[str, str],
) -> tuple[list[str], list[Any]]:
    sql_parts: list[str] = []
    params: list[Any] = []
    for raw_condition in filter_conditions:
        if not isinstance(raw_condition, dict):
            continue
        field_name = str(raw_condition.get('field') or '').strip()
        operator = _normalize_filter_operator(raw_condition.get('operator'))
        logic = _normalize_filter_logic(raw_condition.get('logic'))
        value_text = str(raw_condition.get('value') or '').strip()
        if not field_name:
            continue
        if field_name not in allowed_fields:
            raise ValueError(f'存在非法筛选字段：{field_name}')
        if operator not in {
            'eq', 'ne', 'contains', 'not_contains',
            'gt', 'gte', 'lt', 'lte',
            'is_empty', 'is_not_empty',
        }:
            raise ValueError(f'存在非法筛选条件：{operator}')

        clause = ''
        clause_params: list[Any] = []
        expr = f"REPLACE(CAST({field_name} AS TEXT), '/', '-')" if _is_datetime_column(field_name) else field_name
        is_numeric = _is_numeric_column(field_name, column_types)

        if operator in {'is_empty', 'is_not_empty'}:
            clause = f"({field_name} IS {'NOT ' if operator == 'is_not_empty' else ''}NULL"
            clause += f" {'AND' if operator == 'is_not_empty' else 'OR'} TRIM(CAST({field_name} AS TEXT)) {'<>' if operator == 'is_not_empty' else '='} '')"
        elif value_text == '':
            continue
        elif operator in {'contains', 'not_contains'}:
            clause = f"CAST({field_name} AS TEXT) {'NOT ' if operator == 'not_contains' else ''}LIKE ?"
            clause_params = [f'%{value_text}%']
        elif operator in {'eq', 'ne'}:
            sql_op = '=' if operator == 'eq' else '<>'
            if is_numeric:
                number_value = _clean_numeric_value(value_text)
                clause = f'CAST({field_name} AS REAL) {sql_op} ?'
                clause_params = [number_value]
            elif _is_datetime_column(field_name):
                start_value = _normalize_export_datetime_text(value_text, 'start')
                end_value = _normalize_export_datetime_text(value_text, 'end')
                if operator == 'eq':
                    clause = f'({expr} >= ? AND {expr} <= ?)'
                else:
                    clause = f'({expr} < ? OR {expr} > ?)'
                clause_params = [start_value, end_value]
            else:
                clause = f'CAST({field_name} AS TEXT) {sql_op} ?'
                clause_params = [value_text]
        else:
            sql_op = {'gt': '>', 'gte': '>=', 'lt': '<', 'lte': '<='}[operator]
            if is_numeric:
                clause = f'CAST({field_name} AS REAL) {sql_op} ?'
                clause_params = [_clean_numeric_value(value_text)]
            else:
                boundary = 'end' if operator in {'gt', 'lte'} else 'start'
                compare_value = _normalize_export_datetime_text(value_text, boundary) if _is_datetime_column(field_name) else value_text
                clause = f'{expr if _is_datetime_column(field_name) else f"CAST({field_name} AS TEXT)"} {sql_op} ?'
                clause_params = [compare_value]

        if clause:
            if sql_parts:
                sql_parts.append(logic)
            sql_parts.append(f'({clause})')
            params.extend(clause_params)
    return sql_parts, params


def _build_download_name(table_key: str, start_time: str | None, end_time: str | None) -> str:
    label = EXPORT_TABLE_CONFIG.get(table_key, {}).get('label', '导出数据')

    def safe_part(value: str | None) -> str:
        if not value:
            return '全部时间'
        return re.sub(r'[\\/:*?"<>|\s]+', '_', value)

    return f'{label}_{safe_part(start_time)}_到_{safe_part(end_time)}.xlsx'


def _auto_adjust_excel_columns(worksheet) -> None:
    for index, column_cells in enumerate(worksheet.iter_cols(), start=1):
        max_width = 0
        for cell in column_cells:
            value = '' if cell.value is None else str(cell.value)
            width = sum(2 if ord(ch) > 127 else 1 for ch in value)
            max_width = max(max_width, width)
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(max_width + 2, 10), 40)


def export_data_to_excel(
    table_key: str,
    start_time: str | None,
    end_time: str | None,
    selected_fields: list[str],
    filter_conditions: list[dict[str, Any]] | None = None,
) -> tuple[BytesIO, str]:
    ensure_all_tables()
    cfg = EXPORT_TABLE_CONFIG.get(table_key)
    if not cfg:
        raise ValueError('请选择正确的数据表')
    if not selected_fields:
        raise ValueError('请至少选择一个导出字段')

    allowed_fields = set(cfg['column_types'].keys())
    invalid_fields = [field for field in selected_fields if field not in allowed_fields]
    if invalid_fields:
        raise ValueError(f"存在非法导出字段：{', '.join(invalid_fields)}")

    where_parts: list[str] = []
    params: list[Any] = []
    date_field = cfg['date_field']
    start_value = _normalize_export_datetime_text(start_time, 'start')
    end_value = _normalize_export_datetime_text(end_time, 'end')
    date_expr = f"REPLACE(CAST({date_field} AS TEXT), '/', '-')"
    if start_value:
        where_parts.append(f'{date_expr} >= ?')
        params.append(start_value)
    if end_value:
        where_parts.append(f'{date_expr} <= ?')
        params.append(end_value)

    filter_sql_parts, filter_params = _build_filter_sql_parts(
        filter_conditions or [],
        allowed_fields,
        cfg['column_types'],
    )
    if filter_sql_parts:
        where_parts.append(' '.join(filter_sql_parts))
        params.extend(filter_params)

    field_sql = ', '.join(selected_fields)
    sql = f'SELECT {field_sql} FROM {cfg["source_table"]}'
    if where_parts:
        sql += ' WHERE ' + ' AND '.join(f'({part})' for part in where_parts)
    sql += f' ORDER BY {date_expr} DESC, id DESC'

    with sqlite3.connect(_get_database_path()) as conn:
        df = pd.read_sql_query(sql, conn, params=params)

    header_mapping = {english: chinese for chinese, english in cfg['column_mapping'].items()}
    df = df.rename(columns={field: header_mapping.get(field, field) for field in selected_fields})

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet_name = cfg['label'][:31]
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        _auto_adjust_excel_columns(writer.book[sheet_name])
    output.seek(0)
    return output, _build_download_name(table_key, start_time, end_time)


# ===================== 佣金导出 =====================

KUAISHOU_COMMISSION_AMOUNT_COLUMNS = {
    '达人佣金',
    '团长佣金',
    '其他收费',
    '佣金合计',
    '应开金额',
    '达人佣金(元)',
    '团长佣金(元)',
    '实际结算金额(元)',
}

KUAISHOU_COMMISSION_TEXT_COLUMNS = {
    '订单号',
    '商品ID',
    '达人ID',
    '达人昵称',
    '团长ID',
    '团长id',
    '团长昵称',
    '达人/客户',
    '未匹配类型',
    '未匹配原因',
}


def _table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    row = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,),
    ).fetchone()
    return row is not None


def _normalize_commission_date(value: str | None) -> str:
    text = str(value or '').strip().replace('/', '-')
    if not text:
        raise ValueError('请选择佣金导出的开始日期和结束日期')
    try:
        dt = datetime.strptime(text[:10], '%Y-%m-%d')
    except ValueError as exc:
        raise ValueError('佣金导出日期格式不正确，请使用 YYYY-MM-DD') from exc
    return dt.strftime('%Y-%m-%d')


def _normalize_commission_date_range(start_date: str | None, end_date: str | None) -> tuple[str, str]:
    start_text = _normalize_commission_date(start_date)
    end_text = _normalize_commission_date(end_date)
    if start_text > end_text:
        raise ValueError('佣金导出开始日期不能晚于结束日期')
    return start_text, end_text


def _month_values_between(start_text: str, end_text: str) -> list[tuple[int, int]]:
    start_dt = datetime.strptime(start_text, '%Y-%m-%d')
    end_dt = datetime.strptime(end_text, '%Y-%m-%d')
    values: list[tuple[int, int]] = []
    year, month = start_dt.year, start_dt.month
    while (year, month) <= (end_dt.year, end_dt.month):
        values.append((year, month))
        if month == 12:
            year += 1
            month = 1
        else:
            month += 1
    return values


def _build_commission_month_text(start_text: str, end_text: str, short_year: bool = False) -> str:
    months = _month_values_between(start_text, end_text)
    if not months:
        return ''
    has_multiple_years = len({year for year, _month in months}) > 1
    previous_year: int | None = None
    parts: list[str] = []
    for year, month in months:
        year_text = str(year)[-2:] if short_year else str(year)
        if has_multiple_years or previous_year != year:
            parts.append(f'{year_text}年{month}月')
        else:
            parts.append(f'{month}月')
        previous_year = year
    return ' '.join(parts)


def _safe_download_part(value: Any, fallback: str = '未命名') -> str:
    text = str(value or '').strip()
    text = re.sub(r'[\\/:*?"<>|\r\n\t]+', '_', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text or fallback


def _build_commission_zip_name(prefix: str, start_text: str, end_text: str) -> str:
    month_text = _safe_download_part(_build_commission_month_text(start_text, end_text))
    return f'{prefix}_{month_text}.zip'


def _commission_date_expr(alias: str = '') -> str:
    prefix = f'{alias}.' if alias else ''
    return f"SUBSTR(REPLACE(CAST({prefix}actual_settlement_time AS TEXT), '/', '-'), 1, 10)"


def _ensure_commission_tables(conn: sqlite3.Connection) -> None:
    if not _table_exists(conn, KUAISHOU_FUND_FLOW_TABLE_NAME):
        raise ValueError(f'数据表不存在：{KUAISHOU_FUND_FLOW_TABLE_NAME}，请先导入资金流水表')
    if not _table_exists(conn, KUAISHOU_ORDER_TABLE_NAME):
        raise ValueError(f'数据表不存在：{KUAISHOU_ORDER_TABLE_NAME}，请先导入订单表')


def _normalize_lookup_id(value: Any) -> str:
    return str(value or '').strip().lstrip("'")


def _name_matches_keyword(name: str, id_text: str, keyword: str) -> bool:
    if not keyword:
        return True
    return keyword in name or keyword in id_text


def _load_last_name_map(
    conn: sqlite3.Connection,
    id_column: str,
    name_column: str,
) -> dict[str, str]:
    """按订单表导入顺序取同一 ID 最后一次出现的昵称。"""
    rows = conn.execute(
        f"""
        SELECT {id_column} AS id_text, {name_column} AS name_text
        FROM {KUAISHOU_ORDER_TABLE_NAME}
        WHERE {id_column} IS NOT NULL
          AND TRIM(CAST({id_column} AS TEXT)) <> ''
        ORDER BY id ASC
        """
    ).fetchall()
    result: dict[str, str] = {}
    for row in rows:
        id_text = _normalize_lookup_id(row['id_text'])
        name_text = str(row['name_text'] or '').strip()
        if id_text and name_text and name_text != '-':
            result[id_text] = name_text
    return result


def _load_creator_name_map(conn: sqlite3.Connection) -> dict[str, str]:
    return _load_last_name_map(conn, 'cps_creator_id', 'cps_creator_nickname')


def _load_leader_name_map(conn: sqlite3.Connection) -> dict[str, str]:
    return _load_last_name_map(conn, 'leader_id', 'leader_nickname')


def _read_commission_fund_rows(conn: sqlite3.Connection, start_text: str, end_text: str) -> pd.DataFrame:
    columns_sql = ', '.join(FUND_FLOW_COLUMN_TYPES.keys())
    sql = f"""
        SELECT {columns_sql}
        FROM {KUAISHOU_FUND_FLOW_TABLE_NAME}
        WHERE {_commission_date_expr()} >= ?
          AND {_commission_date_expr()} <= ?
          AND (
              ABS(CAST(COALESCE(creator_commission, 0) AS REAL)) > 0.0001
              OR ABS(CAST(COALESCE(leader_commission, 0) AS REAL)) > 0.0001
              OR ABS(CAST(COALESCE(other_fee, 0) AS REAL)) > 0.0001
          )
        ORDER BY actual_settlement_time ASC, order_no ASC, product_id ASC, id ASC
    """
    df = pd.read_sql_query(sql, conn, params=[start_text, end_text])
    for col in ['creator_commission', 'leader_commission', 'other_fee', 'actual_settlement_amount']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0.0)
    for col in ['creator_id', 'leader_id', 'order_no', 'product_id']:
        if col in df.columns:
            df[col] = df[col].fillna('').map(_normalize_lookup_id)
    return df


def _build_creator_summary_and_unmatched(
    df: pd.DataFrame,
    creator_name_map: dict[str, str],
    keyword: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    rows: list[dict[str, Any]] = []
    unmatched: list[dict[str, Any]] = []
    if df.empty:
        return (
            pd.DataFrame(columns=['达人ID', '达人昵称', '达人佣金', '其他收费', '佣金合计']),
            pd.DataFrame(columns=_unmatched_columns()),
        )

    creator_df = df[
        (df['creator_commission'].abs() > 0.0001)
        | (df['other_fee'].abs() > 0.0001)
    ].copy()
    grouped: dict[str, dict[str, Any]] = {}
    for row in creator_df.itertuples(index=False):
        creator_id = _normalize_lookup_id(getattr(row, 'creator_id', ''))
        creator_commission = float(getattr(row, 'creator_commission', 0) or 0)
        other_fee = float(getattr(row, 'other_fee', 0) or 0)
        if not creator_id:
            unmatched.append(_unmatched_row(row, '达人', '缺少达人ID', creator_commission, 0.0, other_fee))
            continue
        creator_name = creator_name_map.get(creator_id, '')
        if not creator_name:
            unmatched.append(_unmatched_row(row, '达人', '订单表未找到达人昵称', creator_commission, 0.0, other_fee))
            continue
        if not _name_matches_keyword(creator_name, creator_id, keyword):
            continue
        bucket = grouped.setdefault(
            creator_id,
            {
                '达人ID': creator_id,
                '达人昵称': creator_name,
                '达人佣金': 0.0,
                '其他收费': 0.0,
            },
        )
        # 名称映射本身已按订单表最后一次出现取值；这里保持 ID 汇总为一行。
        bucket['达人昵称'] = creator_name
        bucket['达人佣金'] += creator_commission
        bucket['其他收费'] += other_fee

    for bucket in grouped.values():
        bucket['佣金合计'] = float(bucket['达人佣金']) + float(bucket['其他收费'])
        rows.append(bucket)

    summary = pd.DataFrame(rows, columns=['达人ID', '达人昵称', '达人佣金', '其他收费', '佣金合计'])
    if not summary.empty:
        summary = summary.sort_values('佣金合计', ascending=False)
    return summary, pd.DataFrame(unmatched, columns=_unmatched_columns())


def _build_leader_summary_and_unmatched(
    df: pd.DataFrame,
    leader_name_map: dict[str, str],
    keyword: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    rows: list[dict[str, Any]] = []
    unmatched: list[dict[str, Any]] = []
    if df.empty:
        return (
            pd.DataFrame(columns=['团长ID', '团长昵称', '团长佣金']),
            pd.DataFrame(columns=_unmatched_columns()),
        )

    leader_df = df[df['leader_commission'].abs() > 0.0001].copy()
    grouped: dict[str, dict[str, Any]] = {}
    for row in leader_df.itertuples(index=False):
        leader_id = _normalize_lookup_id(getattr(row, 'leader_id', ''))
        leader_commission = float(getattr(row, 'leader_commission', 0) or 0)
        if not leader_id:
            unmatched.append(_unmatched_row(row, '团长', '缺少团长ID', 0.0, leader_commission, 0.0))
            continue
        leader_name = leader_name_map.get(leader_id, '')
        if not leader_name:
            unmatched.append(_unmatched_row(row, '团长', '订单表未找到团长昵称', 0.0, leader_commission, 0.0))
            continue
        if not _name_matches_keyword(leader_name, leader_id, keyword):
            continue
        bucket = grouped.setdefault(
            leader_id,
            {'团长ID': leader_id, '团长昵称': leader_name, '团长佣金': 0.0},
        )
        bucket['团长昵称'] = leader_name
        bucket['团长佣金'] += leader_commission

    rows.extend(grouped.values())
    summary = pd.DataFrame(rows, columns=['团长ID', '团长昵称', '团长佣金'])
    if not summary.empty:
        summary = summary.sort_values('团长佣金', ascending=False)
    return summary, pd.DataFrame(unmatched, columns=_unmatched_columns())


def _unmatched_columns() -> list[str]:
    return [
        '未匹配类型',
        '未匹配原因',
        '实际结算时间',
        '订单号',
        '商品ID',
        '商品名称',
        '达人ID',
        '团长ID',
        '达人佣金',
        '团长佣金',
        '其他收费',
        '其他收费明细',
    ]


def _unmatched_row(
    row: Any,
    kind: str,
    reason: str,
    creator_commission: float,
    leader_commission: float,
    other_fee: float,
) -> dict[str, Any]:
    return {
        '未匹配类型': kind,
        '未匹配原因': reason,
        '实际结算时间': getattr(row, 'actual_settlement_time', ''),
        '订单号': getattr(row, 'order_no', ''),
        '商品ID': getattr(row, 'product_id', ''),
        '商品名称': getattr(row, 'product_name', ''),
        '达人ID': getattr(row, 'creator_id', ''),
        '团长ID': getattr(row, 'leader_id', ''),
        '达人佣金': creator_commission,
        '团长佣金': leader_commission,
        '其他收费': other_fee,
        '其他收费明细': getattr(row, 'other_fee_detail', ''),
    }


def _build_invoice_import_df(creator_df: pd.DataFrame, leader_df: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for _idx, row in creator_df.iterrows():
        rows.append({'达人/客户': row['达人昵称'], '应开金额': float(row['佣金合计'] or 0)})
    for _idx, row in leader_df.iterrows():
        rows.append({'达人/客户': row['团长昵称'], '应开金额': float(row['团长佣金'] or 0)})
    df = pd.DataFrame(rows, columns=['达人/客户', '应开金额'])
    if not df.empty:
        df = df.sort_values('应开金额', ascending=False)
    return df


def _write_dataframe_excel(
    sheets: list[tuple[str, pd.DataFrame]],
    amount_columns: set[str] | None = None,
    text_columns: set[str] | None = None,
) -> BytesIO:
    amount_columns = amount_columns or set()
    text_columns = text_columns or set()
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in sheets:
            safe_sheet = re.sub(r'[\[\]:*?/\\]', '_', str(sheet_name))[:31] or 'Sheet1'
            df.to_excel(writer, index=False, sheet_name=safe_sheet)
            worksheet = writer.sheets[safe_sheet]
            for column_index, column_name in enumerate(df.columns, start=1):
                column_letter = get_column_letter(column_index)
                if column_name in amount_columns:
                    for cell in worksheet[column_letter][1:]:
                        cell.number_format = '0.00'
                if column_name in text_columns:
                    for cell in worksheet[column_letter]:
                        cell.number_format = '@'
            _auto_adjust_excel_columns(worksheet)
    output.seek(0)
    return output


def _fund_flow_chinese_columns(df: pd.DataFrame) -> pd.DataFrame:
    reverse_map = {v: k for k, v in FUND_FLOW_COLUMN_MAPPING.items()}
    return df.rename(columns={col: reverse_map.get(col, col) for col in df.columns})


def _unique_zip_name(used_names: dict[str, int], filename: str) -> str:
    used_count = used_names.get(filename, 0)
    used_names[filename] = used_count + 1
    if not used_count:
        return filename
    stem = filename[:-5] if filename.lower().endswith('.xlsx') else filename
    return f'{stem}_{used_count + 1}.xlsx'


def _prepare_creator_detail_rows(
    df: pd.DataFrame,
    creator_name_map: dict[str, str],
    keyword: str,
) -> pd.DataFrame:
    rows = df[
        (df['creator_commission'].abs() > 0.0001)
        | (df['other_fee'].abs() > 0.0001)
    ].copy()
    if rows.empty:
        return rows
    rows['creator_name'] = rows['creator_id'].map(lambda value: creator_name_map.get(_normalize_lookup_id(value), ''))
    rows = rows[
        rows.apply(
            lambda row: bool(row['creator_name'])
            and _name_matches_keyword(str(row['creator_name']), _normalize_lookup_id(row['creator_id']), keyword),
            axis=1,
        )
    ].copy()
    return rows.sort_values(['creator_name', 'actual_settlement_time', 'order_no', 'product_id'])


def _prepare_leader_detail_rows(
    df: pd.DataFrame,
    leader_name_map: dict[str, str],
    keyword: str,
) -> pd.DataFrame:
    rows = df[df['leader_commission'].abs() > 0.0001].copy()
    if rows.empty:
        return rows
    rows['leader_name'] = rows['leader_id'].map(lambda value: leader_name_map.get(_normalize_lookup_id(value), ''))
    rows = rows[
        rows.apply(
            lambda row: bool(row['leader_name'])
            and _name_matches_keyword(str(row['leader_name']), _normalize_lookup_id(row['leader_id']), keyword),
            axis=1,
        )
    ].copy()
    return rows.sort_values(['leader_name', 'actual_settlement_time', 'order_no', 'product_id'])


def _load_commission_context(
    start_date: str | None,
    end_date: str | None,
    nickname_query: str | None,
) -> tuple[str, str, str, pd.DataFrame, dict[str, str], dict[str, str]]:
    start_text, end_text = _normalize_commission_date_range(start_date, end_date)
    keyword = str(nickname_query or '').strip()
    db_path = _get_database_path()
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_commission_tables(conn)
        fund_df = _read_commission_fund_rows(conn, start_text, end_text)
        creator_name_map = _load_creator_name_map(conn)
        leader_name_map = _load_leader_name_map(conn)
    return start_text, end_text, keyword, fund_df, creator_name_map, leader_name_map


def export_commission_summary_zip(
    start_date: str | None,
    end_date: str | None,
    nickname_query: str | None = None,
) -> tuple[BytesIO, str]:
    start_text, end_text, keyword, fund_df, creator_name_map, leader_name_map = _load_commission_context(
        start_date,
        end_date,
        nickname_query,
    )
    creator_df, creator_unmatched = _build_creator_summary_and_unmatched(fund_df, creator_name_map, keyword)
    leader_df, leader_unmatched = _build_leader_summary_and_unmatched(fund_df, leader_name_map, keyword)
    unmatched_df = pd.concat([creator_unmatched, leader_unmatched], ignore_index=True)

    month_text = _build_commission_month_text(start_text, end_text)
    safe_month = _safe_download_part(month_text)

    summary_sheets: list[tuple[str, pd.DataFrame]] = [
        ('达人汇总', creator_df),
        ('团长汇总', leader_df),
    ]
    if not unmatched_df.empty:
        summary_sheets.append(('未匹配', unmatched_df))

    summary_excel = _write_dataframe_excel(
        summary_sheets,
        amount_columns=KUAISHOU_COMMISSION_AMOUNT_COLUMNS,
        text_columns=KUAISHOU_COMMISSION_TEXT_COLUMNS,
    )
    invoice_excel = _write_dataframe_excel(
        [('应开金额导入', _build_invoice_import_df(creator_df, leader_df))],
        amount_columns={'应开金额'},
        text_columns={'达人/客户'},
    )

    archive = BytesIO()
    with zipfile.ZipFile(archive, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f'{safe_month}佣金汇总.xlsx', summary_excel.getvalue())
        zf.writestr(f'{safe_month}应开金额导入.xlsx', invoice_excel.getvalue())
        if not unmatched_df.empty:
            unmatched_excel = _write_dataframe_excel(
                [('未匹配', unmatched_df)],
                amount_columns=KUAISHOU_COMMISSION_AMOUNT_COLUMNS,
                text_columns=KUAISHOU_COMMISSION_TEXT_COLUMNS,
            )
            zf.writestr(f'{safe_month}未匹配佣金.xlsx', unmatched_excel.getvalue())
    archive.seek(0)
    return archive, _build_commission_zip_name('快手澳柯佣金汇总', start_text, end_text)


def export_commission_detail_zip(
    start_date: str | None,
    end_date: str | None,
    nickname_query: str | None = None,
) -> tuple[BytesIO, str]:
    start_text, end_text, keyword, fund_df, creator_name_map, leader_name_map = _load_commission_context(
        start_date,
        end_date,
        nickname_query,
    )
    creator_rows = _prepare_creator_detail_rows(fund_df, creator_name_map, keyword)
    leader_rows = _prepare_leader_detail_rows(fund_df, leader_name_map, keyword)
    creator_df, creator_unmatched = _build_creator_summary_and_unmatched(fund_df, creator_name_map, keyword)
    leader_df, leader_unmatched = _build_leader_summary_and_unmatched(fund_df, leader_name_map, keyword)
    unmatched_df = pd.concat([creator_unmatched, leader_unmatched], ignore_index=True)
    amount_by_creator_id = {
        str(row['达人ID']): float(row['佣金合计'] or 0)
        for _idx, row in creator_df.iterrows()
    }
    amount_by_leader_id = {
        str(row['团长ID']): float(row['团长佣金'] or 0)
        for _idx, row in leader_df.iterrows()
    }

    month_text = _build_commission_month_text(start_text, end_text, short_year=True)
    archive = BytesIO()
    used_names: dict[str, int] = {}
    with zipfile.ZipFile(archive, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        if not creator_rows.empty:
            for creator_id, group_df in creator_rows.groupby('creator_id', sort=True):
                creator_id_text = _normalize_lookup_id(creator_id)
                creator_name = creator_name_map.get(creator_id_text, creator_id_text)
                amount_sum = amount_by_creator_id.get(creator_id_text, 0.0)
                detail_df = group_df.drop(columns=['creator_name'], errors='ignore').copy()
                detail_df = _fund_flow_chinese_columns(detail_df)
                detail_excel = _write_dataframe_excel(
                    [('明细', detail_df)],
                    amount_columns=KUAISHOU_COMMISSION_AMOUNT_COLUMNS,
                    text_columns={'订单号', '商品ID', '达人ID', '团长id'},
                )
                filename = _safe_download_part(f'{creator_name}_{amount_sum:.2f}_快手{month_text}') + '.xlsx'
                zf.writestr(_unique_zip_name(used_names, filename), detail_excel.getvalue())

        if not leader_rows.empty:
            for leader_id, group_df in leader_rows.groupby('leader_id', sort=True):
                leader_id_text = _normalize_lookup_id(leader_id)
                leader_name = leader_name_map.get(leader_id_text, leader_id_text)
                amount_sum = amount_by_leader_id.get(leader_id_text, 0.0)
                detail_df = group_df.drop(columns=['leader_name'], errors='ignore').copy()
                detail_df = _fund_flow_chinese_columns(detail_df)
                detail_excel = _write_dataframe_excel(
                    [('明细', detail_df)],
                    amount_columns=KUAISHOU_COMMISSION_AMOUNT_COLUMNS,
                    text_columns={'订单号', '商品ID', '达人ID', '团长id'},
                )
                filename = _safe_download_part(f'{leader_name}_{amount_sum:.2f}_团长快手{month_text}') + '.xlsx'
                zf.writestr(_unique_zip_name(used_names, filename), detail_excel.getvalue())

        if not unmatched_df.empty:
            unmatched_excel = _write_dataframe_excel(
                [('未匹配', unmatched_df)],
                amount_columns=KUAISHOU_COMMISSION_AMOUNT_COLUMNS,
                text_columns=KUAISHOU_COMMISSION_TEXT_COLUMNS,
            )
            zf.writestr(f'{_safe_download_part(month_text)}未匹配佣金.xlsx', unmatched_excel.getvalue())

        if not used_names and unmatched_df.empty:
            empty_excel = _write_dataframe_excel([('明细', pd.DataFrame(columns=['提示']))])
            zf.writestr('无佣金明细.xlsx', empty_excel.getvalue())

    archive.seek(0)
    return archive, _build_commission_zip_name('快手澳柯佣金明细', start_text, end_text)
